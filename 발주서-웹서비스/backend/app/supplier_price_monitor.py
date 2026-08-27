from __future__ import annotations

import csv
import logging
import re
import sys
from dataclasses import dataclass, replace
from datetime import datetime, timedelta, timezone
from html import escape, unescape
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import httpx
from openpyxl import load_workbook

from app.config import PBF_PARTNER_ID, PBF_PARTNER_PASSWORD, TEMPLATE_DIR, UPLOAD_DIR

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))
GOOGLE_SHEET_JBT_FRUIT_CSV = (
    "https://docs.google.com/spreadsheets/d/1g0Bxmz773DqPjCfYgBNyBtlYKxYuLJYY2zasqiC5u7Y/export"
    "?format=csv&gid=1683225896"
)
GOOGLE_SHEET_JBT_FARM_CSV = (
    "https://docs.google.com/spreadsheets/d/1g0Bxmz773DqPjCfYgBNyBtlYKxYuLJYY2zasqiC5u7Y/gviz/tq"
    "?tqx=out:csv&sheet=%EB%86%8D%EC%82%B0%EB%AC%BC"
)
GOOGLE_SHEET_JEJUDAPAM_VEGETABLE_CSV = (
    "https://docs.google.com/spreadsheets/d/1hATLzLK5fjum_ehXG-uSrJ3owllSVKeYV-U5Aox5pHY/export"
    "?format=csv&gid=980271705"
)
GOOGLE_SHEET_JEWELRYFRUIT_KG_CSV = (
    "https://docs.google.com/spreadsheets/d/1Gb8feRt2FCCG0fpmhPKXcW1l061OefDD5UIejOhjcxY/gviz/tq"
    "?tqx=out:csv&gid=1831570299"
)
GOOGLE_SHEET_JEWELRYFRUIT_PEACH_CSV = (
    "https://docs.google.com/spreadsheets/d/1Gb8feRt2FCCG0fpmhPKXcW1l061OefDD5UIejOhjcxY/gviz/tq"
    "?tqx=out:csv&gid=1567728493"
)

PRICE_SIGNAL_STYLE = {
    "blue": {
        "font": "2563EB",
        "fill": "DBEAFE",
        "label": "전일 대비 공급가 상승",
    },
    "red": {
        "font": "DC2626",
        "fill": "FEE2E2",
        "label": "전일 대비 공급가 하락",
    },
    "same": {
        "font": "0F172A",
        "fill": "F8FAFC",
        "label": "전일과 동일",
    },
}


@dataclass(frozen=True)
class SupplierSource:
    """옵션 일부를 다른 발주처에서 수집할 때 쓰는 보조 소스.

    발주처가 옵션별로 갈리는 상품(미니밤호박 1kg=제주다팜·3~10kg=쥬얼리,
    백도 딱딱이복숭아 1kg=쥬얼리·2·4kg=제주다팜)을 **한 파일**로 뽑기 위해
    모니터 하나가 여러 발주처를 조회한다. None인 필드는 모니터 기본값을 그대로 쓴다.
    """

    supplier_name: str
    source_type: Literal["adminplus", "google_sheet", "manual"] = "adminplus"
    base_url: str | None = None
    search_value: str | None = None
    product_code: str | None = None
    product_name: str | None = None
    sheet_csv_url: str | None = None
    sheet_product_name: str | None = None
    sheet_product_exclude: tuple[str, ...] | None = None
    sheet_product_column: str | None = None
    sheet_option_column: str | None = None
    sheet_quantity_column: str | None = None
    sheet_vip_column: str | None = None
    sheet_price_fallback_column: str | None = None
    sheet_previous_column: str | None = None

    _OVERRIDABLE = (
        "base_url", "search_value", "product_code", "product_name",
        "sheet_csv_url", "sheet_product_name", "sheet_product_exclude",
        "sheet_product_column", "sheet_option_column", "sheet_quantity_column",
        "sheet_vip_column", "sheet_price_fallback_column", "sheet_previous_column",
    )

    def as_config(self, base: "SupplierMonitorConfig") -> "SupplierMonitorConfig":
        """수집 함수들이 그대로 쓸 수 있도록 모니터 설정에 이 소스를 덮어쓴 사본을 만든다."""
        overrides: dict[str, object] = {
            "supplier_name": self.supplier_name,
            "source_type": self.source_type,
        }
        for field_name in self._OVERRIDABLE:
            value = getattr(self, field_name)
            if value is not None:
                overrides[field_name] = value
        return replace(base, **overrides)


@dataclass(frozen=True)
class SupplierOptionConfig:
    label: str
    supplier_option_name: str
    row: int
    sheet_name: str | None = None
    cell: str = "I"
    # 이 옵션만 다른 발주처에서 수집(None이면 모니터 기본 발주처)
    source: SupplierSource | None = None
    # 기본 발주처를 쓰는 옵션의 발주처 이름(발주처 혼합 모니터에서 행별 표기·검증용)
    supplier_name: str = ""
    # 쿠팡 상품명·옵션 원문. 채워두면 발주 라우팅(order_routing)과 발주처가 일치하는지
    # 테스트가 검증한다 → 발주처가 바뀌면 마진방어 설정도 같이 고치게 강제된다.
    coupang_product: str = ""
    coupang_option: str = ""


@dataclass(frozen=True)
class SupplierMonitorConfig:
    key: str
    supplier_name: str
    product_name: str
    template_path: Path
    output_prefix: str
    options: tuple[SupplierOptionConfig, ...]
    source_type: Literal["adminplus", "google_sheet", "manual"] = "adminplus"
    output_suffix: str = ""
    output_name_pattern: str | None = None
    base_url: str | None = None
    search_value: str | None = None
    product_code: str | None = None  # 설정 시 이름검색 대신 이 상품코드를 직접 사용(동명이상품 오매칭 방지)
    sheet_csv_url: str | None = None
    sheet_product_name: str | None = None
    sheet_product_exclude: tuple[str, ...] = ()  # 상품명에 이 키워드가 있으면 제외(부분일치 충돌 방지, 예: '애플')
    sheet_product_column: str = "D"
    sheet_option_column: str = "E"
    sheet_quantity_column: str | None = None
    sheet_vip_column: str = "H"
    sheet_price_fallback_column: str | None = None
    sheet_previous_column: str | None = None
    skip_missing_options: bool = False
    active_until: str | None = None  # 'YYYY-MM-DD' 이후 자동 만료(한시 모니터)

    @property
    def login_url(self) -> str:
        if not self.base_url:
            raise SupplierMonitorError(f"{self.key} monitor is not configured for adminplus login.")
        return f"{self.base_url}/partner/login.chk.php"

    @property
    def product_list_url(self) -> str:
        if not self.base_url:
            raise SupplierMonitorError(f"{self.key} monitor is not configured for adminplus product list.")
        return f"{self.base_url}/partner/"

    @property
    def product_detail_url(self) -> str:
        if not self.base_url:
            raise SupplierMonitorError(f"{self.key} monitor is not configured for adminplus detail page.")
        # adminplus가 상세 popup을 mod= 라우터로 이관함. 구 경로(/partner/product/prt.grp.detail.pop.php)는
        # 빈 응답(len 0)만 줘서 옵션가를 못 읽었음. actpage/pcode는 _fetch_adminplus_prices가 params로 전달.
        return f"{self.base_url}/partner/"

    def output_filename(self, now: datetime) -> str:
        if self.output_name_pattern:
            return self.output_name_pattern.format(date=now.strftime("%y%m%d"))
        return f"{self.output_prefix}_{now.strftime('%y%m%d')}{self.output_suffix}.xlsx"


# ── 발주처가 옵션별로 갈리는 상품의 보조 소스(제주다팜 adminplus) ──
# 미니밤호박 1kg: pcode 10000015 '제주 미니밤호박 보우짱 로얄과'
JEJU_BAMHOBAK_SOURCE = SupplierSource(
    supplier_name="제주다팜",
    source_type="adminplus",
    base_url="https://kkangta55.adminplus.co.kr",
    product_code="10000015",
    product_name="제주 미니밤호박 보우짱 로얄과",
)
# 백도 딱딱이복숭아 2·4kg: pcode 10001059 '딱딱이복숭아' (옵션명 '딱딱이 복숭아 {등급} {kg}kg', 2026-07-27 실측)
JEJU_BAEKDO_SOURCE = SupplierSource(
    supplier_name="제주다팜",
    source_type="adminplus",
    base_url="https://kkangta55.adminplus.co.kr",
    product_code="10001059",
    product_name="딱딱이복숭아",
)

MONITOR_CONFIGS: dict[str, SupplierMonitorConfig] = {
    "myeongi": SupplierMonitorConfig(
        key="myeongi",
        source_type="adminplus",
        base_url="https://pbfcompany.adminplus.co.kr",
        supplier_name="쥬얼리프룻",
        product_name="명이나물(대명이)",
        search_value="명이",
        template_path=TEMPLATE_DIR / "쥬얼리프룻_명이나물(대명이)_소싱현황_원본.xlsx",
        output_prefix="쥬얼리프룻_명이나물(대명이)_V3.1",
        output_name_pattern="쥬얼리프룻 명이나물(대명이) 소싱현환관리_V3.1_{date}.xlsx",
        options=(
            SupplierOptionConfig("명이나물(대명이) 3kg", "명이나물(대명이) 3kg", 8),
            SupplierOptionConfig("명이나물(대명이) 2kg", "명이나물(대명이) 2kg", 9),
            SupplierOptionConfig("명이나물(대명이) 1kg", "명이나물(대명이) 1kg", 10),
            SupplierOptionConfig("명이나물(대명이) 500g", "명이나물(대명이) 500g", 11),
        ),
    ),
    "apple-corn-jewelryfruit": SupplierMonitorConfig(
        key="apple-corn-jewelryfruit",
        source_type="google_sheet",
        supplier_name="쥬얼리프룻",
        product_name="애플초당옥수수",
        sheet_csv_url=GOOGLE_SHEET_JEWELRYFRUIT_KG_CSV,
        sheet_product_name="애플초당옥수수",
        sheet_product_column="E",
        sheet_option_column="F",
        sheet_vip_column="I",
        sheet_price_fallback_column="G",
        sheet_previous_column="G",
        template_path=TEMPLATE_DIR / "초당옥수수_소싱현황_원본.xlsx",
        output_prefix="애플초당옥수수_쥬얼리프룻_V1",
        output_name_pattern="애플초당옥수수 소싱현황관리_V1_{date}.xlsx",
        options=(
            SupplierOptionConfig("애플초당옥수수 특품 5개", "애플초당옥수수(특품) 5개", 16, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("애플초당옥수수 특품 10개", "애플초당옥수수(특품) 10개", 17, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("애플초당옥수수 특품 15개", "애플초당옥수수(특품) 15개", 18, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("애플초당옥수수 특품 20개", "애플초당옥수수(특품) 20개", 19, sheet_name="쥬얼리프룻"),
        ),
    ),
    "potato-jewelry": SupplierMonitorConfig(
        key="potato-jewelry",
        source_type="google_sheet",
        supplier_name="쥬얼리프룻",
        product_name="햇 홍감자",
        sheet_csv_url=GOOGLE_SHEET_JEWELRYFRUIT_KG_CSV,
        sheet_product_name="햇 홍감자",
        sheet_product_column="E",
        sheet_option_column="F",
        sheet_vip_column="I",
        sheet_price_fallback_column="G",
        sheet_previous_column="G",
        template_path=TEMPLATE_DIR / "홍감자_쥬얼리프룻_소싱현황_원본.xlsx",
        output_prefix="홍감자_쥬얼리프룻_V1",
        output_name_pattern="홍감자 소싱현황관리(쥬얼리)_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            # 현재 쿠팡 판매 5종. 대 주문은 2026-07부터 특대로 발주/마진 계산한다.
            SupplierOptionConfig("홍감자 1kg 중", "중 1kg", 8, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("홍감자 3kg 중", "중 3kg", 9, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("홍감자 3kg 특대", "특대 3kg", 10, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("홍감자 5kg 중", "중 5kg", 11, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("홍감자 5kg 특대", "특대 5kg", 12, sheet_name="쥬얼리프룻"),
        ),
    ),
    # 미니밤호박 — 2026-08-17부터 **전 옵션 제주다팜 발주**(쥬얼리 3·5·10kg 발주 회수).
    # adminplus pcode 10000015, 옵션명 '제주 미니밤호박 보우짱 로얄과 {kg}kg'.
    "bamhobak-jewelry": SupplierMonitorConfig(
        key="bamhobak-jewelry",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="제주 미니밤호박 보우짱 로얄과",
        product_code="10000015",
        template_path=TEMPLATE_DIR / "밤호박_통합_소싱현황_원본.xlsx",
        output_prefix="밤호박_제주다팜_V1",
        output_name_pattern="미니밤호박 소싱현황관리(제주다팜)_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig(
                "미니밤호박 로얄과 1kg", "제주 미니밤호박 보우짱 로얄과 1kg", 8, sheet_name="쥬얼리프룻",
                coupang_product="제주 미니밤호박 보우짱", coupang_option="1박스 로얄 정품 1kg",
            ),
            SupplierOptionConfig(
                "미니밤호박 로얄과 3kg", "제주 미니밤호박 보우짱 로얄과 3kg", 9, sheet_name="쥬얼리프룻",
                coupang_product="제주 미니밤호박 보우짱", coupang_option="1박스 로얄 정품 3kg",
            ),
            SupplierOptionConfig(
                "미니밤호박 로얄과 5kg", "제주 미니밤호박 보우짱 로얄과 5kg", 10, sheet_name="쥬얼리프룻",
                coupang_product="제주 미니밤호박 보우짱", coupang_option="1박스 로얄 정품 5kg",
            ),
            SupplierOptionConfig(
                "미니밤호박 로얄과 10kg", "제주 미니밤호박 보우짱 로얄과 10kg", 11, sheet_name="쥬얼리프룻",
                coupang_product="제주 미니밤호박 보우짱", coupang_option="1박스 로얄 정품 10kg",
            ),
        ),
    ),
    "chamoe-jewelry": SupplierMonitorConfig(
        key="chamoe-jewelry",
        source_type="adminplus",
        base_url="https://pbfcompany.adminplus.co.kr",
        supplier_name="쥬얼리프룻",
        product_name="성주참외",
        search_value="성주참외",
        product_code="10000234",
        template_path=TEMPLATE_DIR / "성주참외_쥬얼리프룻_소싱현황_원본.xlsx",
        output_prefix="성주참외_쥬얼리프룻_V1",
        output_name_pattern="성주참외 소싱현황관리(쥬얼리)_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            # pbfcompany adminplus 옵션명은 첨부 계산기 E열 옵션명과 맞춰 찾는다.
            SupplierOptionConfig("성주참외 가성비 랜덤과 2kg", "2kg (가정용 혼합과)_vip", 8),
            SupplierOptionConfig("성주참외 가성비 랜덤과 3kg", "3kg (가정용 혼합과)_vip", 9),
            SupplierOptionConfig("성주참외 가성비 랜덤과 5kg", "5kg (가정용 혼합과)_vip", 10),
            SupplierOptionConfig("성주참외 로얄과 3kg", "3kg (가정용 로얄과)_vip", 11),
            SupplierOptionConfig("성주참외 로얄과 5kg", "5kg (가정용 로얄과)_vip", 12),
        ),
    ),
    "kolrabi": SupplierMonitorConfig(
        key="kolrabi",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="콜라비(정품)",
        search_value="콜라비",
        template_path=TEMPLATE_DIR / "콜라비_제주다팜_소싱현황_원본.xlsx",
        output_prefix="콜라비_제주다팜_V3.2",
        output_name_pattern="콜라비 소싱현황관리_V3.2_{date}(제주다팜).xlsx",
        options=(
            SupplierOptionConfig("콜라비 정품 3kg", "콜라비(정품 300~750g) 3kg", 8),
            SupplierOptionConfig("콜라비 정품 5kg", "콜라비(정품 300~750g) 5kg", 9),
            SupplierOptionConfig("콜라비 정품 10kg", "콜라비(정품 300~750g) 10kg", 10),
        ),
    ),
    # (2026-07-27) 제주다팜 미니밤호박 1kg 단독 모니터는 bamhobak-jewelry로 통합 — 한 파일에서 비교.
    # 홍감자 (2026-07-14 쥬얼리 품절 → 제주다팜 이관). kkangta55 pcode=10001051 '국산 홍감자',
    # popup 옵션명 '홍감자 {등급}_{KG}KG' (라이브 확인). 소싱시트(V3.0)는 쿠팡 판매옵션 기준 행이고
    # 공급가(J열)는 발주 매핑된 옵션 단가: 중1kg→중_2KG, 대3kg→특_3KG, 대5kg→특_5KG.
    # ※ 이 양식은 공급가가 I가 아닌 J열 → cell="J".
    "potato-jeju": SupplierMonitorConfig(
        key="potato-jeju",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="국산 홍감자",
        product_code="10001051",
        template_path=TEMPLATE_DIR / "홍감자_제주다팜_소싱현황_원본.xlsx",
        output_prefix="홍감자_제주다팜_V3.0",
        output_name_pattern="홍감자_제주다팜 소싱현황관리_V3.0_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig("홍감자 중 1kg(발주 중 2kg)", "홍감자 중_2KG", 8, cell="J"),
            SupplierOptionConfig("홍감자 중 3kg", "홍감자 중_3KG", 9, cell="J"),
            SupplierOptionConfig("홍감자 중 5kg", "홍감자 중_5KG", 10, cell="J"),
            SupplierOptionConfig("홍감자 대 3kg(발주 특 3kg)", "홍감자 특_3KG", 11, cell="J"),
            SupplierOptionConfig("홍감자 대 5kg(발주 특 5kg)", "홍감자 특_5KG", 12, cell="J"),
        ),
    ),
    # 청사과(아오리사과) — 2026-08 제주다팜 신규. adminplus pcode=10001098.
    # 쿠팡 판매는 소과·대과 2·3·4·5kg 8종(중소과는 미판매 — 판매 시작하면 소싱현황 행 추가 후 여기에도 추가).
    # supplier_option_name은 제주다팜 판매옵션 전체 문자열 = kolrabi_order._APPLE_JEJU_OPTIONS와 동일해야 한다.
    "apple-jeju": SupplierMonitorConfig(
        key="apple-jeju",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="청사과",
        product_code="10001098",
        template_path=TEMPLATE_DIR / "청사과_제주다팜_소싱현황_원본.xlsx",
        output_prefix="청사과_제주다팜_V1",
        output_name_pattern="청사과(아오리)_제주다팜 소싱현황관리_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig(
                "청사과 소과 2kg", "청사과 소과(가정용) 포장재포함 2kg(10-13과내외)", 8, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 소과 2kg",
            ),
            SupplierOptionConfig(
                "청사과 소과 3kg", "청사과 소과(가정용) 포장재포함 3kg(15-22과내외)", 9, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 소과 3kg",
            ),
            SupplierOptionConfig(
                "청사과 소과 4kg", "청사과 소과(가정용) 포장재포함 4kg(19-27과내외)", 10, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 소과 4kg",
            ),
            SupplierOptionConfig(
                "청사과 소과 5kg", "청사과 소과(가정용) 포장재포함 5kg(25-40과내외)", 11, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 소과 5kg",
            ),
            SupplierOptionConfig(
                "청사과 대과 2kg", "청사과 대과(가정용) 포장재포함 2kg(6과내)", 12, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 대과 2kg",
            ),
            SupplierOptionConfig(
                "청사과 대과 3kg", "청사과 대과(가정용) 포장재포함 3kg(8-10과내)", 13, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 대과 3kg",
            ),
            SupplierOptionConfig(
                "청사과 대과 4kg", "청사과 대과(가정용) 포장재포함 4kg(10-12과내)", 14, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 대과 4kg",
            ),
            SupplierOptionConfig(
                "청사과 대과 5kg", "청사과 대과(가정용) 포장재포함 5kg(12-16과내)", 15, sheet_name="쥬얼리프룻",
                coupang_product="청사과(아오리사과)", coupang_option="1박스 대과 5kg",
            ),
        ),
    ),
    # 홍로사과(가을햇사과) — 2026-08 제주다팜 신규. adminplus pcode=10001216 '가을햇사과(홍사과)'.
    # 소과·중소과·중대과·대과 × 1.5·2·3·4·5kg = 20종(실측 2026-08-27).
    # supplier_option_name은 kolrabi_order._HONGRO_JEJU_OPTIONS와 같은 문자열이어야 공급가가 붙는다.
    # ※ 쿠폰가(H열)는 소과 1.5kg만 확인됨 — 나머지는 사용자가 소싱현황 H열을 채우면 마진이 계산된다.
    "hongro-jeju": SupplierMonitorConfig(
        key="hongro-jeju",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="가을햇사과(홍사과)",
        product_code="10001216",
        template_path=TEMPLATE_DIR / "홍로사과_제주다팜_소싱현황_원본.xlsx",
        output_prefix="홍로사과_제주다팜_V1",
        output_name_pattern="홍로사과(가을햇사과)_제주다팜 소싱현황관리_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig(
                "홍로사과 소과 1.5kg", "가을햇사과(홍사과) 가정용 소과 포장재포함 1.5kg(7-10과내)", 8, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 소과 1.5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 소과 2kg", "가을햇사과(홍사과)가정용 소과 포장재포함 2kg(11-15과내외)", 9, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 소과 2kg",
            ),
            SupplierOptionConfig(
                "홍로사과 소과 3kg", "가을햇사과(홍사과) 가정용 소과 포장재포함 3kg(17-20과내외)", 10, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 소과 3kg",
            ),
            SupplierOptionConfig(
                "홍로사과 소과 4kg", "가을햇사과(홍사과) 가정용 소과 포장재포함 4kg(20-25과내외)", 11, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 소과 4kg",
            ),
            SupplierOptionConfig(
                "홍로사과 소과 5kg", "가을햇사과(홍사과) 가정용 소과 포장재포함 5kg(29-35과)", 12, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 소과 5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중소과 1.5kg", "가을햇사과(홍사과)가정용 중소과 포장재포함 1.5kg(6과내외)", 13, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중소과 1.5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중소과 2kg", "가을햇사과(홍사과) 가정용 중소과 포장재포함 2kg(7-8과내외)", 14, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중소과 2kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중소과 3kg", "가을햇사과(홍사과) 가정용 중소과 포장재포함 3kg(13-15과내외)", 15, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중소과 3kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중소과 4kg", "가을햇사과(홍사과) 가정용 중소과 포장재포함 4kg(15-17과내외)", 16, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중소과 4kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중소과 5kg", "가을햇사과(홍사과) 가정용 중소과 포장재포함 5kg(23-24과)", 17, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중소과 5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중대과 1.5kg", "가을햇사과(홍사과) 가정용 중대과 포장재포함 1.5kg(4-5과내외)", 18, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중대과 1.5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중대과 2kg", "가을햇사과(홍사과) 가정용 중대과 포장재포함 2kg(7과내외)", 19, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중대과 2kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중대과 3kg", "가을햇사과(홍사과) 가정용 중대과 포장재포함 3kg(11-12과내외)", 20, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중대과 3kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중대과 4kg", "가을햇사과(홍사과) 가정용 중대과 포장재포함 4kg(13-14과내외)", 21, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중대과 4kg",
            ),
            SupplierOptionConfig(
                "홍로사과 중대과 5kg", "가을햇사과(홍사과) 가정용 중대과 포장재포함 5kg(19-22과)", 22, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 중대과 5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 대과 1.5kg", "가을햇사과(홍사과) 가정용 대과 포장재포함 1.5kg(4과내외)", 23, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 대과 1.5kg",
            ),
            SupplierOptionConfig(
                "홍로사과 대과 2kg", "가을햇사과(홍사과) 가정용 대과 포장재포함 2kg(6과내외)", 24, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 대과 2kg",
            ),
            SupplierOptionConfig(
                "홍로사과 대과 3kg", "가을햇사과(홍사과) 가정용 대과 포장재포함 3kg(10과내외)", 25, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 대과 3kg",
            ),
            SupplierOptionConfig(
                "홍로사과 대과 4kg", "가을햇사과(홍사과) 가정용 대과 포장재포함 4kg(12과내외)", 26, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 대과 4kg",
            ),
            SupplierOptionConfig(
                "홍로사과 대과 5kg", "가을햇사과(홍사과) 가정용 대과 포장재포함 5kg(16-18과)", 27, sheet_name="쥬얼리프룻",
                coupang_product="2026 햇 안동 고당도 홍로사과", coupang_option="1박스 대과 5kg",
            ),
        ),
    ),
    "corn-jbt": SupplierMonitorConfig(
        key="corn-jbt",
        source_type="google_sheet",
        supplier_name="쥬얼리프룻",
        product_name="초당옥수수",
        # 2026-06 제주다팜→쥬얼리프룻 발주 이관 → 마진방어도 쥬얼리 시트로.
        # '초당옥수수'는 '애플초당옥수수'와 부분일치하므로 exclude로 분리한다.
        sheet_csv_url=GOOGLE_SHEET_JEWELRYFRUIT_KG_CSV,
        sheet_product_name="초당옥수수",
        sheet_product_exclude=("애플", "미백", "흑찰"),
        sheet_product_column="E",
        sheet_option_column="F",
        sheet_vip_column="I",
        sheet_price_fallback_column="G",
        sheet_previous_column="G",
        template_path=TEMPLATE_DIR / "초당옥수수_소싱현황_원본.xlsx",
        output_prefix="초당옥수수_쥬얼리프룻_V1",
        output_name_pattern="초당옥수수 소싱현황관리(쥬얼리)_V1_{date}.xlsx",
        # 2026-07 중품 4종은 공급 중단(시트 공급가 공란), 특품 4종은 출고중.
        # 옵션 일부가 빠져도 오류 대신 건너뛰고 남은 옵션만 계산. 전 옵션이 빠지면(완전 품절)
        # 조용히 빈 결과 → 오류 알림/재시도 부하 없음. (다른 모니터도 전부 True)
        skip_missing_options=True,
        options=(
            SupplierOptionConfig("초당옥수수 중품 5개", "중품 5개입", 8, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 중품 10개", "중품 10개입", 9, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 중품 15개", "중품 15개입", 10, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 중품 20개", "중품 20개입", 11, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 특품 5개", "특품 5개입", 12, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 특품 10개", "특품 10개입", 13, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 특품 15개", "특품 15개입", 14, sheet_name="쥬얼리프룻"),
            SupplierOptionConfig("초당옥수수 특품 20개", "특품 20개입", 15, sheet_name="쥬얼리프룻"),
        ),
    ),
    "chamoe-jbt": SupplierMonitorConfig(
        key="chamoe-jbt",
        source_type="google_sheet",
        supplier_name="제이비티",
        product_name="성주참외 로얄과",
        sheet_csv_url=GOOGLE_SHEET_JBT_FRUIT_CSV,
        sheet_product_name="참외",
        template_path=TEMPLATE_DIR / "성주참외_제이비티_소싱현황_원본.xlsx",
        output_prefix="성주참외_제이비티_V3",
        output_name_pattern="성주참외 소싱현황관리_V3_{date}.xlsx",
        options=(
            SupplierOptionConfig("성주참외 로얄과 3kg", "가정용참외(로얄과)3kg(8-13과)", 15, sheet_name="제이비티"),
            SupplierOptionConfig("성주참외 로얄과 5kg", "가정용참외(로얄과)5kg(14-23과)", 17, sheet_name="제이비티"),
        ),
    ),
    "tomato-jbt": SupplierMonitorConfig(
        key="tomato-jbt",
        source_type="google_sheet",
        supplier_name="제이비티",
        product_name="대저토마토(고품위)",
        sheet_csv_url=GOOGLE_SHEET_JBT_FRUIT_CSV,
        sheet_product_name="토마토(고품위)",
        template_path=TEMPLATE_DIR / "대저토마토_제이비티_소싱현황_원본.xlsx",
        output_prefix="대저토마토_제이비티_고품위_V3",
        output_name_pattern="대저토마토 소싱현황관리_V3_{date}(로스3%,소득세16%기준).xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig("대저 짭짤이 로얄과 1.5kg", "대저짭짤이특품(S)과 1.5kg", 8, sheet_name="제이비티"),
            SupplierOptionConfig("대저 짭짤이 로얄과 2.5kg", "대저짭짤이특품(S)과2.5kg", 9, sheet_name="제이비티"),
            SupplierOptionConfig("대저토마토 특품 L 2.5kg", "대저토마토특품(L)과2.5kg", 11, sheet_name="제이비티"),
            SupplierOptionConfig("대저토마토 특품 M 1.5kg", "대저토마토특품(M)과1.5kg", 12, sheet_name="제이비티"),
            SupplierOptionConfig("대저토마토 특품 M 2.5kg", "대저토마토특품(M)과2.5kg", 13, sheet_name="제이비티"),
        ),
    ),
    "watermelon-jbt": SupplierMonitorConfig(
        key="watermelon-jbt",
        source_type="google_sheet",
        supplier_name="제이비티",
        product_name="수박",
        sheet_csv_url=GOOGLE_SHEET_JBT_FRUIT_CSV,
        sheet_product_name="수박",
        template_path=TEMPLATE_DIR / "수박_제이비티_소싱현황_원본.xlsx",
        output_prefix="수박_제이비티_V3",
        output_name_pattern="수박 소싱현황관리_V3_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig("하우스수박 상품 6kg 이상", "하우스수박(상품)6kg이상", 8),
            SupplierOptionConfig("하우스수박 상품 7kg 이상", "하우스수박(상품)7kg이상", 10),
            SupplierOptionConfig("하우스수박 상품 8kg 이상", "하우스수박(상품)8kg이상", 12),
        ),
    ),
    "watermelon-jewelry": SupplierMonitorConfig(
        key="watermelon-jewelry",
        source_type="manual",
        supplier_name="쥬얼리프룻",
        product_name="망고수박",
        template_path=TEMPLATE_DIR / "망고수박_쥬얼리프룻_소싱현황_원본.xlsx",
        output_prefix="망고수박_쥬얼리프룻_V3",
        output_name_pattern="망고수박 소싱현황관리(쥬얼리)_V3_{date}.xlsx",
        skip_missing_options=True,
        options=(
            SupplierOptionConfig("망고수박 가정용 2kg", "수박 가정용 2~3kg 내외 (2kg)", 8, sheet_name="제이비티"),
            SupplierOptionConfig("망고수박 가정용 3kg", "수박 가정용 3~4kg 내외 (3kg)", 9, sheet_name="제이비티"),
            SupplierOptionConfig("망고수박 가정용 5kg", "수박 가정용 4~5kg내외(5kg)", 10, sheet_name="제이비티"),
            SupplierOptionConfig("망고수박 가정용 6kg", "수박 가정용 5~6kg내외(6kg)", 11, sheet_name="제이비티"),
        ),
    ),
    # 백도 딱딱이복숭아 — 전량 제주다팜 발주(kkangta55 pcode=10001059, '딱딱이 복숭아 {등급} {kg}kg').
    # 2026-08-10: 쥬얼리프룻 백도가 시즌 종료(어드민 옵션 0개)로 발주 중단 → 쥬얼리 1kg 2행을
    #   설정·템플릿에서 제거(안 빼면 공급가를 못 읽어 매일 오전 알림이 오류로 뜬다).
    #   내년 시즌 재개 시: 아래 JEWELRY 주석의 옵션 2개(중과/대과 1kg)와 템플릿 행을 되살리면 된다.
    #   (쥬얼리 옵션명은 과수 표기가 시즌마다 바뀌므로 과수 없는 짧은 이름으로 부분매칭할 것)
    "baekdo-jewelry": SupplierMonitorConfig(
        key="baekdo-jewelry",
        source_type="adminplus",
        base_url="https://kkangta55.adminplus.co.kr",
        supplier_name="제주다팜",
        product_name="딱딱이복숭아",
        product_code="10001059",
        template_path=TEMPLATE_DIR / "백도복숭아_쥬얼리프룻_소싱현황_원본.xlsx",
        output_prefix="백도복숭아_제주다팜_V1",
        output_name_pattern="백도복숭아 소싱현황관리_V1_{date}.xlsx",
        skip_missing_options=True,
        options=(
            # 템플릿 행8~11 = 제주다팜 4옵션 (쥬얼리 1kg 시즌아웃으로 행 정리됨)
            SupplierOptionConfig(
                "백도 딱딱이복숭아 중과 2kg", "딱딱이 복숭아 중과 2kg", 8, sheet_name="쥬얼리프룻",
                coupang_product="햇 백도 딱딱이복숭아", coupang_option="1박스 중과 2kg",
            ),
            SupplierOptionConfig(
                "백도 딱딱이복숭아 중과 4kg", "딱딱이 복숭아 중과 4kg", 9, sheet_name="쥬얼리프룻",
                coupang_product="햇 백도 딱딱이복숭아", coupang_option="1박스 중과 4kg",
            ),
            SupplierOptionConfig(
                "백도 딱딱이복숭아 대과 2kg", "딱딱이 복숭아 대과 2kg", 10, sheet_name="쥬얼리프룻",
                coupang_product="햇 백도 딱딱이복숭아", coupang_option="1박스 대과 2kg",
            ),
            SupplierOptionConfig(
                "백도 딱딱이복숭아 대과 4kg", "딱딱이 복숭아 대과 4kg", 11, sheet_name="쥬얼리프룻",
                coupang_product="햇 백도 딱딱이복숭아", coupang_option="1박스 대과 4kg",
            ),
            # JEWELRY(시즌 재개 시 복구): pbfcompany pcode=10000549, source=SupplierSource(
            #   supplier_name="쥬얼리프룻", base_url="https://pbfcompany.adminplus.co.kr",
            #   product_code="10000549", product_name="백도 딱딱이 복숭아")
            #   옵션: "딱딱이 백도 복숭아 중과 1kg" / "딱딱이 백도 복숭아 대과 1kg"
        ),
    ),
    # ── 신비복숭아 한시 모니터 (2026-06-15 ~ 2026-06-29, 2주) — 2026-07 품절로 마진방어 UI에서 제외 ──
    "peach-jewelry": SupplierMonitorConfig(
        key="peach-jewelry",
        source_type="google_sheet",
        supplier_name="쥬얼리프룻",
        product_name="신비복숭아",
        sheet_csv_url=GOOGLE_SHEET_JEWELRYFRUIT_PEACH_CSV,
        sheet_product_name="신비 복숭아",
        sheet_product_column="E",
        sheet_option_column="F",
        sheet_vip_column="I",            # 변경단가(VIP) 우선
        sheet_price_fallback_column="G",  # 비어있으면 일반공급가
        sheet_previous_column="G",
        template_path=TEMPLATE_DIR / "신비복숭아_쥬얼리프룻_소싱현황_원본.xlsx",
        output_prefix="신비복숭아_쥬얼리프룻_V1",
        output_name_pattern="신비복숭아 소싱현황관리(쥬얼리)_V1_{date}.xlsx",
        active_until="2026-06-29",
        skip_missing_options=True,
        options=(
            # 쥬얼리 구글시트(gid=1567728493) 신비 복숭아 옵션(F열)의 변경단가(I열). 쿠팡 판매 4종, 소싱현황 행순서.
            SupplierOptionConfig("신비복숭아 1kg 중소과", "1kg (15과 내외)", 8),
            SupplierOptionConfig("신비복숭아 2kg 중소과", "2kg (30과 내외)", 9),
            SupplierOptionConfig("신비복숭아 1kg 대과", "1kg (11과 내외)", 10),
            SupplierOptionConfig("신비복숭아 2kg 대과", "2kg (22과 내외)", 11),
        ),
    ),
    "peach-jbt": SupplierMonitorConfig(
        key="peach-jbt",
        source_type="google_sheet",
        supplier_name="제이비티",
        product_name="신비복숭아",
        sheet_csv_url=GOOGLE_SHEET_JBT_FRUIT_CSV,
        sheet_product_name="복숭아",
        template_path=TEMPLATE_DIR / "신비복숭아_제이비티_소싱현황_원본.xlsx",
        output_prefix="신비복숭아_제이비티_V1",
        output_name_pattern="신비복숭아 소싱현황관리(제이비티)_V1_{date}.xlsx",
        active_until="2026-06-29",
        skip_missing_options=True,
        options=(
            # 제이비티 발주 신비복숭아 3·4kg은 중소과. 시트에 등급별로 있어 '중소과'로 특정해야 매칭됨.
            SupplierOptionConfig("신비복숭아 3kg 중소과", "신비복숭아 3kg 중소과", 8, sheet_name="제이비티"),
            SupplierOptionConfig("신비복숭아 4kg 중소과", "신비복숭아 4kg 중소과", 9, sheet_name="제이비티"),
        ),
    ),
    "dureup-jbt": SupplierMonitorConfig(
        key="dureup-jbt",
        source_type="google_sheet",
        supplier_name="제이비티",
        product_name="남해땅두릅(특품)",
        sheet_csv_url=GOOGLE_SHEET_JBT_FARM_CSV,
        sheet_product_name="두릅(고품위)",
        template_path=TEMPLATE_DIR / "남해땅두릅_제이비티_소싱현황_원본.xlsx",
        output_prefix="남해땅두릅_제이비티_특품_V1",
        output_name_pattern="땅두릅(남해) 소싱현황관리_V1_{date}.xlsx",
        options=(
            SupplierOptionConfig("남해땅두릅 특품 500g", "남해땅두릅(특품)500g", 8, sheet_name="제이비티"),
            SupplierOptionConfig("남해땅두릅 특품 1kg", "남해땅두릅(특품)1kg", 9, sheet_name="제이비티"),
            SupplierOptionConfig("남해땅두릅 특품 2kg", "남해땅두릅(특품)2kg", 10, sheet_name="제이비티"),
        ),
    ),
}

# 일시중지(숨김) 모니터. 여기 추가하면 스냅샷/알림에서 'paused'로 빠져 오류 안 뜸.
#  - watermelon-jbt·chamoe-jbt: 발주처 제이비티→쥬얼리 이전으로 무용
#  - apple-corn-jewelryfruit: 2026-07 애플초당옥수수 품절(구글시트에서 VIP 공급가 빠짐). 시즌 재개 시 제거.
#  - potato-jewelry: 2026-07-13 쥬얼리 홍감자 품절 → 발주 제주다팜 이관(마진 시트는 쥬얼리 기준이라 무용).
# ※ 품절/시즌아웃 상품은 여기 추가 + 프론트 SUPPLIER_MONITOR_KEYS에서 제거로 마진방어에서 자동 숨김.
PAUSED_SUPPLIER_MONITOR_KEYS = {"myeongi", "dureup-jbt", "watermelon-jbt", "chamoe-jbt", "apple-corn-jewelryfruit", "potato-jewelry"}

MONITOR_CONFIGS["dureup-jbt"] = replace(
    MONITOR_CONFIGS["dureup-jbt"],
    sheet_vip_column="H",
    options=(
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud2b9\ud488 500g",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud2b9\ud488)500g",
            8,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud2b9\ud488 1kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud2b9\ud488)1kg",
            9,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud2b9\ud488 2kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud2b9\ud488)2kg",
            10,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \uc7a5\uc544\ucc0c\uc6a9 500g",
            "\ub0a8\ud574\ub545\ub450\ub985(\uc7a5\uc544\ucc0c\uc6a9)500g",
            11,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \uc7a5\uc544\ucc0c\uc6a9 1kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\uc7a5\uc544\ucc0c\uc6a9)1kg",
            12,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \uc7a5\uc544\ucc0c\uc6a9 2kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\uc7a5\uc544\ucc0c\uc6a9)2kg",
            13,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud280\uae40\uc6a9 500g",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud280\uae40\uc6a9)500g",
            14,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud280\uae40\uc6a9 1kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud280\uae40\uc6a9)1kg",
            15,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
        SupplierOptionConfig(
            "\ub0a8\ud574\ub545\ub450\ub985 \ud280\uae40\uc6a9 2kg",
            "\ub0a8\ud574\ub545\ub450\ub985(\ud280\uae40\uc6a9)2kg",
            16,
            sheet_name="\uc81c\uc774\ube44\ud2f0",
        ),
    ),
)


class SupplierMonitorError(RuntimeError):
    pass


def option_supplier_name(option: SupplierOptionConfig, config: SupplierMonitorConfig) -> str:
    """옵션 한 줄의 실제 발주처 (보조 소스 > 옵션 지정 > 모니터 기본)."""
    if option.source:
        return option.source.supplier_name
    return option.supplier_name or config.supplier_name


def _normalize_option_name(raw: str) -> str:
    text = re.sub(r"<[^>]+>", "", raw or "")
    text = unescape(text).replace("\xa0", " ")
    text = re.sub(r"^\[\d+\]\s*", "", text)
    return re.sub(r"\s+", " ", text).strip()


def _parse_price(raw: object) -> int:
    if raw is None:
        raise SupplierMonitorError("공급가를 읽을 수 없습니다.")
    if isinstance(raw, (int, float)):
        return int(raw)
    digits = re.sub(r"[^\d]", "", str(raw))
    if not digits:
        raise SupplierMonitorError(f"숫자 공급가를 찾지 못했습니다: {raw}")
    return int(digits)


def _price_signal(previous_supplier_price: int, supplier_price: int) -> str:
    if previous_supplier_price < supplier_price:
        return "blue"
    if previous_supplier_price > supplier_price:
        return "red"
    return "same"


def _formula_rate(cell_value, default: float = 0.0) -> float:
    """'=H8*5%' 류 수식에서 퍼센트를 비율(0.05)로 추출."""
    m = re.search(r"\*\s*([\d.]+)\s*%", str(cell_value or ""))
    return float(m.group(1)) / 100 if m else default


def _compute_margin(ws, row: int, supplier_price: int) -> int | None:
    """소싱현황 정식 양식 기준 마진(U, 원)을 계산. H(쿠폰가) 없으면 None.

    U = 쿠폰가 - 공급가 - 마진방어 - 쿠팡수수료 - 소득세 - cs (템플릿 U열 수식과 동일).
    수수료율(K)은 템플릿 값을 그대로 사용하므로 카테고리별(예: 홍감자 8.6%) 자동 반영.
    """
    def num(col: str) -> float:
        try:
            return float(ws[f"{col}{row}"].value)
        except (TypeError, ValueError):
            return 0.0

    coupon = num("H")  # 쿠폰가
    if coupon <= 0:
        return None
    fee_rate = num("K") or 0.1166       # 쿠팡 수수료율(템플릿 값)
    supply = float(supplier_price)      # 공급가(vip, 실시간)
    logistics = num("L")
    extra = num("M")
    defense = coupon * _formula_rate(ws[f"N{row}"].value)  # 마진방어
    sale = coupon + 500                                    # 판매가
    coupang_fee = (sale * fee_rate) + (coupon * 0.0363)    # 쿠팡수수료
    subtotal = coupon - supply - logistics - extra - defense - coupang_fee  # 1차 합계(R)
    income_tax = subtotal * 0.16
    cs = sale * 0.01
    return round(coupon - supply - logistics - extra - defense - coupang_fee - income_tax - cs)


def _formula_cache_updates(ws, row: int, supplier_price: int) -> list[tuple[str, str, int | float | None]]:
    """Return formula cells with calculated cache values for the sourcing workbook.

    The monitor patches xlsx XML directly so Excel formulas stay intact. Formula
    cache values must also be written, otherwise downloaded files look blank
    until Excel recalculates them locally.
    """

    def num(col: str) -> float:
        try:
            return float(ws[f"{col}{row}"].value)
        except (TypeError, ValueError):
            return 0.0

    coupon = num("H")
    if coupon <= 0:
        return []

    fee_rate = num("K") or 0.1166
    supply = float(supplier_price)
    logistics = num("L")
    extra = num("M")
    sale = coupon + 500
    defense = coupon * _formula_rate(ws[f"N{row}"].value)
    coupang_fee = (sale * fee_rate) + (coupon * 0.0363)
    vat = (coupon - (supply + logistics + extra + defense + coupang_fee)) * 0
    subtotal = coupon - supply - logistics - extra - defense - coupang_fee - vat
    income_tax = subtotal * 0.16
    cs = sale * 0.01
    margin = coupon - supply - logistics - extra - defense - coupang_fee - vat - income_tax - cs
    margin_rate = None if sale == 0 or margin == 0 else margin / sale

    return [
        (f"G{row}", f"=H{row}+500", sale),
        (f"N{row}", str(ws[f"N{row}"].value or f"=H{row}*3%"), defense),
        (f"P{row}", f"=(G{row}*K{row})+(H{row}*3.63%)", coupang_fee),
        (f"Q{row}", f"=((H{row})-(I{row}+L{row}+M{row}+N{row}+P{row}))*0%", vat),
        (f"R{row}", f"=H{row}-I{row}-L{row}-M{row}-N{row}-P{row}-Q{row}", subtotal),
        (f"S{row}", f"=R{row}*16%", income_tax),
        (f"T{row}", f"=G{row}*1%", cs),
        (f"U{row}", f"=H{row}-I{row}-L{row}-M{row}-N{row}-P{row}-Q{row}-S{row}-T{row}", margin),
        (f"V{row}", f'=IF(U{row}=0,"",U{row}/G{row})', margin_rate),
    ]


def _snapshot_lookup_keys(option_name: str, supplier_option_name: str) -> set[tuple[str, str]]:
    return {
        (option_name, supplier_option_name),
        (_normalize_option_name(option_name), _normalize_option_name(supplier_option_name)),
        (
            re.sub(r"\s+", "", _normalize_option_name(option_name)),
            re.sub(r"\s+", "", _normalize_option_name(supplier_option_name)),
        ),
    }


def _supplier_price_lookup(prices: dict[str, int], supplier_option_name: str) -> int | None:
    normalized_name = _normalize_option_name(supplier_option_name)
    candidates = _supplier_price_lookup_candidates(normalized_name)
    compact_candidates = [re.sub(r"\s+", "", candidate) for candidate in candidates]
    for key in [supplier_option_name, *candidates, *compact_candidates]:
        if key in prices:
            return prices[key]
    for key, price in prices.items():
        compact_key = re.sub(r"\s+", "", _normalize_option_name(key))
        if compact_key in compact_candidates:
            return price
    partial_matches = []
    for key, price in prices.items():
        compact_key = re.sub(r"\s+", "", _normalize_option_name(key))
        for compact_name in compact_candidates:
            if compact_name and (compact_name in compact_key or compact_key in compact_name):
                partial_matches.append(price)
                break
    if len(set(partial_matches)) == 1:
        return partial_matches[0]
    return None


def _supplier_price_lookup_candidates(option_name: str) -> list[str]:
    candidates: list[str] = []

    def add(value: str) -> None:
        normalized = _normalize_option_name(value)
        if normalized and normalized not in candidates:
            candidates.append(normalized)

    normalized = _normalize_option_name(option_name)
    add(normalized)
    without_vip = re.sub(r"\s*_vip\s*$", "", normalized, flags=re.IGNORECASE)
    add(without_vip)

    # pbfcompany 성주참외 adminplus 옵션은 계산기 E열과 표기가 다르다.
    # 예: "2kg (가정용 혼합과)_vip" -> "성주참외 소과 2kg (R)"
    kg_match = re.search(r"(\d+(?:\.\d+)?)\s*kg", without_vip, flags=re.IGNORECASE)
    if kg_match and "가정용" in without_vip:
        kg = kg_match.group(1)
        if kg.endswith(".0"):
            kg = kg[:-2]
        grade = ""
        if "로얄과" in without_vip:
            grade = "로얄과"
        elif any(token in without_vip for token in ("혼합과", "랜덤과", "가성비")):
            grade = "소과"
        if grade:
            add(f"성주참외 {grade} {kg}kg (R)")
            add(f"성주참외 {grade} {kg}kg")
            add(f"{grade} {kg}kg (R)")
            add(f"{grade} {kg}kg")

    return candidates


def _sheet_product_matches(actual: str, expected: str) -> bool:
    actual_name = _normalize_option_name(actual)
    expected_name = _normalize_option_name(expected)
    if actual_name == expected_name:
        return True
    actual_compact = re.sub(r"\s+", "", actual_name)
    expected_compact = re.sub(r"\s+", "", expected_name)
    return bool(
        actual_compact
        and expected_compact
        and (expected_compact in actual_compact or actual_compact in expected_compact)
    )


def _google_sheet_option_keys(product_name: str, option_name: str, quantity_name: str = "") -> list[str]:
    product = _normalize_option_name(product_name)
    option = _normalize_option_name(option_name)
    quantity = _normalize_option_name(quantity_name)
    keys: list[str] = []

    def add(value: str) -> None:
        cleaned = _normalize_option_name(value)
        if cleaned and cleaned not in keys:
            keys.append(cleaned)

    add(option)
    if option and quantity:
        add(f"{option} {quantity}")

    product_compact = re.sub(r"\s+", "", product)
    product_aliases = [product_compact] if product_compact else []
    for alias in ("애플초당옥수수", "초당옥수수"):
        if alias in product_compact and alias not in product_aliases:
            product_aliases.append(alias)

    grade = ""
    for candidate in ("중품", "특품", "상품", "정품", "대중과", "랜덤과", "로얄과", "중소과"):
        if candidate in option:
            grade = candidate
            break
    count_match = re.search(r"(\d+)\s*개", f"{option} {quantity}")
    count_name = f"{count_match.group(1)}개" if count_match else ""
    normalized_quantity = count_name or quantity
    for product_alias in product_aliases:
        if grade and normalized_quantity:
            add(f"{product_alias}({grade}) {normalized_quantity}")
            add(f"{product_alias} {grade} {normalized_quantity}")
        if option and quantity:
            add(f"{product_alias} {option} {quantity}")
    return keys


def _persist_output(filename: str, payload: bytes) -> Path:
    output_dir = UPLOAD_DIR / "supplier-price-monitor"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename
    output_path.write_bytes(payload)
    return output_path


def _select_sheet(workbook, sheet_name: str | None):
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _column_index(column_name: str) -> int:
    value = 0
    for char in column_name.upper():
        if not ("A" <= char <= "Z"):
            raise SupplierMonitorError(f"Invalid column name: {column_name}")
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1


async def _login(client: httpx.AsyncClient, config: SupplierMonitorConfig) -> None:
    if not PBF_PARTNER_ID or not PBF_PARTNER_PASSWORD:
        raise SupplierMonitorError("PBF_PARTNER_ID 또는 PBF_PARTNER_PASSWORD가 설정되지 않았습니다.")

    response = await client.post(
        config.login_url,
        data={"admid": PBF_PARTNER_ID, "admpwd": PBF_PARTNER_PASSWORD},
    )
    response.raise_for_status()
    if response.text.strip().lower() != "ok":
        raise SupplierMonitorError("관리자 사이트 로그인에 실패했습니다.")


async def _find_product_code(client: httpx.AsyncClient, config: SupplierMonitorConfig) -> str:
    response = await client.get(
        config.product_list_url,
        params={
            "mod": "product/json",
            "actpage": "prt.list.proc",
            "searchval": config.search_value,
            "page": "1",
            "order": "",
            "by": "",
        },
    )
    response.raise_for_status()
    matches = re.findall(
        r"""prtView\("(?P<pcode>\d+)","2","1"\).*?<div class='pname'>(?P<name>[^<]+)</div>""",
        response.text,
        flags=re.DOTALL,
    )
    for pcode, product_name in matches:
        if _normalize_option_name(product_name) == _normalize_option_name(config.product_name):
            return pcode
    target = re.sub(r"\s+", "", _normalize_option_name(config.product_name))
    partial_matches = []
    for pcode, product_name in matches:
        candidate = re.sub(r"\s+", "", _normalize_option_name(product_name))
        if target and (target in candidate or candidate in target):
            partial_matches.append((pcode, product_name))
    if len(partial_matches) == 1:
        return partial_matches[0][0]
    raise SupplierMonitorError(f"{config.product_name} 상품 코드를 찾지 못했습니다.")


def _parse_adminplus_popup_prices(html: str) -> dict[str, int]:
    """adminplus 상세 popup에서 옵션명→공급가를 뽑는다 (신·구 마크업 모두).

    2026-08 adminplus가 popup을 리뉴얼함: 구성상품 테이블이
    <tr class="product_set_row"> / <p class="set_item_name">옵션명</p> /
    <td class="col_price">공급가</td><td class="col_price">판매가</td> 구조로 변경
    (헤더 순서: 제품명·재고·공급가·판매가·과세·배송비 — 첫 col_price가 공급가).
    구 마크업(<td style='border-left:0px'>…)은 폴백으로 유지.
    """
    # 클래스는 부분일치로 본다 — 실제 셀은 class="col_price set_meta_value is_price"처럼
    # 보조 클래스가 붙는다(정확일치 정규식이 0건 매칭하던 원인, 2026-08-03 실측).
    prices: dict[str, int] = {}
    row_re = re.compile(r"<tr[^>]*class=[\"'][^\"']*product_set_row[^\"']*[\"'][^>]*>(.*?)</tr>", flags=re.DOTALL)
    name_re = re.compile(r"<p[^>]*class=[\"'][^\"']*set_item_name[^\"']*[\"'][^>]*>(.*?)</p>", flags=re.DOTALL)
    price_re = re.compile(r"<td[^>]*class=[\"'][^\"']*col_price[^\"']*[\"'][^>]*>(.*?)</td>", flags=re.DOTALL)
    supply_re = re.compile(r"<td[^>]*class=[\"'][^\"']*is_price[^\"']*[\"'][^>]*>(.*?)</td>", flags=re.DOTALL)
    for row in row_re.findall(html):
        name_match = name_re.search(row)
        if not name_match:
            continue
        # 공급가 셀: is_price 표시가 있으면 그것, 없으면 첫 col_price(헤더 순서: 공급가→판매가)
        supply_match = supply_re.search(row)
        price_match = supply_match or price_re.search(row)
        if not price_match:
            continue
        normalized_name = _normalize_option_name(name_match.group(1))
        if not normalized_name:
            continue
        try:
            prices[normalized_name] = _parse_price(price_match.group(1))
        except SupplierMonitorError:
            continue  # 품절 등 비숫자 공급가 옵션은 건너뜀
    if prices:
        return prices

    matches = re.findall(
        r"""<tr[^>]*>\s*<td style='border-left:0px'>(?P<option>.*?)</td>.*?<td style='text-align:right;color:black;font-weight:bold'>(?P<price>.*?)</td>""",
        html,
        flags=re.DOTALL,
    )
    for option_name, price in matches:
        normalized_name = _normalize_option_name(option_name)
        if normalized_name:
            prices[normalized_name] = _parse_price(price)
    return prices


async def _fetch_adminplus_prices(client: httpx.AsyncClient, config: SupplierMonitorConfig, product_code: str) -> dict[str, int]:
    response = await client.get(
        config.product_detail_url,
        params={"mod": "product", "actpage": "prt.grp.detail.pop", "pcode": product_code},
    )
    response.raise_for_status()

    prices = _parse_adminplus_popup_prices(response.text)
    if not prices:
        raise SupplierMonitorError(f"{config.product_name} 옵션 공급가를 읽지 못했습니다.")
    return prices


async def _fetch_google_sheet_prices(
    client: httpx.AsyncClient,
    config: SupplierMonitorConfig,
) -> tuple[dict[str, int], dict[str, int]]:
    if not config.sheet_csv_url or not config.sheet_product_name:
        raise SupplierMonitorError(f"{config.key} monitor is missing google sheet configuration.")

    response = await client.get(config.sheet_csv_url)
    response.raise_for_status()

    rows = list(csv.reader(StringIO(response.text)))
    if not rows:
        raise SupplierMonitorError("구글시트 데이터가 비어 있습니다.")

    product_index = _column_index(config.sheet_product_column)
    option_index = _column_index(config.sheet_option_column)
    vip_index = _column_index(config.sheet_vip_column)
    quantity_index = _column_index(config.sheet_quantity_column) if config.sheet_quantity_column else None
    fallback_index = _column_index(config.sheet_price_fallback_column) if config.sheet_price_fallback_column else None
    previous_index = _column_index(config.sheet_previous_column) if config.sheet_previous_column else None

    prices: dict[str, int] = {}
    previous_prices: dict[str, int] = {}
    current_product = ""
    current_option = ""
    for row in rows:
        required_indexes = [product_index, option_index, vip_index]
        if quantity_index is not None:
            required_indexes.append(quantity_index)
        if fallback_index is not None:
            required_indexes.append(fallback_index)
        if previous_index is not None:
            required_indexes.append(previous_index)
        if max(required_indexes) >= len(row):
            continue
        product_cell = _normalize_option_name(row[product_index])
        option_cell = _normalize_option_name(row[option_index])
        if product_cell:
            current_product = product_cell
        if option_cell:
            current_option = option_cell
        product_name = current_product
        option_name = current_option
        quantity_name = _normalize_option_name(row[quantity_index]) if quantity_index is not None else ""
        vip_price = row[vip_index]
        if fallback_index is not None and not vip_price.strip():
            vip_price = row[fallback_index]
        if config.sheet_product_exclude and any(ex in product_name for ex in config.sheet_product_exclude):
            continue  # '초당옥수수' 필터가 '애플초당옥수수' 등 부분일치로 잘못 잡는 것 방지
        if not _sheet_product_matches(product_name, config.sheet_product_name) or not option_name or not vip_price.strip():
            continue
        option_keys = _google_sheet_option_keys(product_name, option_name, quantity_name)
        if not option_keys:
            continue
        parsed_price = _parse_price(vip_price)
        previous_price = None
        if previous_index is not None and row[previous_index].strip():
            try:
                previous_price = _parse_price(row[previous_index])
            except SupplierMonitorError:
                previous_price = None
        for option_key in option_keys:
            prices[option_key] = parsed_price
            if previous_price is not None:
                previous_prices[option_key] = previous_price

    if not prices:
        if config.skip_missing_options:
            return {}, {}
        raise SupplierMonitorError(f"구글시트에서 {config.sheet_product_name} VIP 공급가를 찾지 못했습니다.")
    return prices, previous_prices


def _load_workbooks(config: SupplierMonitorConfig):
    if not config.template_path.exists():
        raise SupplierMonitorError(f"기준 엑셀 파일이 없습니다: {config.template_path.name}")
    editable_wb = load_workbook(filename=str(config.template_path))
    data_wb = load_workbook(filename=str(config.template_path), data_only=True)
    return editable_wb, data_wb


def _worksheet_xml_paths_from_zip(workbook_zip: ZipFile) -> dict[str, str]:
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))

    relationships = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall(f"{package_rel_ns}Relationship")
    }

    paths: dict[str, str] = {}
    for sheet in workbook_root.findall(f"{main_ns}sheets/{main_ns}sheet"):
        sheet_name = sheet.attrib["name"]
        relationship_id = sheet.attrib[f"{rel_ns}id"]
        target = relationships[relationship_id].lstrip("/")
        paths[sheet_name] = target if target.startswith("xl/") else f"xl/{target}"
    return paths


def _excel_number(value: int | float) -> str:
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value)


_CELL_VALUE_RE = re.compile(r"<v\b[^>]*/>|<v>.*?</v>", flags=re.DOTALL)


def _replace_formula_cell_value(
    sheet_xml: bytes,
    cell_ref: str,
    formula: str,
    value: int | float | None,
) -> bytes:
    xml_text = sheet_xml.decode("utf-8")
    pattern = re.compile(
        rf'(<c\b(?=[^>]*\br="{re.escape(cell_ref)}")[^>]*>)(.*?)(</c>)',
        flags=re.DOTALL,
    )
    formula_text = escape(formula[1:] if formula.startswith("=") else formula, quote=False)

    def replace_cell(match: re.Match[str]) -> str:
        opening_tag = re.sub(r'\s+t="[^"]*"', "", match.group(1))
        body = match.group(2)
        body = re.sub(r"<is\b[^>]*>.*?</is>", "", body, flags=re.DOTALL)
        if re.search(r"<f\b[^>]*>.*?</f>", body, flags=re.DOTALL):
            body = re.sub(r"<f\b[^>]*>.*?</f>", f"<f>{formula_text}</f>", body, count=1, flags=re.DOTALL)
        else:
            body = f"<f>{formula_text}</f>{body}"

        body = _CELL_VALUE_RE.sub("", body)
        if value is not None:
            value_text = _excel_number(value)
            body = f"{body}<v>{value_text}</v>"
        return f"{opening_tag}{body}{match.group(3)}"

    updated_xml, changed_count = pattern.subn(replace_cell, xml_text, count=1)
    if changed_count != 1:
        raise SupplierMonitorError(f"기준 엑셀 템플릿에서 {cell_ref} 셀을 찾지 못했습니다.")
    return updated_xml.encode("utf-8")


def _force_excel_recalculation(workbook_xml: bytes) -> bytes:
    xml_text = workbook_xml.decode("utf-8")

    def set_attr(tag: str, attr: str, value: str) -> str:
        if re.search(rf'\s{attr}="[^"]*"', tag):
            return re.sub(rf'\s{attr}="[^"]*"', f' {attr}="{value}"', tag)
        if tag.endswith("/>"):
            return f'{tag[:-2]} {attr}="{value}"/>'
        return f'{tag[:-1]} {attr}="{value}">'

    def replace_calc_pr(match: re.Match[str]) -> str:
        tag = match.group(0)
        for attr, value in (("calcMode", "auto"), ("fullCalcOnLoad", "1"), ("forceFullCalc", "1")):
            tag = set_attr(tag, attr, value)
        return tag

    updated_xml, changed_count = re.subn(r"<calcPr\b[^>]*/?>", replace_calc_pr, xml_text, count=1)
    if changed_count:
        return updated_xml.encode("utf-8")
    return xml_text.replace(
        "</workbook>",
        '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>',
    ).encode("utf-8")


def _remove_calc_chain_relationships(rels_xml: bytes) -> bytes:
    xml_text = rels_xml.decode("utf-8")
    return re.sub(
        r'<Relationship\b(?=[^>]*\bType="http://schemas\.openxmlformats\.org/officeDocument/2006/relationships/calcChain")[^>]*/>',
        "",
        xml_text,
    ).encode("utf-8")


def _remove_calc_chain_content_type(content_types_xml: bytes) -> bytes:
    xml_text = content_types_xml.decode("utf-8")
    return re.sub(
        r'<Override\b(?=[^>]*\bPartName="/xl/calcChain\.xml")[^>]*/>',
        "",
        xml_text,
    ).encode("utf-8")


def _patch_workbook_values(
    config: SupplierMonitorConfig,
    updates: list[tuple[str, str, int]],
    formula_updates: list[tuple[str, str, str, int | float | None]] | None = None,
) -> bytes:
    workbook = load_workbook(filename=str(config.template_path))
    for sheet_name, cell_ref, value in updates:
        if sheet_name not in workbook.sheetnames:
            raise SupplierMonitorError(f"엑셀 템플릿에서 시트를 찾지 못했습니다: {sheet_name}")
        workbook[sheet_name][cell_ref].value = value

    for sheet_name, cell_ref, formula, _ in formula_updates or []:
        if sheet_name not in workbook.sheetnames:
            raise SupplierMonitorError(f"엑셀 템플릿에서 시트를 찾지 못했습니다: {sheet_name}")
        workbook[sheet_name][cell_ref].value = formula if formula.startswith("=") else f"={formula}"

    if hasattr(workbook, "calculation"):
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    clean_buffer = BytesIO()
    workbook.save(clean_buffer)
    clean_bytes = clean_buffer.getvalue()

    formula_updates_by_path: dict[str, list[tuple[str, str, int | float | None]]] = {}
    with ZipFile(BytesIO(clean_bytes), "r") as clean_zip:
        sheet_paths = _worksheet_xml_paths_from_zip(clean_zip)
    for sheet_name, cell_ref, formula, value in formula_updates or []:
        formula_updates_by_path.setdefault(sheet_paths[sheet_name], []).append((cell_ref, formula, value))

    output_buffer = BytesIO()
    with ZipFile(BytesIO(clean_bytes), "r") as source_zip, ZipFile(output_buffer, "w") as target_zip:
        for item in source_zip.infolist():
            if item.filename == "xl/calcChain.xml":
                continue
            item_data = source_zip.read(item.filename)
            if item.filename in formula_updates_by_path:
                for cell_ref, formula, value in formula_updates_by_path[item.filename]:
                    item_data = _replace_formula_cell_value(item_data, cell_ref, formula, value)
            elif item.filename == "xl/_rels/workbook.xml.rels":
                item_data = _remove_calc_chain_relationships(item_data)
            elif item.filename == "[Content_Types].xml":
                item_data = _remove_calc_chain_content_type(item_data)
            elif item.filename == "xl/workbook.xml":
                item_data = _force_excel_recalculation(item_data)
            target_zip.writestr(item, item_data)

    return output_buffer.getvalue()


async def _collect_all_supplier_prices(
    config: SupplierMonitorConfig,
) -> dict[SupplierSource | None, tuple[dict[str, int], dict[str, int]]]:
    """옵션들이 참조하는 발주처별로 공급가를 수집한다(기본 발주처 = None 키)."""
    sources: list[SupplierSource | None] = []
    for option in config.options:
        if option.source not in sources:
            sources.append(option.source)
    collected: dict[SupplierSource | None, tuple[dict[str, int], dict[str, int]]] = {}
    for source in sources:
        target = source.as_config(config) if source else config
        collected[source] = await _collect_supplier_prices(target)
    return collected


async def _collect_supplier_prices(config: SupplierMonitorConfig) -> tuple[dict[str, int], dict[str, int]]:
    if config.source_type == "manual":
        return {}, {}
    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        if config.source_type == "google_sheet":
            return await _fetch_google_sheet_prices(client, config)

        await _login(client, config)
        product_code = config.product_code or await _find_product_code(client, config)
        return await _fetch_adminplus_prices(client, config, product_code), {}


def _find_monitor_config(key: str) -> SupplierMonitorConfig | None:
    aliases = {key, key.replace("_", "-"), key.replace("-", "_")}
    for alias in aliases:
        if alias in MONITOR_CONFIGS:
            return MONITOR_CONFIGS[alias]
    return None


def is_supplier_monitor_expired(config: SupplierMonitorConfig | None, *, today: str | None = None) -> bool:
    """한시 모니터(active_until)가 만료됐는지. (오늘 > active_until)"""
    if not config or not config.active_until:
        return False
    current = today or datetime.now(KST).date().isoformat()
    return current > config.active_until


def is_supplier_monitor_paused(key: str) -> bool:
    config = _find_monitor_config(key)
    monitor_key = config.key if config else key.replace("_", "-")
    if monitor_key in PAUSED_SUPPLIER_MONITOR_KEYS:
        return True
    # 한시 모니터 만료 시 일시중지와 동일 취급
    return is_supplier_monitor_expired(config)


def active_supplier_monitor_keys() -> tuple[str, ...]:
    return tuple(
        key
        for key, config in MONITOR_CONFIGS.items()
        if key not in PAUSED_SUPPLIER_MONITOR_KEYS and not is_supplier_monitor_expired(config)
    )


async def run_supplier_monitor(key: str) -> tuple[dict, bytes, str]:
    config = _find_monitor_config(key)
    if not config:
        raise SupplierMonitorError(f"지원하지 않는 공급가 모니터입니다: {key}")
    if is_supplier_monitor_expired(config):
        raise SupplierMonitorError(
            f"{config.product_name} 공급가 모니터링은 한시 운영 기간(~{config.active_until})이 만료되었습니다."
        )
    if is_supplier_monitor_paused(config.key):
        raise SupplierMonitorError(f"{config.product_name} 공급가 모니터링은 품절로 일시중지 중입니다.")

    price_sets = await _collect_all_supplier_prices(config)
    editable_wb, data_wb = _load_workbooks(config)
    now = datetime.now(KST)
    run_date = now.date().isoformat()

    from app import db as database

    previous_snapshot_rows = await database.latest_supplier_price_snapshots_before_run_date(
        config.key,
        run_date,
    )
    previous_price_map: dict[tuple[str, str], int] = {}
    for snapshot in previous_snapshot_rows:
        option_name = str(snapshot.get("option_name") or "")
        supplier_option_name = str(snapshot.get("supplier_option_name") or option_name)
        supplier_price = int(snapshot.get("supplier_price") or 0)
        if supplier_price <= 0:
            continue
        for lookup_key in _snapshot_lookup_keys(option_name, supplier_option_name):
            previous_price_map.setdefault(lookup_key, supplier_price)

    rows: list[dict] = []
    missing_options: list[str] = []  # 공급가를 못 찾아 빠진 옵션(조용한 누락 방지)
    workbook_updates: list[tuple[str, str, int]] = []
    formula_updates: list[tuple[str, str, str, int | float | None]] = []
    for option in config.options:
        editable_ws = _select_sheet(editable_wb, option.sheet_name)
        data_ws = _select_sheet(data_wb, option.sheet_name)
        cell_ref = f"{option.cell}{option.row}"
        try:
            workbook_price = _parse_price(data_ws[cell_ref].value)
        except SupplierMonitorError:
            # 공급가가 '예정' 등 비숫자면 skip 허용 모니터는 조용히 건너뜀 (불필요한 오류 방지)
            if config.skip_missing_options:
                logger.info(
                    "Skip non-numeric workbook price for %s: %s (%r)",
                    config.key, option.supplier_option_name, data_ws[cell_ref].value,
                )
                missing_options.append(f"{option.label}(양식 공급가 비숫자)")
                continue
            raise
        supplier_prices, source_previous_prices = price_sets.get(option.source, ({}, {}))
        option_source_type = option.source.source_type if option.source else config.source_type
        supplier_price = _supplier_price_lookup(supplier_prices, option.supplier_option_name)
        # manual 모니터: 외부 수집값이 없으므로 템플릿에 적힌 공급가를 그대로 사용
        if supplier_price is None and option_source_type == "manual":
            supplier_price = workbook_price
        if supplier_price is None:
            if config.skip_missing_options:
                logger.warning(
                    "Skip missing supplier option for %s: %s (%s)",
                    config.key,
                    option.supplier_option_name,
                    option_supplier_name(option, config),
                )
                missing_options.append(f"{option.label}({option_supplier_name(option, config)} 공급가 없음)")
                continue
            raise SupplierMonitorError(f"옵션 공급가를 찾지 못했습니다: {option.supplier_option_name}")

        previous_supplier_price = None
        for lookup_key in _snapshot_lookup_keys(option.label, option.supplier_option_name):
            if lookup_key in previous_price_map:
                previous_supplier_price = previous_price_map[lookup_key]
                break
        if previous_supplier_price is None:
            previous_supplier_price = _supplier_price_lookup(source_previous_prices, option.supplier_option_name)
        if previous_supplier_price is None:
            previous_supplier_price = supplier_price

        signal = _price_signal(previous_supplier_price, supplier_price)
        style = PRICE_SIGNAL_STYLE[signal]
        workbook_updates.append((editable_ws.title, cell_ref, supplier_price))
        formula_updates.extend(
            (editable_ws.title, formula_cell, formula, cached_value)
            for formula_cell, formula, cached_value in _formula_cache_updates(
                editable_ws,
                option.row,
                supplier_price,
            )
        )

        margin = _compute_margin(editable_ws, option.row, supplier_price)

        rows.append(
            {
                "option_name": option.label,
                "supplier_option_name": option.supplier_option_name,
                "supplier_name": option_supplier_name(option, config),
                "sheet_name": option.sheet_name or editable_ws.title,
                "sheet_row": option.row,
                "cell": cell_ref,
                "spreadsheet_price": previous_supplier_price,
                "workbook_price": workbook_price,
                "comparison_basis": "previous_supplier_snapshot",
                "has_previous_snapshot": bool(previous_snapshot_rows),
                "supplier_price": supplier_price,
                "diff": supplier_price - previous_supplier_price,
                "signal": signal,
                "signal_label": style["label"],
                "margin": margin,
                "margin_negative": margin is not None and margin < 0,
            }
        )

    output_filename = config.output_filename(now)
    output_bytes = _patch_workbook_values(config, workbook_updates, formula_updates)
    _persist_output(output_filename, output_bytes)

    summary = {
        "key": config.key,
        "title": "공급가 단가변경 바로 알림",
        "supplier_name": config.supplier_name,
        "product_name": config.product_name,
        "checked_at": now.isoformat(),
        "output_filename": output_filename,
        "total_items": len(rows),
        "changed_items": sum(1 for row in rows if row["diff"] != 0),
        "blue_count": sum(1 for row in rows if row["signal"] == "blue"),
        "red_count": sum(1 for row in rows if row["signal"] == "red"),
        "same_count": sum(1 for row in rows if row["signal"] == "same"),
        "negative_margin_count": sum(1 for row in rows if row.get("margin_negative")),
        "missing_options": missing_options,
        "rows": rows,
    }
    await database.save_supplier_price_snapshot(summary)
    await database.save_supplier_price_monitor_run(config.key, summary)
    return summary, output_bytes, output_filename


async def refresh_supplier_price_snapshots(keys: tuple[str, ...] | None = None) -> dict[str, str]:
    target_keys = keys or active_supplier_monitor_keys()
    results: dict[str, str] = {}
    from app import db as database

    for key in target_keys:
        if is_supplier_monitor_paused(key):
            results[key] = "paused"
            continue
        try:
            await run_supplier_monitor(key)
            results[key] = "ok"
        except Exception as exc:
            logger.warning("Supplier price snapshot refresh failed for %s: %s", key, exc)
            await database.save_supplier_price_monitor_run(key, status="error", error_message=str(exc))
            results[key] = f"error: {exc}"
    return results


async def run_myeongi_monitor() -> tuple[dict, bytes, str]:
    return await run_supplier_monitor("myeongi")


async def run_kolrabi_monitor() -> tuple[dict, bytes, str]:
    return await run_supplier_monitor("kolrabi")


if __name__ == "__main__":
    import asyncio
    import json

    monitor_key = sys.argv[1] if len(sys.argv) > 1 else "myeongi"
    summary, _, _ = asyncio.run(run_supplier_monitor(monitor_key))
    print(json.dumps(summary, ensure_ascii=False))
