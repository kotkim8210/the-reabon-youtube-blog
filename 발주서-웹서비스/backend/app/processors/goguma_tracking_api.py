"""Register goguma tracking numbers to Coupang via API."""

from __future__ import annotations

import asyncio
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.coupang.client import coupang_goguma_client
from app.processors.goguma_order import transform_option
from app.processors.haedal_tracking_parser import (
    detect_haedal_columns,
    find_courier_in_row,
    find_tracking_in_row,
)
from app.processors.tracking_match import (
    name_counts,
    option_key_set,
    options_match,
    requires_option_guard,
)

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

GOGUMA_KEYWORDS = ["고구마", "꿀고구마"]
DELIVERY_COMPANY_CODE = "HANJIN"  # 기본(해달은 보통 한진). 파일에 택배사가 명시되면 그걸 우선.

# 택배사명 → 쿠팡 deliveryCompanyCode. 미인식 시 기본(한진).
_COURIER_NAME_TO_CODE: list[tuple[tuple[str, ...], str]] = [
    (("대한통운", "cj"), "CJGLS"),
    (("한진",), "HANJIN"),
    (("롯데", "현대"), "HYUNDAI"),   # 롯데택배 = 쿠팡 코드 HYUNDAI(구 현대택배)
    (("우체국", "epost"), "EPOST"),
    (("로젠", "logen"), "KGB"),
]


def courier_code(name: object, default: str = DELIVERY_COMPANY_CODE) -> str:
    """택배사명 문자열 → 쿠팡 deliveryCompanyCode. 못 알아보면 기본값."""
    compact = re.sub(r"\s+", "", str(name or "")).lower()
    if not compact:
        return default
    for keys, code in _COURIER_NAME_TO_CODE:
        if any(k in compact for k in keys):
            return code
    return default


def normalize(value: object) -> str:
    """Remove all whitespace for matching."""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def phone_digits(value: object) -> str:
    """전화번호에서 숫자만 추출 (하이픈/공백 형식 차이 무시)"""
    if value is None:
        return ""
    return re.sub(r"\D", "", str(value))


def parse_haedal_file(haedal_bytes: bytes) -> list[dict]:
    """Parse shipping workbook and extract rows with tracking numbers."""
    wb = load_workbook(filename=BytesIO(haedal_bytes), data_only=True)
    ws = wb.active
    cols = detect_haedal_columns(ws)

    entries = []
    missing: list[str] = []
    for row_idx in range(cols.start_row, ws.max_row + 1):
        name = normalize(ws.cell(row=row_idx, column=cols.name).value)
        phone = phone_digits(ws.cell(row=row_idx, column=cols.phone).value)
        address = normalize(ws.cell(row=row_idx, column=cols.address).value)
        product = ws.cell(row=row_idx, column=cols.product).value
        if not name:
            continue

        tracking = find_tracking_in_row(ws, row_idx, cols.tracking)
        if not tracking:
            missing.append(f"{row_idx}행 {name}")
        if tracking:
            skip_cols = tuple(
                c for c in (cols.name, cols.phone, cols.address, cols.product, cols.tracking) if c
            )
            courier_name = find_courier_in_row(ws, row_idx, cols.courier, skip_cols)
            entries.append(
                {
                    "name": name,
                    "phone": phone,
                    "address": address,
                    "tracking": tracking,
                    "delivery_company_code": courier_code(courier_name),
                    "option_keys": option_key_set(product, transform_option(str(product or ""))),
                }
            )

    # 조용한 누락 금지: 수취인은 있는데 송장이 없는 행은 로그로 남긴다.
    if missing:
        logger.warning(
            "해달 회신에 송장 없는 행 %s건 (송장열=%s): %s",
            len(missing),
            cols.tracking,
            ", ".join(missing[:20]),
        )

    return entries


def is_goguma_order(item: dict) -> bool:
    """Check if an order item is a goguma product."""
    seller_name = (item.get("sellerProductName") or "").lower()
    vendor_name = (item.get("vendorItemName") or "").lower()
    return any(keyword in seller_name or keyword in vendor_name for keyword in GOGUMA_KEYWORDS)


async def _fetch_orders_by_status(order_status: str, from_date: str, to_date: str) -> list[dict]:
    rows: list[dict] = []
    next_token = None

    while True:
        result = await coupang_goguma_client.get_order_sheets(
            created_at_from=from_date,
            created_at_to=to_date,
            order_status=order_status,
            max_per_page=50,
            next_token=next_token,
        )
        if not result:
            break

        shipments = result.get("data", [])
        if not shipments:
            break

        for shipment in shipments:
            receiver = shipment.get("receiver") or {}
            shipment_box_id = shipment.get("shipmentBoxId")
            order_id = shipment.get("orderId")

            for item in shipment.get("orderItems", []):
                if not is_goguma_order(item) or not shipment_box_id:
                    continue

                phone_raw = (
                    receiver.get("safeNumber")
                    or receiver.get("receiverPhoneNumber1")
                    or ""
                )
                address_raw = f"{receiver.get('addr1') or ''} {receiver.get('addr2') or ''}".strip()

                rows.append(
                    {
                        "shipment_box_id": shipment_box_id,
                        "order_id": order_id,
                        "vendor_item_id": item.get("vendorItemId"),
                        "name": normalize(receiver.get("name") or ""),
                        "phone": phone_digits(phone_raw),
                        "address": normalize(address_raw),
                        "name_display": receiver.get("name") or "",
                        "order_status": order_status,
                        "option_keys": option_key_set(
                            item.get("vendorItemName"),
                            transform_option(item.get("vendorItemName") or ""),
                        )
                        or option_key_set(item.get("sellerProductName")),
                    }
                )

        next_token = result.get("nextToken")
        if not next_token:
            break

    return rows


async def fetch_pending_orders() -> list[dict]:
    """Fetch ACCEPT + INSTRUCT goguma orders from Coupang (조회만, 상태 변경 없음).

    결제완료(ACCEPT) 주문의 확인 처리(→상품준비중)는 여기서 하지 않는다 —
    해달 운송장과 실제로 매칭된 주문만 process_tracking_api에서 확인 처리한다.
    (매칭 안 된 결제완료 주문까지 넘어가면 고객 즉시취소가 막히는 부작용)
    """
    now = datetime.now(KST)
    from_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    to_date = now.strftime("%Y-%m-%d")

    instruct_orders = await _fetch_orders_by_status("INSTRUCT", from_date, to_date)
    accept_orders = await _fetch_orders_by_status("ACCEPT", from_date, to_date)

    combined: dict[tuple[int, int], dict] = {}
    for row in instruct_orders + accept_orders:
        combined[(row["shipment_box_id"], row["vendor_item_id"])] = row

    all_orders = list(combined.values())
    logger.info(
        "고구마 주문 조회 완료: total=%s, instruct=%s, accept=%s",
        len(all_orders),
        len(instruct_orders),
        len(accept_orders),
    )
    return all_orders


async def process_tracking_api(haedal_bytes: bytes) -> dict:
    """Parse tracking file, match Coupang orders, and register invoices."""
    haedal_entries = parse_haedal_file(haedal_bytes)
    if not haedal_entries:
        raise ValueError("해달 발주서에서 운송장번호를 찾을 수 없습니다.")

    coupang_orders = await fetch_pending_orders()
    if not coupang_orders:
        raise ValueError("결제완료 또는 상품준비중인 고구마 주문이 없습니다.")

    entry_by_name = defaultdict(list)
    entry_by_phone = defaultdict(list)
    for entry in haedal_entries:
        entry_by_name[entry["name"]].append(entry)
        if entry["phone"]:
            entry_by_phone[entry["phone"]].append(entry)

    used_entries = set()
    matched_pairs = []
    matched_via_phone: dict[int, str] = {}  # id(order) -> 해달측 이름 (fallback 표시용)
    results = []
    skip_count = 0
    coupang_name_counts = name_counts(order["name"] for order in coupang_orders)

    for order in coupang_orders:
        candidates = entry_by_name.get(order["name"], [])
        available = [c for c in candidates if id(c) not in used_entries]
        if requires_option_guard(order["name"], coupang_name_counts, len(candidates)):
            available = [
                c for c in available
                if options_match(c.get("option_keys"), order.get("option_keys"))
            ]

        matched = None
        matched_via_phone_only = False

        if available:
            for candidate in available:
                if candidate["phone"] == order["phone"] and candidate["address"] == order["address"]:
                    matched = candidate
                    break
            if matched is None:
                for candidate in available:
                    if candidate["phone"] == order["phone"]:
                        matched = candidate
                        break
            if matched is None:
                for candidate in available:
                    if candidate["address"] == order["address"]:
                        matched = candidate
                        break
            if matched is None:
                matched = available[0]

        # Fallback: 이름 매칭 실패 시 전화번호로 재시도
        # (CS로 수취인/주소가 바뀐 케이스 대응 — 전화번호는 보통 유지됨)
        if matched is None and order["phone"]:
            phone_pool = [
                c for c in entry_by_phone.get(order["phone"], [])
                if id(c) not in used_entries
            ]
            if requires_option_guard(order["name"], coupang_name_counts, len(phone_pool)):
                phone_pool = [
                    c for c in phone_pool
                    if options_match(c.get("option_keys"), order.get("option_keys"))
                ]
            if len(phone_pool) == 1:
                matched = phone_pool[0]
                matched_via_phone_only = True
            elif len(phone_pool) > 1:
                for c in phone_pool:
                    if c["address"] == order["address"]:
                        matched = c
                        matched_via_phone_only = True
                        break

        if matched is None:
            skip_count += 1
            results.append(
                {
                    "order_id": str(order["order_id"]),
                    "name": order["name_display"],
                    "status": "skip",
                    "message": "매칭되는 수령인 없음 (이름·전화번호 모두 불일치)",
                }
            )
            continue

        used_entries.add(id(matched))
        order["delivery_code"] = matched.get("delivery_company_code") or DELIVERY_COMPANY_CODE
        matched_pairs.append((order, matched["tracking"]))
        if matched_via_phone_only:
            matched_via_phone[id(order)] = matched["name"]

    success_count = 0
    fail_count = 0

    def _build_dtos(pairs: list[tuple[dict, str]]) -> list[dict]:
        return [
            {
                "shipmentBoxId": order["shipment_box_id"],
                "orderId": order["order_id"],
                "vendorItemId": order["vendor_item_id"],
                "deliveryCompanyCode": order.get("delivery_code") or DELIVERY_COMPANY_CODE,
                "invoiceNumber": tracking,
            }
            for order, tracking in pairs
        ]

    def _is_accept_related_error(msg: str, code) -> bool:
        if code in ("ORDER_NOT_ACCEPTED", "INVALID_STATUS"):
            return True
        lowered = str(msg or "").lower()
        return (
            "상품준비중" in str(msg)
            or "결제완료" in str(msg)
            or "status" in lowered
            or "state" in lowered
            or "accept" in lowered
        )

    def _extract_response_list(api_result: dict) -> list[dict]:
        """쿠팡 송장업로드 응답에서 결과 항목 리스트를 꺼낸다.

        응답 구조가 버전/케이스별로 달라서 모두 대응:
          - {"data": {"responseList": [...]}}
          - {"data": [...]}                  (data 자체가 리스트)
          - {"responseList": [...]}          (data 래핑 없음)
        """
        if not isinstance(api_result, dict):
            return []
        data = api_result.get("data")
        if isinstance(data, list):
            return [x for x in data if isinstance(x, dict)]
        if isinstance(data, dict):
            rl = data.get("responseList")
            if isinstance(rl, list):
                return [x for x in rl if isinstance(x, dict)]
        rl = api_result.get("responseList")
        if isinstance(rl, list):
            return [x for x in rl if isinstance(x, dict)]
        return []

    def _item_failed(item: dict) -> tuple[bool, str, object]:
        """결과 항목이 '실패'인지 판정. (failed, message, result_code)

        쿠팡은 성공 시 succeed=true 또는 resultCode가 SUCCESS류이거나
        resultMessage가 비어있다. 명확한 실패 신호가 있을 때만 실패로 본다.
        """
        result_code = item.get("resultCode")
        result_message = item.get("resultMessage") or ""
        succeed = item.get("succeed")

        # 명시적 성공 신호
        if succeed is True:
            return (False, "", result_code)
        code_str = str(result_code or "").upper()
        if code_str in ("SUCCESS", "200", "0", "OK"):
            return (False, "", result_code)
        # 명시적 실패 신호 (succeed가 False이거나 실패 메시지/코드가 있음)
        if succeed is False:
            return (True, result_message or code_str or "등록 실패", result_code)
        if result_message and code_str not in ("", "SUCCESS", "200", "0", "OK"):
            return (True, result_message, result_code)
        if code_str and code_str not in ("SUCCESS", "200", "0", "OK"):
            return (True, result_message or code_str, result_code)
        # 신호 없음 → 배치 code가 200이면 성공으로 간주(상위에서 처리)
        return (False, "", result_code)

    async def _upload_and_parse(pairs: list[tuple[dict, str]]):
        dtos = _build_dtos(pairs)
        try:
            api_result = await coupang_goguma_client.upload_invoices(dtos)
            logger.info("쿠팡 송장업로드 API 응답: %s", api_result)
        except Exception as exc:
            logger.error("쿠팡 송장업로드 API 오류: %s", exc)
            return [], [(order, tracking, str(exc), None) for order, tracking in pairs]

        api_result = api_result or {}
        top_code = api_result.get("code")
        response_list = _extract_response_list(api_result)

        # 배치 자체가 거부된 경우 (HTTP 200 envelope이지만 code가 에러)
        batch_ok = str(top_code or "").upper() in ("200", "0", "OK", "SUCCESS") or top_code is None

        # 결과 항목이 없으면: 배치 성공이면 전체 성공, 아니면 전체 실패
        if not response_list:
            if batch_ok:
                return list(pairs), []
            msg = api_result.get("message") or "쿠팡 응답에 결과 목록이 없습니다."
            return [], [(order, tracking, str(msg), top_code) for order, tracking in pairs]

        success_pairs: list[tuple[dict, str]] = []
        fail_pairs: list[tuple[dict, str, str, object]] = []

        # shipmentBoxId 기준 매칭이 가능하면 우선 사용, 안 되면 순서(zip) 기반.
        # 쿠팡은 요청 순서대로 결과를 돌려주므로 zip이 안전한 기본값.
        def _norm_id(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return None

        resp_by_box: dict[int, dict] = {}
        for item in response_list:
            bid = _norm_id(item.get("shipmentBoxId"))
            if bid is not None and bid != 0:
                resp_by_box.setdefault(bid, item)

        use_box_match = len(resp_by_box) >= len(pairs) > 0

        for idx, (order, tracking) in enumerate(pairs):
            item = None
            if use_box_match:
                item = resp_by_box.get(_norm_id(order["shipment_box_id"]))
            if item is None and idx < len(response_list):
                item = response_list[idx]  # 순서 기반 fallback
            if item is None:
                # 결과 항목을 못 찾았지만 배치가 성공이면 성공으로 간주
                if batch_ok:
                    success_pairs.append((order, tracking))
                else:
                    fail_pairs.append((order, tracking, "결과 항목 없음", top_code))
                continue

            failed, message, result_code = _item_failed(item)
            if failed:
                fail_pairs.append((order, tracking, str(message or "등록 실패"), result_code))
            else:
                success_pairs.append((order, tracking))

        return success_pairs, fail_pairs

    if matched_pairs:
        # 매칭된 주문 중 결제완료(ACCEPT)만 확인 처리 — 송장 등록은 상품준비중에서만 가능.
        # 매칭 안 된 결제완료 주문은 건드리지 않는다.
        matched_accept_boxes = list({
            order["shipment_box_id"]
            for order, _ in matched_pairs
            if order.get("order_status") == "ACCEPT"
        })
        if matched_accept_boxes:
            logger.info("매칭된 결제완료 주문 %s건 확인 처리(상품준비중 전환)", len(matched_accept_boxes))
            try:
                confirm_result = await coupang_goguma_client.confirm_orders(matched_accept_boxes)
                code = (confirm_result or {}).get("code")
                if code not in (200, "200", 0, None):
                    logger.warning("쿠팡 주문 확인 응답 이상: %s", confirm_result)
            except Exception as exc:
                logger.error("쿠팡 주문 확인 실패 (송장 등록은 계속 시도): %s", exc)
            await asyncio.sleep(3)

        success_pairs, fail_items = await _upload_and_parse(matched_pairs)

        retry_pairs = [(order, tracking) for order, tracking, msg, code in fail_items if _is_accept_related_error(msg, code)]
        retry_boxes = list({order["shipment_box_id"] for order, _ in retry_pairs})
        if retry_pairs:
            logger.info("상태 전이 의심 %s건 재시도 시작", len(retry_pairs))
            try:
                confirm_result = await coupang_goguma_client.confirm_orders(retry_boxes)
                logger.info("재시도용 주문 확인 응답: %s", confirm_result)
            except Exception as exc:
                logger.error("재시도용 주문 확인 실패: %s", exc)
            await asyncio.sleep(3)

            retry_success, retry_fail = await _upload_and_parse(retry_pairs)
            success_pairs.extend(retry_success)
            fail_items = [
                item for item in fail_items if not _is_accept_related_error(item[2], item[3])
            ] + retry_fail

        for order, tracking in success_pairs:
            success_count += 1
            haedal_name = matched_via_phone.get(id(order))
            if haedal_name:
                message = f"운송장 등록 완료 (전화번호 매칭: 해달 '{haedal_name}' → 쿠팡 '{order['name_display']}')"
            else:
                message = "운송장 등록 완료"
            results.append(
                {
                    "order_id": str(order["order_id"]),
                    "name": order["name_display"],
                    "tracking": tracking,
                    "status": "success",
                    "message": message,
                }
            )
        for order, tracking, message, _ in fail_items:
            fail_count += 1
            results.append(
                {
                    "order_id": str(order["order_id"]),
                    "name": order["name_display"],
                    "tracking": tracking,
                    "status": "fail",
                    "message": message,
                }
            )

    return {
        "haedal_entries": len(haedal_entries),
        "coupang_orders": len(coupang_orders),
        "success": success_count,
        "fail": fail_count,
        "skip": skip_count,
        "results": results,
    }
