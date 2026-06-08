import logging
import re
from hashlib import sha1
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR

logger = logging.getLogger(__name__)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def stable_order_id(prefix: str, *values) -> str:
    raw = "|".join(normalize(value) for value in values if normalize(value))
    if not raw:
        return ""
    return f"{prefix}:{sha1(raw.encode('utf-8')).hexdigest()[:16]}"


def toss_order_id(item: dict) -> str:
    for key in ("orderNo", "orderId", "orderNumber", "orderSheetId", "paymentKey"):
        value = item.get(key)
        if value:
            return str(value)
    return ""


def transform_option(option_text: str) -> str:
    """Transform option text for 고구마 orders.

    Remove '1박스 황금 ' prefix if present.
    """
    text = normalize(option_text)
    # Remove "1박스 황금 " prefix
    text = re.sub(r"^1박스\s*황금\s*", "", text)
    return text


def transform_alwayz_option(option_text: str) -> str:
    """Transform 올웨이즈 option text for 고구마 orders.

    Example: '0. 해남 황금 꿀고구마: 5Kg 중상[추천]' -> '꿀고구마 5Kg (중상)'
             '0. 해남 황금 꿀고구마: 2Kg 한입' -> '꿀고구마 2Kg (한입)'
    """
    text = normalize(option_text)
    # Pattern: "숫자. 해남 황금 꿀고구마: {weight} {grade}[추천]"
    m = re.match(
        r"^\d+\.\s*해남\s*황금\s*꿀고구마\s*:\s*(\d+\s*[Kk]g)\s*(.+)$", text
    )
    if m:
        weight = m.group(1).strip()
        grade = m.group(2).strip()
        # Remove [추천] suffix
        grade = re.sub(r"\[추천\]$", "", grade).strip()
        return f"꿀고구마 {weight} ({grade})"
    return text


def canonical_goguma_option(*values: str) -> str:
    """Normalize marketplace 고구마 option text to the 해달 발주서 품목명."""
    texts = [normalize(value) for value in values if normalize(value)]
    text = " ".join(texts)
    if not text:
        return ""

    compact = re.sub(r"\s+", "", text).lower()
    weight_match = re.search(r"(10|5|3|2)\s*(?:kg|키로)", text, re.IGNORECASE)
    weight = weight_match.group(1) if weight_match else ""

    grade = ""
    grade_patterns = [
        ("중상", r"중\s*상"),
        ("특상", r"특\s*상"),
        ("한입", r"한\s*입|꼬마|미니"),
    ]
    for label, pattern in grade_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            grade = label
            break

    if not grade:
        # '대'는 상품명 안에 자주 들어가는 조사/단어와 섞이므로 무게 주변의 등급 표기만 인정한다.
        large_patterns = [
            r"\(\s*대\s*\)",
            r"(?:^|[\s:/_\-\[\]])대(?:$|[\s:/_\-\]\[])",
            r"(?:kg|키로)\s*대(?:$|[\s:/_\-\]\[])",
        ]
        if any(re.search(pattern, text, re.IGNORECASE) for pattern in large_patterns):
            grade = "대"

    if weight and grade:
        return f"꿀고구마 {weight}Kg ({grade})"

    # Fallback: keep legacy cleanup, but remove noisy market prefixes.
    cleaned = re.sub(r"^\d+\.\s*", "", text)
    cleaned = re.sub(r"^1박스\s*황금\s*", "", cleaned)
    cleaned = re.sub(r"\[추천\]$", "", cleaned).strip()
    return cleaned


def build_alwayz_memo(delivery_method: str, door_password: str) -> str:
    """Build memo from 올웨이즈 수령방법(R열) + 공동현관비밀번호(Q열).

    Example: R='문 앞', Q='#5165' -> '문앞(공동현관비번#5165)'
             R='문 앞', Q=''      -> '문앞'
    """
    method = normalize(delivery_method).replace(" ", "")
    pw = normalize(door_password)
    if pw:
        return f"{method}(공동현관비번{pw})"
    return method


GOGUMA_KEYWORDS = ["고구마", "꿀고구마"]

HAEDAL_HEADERS = [
    "받으시는 분",
    "받으시는 분 전화",
    "받는분담당자(선택)",
    "받는분핸드폰(선택)",
    "받는분우편번호(선택)",
    "받는분총주소",
    "보내시는 분",
    "보내시는 분 전화",
    "보내는분담당자(선택)",
    "보내는분담당자HP(선택)",
    "보내는분우편번호(선택)",
    "보내는분총주소",
    "수량",
    "품목명",
    "운임Type",
    "지불조건",
    "출고번호",
    "특기사항",
    "메모1",
    "메모2",
    "메모3",
    "메모4",
]

HAEDAL_COLUMN_WIDTHS = {
    "A": 18,
    "B": 27,
    "C": 18,
    "D": 18,
    "E": 21,
    "F": 18,
    "G": 13,
    "H": 27,
    "I": 21,
    "J": 27,
    "K": 24,
    "L": 21,
    "M": 10,
    "N": 10,
    "O": 18,
    "P": 12,
    "Q": 13,
    "R": 13,
    "S": 10,
    "T": 13,
    "U": 13,
    "V": 13,
}


def ensure_haedal_header(ws) -> None:
    """해달 한진 발주서 1행이 다른 양식에 의해 깨져도 출력 직전에 복구한다."""
    header_font = Font(size=11, bold=True)
    for col, value in enumerate(HAEDAL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = header_font
    for column_letter, width in HAEDAL_COLUMN_WIDTHS.items():
        ws.column_dimensions[column_letter].width = width


def is_goguma_order(item: dict) -> bool:
    """Check if a Toss order item is a 고구마 product."""
    product_name = (item.get("productName") or "").lower()
    option_name = (item.get("optionName") or "").lower()
    for kw in GOGUMA_KEYWORDS:
        if kw in product_name or kw in option_name:
            return True
    return False


def transform_toss_option(product_name: str, option_name: str) -> str:
    """Transform Toss option text for 고구마 orders.

    Uses optionName if available, falls back to productName.
    Applies same normalization as other sources.
    """
    # Toss optionName is usually the selected option, so prefer it over the listing title.
    return canonical_goguma_option(option_name, product_name)


def parse_toss_excel(file_bytes: bytes) -> list[dict]:
    """Parse 토스 주문 엑셀 다운로드 파일.

    열 매핑:
      F(6): 수령인명, G(7): 주소, H(8): 우편번호,
      I(9): 수령인 연락처, N(14): 옵션, Q(17): 수량
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=3):  # 1행=설명, 2행=헤더, 3행부터 데이터
        name = normalize(row[5].value) if len(row) > 5 else ""  # F = index 5
        if not name:
            continue
        zipcode = normalize(row[7].value) if len(row) > 7 else ""  # H = index 7
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = str(zipcode).zfill(5)

        option = normalize(row[13].value) if len(row) > 13 else ""  # N = index 13
        product = transform_toss_option("", option)
        order_id = normalize(row[0].value) if len(row) > 0 else ""

        entries.append({
            "name": name,
            "phone": normalize(row[8].value) if len(row) > 8 else "",   # I = index 8
            "zipcode": zipcode,
            "address": normalize(row[6].value) if len(row) > 6 else "",  # G = index 6
            "qty": normalize(row[16].value) if len(row) > 16 else "",    # Q = index 16
            "product": product,
            "memo": "",
            "order_id": order_id or stable_order_id("toss-excel", name, row[8].value, row[6].value, product, row[16].value),
        })

    logger.info(f"토스 엑셀 파싱: {len(entries)}건")
    return entries


async def collect_toss_orders(from_date: str, to_date: str) -> list[dict]:
    """Collect 고구마 orders from Toss Shopping API.

    Args:
        from_date: Start date in YYYY-MM-DD format.
        to_date: End date in YYYY-MM-DD format.

    Returns:
        List of order entry dicts.
    """
    from app.toss.client import toss_client

    orders = await toss_client.get_orders(
        start_date=from_date,
        end_date=to_date,
        status="PAID",
    )

    entries = []
    for item in orders:
        if not is_goguma_order(item):
            continue

        address = item.get("address") or ""
        detail_addr = item.get("detailAddress") or ""
        full_address = f"{address} {detail_addr}".strip()

        zipcode = item.get("zipCode") or ""
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = str(zipcode).zfill(5)

        product = transform_toss_option(
            item.get("productName") or "",
            item.get("optionName") or "",
        )

        entries.append({
            "name": item.get("receiverName") or "",
            "phone": item.get("receiverRealPhone") or item.get("receiverPhone") or "",
            "zipcode": zipcode,
            "address": full_address,
            "qty": str(item.get("quantity") or 1),
            "product": product,
            "memo": item.get("shippingNote") or "",
            "order_id": toss_order_id(item) or stable_order_id(
                "toss-api",
                item.get("receiverName"),
                item.get("receiverRealPhone") or item.get("receiverPhone"),
                full_address,
                product,
                item.get("quantity") or 1,
            ),
        })

    logger.info(f"토스 고구마 주문 수집: {len(entries)}건 ({from_date} ~ {to_date})")
    return entries


def process(
    delivery_file_bytes: bytes,
    template_file_bytes: bytes = None,
    alwayz_file_bytes: bytes = None,
    toss_entries: list[dict] | None = None,
    toss_file_bytes: bytes = None,
) -> tuple[bytes, str, dict]:
    """Process delivery list for 고구마 orders.

    Args:
        delivery_file_bytes: Raw bytes of the DeliveryList Excel file.
        template_file_bytes: Optional raw bytes of the template file.
            If None, loads from TEMPLATE_DIR / "해달_발주서_한진양식.xlsx".
        alwayz_file_bytes: Optional raw bytes of the 올웨이즈 주문내역 Excel file.
        toss_entries: Optional list of Toss order entries (pre-fetched from API).
        toss_file_bytes: Optional raw bytes of the 토스 주문 엑셀 다운로드 file.

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    if toss_entries is None:
        toss_entries = []
    # 토스 엑셀 파일이 있으면 파싱하여 toss_entries에 추가
    if toss_file_bytes:
        toss_excel_entries = parse_toss_excel(toss_file_bytes)
        toss_entries = toss_entries + toss_excel_entries
    # Load delivery list
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    # Extract Coupang sweet-potato rows only. Some bulk workflows upload a
    # mixed DeliveryList, so do not let unrelated products leak into this file.
    filtered_rows = []
    for row in dl_ws.iter_rows(min_row=2):
        name_val = normalize(row[26].value) if len(row) > 26 else ""  # AA = index 26
        product_name = normalize(row[10].value) if len(row) > 10 else ""  # K = index 10
        if name_val and "고구마" in product_name:
            filtered_rows.append(row)

    # Parse 올웨이즈 orders if provided
    alwayz_entries = []
    if alwayz_file_bytes:
        az_wb = load_workbook(filename=BytesIO(alwayz_file_bytes), data_only=True)
        az_ws = az_wb.active
        for az_row in az_ws.iter_rows(min_row=2):
            name = normalize(az_row[18].value) if len(az_row) > 18 else ""  # S(19)=idx18
            if not name:
                continue
            phone = normalize(az_row[19].value) if len(az_row) > 19 else ""  # T(20)=idx19
            zipcode = normalize(az_row[15].value) if len(az_row) > 15 else ""  # P(16)=idx15
            address = normalize(az_row[14].value) if len(az_row) > 14 else ""  # O(15)=idx14
            qty = normalize(az_row[6].value) if len(az_row) > 6 else ""  # G(7)=idx6
            option = normalize(az_row[5].value) if len(az_row) > 5 else ""  # F(6)=idx5
            door_pw = normalize(az_row[16].value) if len(az_row) > 16 else ""  # Q(17)=idx16
            recv_method = normalize(az_row[17].value) if len(az_row) > 17 else ""  # R(18)=idx17
            order_id = normalize(az_row[0].value) if len(az_row) > 0 else ""

            product = transform_alwayz_option(option)
            memo = build_alwayz_memo(recv_method, door_pw)

            if zipcode:
                try:
                    zipcode = str(int(float(zipcode))).zfill(5)
                except (ValueError, TypeError):
                    zipcode = zipcode.zfill(5)

            alwayz_entries.append({
                "name": name,
                "phone": phone,
                "zipcode": zipcode,
                "address": address,
                "qty": qty,
                "product": product,
                "memo": memo,
                "order_id": order_id or stable_order_id("alwayz", name, phone, address, product, qty),
            })

    # Load template
    if template_file_bytes is not None:
        tmpl_wb = load_workbook(filename=BytesIO(template_file_bytes))
    else:
        template_path = TEMPLATE_DIR / "해달_발주서_한진양식.xlsx"
        if not template_path.exists():
            raise FileNotFoundError(
                f"Template not found: {template_path}. "
                f"Please place the template file in the templates directory."
            )
        tmpl_wb = load_workbook(filename=str(template_path))

    # Keep only first sheet
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]
    ensure_haedal_header(ws)

    font11 = Font(size=11)

    # Find first empty row in template (look at column 1 for name)
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}

    # Write DeliveryList rows
    row_idx = start_row
    for row in filtered_rows:
        # Source columns (0-indexed):
        # AA=26 -> col 1 (name)
        # AB=27 -> col 2 (phone)
        # AC=28 -> col 5 (zipcode, zfill 5)
        # AD=29 -> col 6 (address)
        # W=22  -> col 13 (qty)
        # L=11  -> col 14 (product, transformed)
        # AE=30 -> col 19 (memo)

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        zipcode = normalize(row[28].value) if len(row) > 28 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)
        option = normalize(row[11].value) if len(row) > 11 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        product = transform_option(option)
        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        # Zipcode with zero-fill to 5 digits
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = zipcode.zfill(5)

        mapping = {
            1: name,                                     # name
            2: phone,                                    # phone
            5: zipcode,                                  # zipcode
            6: address,                                  # address
            7: "식품애착",                                 # fixed sender
            8: "010-5700-7756",                          # fixed phone
            12: "전라남도 해남군 산이면 새상골길 ",           # fixed origin address
            13: qty,                                     # quantity
            14: product,                                 # product (transformed)
            16: "선불",                                   # prepaid
            19: memo,                                    # memo
        }

        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11
        row_idx += 1

        bucket = option_totals.setdefault(
            option or product or "고구마",
            {
                "coupang_option_keyword": option or product or "고구마",
                "vendor_option_name": product or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    # Write 올웨이즈 rows
    for entry in alwayz_entries:
        mapping = {
            1: entry["name"],                            # A: 받으시는 분
            2: entry["phone"],                           # B: 받으시는 분 전화
            5: entry["zipcode"],                         # E: 우편번호
            6: entry["address"],                         # F: 주소
            7: "식품애착",                                 # G: fixed sender
            8: "010-5700-7756",                          # H: fixed phone
            12: "전라남도 해남군 산이면 새상골길 ",           # L: fixed origin address
            13: entry["qty"],                            # M: 수량
            14: entry["product"],                        # N: 품목명
            16: "선불",                                   # P: 지불조건
            19: entry["memo"],                           # S: 메모1
        }

        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11
        row_idx += 1

        try:
            qty_int = int(float(entry["qty"])) if entry["qty"] not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        bucket = option_totals.setdefault(
            entry["product"] or "고구마",
            {
                "coupang_option_keyword": entry["product"] or "고구마",
                "vendor_option_name": entry["product"] or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    # Write 토스 rows
    for entry in toss_entries:
        mapping = {
            1: entry["name"],
            2: entry["phone"],
            5: entry["zipcode"],
            6: entry["address"],
            7: "식품애착",
            8: "010-5700-7756",
            12: "전라남도 해남군 산이면 새상골길 ",
            13: entry["qty"],
            14: entry["product"],
            16: "선불",
            19: entry["memo"],
        }

        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11
        row_idx += 1

        try:
            qty_int = int(float(entry["qty"])) if entry["qty"] not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        bucket = option_totals.setdefault(
            entry["product"] or "고구마",
            {
                "coupang_option_keyword": entry["product"] or "고구마",
                "vendor_option_name": entry["product"] or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    # Save to bytes
    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    total = len(filtered_rows) + len(alwayz_entries) + len(toss_entries)
    filename = f"해달 발주서 한진양식_알제이시스템즈({now.strftime('%y%m%d')}).xlsx"
    stats = {
        "total": total,
        "coupang": len(filtered_rows),
        "alwayz": len(alwayz_entries),
        "toss": len(toss_entries),
        "product": "고구마",
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats
