"""사방넷 주문 → 쿠팡 DeliveryList 호환 엑셀 변환 + 회신(orderlist) 송장 파싱.

기존 myeongi(쥬얼리)/kolrabi(제주다팜) 발주 파이프라인은 쿠팡 DeliveryList 엑셀의
고정 컬럼을 읽는다. 사방넷 수집 주문을 그 컬럼 배치 그대로 합성하면
발주서 생성·발주이력 중복방지(filter_delivery_by_issued: C열 주문번호, AA열 수취인)를
코드 수정 없이 재사용할 수 있다.

컬럼 매핑 (myeongi_order.process / issued_orders 기준):
- C(3)  주문번호   ← 사방넷 IDX  (발주서 M열로 전달 → 회신 orderlist D열로 돌아옴 → 송장전송 키)
- K(11) 상품명     ← PRODUCT_NAME (없으면 P_PRODUCT_NAME)
- L(12) 옵션       ← SKU_VALUE (없으면 P_SKU_VALUE)
- W(23) 수량       ← SALE_CNT (없으면 P_EA, 기본 1)
- AA(27) 수취인명  ← RECEIVE_NAME
- AB(28) 전화      ← RECEIVE_CEL (없으면 RECEIVE_TEL)
- AD(30) 주소      ← RECEIVE_ADDR
- AE(31) 배송메모  ← DELV_MSG
"""

import logging
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

_LAST_COL = 31  # AE열까지 확보 (iter_rows가 row[30]까지 돌려주도록)

_HEADERS = {
    1: "번호",
    3: "주문번호",
    11: "노출상품명(옵션명)",
    12: "등록옵션명",
    23: "구매수(수량)",
    27: "수취인이름",
    28: "수취인전화번호",
    30: "수취인 주소",
    31: "배송메세지",
}


def _pick(order: dict, *keys: str) -> str:
    for key in keys:
        val = str(order.get(key) or "").strip()
        if val:
            return val
    return ""


def orders_to_delivery_xlsx(orders: list[dict]) -> bytes:
    """사방넷 주문 목록(소문자 키 dict)을 DeliveryList 호환 엑셀 bytes로 변환."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DeliveryList"

    for col in range(1, _LAST_COL + 1):
        ws.cell(row=1, column=col, value=_HEADERS.get(col, f"col{col}"))

    for i, order in enumerate(orders):
        r = i + 2
        qty = _pick(order, "sale_cnt", "p_ea") or "1"
        row_map = {
            1: i + 1,
            3: _pick(order, "idx"),
            11: _pick(order, "product_name", "p_product_name"),
            12: _pick(order, "sku_value", "p_sku_value"),
            23: qty,
            27: _pick(order, "receive_name"),
            28: _pick(order, "receive_cel", "receive_tel"),
            30: _pick(order, "receive_addr"),
            31: _pick(order, "delv_msg"),
        }
        for col in range(1, _LAST_COL + 1):
            ws.cell(row=r, column=col, value=row_map.get(col, ""))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_lotte_invoice(invoice: str) -> bool:
    """롯데택배 운송장 체크섬: 12자리, 앞 11자리 정수 % 7 == 마지막 자리."""
    return (
        len(invoice) == 12
        and invoice.isdigit()
        and int(invoice[:11]) % 7 == int(invoice[11])
    )


def summarize_courier_codes(orders: list[dict]) -> list[dict]:
    """출고완료 주문들의 DELIVERY_ID(택배사코드)를 집계하고 송장 패턴으로 택배사를 추정.

    반환: [{code, count, sample, guess}] — 건수 많은 순.
    guess는 롯데 체크섬 통과율 70% 이상일 때만 '롯데택배(추정)'.
    """
    groups: dict[str, list[str]] = {}
    for order in orders:
        code = str(order.get("delivery_id") or "").strip()
        invoice = re.sub(r"\D", "", str(order.get("invoice_no") or ""))
        if code:
            groups.setdefault(code, []).append(invoice)

    result = []
    for code, invoices in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        valid = [v for v in invoices if v]
        lotte_hits = sum(1 for v in valid if _is_lotte_invoice(v))
        guess = ""
        if len(valid) >= 2 and lotte_hits / len(valid) >= 0.7:
            guess = "롯데택배(추정)"
        result.append({
            "code": code,
            "count": len(invoices),
            "sample": valid[0] if valid else "",
            "guess": guess,
        })
    return result


def parse_orderlist_for_sabang(orderlist_bytes: bytes) -> list[dict]:
    """거래처 회신(orderlist)에서 (사방넷 IDX, 운송장번호) 추출.

    orderlist 형식(쥬얼리/제주다팜 회신, myeongi_tracking과 동일 좌표):
    - D열(row[3]) = 주문번호 → 사방넷 수집 발주서라면 사방넷 IDX
    - R열(row[17]) = 운송장번호 (숫자만 남김)
    """
    wb = load_workbook(filename=BytesIO(orderlist_bytes), data_only=True)
    ws = wb.active

    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2):
        idx = str(row[3].value or "").strip() if len(row) > 3 else ""
        tracking_raw = str(row[17].value or "").strip() if len(row) > 17 else ""
        name = str(row[2].value or "").strip() if len(row) > 2 else ""
        tracking = re.sub(r"\D", "", tracking_raw)
        if not idx or not tracking:
            continue
        # 엑셀 숫자 셀이 12345.0으로 읽히는 경우 정리
        if idx.endswith(".0"):
            idx = idx[:-2]
        entries.append({"idx": idx, "tracking": tracking, "name": name})
    return entries
