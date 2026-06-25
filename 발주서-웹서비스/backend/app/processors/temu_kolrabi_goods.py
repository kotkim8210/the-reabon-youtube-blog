"""Generate Temu bulk goods registration workbook for kolrabi."""

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.config import TEMPLATE_DIR
from app.processors.kolrabi_order import KST


TEMPLATE_NAME = "temu_kolrabi_goods_template.xlsx"

KOLRABI_GOODS_ROWS = [
    {
        "row": 5,
        "category": 43690,
        "product_type": "일반 제품",
        "product_name": "제주 콜라비 2kg",
        "out_goods_sn": "KOLRABI-2KG",
        "out_sku_sn": "KOLRABI-2KG-01",
    },
    {
        "row": 6,
        "category": 43690,
        "product_type": "일반 제품",
        "product_name": "제주 콜라비 3kg",
        "out_goods_sn": "KOLRABI-3KG",
        "out_sku_sn": "KOLRABI-3KG-01",
    },
    {
        "row": 7,
        "category": 43690,
        "product_type": "일반 제품",
        "product_name": "제주 콜라비 5kg",
        "out_goods_sn": "KOLRABI-5KG",
        "out_sku_sn": "KOLRABI-5KG-01",
    },
]


def generate_workbook() -> tuple[bytes, str, dict]:
    template_path = TEMPLATE_DIR / TEMPLATE_NAME
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(filename=str(template_path))
    if "Template" not in wb.sheetnames:
        raise ValueError("테무 콜라비 템플릿에서 Template 시트를 찾을 수 없습니다.")

    ws = wb["Template"]
    for item in KOLRABI_GOODS_ROWS:
        row = item["row"]
        ws.cell(row=row, column=5, value=item["category"])
        # Column 6 has Temu's category-name lookup formula; keep it intact.
        ws.cell(row=row, column=7, value=item["product_type"])
        ws.cell(row=row, column=12, value=item["product_name"])
        ws.cell(row=row, column=13, value=item["out_goods_sn"])
        ws.cell(row=row, column=14, value=item["out_sku_sn"])
        ws.cell(row=row, column=15, value="Add")
        ws.cell(row=row, column=20, value="←상품 상세설명 입력")
        ws.cell(row=row, column=21, value="←핵심설명1")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now(KST).strftime("%Y%m%d")
    filename = f"TEMU_콜라비_대량상품등록_작성본_{today}.xlsx"
    stats = {
        "status": "ok",
        "product": "콜라비",
        "platform": "Temu",
        "total": len(KOLRABI_GOODS_ROWS),
        "options": ["2kg", "3kg", "5kg"],
        "template": TEMPLATE_NAME,
    }
    return output.read(), filename, stats
