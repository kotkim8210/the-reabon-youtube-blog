"""테무 송장 대량등록.

입력:
  - 테무 주문목록 export (order_export_*.xlsx, 'Order report' 시트, 헤더에 '주문 ID'/'주문 상품 ID')
  - 택배 송장파일 1~2개 (수령인/주문번호 + 송장번호 포함, 헤더 자동 감지)

처리:
  - 주문목록 각 행(=주문 상품)에 대해 송장파일에서 운송장번호를 찾아 매칭
    1) 주문번호(=주문 ID) 일치 우선
    2) 실패 시 수령인 이름 일치(동명이인은 전화/주소로 구분)
  - 배송사는 kg로 자동 결정: 2/3/5kg → 롯데택배, 6/7/8kg → Hanjin
  - 배송 출발지 = 식품애착

출력:
  - 테무 '배송 확인' 업로드 템플릿(주문 ID / 주문 상품 id / 수량 / 배송 출발지: / 배송사 / 추적 번호)
"""

import csv
import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO, StringIO
from zipfile import BadZipFile

from openpyxl import load_workbook

from app.config import TEMPLATE_DIR

KST = timezone(timedelta(hours=9))

TEMPLATE_NAME = "테무_배송확인_템플릿.xlsx"
SHIP_SHEET = "배송 확인"
SHIP_FROM = "식품애착"           # 배송 출발지(테무 배송설정 등록 이름)
CARRIER_SMALL = "롯데택배"       # 2/3/5kg (쥬얼리프룻)
CARRIER_LARGE = "롯데택배"       # 6/7/8kg — 2026-06부터 제이비티→쥬얼리프룻 발주 전환, 기본 롯데
                                  # (실제 송장파일에 택배사가 있으면 그 값이 우선 적용됨)
SMALL_KGS = {2, 3, 5}
LARGE_KGS = {6, 7, 8}


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v).strip())


def _text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def _digits(v) -> str:
    return re.sub(r"\D", "", str(v)) if v is not None else ""


def _kg_from_text(text) -> int | None:
    if not text:
        return None
    for m in re.finditer(r"(\d+(?:\.\d+)?)\s*kg", str(text), re.IGNORECASE):
        try:
            kg = float(m.group(1))
        except ValueError:
            continue
        if kg.is_integer():
            return int(kg)
    return None


def _carrier_for_kg(kg) -> str | None:
    if kg in SMALL_KGS:
        return CARRIER_SMALL
    if kg in LARGE_KGS:
        return CARRIER_LARGE
    return None


def _temu_carrier_name(value) -> str:
    text = _text(value)
    compact = _norm(text).lower()
    if not compact:
        return ""
    if "한진" in compact or "hanjin" in compact:
        return CARRIER_LARGE
    if "롯데" in compact or "lotte" in compact:
        return CARRIER_SMALL
    if "cj" in compact or "대한통운" in compact:
        return "CJ 대한통운"
    return text


def _to_qty(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------------------------
# 테무 주문목록 파싱
# ----------------------------------------------------------------------------
def _decode_csv_text(csv_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp949", "euc-kr"):
        try:
            return csv_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("테무 주문목록 CSV 파일 인코딩을 읽지 못했습니다. UTF-8 또는 CP949 CSV로 저장해 다시 올려주세요.")


def _read_temu_export_rows(order_bytes: bytes, filename: str = "") -> list[tuple]:
    if filename.lower().endswith(".csv"):
        text = _decode_csv_text(order_bytes)
        return [tuple(row) for row in csv.reader(StringIO(text)) if any(_text(value) for value in row)]

    try:
        wb = load_workbook(filename=BytesIO(order_bytes), data_only=True, read_only=True)
    except BadZipFile:
        text = _decode_csv_text(order_bytes)
        return [tuple(row) for row in csv.reader(StringIO(text)) if any(_text(value) for value in row)]

    sheet = "Order report" if "Order report" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _parse_temu_export(order_bytes: bytes, filename: str = "") -> list[dict]:
    rows = _read_temu_export_rows(order_bytes, filename)

    # 헤더 행 탐색 ('주문 ID' + '주문 상품 ID' 동시 포함)
    header_idx = None
    for i, row in enumerate(rows[:20]):
        labels = {_norm(c) for c in row if c is not None}
        if "주문ID" in labels and "주문상품ID" in labels:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "테무 주문목록(order_export) 형식이 아닙니다. "
            "'주문 ID' / '주문 상품 ID' 헤더가 있는 엑셀 또는 CSV 파일을 올려주세요."
        )

    header = rows[header_idx]
    idx: dict[str, int] = {}
    for j, v in enumerate(header):
        key = _norm(v)
        if key and key not in idx:
            idx[key] = j

    def g(row, *labels):
        for label in labels:
            j = idx.get(_norm(label))
            if j is not None and j < len(row):
                val = row[j]
                if val is not None and str(val).strip() != "":
                    return val
        return None

    orders: list[dict] = []
    for row in rows[header_idx + 1:]:
        order_id = _norm(g(row, "주문 ID"))
        item_id = _norm(g(row, "주문 상품 ID"))
        if not order_id and not item_id:
            continue

        status_text = f"{_norm(g(row, '주문 상태'))} {_norm(g(row, '주문 상품 상태'))}"
        if any(keyword in status_text for keyword in ("취소", "환불", "반품")):
            continue

        qty = _to_qty(g(row, "발송할 수량")) or 0
        if qty <= 0:
            continue

        name_full = _norm(g(row, "수령인 이름"))
        sung = _norm(g(row, "수령인 성"))
        name_cands = set()
        if name_full:
            name_cands.add(name_full)
        if sung:
            name_cands.add(_norm(sung + name_full))
            name_cands.add(_norm(name_full + sung))

        option = g(row, "선택 사항") or ""
        product = g(row, "제품 이름", "고객 주문별 제품 이름") or ""
        kg = _kg_from_text(option) or _kg_from_text(product)

        addr = " ".join(
            _norm(g(row, a)) for a in ("배송 주소 1", "배송 주소 2", "배송 주소 3") if g(row, a)
        )

        orders.append(
            {
                "order_id": order_id,
                "item_id": item_id or order_id,
                "qty": qty,
                "name": name_full or _norm(sung + name_full),
                "name_cands": {n for n in name_cands if n},
                "phone": _digits(g(row, "수령인 전화번호")),
                "address": _norm(addr),
                "kg": kg,
                "option": str(option),
            }
        )
    return orders


# ----------------------------------------------------------------------------
# 택배 송장파일 파싱 (헤더 자동 감지)
# ----------------------------------------------------------------------------
def _parse_tracking_files(tracking_files: list[tuple[str, bytes]]):
    by_order: dict[str, dict] = {}
    by_name: dict[str, list[dict]] = defaultdict(list)
    total = 0

    for _fname, data in tracking_files:
        if not data:
            continue
        wb = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            rows = list(wb[sheet].iter_rows(values_only=True))
            if not rows:
                continue

            header_i = None
            for i, row in enumerate(rows[:10]):
                joined = " ".join(_norm(c) for c in row if c is not None)
                if any(k in joined for k in ("송장", "운송장", "추적", "출고번호")) and any(
                    k in joined for k in ("수령인", "이름", "성함", "주문", "받는")
                ):
                    header_i = i
                    break
            if header_i is None:
                continue

            col: dict[str, int] = {}
            for j, v in enumerate(rows[header_i]):
                key = _norm(v)
                if not key:
                    continue
                if "tracking" not in col and any(k in key for k in ("송장", "운송장", "추적", "출고번호")) and "주문" not in key:
                    col["tracking"] = j
                if "order" not in col and any(k in key for k in ("주문번호", "주문id", "orderid", "orderno")):
                    col["order"] = j
                if "courier" not in col and any(
                    k in key for k in ("택배사", "배송사", "운송사", "지불조건", "courier", "carrier")
                ):
                    col["courier"] = j
                if "name" not in col and (
                    any(k in key for k in ("수령인", "성함", "받는", "받으시는분")) or key == "이름"
                ):
                    col["name"] = j
                if "phone" not in col and any(k in key for k in ("연락처", "전화", "핸드폰", "휴대폰", "이동통신")):
                    col["phone"] = j
                if "addr" not in col and any(k in key for k in ("주소", "총주소")):
                    col["addr"] = j
                if "product" not in col and any(k in key for k in ("품목명", "상품명", "제품명", "상품")):
                    col["product"] = j

            t = col.get("tracking")
            if t is None:
                continue

            for row in rows[header_i + 1:]:
                if t >= len(row):
                    continue
                tracking = _digits(row[t]) or _norm(row[t])
                if not tracking:
                    continue
                entry = {
                    "tracking": tracking,
                    "order_id": _norm(row[col["order"]]) if col.get("order") is not None and col["order"] < len(row) else "",
                    "courier": _temu_carrier_name(row[col["courier"]])
                    if col.get("courier") is not None and col["courier"] < len(row)
                    else "",
                    "name": _norm(row[col["name"]]) if col.get("name") is not None and col["name"] < len(row) else "",
                    "phone": _digits(row[col["phone"]]) if col.get("phone") is not None and col["phone"] < len(row) else "",
                    "address": _norm(row[col["addr"]]) if col.get("addr") is not None and col["addr"] < len(row) else "",
                    "product": _norm(row[col["product"]]) if col.get("product") is not None and col["product"] < len(row) else "",
                }
                total += 1
                if entry["order_id"]:
                    by_order[_digits(entry["order_id"]) or entry["order_id"]] = entry
                if entry["name"]:
                    by_name[entry["name"]].append(entry)
        wb.close()

    return by_order, by_name, total


def _match(order: dict, by_order: dict, by_name: dict) -> dict | None:
    # 1) 주문번호 일치
    key = _digits(order["order_id"]) or order["order_id"]
    if key and key in by_order:
        return by_order[key]
    # 2) 수령인 이름 일치
    for nm in (order["name_cands"] or {order["name"]}):
        cands = by_name.get(nm)
        if not cands:
            continue
        if len(cands) == 1:
            return cands[0]
        # 동명이인: 전화 → 주소 순으로 구분
        if order["phone"]:
            pick = [c for c in cands if c["phone"] and c["phone"][-8:] == order["phone"][-8:]]
            if len(pick) == 1:
                return pick[0]
        if order["address"]:
            pick = [
                c for c in cands
                if c["address"] and (c["address"] in order["address"] or order["address"] in c["address"])
            ]
            if len(pick) == 1:
                return pick[0]
        option_key = _norm(order.get("option"))
        if option_key:
            pick = [c for c in cands if c.get("product") and option_key in c["product"]]
            if len(pick) == 1:
                return pick[0]
    return None


# ----------------------------------------------------------------------------
# 메인
# ----------------------------------------------------------------------------
def process(order_bytes: bytes, tracking_files: list[tuple[str, bytes]], order_filename: str = "") -> tuple[bytes, str, dict]:
    orders = _parse_temu_export(order_bytes, order_filename)
    if not orders:
        raise ValueError(
            "테무 주문목록에 처리할 주문이 없습니다. "
            "미발송 주문이 있는 상태에서 주문 내보내기 후 다시 시도해주세요."
        )

    by_order, by_name, tracking_total = _parse_tracking_files(tracking_files)
    if tracking_total == 0:
        raise ValueError(
            "택배 송장파일에서 운송장번호를 찾지 못했습니다. "
            "송장(운송장)번호와 수령인/주문번호가 들어있는 파일인지 확인해주세요."
        )

    template_path = TEMPLATE_DIR / TEMPLATE_NAME
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(filename=str(template_path))
    ws = wb[SHIP_SHEET] if SHIP_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    out_row = 2
    filled = 0
    warnings: list[str] = []
    carrier_counts: dict[str, int] = defaultdict(int)

    for o in orders:
        entry = _match(o, by_order, by_name)
        if entry is None:
            warnings.append(f"미매칭(운송장 없음): 주문 {o['order_id']} / {o['name'] or '이름미상'}")
            continue

        carrier = _temu_carrier_name(entry.get("courier")) or _carrier_for_kg(o["kg"])
        if carrier is None:
            carrier = CARRIER_LARGE
            warnings.append(
                f"kg 판별 실패 → 기본 '{CARRIER_LARGE}' 적용: 주문 {o['order_id']} (옵션='{o['option']}')"
            )

        ws.cell(row=out_row, column=1, value=o["order_id"])
        ws.cell(row=out_row, column=2, value=o["item_id"])
        ws.cell(row=out_row, column=3, value=o["qty"])
        ws.cell(row=out_row, column=4, value=SHIP_FROM)
        ws.cell(row=out_row, column=5, value=carrier)
        ws.cell(row=out_row, column=6, value=entry["tracking"])
        carrier_counts[carrier] += 1
        out_row += 1
        filled += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"테무_배송확인_{now.strftime('%Y%m%d')}.xlsx"
    stats: dict = {
        "total_orders": len(orders),
        "filled": filled,
        "unmatched": len(orders) - filled,
        "tracking_rows": tracking_total,
    }
    for c, n in sorted(carrier_counts.items()):
        stats[f"배송사:{c}"] = n
    if warnings:
        stats["warnings"] = warnings[:50]

    return output.read(), filename, stats
