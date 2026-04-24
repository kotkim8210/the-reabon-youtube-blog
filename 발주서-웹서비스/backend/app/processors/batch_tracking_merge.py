from copy import copy
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook


KST = timezone(timedelta(hours=9))


def process(result_files: list[tuple[str, bytes]]) -> tuple[bytes, str, dict]:
    """Merge processed DeliveryList workbooks into one combined workbook."""
    valid_files = [(name, data) for name, data in result_files if data]
    if not valid_files:
        raise ValueError("병합할 송장입력 파일이 없습니다.")

    first_name, first_bytes = valid_files[0]
    merged_wb = load_workbook(filename=BytesIO(first_bytes))
    merged_sheet_name = merged_wb.sheetnames[0]
    for sheet_name in merged_wb.sheetnames[1:]:
        del merged_wb[sheet_name]
    merged_ws = merged_wb[merged_sheet_name]

    merged_rows = max(merged_ws.max_row - 1, 0)
    source_files = [first_name]

    for filename, file_bytes in valid_files[1:]:
        wb = load_workbook(filename=BytesIO(file_bytes))
        sheet_name = wb.sheetnames[0]
        for extra_sheet in wb.sheetnames[1:]:
            del wb[extra_sheet]
        ws = wb[sheet_name]
        source_files.append(filename)

        for row_idx in range(2, ws.max_row + 1):
            target_row_idx = merged_ws.max_row + 1
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell = merged_ws.cell(
                    row=target_row_idx,
                    column=col_idx,
                    value=source_cell.value,
                )
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

            row_dimension = ws.row_dimensions.get(row_idx)
            if row_dimension and row_dimension.height:
                merged_ws.row_dimensions[target_row_idx].height = row_dimension.height

        merged_rows += max(ws.max_row - 1, 0)

    output = BytesIO()
    merged_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"송장입력파일_통합({now.strftime('%Y%m%d_%H%M')}).xlsx"
    stats = {
        "source_files": len(valid_files),
        "merged_rows": merged_rows,
        "sources": source_files,
    }
    return output.read(), filename, stats
