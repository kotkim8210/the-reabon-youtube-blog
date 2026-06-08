import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.processors.chamdureup_order import convert_option
from app.processors.tracking_match import (
    name_counts,
    normalize_courier_name,
    option_key_set,
    options_match,
    requires_option_guard,
)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    """공백 완전 제거 후 문자열 반환 (매칭용)"""
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def process(
    orderlist_bytes: bytes,
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """참두릅 orderlist의 운송장번호를 DeliveryList에 매핑.

    1차 매칭: 주문번호 (orderlist D열 = DeliveryList C열)
    폴백: 이름/전화/주소 매칭
    """
    ol_wb = load_workbook(filename=BytesIO(orderlist_bytes), data_only=True)
    ol_ws = ol_wb.active

    order_entries = []
    for row in ol_ws.iter_rows(min_row=2):
        order_no = normalize(row[3].value) if len(row) > 3 else ""
        name = normalize(row[10].value) if len(row) > 10 else ""
        phone = normalize(row[12].value) if len(row) > 12 else ""
        address = normalize(row[14].value) if len(row) > 14 else ""
        courier = normalize(row[16].value) if len(row) > 16 else ""
        tracking = normalize(row[17].value) if len(row) > 17 else ""
        option_keys = option_key_set(
            row[6].value if len(row) > 6 else "",
            row[8].value if len(row) > 8 else "",
            row[11].value if len(row) > 11 else "",
        )

        if tracking:
            order_entries.append({
                "order_no": order_no,
                "name": name,
                "phone": phone,
                "address": address,
                "courier": courier,
                "tracking": tracking,
                "option_keys": option_keys,
            })

    order_by_order_no = {}
    order_by_name = defaultdict(list)
    for entry in order_entries:
        if entry["order_no"]:
            order_by_order_no[entry["order_no"]] = entry
        if entry["name"]:
            order_by_name[entry["name"]].append(entry)

    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    dl_ws = dl_wb.active

    first_sheet_name = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet_name]

    used_entries = set()
    filled = 0
    skipped = 0
    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )

    for row_idx in range(2, dl_ws.max_row + 1):
        e_cell = dl_ws.cell(row=row_idx, column=5)

        if e_cell.value is not None and normalize(e_cell.value) != "":
            continue

        dl_order_no = normalize(dl_ws.cell(row=row_idx, column=3).value)
        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)
        dl_phone = normalize(dl_ws.cell(row=row_idx, column=28).value)
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)
        dl_product = dl_ws.cell(row=row_idx, column=11).value
        dl_option = dl_ws.cell(row=row_idx, column=12).value
        dl_option_keys = option_key_set(dl_option, convert_option(str(dl_option or "")))
        if not dl_option_keys:
            dl_option_keys = option_key_set(dl_product)

        if not dl_name and not dl_order_no:
            continue

        matched = None

        if dl_order_no and dl_order_no in order_by_order_no:
            candidate = order_by_order_no[dl_order_no]
            if id(candidate) not in used_entries:
                matched = candidate

        if matched is None and dl_name:
            candidates = order_by_name.get(dl_name, [])

            if not candidates:
                MIN_PREFIX_LEN = 8
                for order_name, entries in order_by_name.items():
                    if len(order_name) < MIN_PREFIX_LEN:
                        continue
                    if dl_name.startswith(order_name) or order_name.startswith(dl_name):
                        candidates = entries
                        break

            available = [c for c in candidates if id(c) not in used_entries]

            if available:
                if requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
                    available = [
                        c for c in available
                        if options_match(c.get("option_keys"), dl_option_keys)
                    ]
                    if not available:
                        skipped += 1
                        continue

                for c in available:
                    if c["phone"] == dl_phone and c["address"] == dl_address:
                        matched = c
                        break
                if matched is None:
                    for c in available:
                        if c["phone"] == dl_phone:
                            matched = c
                            break
                if matched is None:
                    for c in available:
                        if c["address"] == dl_address:
                            matched = c
                            break
                if matched is None:
                    matched = available[0]

        if matched:
            e_cell.value = matched["tracking"]
            d_cell = dl_ws.cell(row=row_idx, column=4)
            if matched["courier"]:
                d_cell.value = normalize_courier_name(matched["courier"])
            used_entries.add(id(matched))
            filled += 1
        else:
            skipped += 1

    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_참두릅_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {"filled": filled, "skipped": skipped}

    return output.read(), filename, stats
