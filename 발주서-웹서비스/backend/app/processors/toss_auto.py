"""Toss-only: 토스 주문 수집 → 해달 발주서 생성 + 토스 운송장 등록."""

import re
import logging
from hashlib import sha1
from datetime import datetime, timezone, timedelta
from io import BytesIO
from collections import defaultdict

import httpx
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR
from app.processors.tracking_match import (
    name_counts,
    option_key_set,
    options_match,
    requires_option_guard,
)
from app.processors.goguma_order import transform_toss_option as normalize_goguma_toss_option
from app.processors.haedal_tracking_parser import (
    detect_haedal_columns,
    find_courier_in_row,
    find_tracking_in_row,
)
from app.toss.client import toss_client

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

GOGUMA_KEYWORDS = ["고구마", "꿀고구마"]
TOSS_DELIVERY_COMPANY = "한진택배"
TRACKABLE_TOSS_STATUSES = {"PAID", "PREPARING_PRODUCT", "DELIVERING"}


def normalize_toss_courier(name: object, default: str | None = None) -> str:
    """택배사명을 토스가 인식하는 정식명으로 정규화. (해달 파일이 '한진'처럼 약식으로 적어도)

    토스 deliveryCompany는 정식 택배사명을 요구한다. '한진' → '한진택배' 처럼 맞춰야 등록된다.
    default: 값이 비었을 때 쓸 택배사. 미지정 시 고구마(해달) 기본인 한진택배.
    발주처별 실제 발송 택배사가 다르므로 호출부가 명시해야 한다
    (쥬얼리프룻·제주다팜=롯데택배 — 한진 기본값 탓에 수박 건이 토스에서
    배송추적 안 되던 2026-06-15 사고 재발 방지).
    """
    fallback = default or TOSS_DELIVERY_COMPANY
    compact = re.sub(r"\s+", "", str(name or "")).lower()
    if not compact:
        return fallback
    if "대한통운" in compact or compact.startswith("cj"):
        return "CJ대한통운"
    if "한진" in compact:
        return "한진택배"
    if "롯데" in compact or "현대" in compact:
        return "롯데택배"
    if "우체국" in compact or "epost" in compact:
        return "우체국택배"
    if "로젠" in compact or "logen" in compact:
        return "로젠택배"
    return str(name).strip() or fallback


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def stable_order_id(prefix: str, *values) -> str:
    raw = "|".join(normalize(value) for value in values if normalize(value))
    if not raw:
        return ""
    return f"{prefix}:{sha1(raw.encode('utf-8')).hexdigest()[:16]}"


def toss_order_id(item: dict) -> str:
    for key in ("orderNo", "orderId", "orderNumber", "orderSheetId", "paymentKey"):
        value = item.get(key)
        if value:
            return str(value)
    return ""


def normalize_strip(value) -> str:
    """공백 완전 제거 (매칭용)."""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def transform_toss_option(product_name: str, option_name: str) -> str:
    return normalize_goguma_toss_option(product_name, option_name)


def is_goguma_order(item: dict) -> bool:
    product_name = (item.get("productName") or "").lower()
    option_name = (item.get("optionName") or "").lower()
    return any(kw in product_name or kw in option_name for kw in GOGUMA_KEYWORDS)


# ---------------------------------------------------------------------------
# 1) 토스 주문 수집 → 발주서 생성
# ---------------------------------------------------------------------------

async def process_toss_order(from_date: str, to_date: str) -> tuple[bytes, str, dict]:
    """Collect goguma orders from Toss API and generate 해달 발주서.

    Args:
        from_date: Start date (YYYY-MM-DD).
        to_date: End date (YYYY-MM-DD).

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    orders = await toss_client.get_orders(
        start_date=from_date,
        end_date=to_date,
        status="PAID",
    )

    entries = []
    for item in orders:
        if not is_goguma_order(item):
            continue

        address = item.get("address") or ""
        detail_addr = item.get("detailAddress") or ""
        full_address = f"{address} {detail_addr}".strip()

        zipcode = item.get("zipCode") or ""
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = str(zipcode).zfill(5)

        product = transform_toss_option(
            item.get("productName") or "",
            item.get("optionName") or "",
        )

        entries.append({
            "name": item.get("receiverName") or "",
            "phone": item.get("receiverRealPhone") or item.get("receiverPhone") or "",
            "zipcode": zipcode,
            "address": full_address,
            "qty": str(item.get("quantity") or 1),
            "product": product,
            "memo": item.get("shippingNote") or "",
            "order_id": toss_order_id(item) or stable_order_id(
                "toss-api",
                item.get("receiverName"),
                item.get("receiverRealPhone") or item.get("receiverPhone"),
                full_address,
                product,
                item.get("quantity") or 1,
            ),
        })

    if not entries:
        raise ValueError("해당 기간에 토스 고구마 주문이 없습니다.")

    # Load template
    template_path = TEMPLATE_DIR / "해달_발주서_한진양식.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"템플릿 파일을 찾을 수 없습니다: {template_path}. "
            f"templates 디렉토리에 해달_발주서_한진양식.xlsx 파일을 넣어주세요."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet]

    font11 = Font(size=11)

    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}

    for i, entry in enumerate(entries):
        row_idx = start_row + i
        mapping = {
            1: entry["name"],
            2: entry["phone"],
            5: entry["zipcode"],
            6: entry["address"],
            7: "식품애착",
            8: "010-5700-7756",
            12: "전라남도 해남군 산이면 새상골길 ",
            13: entry["qty"],
            14: entry["product"],
            16: "선불",
            19: entry["memo"],
        }
        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11

        try:
            qty_int = int(float(entry["qty"])) if entry["qty"] not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        bucket = option_totals.setdefault(
            entry["product"] or "고구마",
            {
                "coupang_option_keyword": entry["product"] or "고구마",
                "vendor_option_name": entry["product"] or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"해달 발주서 한진양식_토스_알제이시스템즈({now.strftime('%y%m%d')}).xlsx"
    stats = {
        "total": len(entries),
        "period": f"{from_date} ~ {to_date}",
        "product": "고구마",
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats


# ---------------------------------------------------------------------------
# 2) 토스 운송장 등록
# ---------------------------------------------------------------------------

def parse_haedal_file(haedal_bytes: bytes) -> list[dict]:
    """Parse 해달 발주서 and extract entries with tracking numbers."""
    wb = load_workbook(filename=BytesIO(haedal_bytes), data_only=True)
    ws = wb.active

    cols = detect_haedal_columns(ws)

    entries = []
    missing: list[str] = []
    for row_idx in range(cols.start_row, ws.max_row + 1):
        name = normalize_strip(ws.cell(row=row_idx, column=cols.name).value)
        phone = normalize_strip(ws.cell(row=row_idx, column=cols.phone).value)
        address = normalize_strip(ws.cell(row=row_idx, column=cols.address).value)
        product = ws.cell(row=row_idx, column=cols.product).value

        if not name:
            continue

        tracking = find_tracking_in_row(ws, row_idx, cols.tracking)
        if not tracking:
            missing.append(f"{row_idx}행 {name}")
        if tracking:
            skip_cols = tuple(
                c for c in (cols.name, cols.phone, cols.address, cols.product, cols.tracking) if c
            )
            courier = find_courier_in_row(ws, row_idx, cols.courier, skip_cols)
            entries.append({
                "name": name,
                "phone": phone,
                "address": address,
                "tracking": tracking,
                "courier": courier,
                "option_keys": option_key_set(product, transform_toss_option("", str(product or ""))),
            })

    # 조용한 누락 금지: 수취인은 있는데 송장이 없는 행은 로그로 남긴다.
    if missing:
        logger.warning(
            "해달 회신에 송장 없는 행 %s건 (송장열=%s): %s",
            len(missing),
            cols.tracking,
            ", ".join(missing[:20]),
        )

    return entries


def _masked_name_match(masked: str, full_names) -> list[str]:
    """토스가 가린 수취인명('정*순')을 해달 풀네임('정정순')과 매칭.

    같은 글자수이고 '*' 위치를 뺀 나머지 글자가 모두 일치하면 같은 사람으로 본다.
    토스 Shopping API는 일부 주문의 receiverName을 별표로 가려서 내려준다.
    """
    if "*" not in masked:
        return []
    return [
        name
        for name in full_names
        if len(name) == len(masked)
        and all(m == "*" or m == c for m, c in zip(masked, name))
    ]


async def process_toss_tracking(haedal_bytes: bytes) -> dict:
    """Parse tracking file, match Toss orders, register via API.

    Returns:
        Dict with counts and per-order results.
    """
    # 1. Parse 해달 file
    haedal_entries = parse_haedal_file(haedal_bytes)
    if not haedal_entries:
        raise ValueError("해달 발주서에서 운송장번호를 찾을 수 없습니다.")

    # 2. Fetch Toss orders and keep only statuses that can still accept
    # delivery info updates. Fetching all recent orders lets us return
    # clear skip messages instead of a hard error when orders already moved.
    now = datetime.now(KST)
    from_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    to_date = now.strftime("%Y-%m-%d")

    try:
        toss_orders_raw = await toss_client.get_orders(
            start_date=from_date,
            end_date=to_date,
            status=None,
        )
    except Exception as e:
        raise ValueError(f"토스 API 조회 실패: {e}") from e

    # Filter goguma orders
    toss_orders = []
    pre_skipped_results = []
    seen_order_product_ids = set()
    for item in toss_orders_raw:
        if not is_goguma_order(item):
            continue

        order_product_id = item.get("orderProductId")
        if not order_product_id or order_product_id in seen_order_product_ids:
            continue
        seen_order_product_ids.add(order_product_id)

        receiver_name = item.get("receiverName") or ""
        receiver_phone = item.get("receiverRealPhone") or item.get("receiverPhone") or ""
        address = item.get("address") or ""
        detail_addr = item.get("detailAddress") or ""
        full_address = f"{address} {detail_addr}".strip()
        order_status = item.get("orderProductStatus") or item.get("status") or ""
        existing_tracking = normalize_strip(item.get("shippingTrackingNumber") or "")
        order_id = item.get("orderId") or item.get("orderNumber") or ""

        if existing_tracking:
            pre_skipped_results.append({
                "order_id": str(order_id),
                "name": receiver_name,
                "tracking": existing_tracking,
                "status": "skip",
                "message": "이미 운송장번호가 등록된 주문",
            })
            continue

        if order_status not in TRACKABLE_TOSS_STATUSES:
            pre_skipped_results.append({
                "order_id": str(order_id),
                "name": receiver_name,
                "status": "skip",
                "message": f"운송장 등록 불가 상태: {order_status or 'UNKNOWN'}",
            })
            continue

        toss_orders.append({
            "order_product_id": order_product_id,
            "order_id": order_id,
            "name": normalize_strip(receiver_name),
            "phone": normalize_strip(receiver_phone),
            "address": normalize_strip(full_address),
            "name_display": receiver_name,
            "order_status": order_status,
            "option_keys": option_key_set(
                item.get("optionName"),
                transform_toss_option(item.get("productName") or "", item.get("optionName") or ""),
            )
            or option_key_set(item.get("productName")),
        })

    if not toss_orders:
        goguma_total = len(seen_order_product_ids)
        if pre_skipped_results:
            return {
                "haedal_entries": len(haedal_entries),
                "toss_orders": goguma_total,
                "success": 0,
                "fail": 0,
                "skip": len(pre_skipped_results),
                "results": pre_skipped_results,
            }
        if toss_orders_raw:
            raise ValueError(
                f"토스에서 {len(toss_orders_raw)}건의 주문을 찾았지만 "
                f"고구마/꿀고구마 키워드가 포함된 주문이 없습니다."
            )
        raise ValueError(
            f"토스에서 고구마 주문을 찾을 수 없습니다. "
            f"(조회 기간: {from_date} ~ {to_date})"
        )

    # 3. Match
    entry_by_name = defaultdict(list)
    for entry in haedal_entries:
        entry_by_name[entry["name"]].append(entry)

    used_entries = set()
    matched_pairs = []
    results = list(pre_skipped_results)
    skip_count = len(pre_skipped_results)
    toss_name_counts = name_counts(order["name"] for order in toss_orders)

    for order in toss_orders:
        candidates = entry_by_name.get(order["name"], [])
        # 토스가 수취인명을 가린 경우('정*순')에도 풀네임 해달 항목과 와일드카드 매칭
        if not candidates and "*" in order["name"]:
            for masked_name in _masked_name_match(order["name"], entry_by_name.keys()):
                candidates = candidates + entry_by_name[masked_name]
        if not candidates:
            skip_count += 1
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "status": "skip",
                "message": "매칭되는 수령인 없음",
            })
            continue

        available = [
            (i, c) for i, c in enumerate(candidates)
            if id(c) not in used_entries
        ]
        if requires_option_guard(order["name"], toss_name_counts, len(candidates)):
            available = [
                (i, c) for i, c in available
                if options_match(c.get("option_keys"), order.get("option_keys"))
            ]
        if not available:
            skip_count += 1
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "status": "skip",
                "message": "사용 가능한 운송장번호 없음",
            })
            continue

        matched = None

        for _, c in available:
            if c["phone"] == order["phone"] and c["address"] == order["address"]:
                matched = c
                break

        if matched is None:
            for _, c in available:
                if c["phone"] == order["phone"]:
                    matched = c
                    break

        if matched is None:
            for _, c in available:
                if c["address"] == order["address"]:
                    matched = c
                    break

        if matched is None:
            matched = available[0][1]

        used_entries.add(id(matched))
        matched_pairs.append((order, matched["tracking"], matched.get("courier") or ""))

    # 4. Register tracking numbers via Toss API
    success_count = 0
    fail_count = 0

    for order, tracking, courier in matched_pairs:
        order_product_id = order.get("order_product_id")
        if not order_product_id:
            fail_count += 1
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "tracking": tracking,
                "status": "fail",
                "message": "orderProductId 없음",
            })
            continue

        # 해달 파일에 적힌 실제 택배사명 그대로 등록(CJ대한통운/한진택배 등). 없으면 기본값.
        delivery_company = normalize_toss_courier(courier)
        try:
            resp = await toss_client.register_tracking(
                order_product_id=order_product_id,
                delivery_company=delivery_company,
                tracking_number=tracking,
            )
            success_count += 1
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "tracking": tracking,
                "status": "success",
                "message": "운송장 등록 완료",
            })
        except httpx.HTTPStatusError as e:
            fail_count += 1
            body = ""
            try:
                body = e.response.text[:500]
            except Exception:
                pass
            err_msg = f"HTTP {e.response.status_code}: {body}" if body else f"HTTP {e.response.status_code}"
            logger.error(
                f"토스 운송장 등록 실패 (orderProductId={order_product_id}): {err_msg}"
            )
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "tracking": tracking,
                "status": "fail",
                "message": err_msg,
            })
        except Exception as e:
            fail_count += 1
            err_msg = f"{type(e).__name__}: {e}"
            logger.error(
                f"토스 운송장 등록 실패 (orderProductId={order_product_id}): {err_msg}"
            )
            results.append({
                "order_id": str(order["order_id"]),
                "name": order["name_display"],
                "tracking": tracking,
                "status": "fail",
                "message": err_msg,
            })

    return {
        "haedal_entries": len(haedal_entries),
        "toss_orders": len(toss_orders),
        "success": success_count,
        "fail": fail_count,
        "skip": skip_count,
        "results": results,
    }
