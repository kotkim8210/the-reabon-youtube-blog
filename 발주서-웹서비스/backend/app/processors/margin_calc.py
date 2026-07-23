"""마진계산기 (소싱현황) — 판매가·공급가 입력 → 정식 양식 마진 산출.

정식 양식(개인 소싱현황관리 표준, 2026-06 실파일로 검증)을 그대로 따른다.
축소양식(G 하드코딩·단순 마진식) 금지 — 마진방어·소득세16%·cs로스까지 반영해야
원가율 높은 품목이 마이너스로 정확히 잡힌다.

컬럼/수식 (행 r):
  G 판매가   = H+500 (쿠폰가+500)
  H 쿠폰가   = 입력값(쿠팡 판매가)
  I 공급가   = 입력값(매입 원가)
  K 수수료율 = 0.1166 (기본)
  N 마진방어 = H*5%(1·2kg) / H*3%(3kg↑)
  P 쿠팡수수료 = (G*K)+(H*3.63%)
  Q 부가세   = ((G+H)-(I+L+M+N+P))*0%  (→ 0)
  R 1차합계  = G+H-I-L-M-N-P-Q
  S 소득세   = R*16%
  T cs·로스  = G*1%
  U 마진     = H-I-L-M-N-P-Q-S-T   (G 미포함)
  V 마진율   = U/G
"""

import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

KST = timezone(timedelta(hours=9))

DEFAULT_FEE_RATE = 0.1166  # 쿠팡 수수료율 기본값
_KG_RE = re.compile(r"(\d+(?:\.\d+)?)\s*kg", re.IGNORECASE)

# 정식 양식 헤더 (행 7). 컬럼 letter → 라벨. D·O는 비워둠(원본과 동일).
HEADER_ROW = 7
DATA_START_ROW = 8
_HEADERS = {
    "B": "NO", "C": "상품명(소싱처)", "E": "옵션", "F": "수량",
    "G": "판매가", "H": "쿠폰가(판매가)", "I": "공급가", "J": "원가(위안)",
    "K": "수수료율", "L": "물류비", "M": "포장/부자재", "N": "마진방어",
    "P": "쿠팡수수료", "Q": "부가세", "R": "1차합계", "S": "소득세",
    "T": "cs·로스비용", "U": "마진", "V": "마진율", "X": "공급처URL",
}


def _extract_kg(text: str) -> float | None:
    m = _KG_RE.search(str(text or ""))
    return float(m.group(1)) if m else None


def _margin_defense_rate(option: str) -> float:
    """마진방어율: 1·2kg=5%, 3kg 이상=3%. kg 불명이면 5%(보수적)."""
    kg = _extract_kg(option)
    if kg is not None and kg >= 3:
        return 0.03
    return 0.05


def _num(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return float(s) if s not in ("", "-", ".") else 0.0
    except ValueError:
        return 0.0


def compute_margin(sale_price: float, supply_price: float, option: str = "",
                   logistics: float = 0.0, packaging: float = 0.0,
                   fee_rate: float = DEFAULT_FEE_RATE) -> dict:
    """정식 양식 수식으로 한 행의 마진을 계산. 반환: 각 컬럼 값 dict."""
    h = float(sale_price)
    i = float(supply_price)
    ll = float(logistics)
    m = float(packaging)
    g = h + 500
    k = fee_rate
    n = h * _margin_defense_rate(option)
    p = (g * k) + (h * 0.0363)
    q = ((g + h) - (i + ll + m + n + p)) * 0.0  # 항상 0 (원본 *0%)
    r = g + h - i - ll - m - n - p - q
    s = r * 0.16
    t = g * 0.01
    u = h - i - ll - m - n - p - q - s - t
    v = (u / g) if g else 0.0
    return {"G": g, "H": h, "I": i, "K": k, "L": ll, "M": m, "N": n,
            "P": p, "Q": q, "R": r, "S": s, "T": t, "U": u, "V": v}


def build_sourcing_sheet(rows: list[dict], title: str = "소싱현황") -> tuple[bytes, str, dict]:
    """입력 행 → 정식 양식 소싱현황 엑셀(수식 포함) + 통계.

    rows: [{product, option, qty, sale_price, supply_price, logistics?, packaging?, fee_rate?, url?}]
    엑셀 셀에는 원본과 동일한 수식을 넣어 엑셀이 열 때 자동 계산되게 하고,
    통계(마이너스 마진 건수 등)는 파이썬 계산값으로 낸다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "소싱현황"

    title_font = Font(size=14, bold=True)
    head_font = Font(size=10, bold=True)
    head_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center")

    ws["B2"] = "소싱현황 관리"
    ws["B2"].font = title_font

    for col, label in _HEADERS.items():
        c = ws[f"{col}{HEADER_ROW}"]
        c.value = label
        c.font = head_font
        c.fill = head_fill
        c.alignment = center

    neg_fill = PatternFill("solid", fgColor="FFC7CE")  # 마이너스 마진 강조
    minus_count = 0
    margin_rates: list[float] = []

    for idx, row in enumerate(rows):
        r = DATA_START_ROW + idx
        option = str(row.get("option") or "")
        sale = _num(row.get("sale_price"))
        supply = _num(row.get("supply_price"))
        logistics = _num(row.get("logistics"))
        packaging = _num(row.get("packaging"))
        fee_rate = row.get("fee_rate")
        fee_rate = DEFAULT_FEE_RATE if fee_rate in (None, "", 0) else float(fee_rate)
        defense_pct = int(round(_margin_defense_rate(option) * 100))

        ws[f"B{r}"] = idx + 1
        ws[f"C{r}"] = str(row.get("product") or "")
        ws[f"E{r}"] = option
        ws[f"F{r}"] = row.get("qty") or 1
        ws[f"G{r}"] = f"=H{r}+500"
        ws[f"H{r}"] = sale
        ws[f"I{r}"] = supply
        ws[f"J{r}"] = 0
        ws[f"K{r}"] = fee_rate
        ws[f"L{r}"] = logistics
        ws[f"M{r}"] = packaging
        ws[f"N{r}"] = f"=H{r}*{defense_pct}%"
        ws[f"P{r}"] = f"=(G{r}*K{r})+(H{r}*3.63%)"
        ws[f"Q{r}"] = f"=((G{r}+H{r})-(I{r}+L{r}+M{r}+N{r}+P{r}))*0%"
        ws[f"R{r}"] = f"=G{r}+H{r}-I{r}-L{r}-M{r}-N{r}-P{r}-Q{r}"
        ws[f"S{r}"] = f"=R{r}*16%"
        ws[f"T{r}"] = f"=G{r}*1%"
        ws[f"U{r}"] = f"=H{r}-I{r}-L{r}-M{r}-N{r}-P{r}-Q{r}-S{r}-T{r}"
        ws[f"V{r}"] = f'=IF(U{r}=0,"",U{r}/G{r})'
        ws[f"X{r}"] = str(row.get("url") or "")
        ws[f"V{r}"].number_format = "0.0%"

        calc = compute_margin(sale, supply, option, logistics, packaging, fee_rate)
        if calc["U"] < 0:
            minus_count += 1
            for col in ("C", "H", "I", "U", "V"):
                ws[f"{col}{r}"].fill = neg_fill
        if calc["G"]:
            margin_rates.append(calc["V"])

    widths = {"B": 5, "C": 16, "E": 16, "F": 6, "G": 10, "H": 10, "I": 10,
              "K": 8, "N": 10, "P": 11, "R": 11, "S": 9, "T": 9, "U": 11, "V": 9, "X": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    now = datetime.now(KST)
    safe_title = re.sub(r"[\\/:*?\"<>|]", "_", title).strip() or "소싱현황"
    filename = f"{safe_title} 소싱현황관리_{now.strftime('%y%m%d')}.xlsx"
    avg_rate = round(sum(margin_rates) / len(margin_rates) * 100, 1) if margin_rates else 0.0
    stats = {
        "총_상품수": len(rows),
        "마진_마이너스": minus_count,
        "평균_마진율(%)": avg_rate,
    }
    return buf.getvalue(), filename, stats


# ── 고객 업로드 파일 파싱 (판매가·공급가 자동 인식) ─────────────────
_COL_ALIASES = {
    "product": ["상품명", "상품", "소싱처", "품목", "품명"],
    "option": ["옵션", "규격", "중량", "옵션명"],
    "qty": ["수량"],
    "sale_price": ["쿠폰가", "판매가", "판가", "쿠팡가", "판매금액"],
    "supply_price": ["공급가", "원가", "매입가", "사입가", "매입", "공급단가"],
    "logistics": ["물류비", "배송비"],
    "packaging": ["포장", "부자재", "포장비"],
    "fee_rate": ["수수료율", "수수료"],
    "url": ["url", "공급처url", "링크", "거래처url"],
}


def _match_col(header: str) -> str | None:
    h = re.sub(r"\s+", "", str(header or "")).lower()
    if not h:
        return None
    for field, aliases in _COL_ALIASES.items():
        for a in aliases:
            if a.lower() in h:
                return field
    return None


def parse_input_file(file_bytes: bytes) -> list[dict]:
    """고객 파일에서 헤더를 자동 인식해 계산 입력 행으로 변환.

    최소 판매가(쿠폰가/판매가)와 공급가(원가/매입가) 컬럼이 있어야 한다.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = None
    col_map: dict[int, str] = {}
    for r in range(1, min(ws.max_row, 15) + 1):
        found: dict[int, str] = {}
        for c in range(1, min(ws.max_column, 40) + 1):
            field = _match_col(ws.cell(r, c).value)
            if field and field not in found.values():
                found[c] = field
        if "sale_price" in found.values() and "supply_price" in found.values():
            header_row = r
            col_map = found
            break

    if header_row is None:
        raise ValueError(
            "판매가(쿠폰가)와 공급가(원가) 컬럼을 찾지 못했습니다. "
            "상품명·옵션·판매가·공급가 열이 있는 파일을 올리거나, 제공되는 빈 양식을 사용하세요."
        )

    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        record: dict = {}
        for c, field in col_map.items():
            record[field] = ws.cell(r, c).value
        if _num(record.get("sale_price")) <= 0 and _num(record.get("supply_price")) <= 0:
            continue
        rows.append(record)
    return rows


def blank_template_bytes() -> bytes:
    """고객이 판매가·공급가만 채우면 되는 빈 정식 양식(예시 2행)."""
    example = [
        {"product": "예시상품", "option": "1kg정품", "qty": 1, "sale_price": 10800, "supply_price": 5900},
        {"product": "예시상품", "option": "3kg정품", "qty": 1, "sale_price": 19900, "supply_price": 9200},
    ]
    data, _fn, _stats = build_sourcing_sheet(example, title="빈양식")
    return data
