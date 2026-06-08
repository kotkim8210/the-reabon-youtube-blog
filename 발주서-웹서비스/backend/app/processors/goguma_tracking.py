import re
from datetime import datetime, timezone, timedelta
from io import BytesIO
from collections import defaultdict

from openpyxl import load_workbook

from app.config import TEMPLATE_DIR
from app.processors.goguma_order import transform_option
from app.processors.haedal_tracking_parser import detect_haedal_columns, find_tracking_in_row
from app.processors.tracking_match import (
    name_counts,
    option_key_set,
    options_match,
    requires_option_guard,
)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    """공백 완전 제거 후 문자열 반환 (이름 매칭용)"""
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def phone_digits(value) -> str:
    """전화번호에서 숫자만 추출 (하이픈/공백 형식 차이 무시)"""
    if value is None:
        return ""
    return re.sub(r'\D', '', str(value))

def process(
    haedal_bytes: bytes,
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """Input tracking numbers from 해달 발주서 into DeliveryList.

    Args:
        haedal_bytes: Raw bytes of the 해달 발주서 Excel file.
            Columns: A=name, B=phone, F=address, P~V=tracking (10-14 digit number)
        delivery_bytes: Raw bytes of the DeliveryList Excel file.

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    # Load 해달 발주서
    hd_wb = load_workbook(filename=BytesIO(haedal_bytes), data_only=True)
    hd_ws = hd_wb.active

    # Header-based column detection supports both old and new 해달/한진 reply formats.
    cols = detect_haedal_columns(hd_ws)

    # Extract entries from 해달
    haedal_entries = []
    for row_idx in range(cols.start_row, hd_ws.max_row + 1):
        name = normalize(hd_ws.cell(row=row_idx, column=cols.name).value)
        phone = phone_digits(hd_ws.cell(row=row_idx, column=cols.phone).value)
        address = normalize(hd_ws.cell(row=row_idx, column=cols.address).value)
        product = hd_ws.cell(row=row_idx, column=cols.product).value

        if not name:
            continue

        tracking = find_tracking_in_row(hd_ws, row_idx, cols.tracking)

        if tracking:
            haedal_entries.append({
                "name": name,
                "phone": phone,
                "address": address,
                "tracking": tracking,
                "option_keys": option_key_set(product, transform_option(str(product or ""))),
            })

    # Load DeliveryList
    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    dl_ws = dl_wb.active

    # Keep only first sheet
    first_sheet_name = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet_name]

    # Build indices by name and by phone (phone index 용 fallback 매칭)
    entry_by_name = defaultdict(list)
    entry_by_phone = defaultdict(list)
    for entry in haedal_entries:
        entry_by_name[entry["name"]].append(entry)
        if entry["phone"]:
            entry_by_phone[entry["phone"]].append(entry)

    used_entries = set()
    filled = 0
    skipped = 0
    skipped_names: list[str] = []
    phone_fallback_notes: list[str] = []
    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )

    for row_idx in range(2, dl_ws.max_row + 1):
        # Column E = 5 (tracking number destination)
        e_cell = dl_ws.cell(row=row_idx, column=5)

        # Only fill empty E cells
        if e_cell.value is not None and normalize(e_cell.value) != "":
            continue

        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)     # AA
        dl_phone = phone_digits(dl_ws.cell(row=row_idx, column=28).value)    # AB
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)  # AD
        dl_option_raw = dl_ws.cell(row=row_idx, column=12).value          # L
        dl_option_keys = option_key_set(
            dl_option_raw,
            transform_option(str(dl_option_raw or "")),
        )
        if not dl_option_keys:
            dl_option_keys = option_key_set(dl_ws.cell(row=row_idx, column=11).value)

        if not dl_name:
            continue

        candidates = entry_by_name.get(dl_name, [])
        available = [c for c in candidates if id(c) not in used_entries]

        matched = None
        matched_via_phone_only = False

        if available:
            if requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
                available = [
                    c for c in available
                    if options_match(c.get("option_keys"), dl_option_keys)
                ]
                if not available:
                    skipped += 1
                    skipped_names.append(dl_name)
                    continue

            # Try phone + address match first
            for c in available:
                if c["phone"] == dl_phone and c["address"] == dl_address:
                    matched = c
                    break

            # Try phone only match
            if matched is None:
                for c in available:
                    if c["phone"] == dl_phone:
                        matched = c
                        break

            # Try address only match
            if matched is None:
                for c in available:
                    if c["address"] == dl_address:
                        matched = c
                        break

            # Fall back to first available
            if matched is None:
                matched = available[0]

        # Fallback: 이름 매칭 실패 시 전화번호로 재시도
        # (CS로 수취인/주소가 바뀐 경우 — 전화번호는 보통 그대로 유지)
        if matched is None and dl_phone:
            phone_candidates = [
                c for c in entry_by_phone.get(dl_phone, [])
                if id(c) not in used_entries
            ]
            if requires_option_guard(dl_name, delivery_name_counts, len(phone_candidates)):
                phone_candidates = [
                    c for c in phone_candidates
                    if options_match(c.get("option_keys"), dl_option_keys)
                ]
            if len(phone_candidates) == 1:
                matched = phone_candidates[0]
                matched_via_phone_only = True
            elif len(phone_candidates) > 1:
                # 같은 전화번호 후보가 여러 개면 주소로 좁히기
                for c in phone_candidates:
                    if c["address"] == dl_address:
                        matched = c
                        matched_via_phone_only = True
                        break

        if matched:
            e_cell.value = matched["tracking"]
            used_entries.add(id(matched))
            filled += 1
            if matched_via_phone_only:
                phone_fallback_notes.append(
                    f"{matched['name']}→{dl_name}(전화번호 일치)"
                )
        else:
            skipped += 1
            skipped_names.append(dl_name)

    # Save to bytes
    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList송장파일_고구마({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "haedal_entries": len(haedal_entries),
    }
    if phone_fallback_notes:
        stats["전화번호매칭"] = ", ".join(phone_fallback_notes)
    if skipped_names:
        stats["매칭실패"] = ", ".join(skipped_names)

    return output.read(), filename, stats
