"""Fill DeliveryList tracking numbers from a returned Danharu sheet."""

from __future__ import annotations

import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.processors.danharu_order import build_platform_label
from app.processors.tracking_match import name_counts, options_match, requires_option_guard

KST = timezone(timedelta(hours=9))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_phone(value: object) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def normalize_address(value: object) -> str:
    return normalize_key(str(value or "").replace("*", ""))


def parse_quantity(value: object) -> int:
    try:
        return max(1, int(float(str(value).strip())))
    except (TypeError, ValueError):
        return 1


def _extract_reply_entries(reply_bytes: bytes) -> list[dict]:
    reply_wb = load_workbook(filename=BytesIO(reply_bytes), data_only=True)
    reply_ws = reply_wb.active

    entries: list[dict] = []
    for row_index in range(2, reply_ws.max_row + 1):
        tracking_number = normalize_text(reply_ws.cell(row=row_index, column=13).value)
        if not tracking_number:
            continue

        receiver_name = normalize_text(reply_ws.cell(row=row_index, column=5).value)
        if not receiver_name:
            continue

        entries.append(
            {
                "tracking_number": tracking_number,
                "receiver_name_key": normalize_key(receiver_name),
                "receiver_phone_key": normalize_phone(reply_ws.cell(row=row_index, column=6).value),
                "receiver_address_key": normalize_address(reply_ws.cell(row=row_index, column=9).value),
                "platform_label_key": normalize_key(reply_ws.cell(row=row_index, column=11).value),
                "supply_option_key": normalize_key(reply_ws.cell(row=row_index, column=12).value),
                "quantity": parse_quantity(reply_ws.cell(row=row_index, column=10).value),
            }
        )

    return entries


def _score_match(
    entry: dict,
    phone_key: str,
    address_key: str,
    platform_label_key: str,
    option_key: str,
    quantity: int,
) -> int:
    score = 0
    if entry["receiver_phone_key"] and phone_key and entry["receiver_phone_key"] == phone_key:
        score += 4
    if entry["receiver_address_key"] and address_key and entry["receiver_address_key"] == address_key:
        score += 3
    if entry["platform_label_key"] and platform_label_key and entry["platform_label_key"] == platform_label_key:
        score += 2
    if entry["supply_option_key"] and option_key and entry["supply_option_key"] == option_key:
        score += 2
    if quantity > 0 and entry["quantity"] == quantity:
        score += 1
    return score


def process(
    reply_bytes: bytes | list[bytes],
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """Write Danharu tracking numbers into DeliveryList column E only.

    `reply_bytes` 는 단일 파일(bytes) 또는 여러 파일(list[bytes])을 받는다.
    여러 회신 파일에서 추출한 항목을 하나로 합쳐 DeliveryList에 적용한다.
    """
    if isinstance(reply_bytes, (bytes, bytearray)):
        reply_files = [bytes(reply_bytes)]
    else:
        reply_files = [bytes(b) for b in reply_bytes]

    reply_entries: list[dict] = []
    for raw in reply_files:
        reply_entries.extend(_extract_reply_entries(raw))

    if not reply_entries:
        raise ValueError("회신 파일에서 M열 송장번호가 입력된 주문을 찾지 못했습니다.")

    delivery_wb = load_workbook(filename=BytesIO(delivery_bytes))
    first_sheet_name = delivery_wb.sheetnames[0]
    for sheet_name in delivery_wb.sheetnames[1:]:
        del delivery_wb[sheet_name]
    delivery_ws = delivery_wb[first_sheet_name]

    entries_by_name: dict[str, list[dict]] = {}
    for entry in reply_entries:
        entries_by_name.setdefault(entry["receiver_name_key"], []).append(entry)

    used_entries: set[int] = set()
    filled = 0
    skipped = 0
    delivery_name_counts = name_counts(
        normalize_key(delivery_ws.cell(row=row_index, column=27).value)
        for row_index in range(2, delivery_ws.max_row + 1)
    )

    for row_index in range(2, delivery_ws.max_row + 1):
        tracking_cell = delivery_ws.cell(row=row_index, column=5)
        if normalize_text(tracking_cell.value):
            continue

        receiver_name_key = normalize_key(delivery_ws.cell(row=row_index, column=27).value)
        if not receiver_name_key:
            continue

        receiver_phone_key = normalize_phone(delivery_ws.cell(row=row_index, column=28).value)
        receiver_address_key = normalize_address(delivery_ws.cell(row=row_index, column=30).value)
        product_name = normalize_text(delivery_ws.cell(row=row_index, column=11).value)
        option_name = normalize_text(delivery_ws.cell(row=row_index, column=12).value)
        platform_label_key = normalize_key(build_platform_label(product_name, option_name))
        option_key = normalize_key(option_name)
        quantity = parse_quantity(delivery_ws.cell(row=row_index, column=23).value)

        candidates = [
            entry for entry in entries_by_name.get(receiver_name_key, [])
            if id(entry) not in used_entries
        ]
        if not candidates:
            skipped += 1
            continue

        if requires_option_guard(
            receiver_name_key,
            delivery_name_counts,
            len(entries_by_name.get(receiver_name_key, [])),
        ):
            target_option_keys = {key for key in (platform_label_key, option_key) if key}
            candidates = [
                entry for entry in candidates
                if options_match(
                    {entry.get("platform_label_key", ""), entry.get("supply_option_key", "")},
                    target_option_keys,
                )
            ]
            if not candidates:
                skipped += 1
                continue

        scored_candidates = [
            (
                _score_match(
                    entry,
                    receiver_phone_key,
                    receiver_address_key,
                    platform_label_key,
                    option_key,
                    quantity,
                ),
                entry,
            )
            for entry in candidates
        ]
        scored_candidates.sort(key=lambda item: item[0], reverse=True)
        best_score, best_entry = scored_candidates[0]

        if best_score <= 0 and len(candidates) > 1:
            skipped += 1
            continue

        tracking_cell.value = best_entry["tracking_number"]
        used_entries.add(id(best_entry))
        filled += 1

    output = BytesIO()
    delivery_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_단하루_송장입력_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "reply_rows": len(reply_entries),
        "reply_files": len(reply_files),
    }
    return output.read(), filename, stats
