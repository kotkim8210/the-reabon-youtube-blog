import re
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


KST = timezone(timedelta(hours=9))

COL_MAP = {
    "order_no": 3,    # C열 (1-indexed)
    "name": 27,       # AA열
    "phone": 28,      # AB열
    "address": 30,    # AD열
    "product": 11,    # K열
    "option": 12,     # L열
    "qty": 23,        # W열
    "memo": 31,       # AE열
}

DELIVERY_HEADERS = [
    "번호",
    "묶음배송번호",
    "주문번호",
    "택배사",
    "운송장번호",
    "분리배송 Y/N",
    "분리배송 출고예정일",
    "주문시 출고예정일",
    "출고일(발송일)",
    "주문일",
    "등록상품명",
    "등록옵션명",
    "노출상품명(옵션명)",
    "노출상품ID",
    "옵션ID",
    "최초등록등록상품명/옵션명",
    "업체상품코드",
    "바코드",
    "결제액",
    "배송비구분",
    "배송비",
    "도서산간 추가배송비",
    "구매수(수량)",
    "옵션판매가(판매단가)",
    "구매자",
    "구매자전화번호",
    "수취인이름",
    "수취인전화번호",
    "우편번호",
    "수취인 주소",
    "배송메세지",
    "상품별 추가메시지",
    "주문자 추가메시지",
    "배송완료일",
    "구매확정일자",
    "개인통관번호(PCCC)",
    "통관용수취인전화번호",
    "기타",
    "결제위치",
    "배송유형",
]

GAEGEOLMU_PRODUCT_NAME = "식품애착 게걸무씨앗기름 폐 기침 기관지"


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_order_no(value) -> str:
    text = normalize(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _format_excel_datetime(value, *, date_only: bool = False) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d" if date_only else "%Y-%m-%d %H:%M:%S")

    text = normalize(value)
    if not text:
        return ""

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d %H:%M:%S", "%Y.%m.%d"):
        try:
            parsed = datetime.strptime(text[:19], fmt)
            return parsed.strftime("%Y-%m-%d" if date_only else "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text[:10] if date_only and len(text) >= 10 else text


def _cell_by_header(ws, headers: dict[str, int], row_number: int, header: str):
    col = headers.get(header)
    if not col:
        return None
    return ws.cell(row=row_number, column=col).value


def _normalize_carrier(value) -> str:
    text = normalize(value)
    compact = text.replace(" ", "")
    if "CJ" in compact.upper() or "대한통운" in compact:
        return "CJ 대한통운"
    if "롯데" in compact:
        return "롯데택배"
    if "한진" in compact:
        return "한진택배"
    return text


def _number_text(value) -> str:
    text = normalize_order_no(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _money_text(value) -> str:
    text = normalize_order_no(value)
    if not text:
        return "0"
    try:
        return str(int(float(text.replace(",", ""))))
    except ValueError:
        return text


def _gaegeolmu_option_display(option_name: str) -> tuple[str, str]:
    if "2병" in option_name:
        return "2개 180ml", f"{GAEGEOLMU_PRODUCT_NAME}, 2개, 180ml (180ml 2개)"
    return "1개 180ml", f"{GAEGEOLMU_PRODUCT_NAME}, 1개, 180ml (180ml 1개)"


def _apply_delivery_sheet_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="1F2937")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        1: 8,
        2: 18,
        3: 18,
        4: 14,
        5: 18,
        10: 20,
        11: 34,
        12: 16,
        13: 46,
        16: 40,
        19: 12,
        23: 12,
        24: 16,
        25: 12,
        26: 18,
        27: 12,
        28: 18,
        29: 12,
        30: 48,
        31: 28,
        39: 16,
        40: 14,
    }
    for col_idx in range(1, len(DELIVERY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)
    ws.freeze_panes = "A2"


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize(ws.cell(row=1, column=col).value)
        if header:
            headers[header] = col
    return headers


def _is_gmarket_sheet(ws) -> bool:
    headers = _header_map(ws)
    return all(key in headers for key in ("판매아이디", "주문번호", "상품명", "수령인명"))


def canonical_gaegeolmu_option(product_text: str, option_text: str = "") -> str:
    raw = f"{normalize(product_text)} {normalize(option_text)}"
    compact = re.sub(r"\s+", "", raw)
    if (
        "1+1" in compact
        or "2병" in raw
        or "2 병" in raw
        or "2개" in raw
        or "2 개" in raw
        or "2개입" in compact
        or "2세트" in raw
        or "2 세트" in raw
    ):
        return "게걸무씨앗기름 2병"
    if "게걸무" in raw:
        return "게걸무씨앗기름 1병"
    return normalize(option_text) or normalize(product_text) or "게걸무씨앗기름"


def _looks_like_header_row(ws, row_number: int) -> bool:
    order_no = normalize(ws.cell(row=row_number, column=COL_MAP["order_no"]).value)
    product = normalize(ws.cell(row=row_number, column=COL_MAP["product"]).value)
    option = normalize(ws.cell(row=row_number, column=COL_MAP["option"]).value)
    header_text = f"{order_no} {product} {option}"
    return any(token in header_text for token in ("주문번호", "등록상품명", "쿠팡옵션", "옵션명"))


def _collect_option_total(option_totals: dict[str, dict], option_name: str, order_no: str, qty_int: int) -> None:
    bucket = option_totals.setdefault(
        option_name or "게걸무씨앗기름",
        {
            "coupang_option_keyword": option_name or "게걸무씨앗기름",
            "vendor_option_name": option_name or "게걸무씨앗기름",
            "quantity": 0,
            "orders": [],
        },
    )
    bucket["quantity"] += qty_int
    if order_no:
        bucket["orders"].append({"order_id": order_no, "quantity": qty_int})


def _process_coupang_delivery(dl_ws) -> tuple[int, list[dict]]:
    matching_row_numbers: list[int] = []
    option_totals: dict[str, dict] = {}
    header_rows: set[int] = set()

    for row_number in range(1, dl_ws.max_row + 1):
        if _looks_like_header_row(dl_ws, row_number):
            header_rows.add(row_number)
            continue

        product_text = normalize(dl_ws.cell(row=row_number, column=COL_MAP["product"]).value)
        if "게걸무" not in product_text:
            continue

        matching_row_numbers.append(row_number)
        order_no = normalize_order_no(dl_ws.cell(row=row_number, column=COL_MAP["order_no"]).value)
        option_text = normalize(dl_ws.cell(row=row_number, column=COL_MAP["option"]).value)
        qty_text = normalize(dl_ws.cell(row=row_number, column=COL_MAP["qty"]).value)
        try:
            qty_int = int(float(qty_text)) if qty_text else 1
        except (ValueError, TypeError):
            qty_int = 1
        option_name = canonical_gaegeolmu_option(product_text, option_text)
        _collect_option_total(option_totals, option_name, order_no, qty_int)

    matching_set = set(matching_row_numbers) | header_rows
    for row_number in range(dl_ws.max_row, 0, -1):
        if row_number not in matching_set:
            dl_ws.delete_rows(row_number, 1)

    return len(matching_row_numbers), list(option_totals.values())


def _process_gmarket(ws) -> tuple[Workbook, int, list[dict]]:
    """지마켓 신규주문 시트 단독 처리 — 쿠팡 DeliveryList 양식 워크북으로 변환."""
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Delivery"
    out_ws.append(DELIVERY_HEADERS)
    _apply_delivery_sheet_style(out_ws)

    added, option_totals = _append_gmarket_rows(ws, out_ws, start_index=0)
    return out_wb, added, list(option_totals.values())


def _append_gmarket_rows(
    ws,
    out_ws,
    start_index: int = 0,
    option_totals: dict[str, dict] | None = None,
) -> tuple[int, dict[str, dict]]:
    """지마켓 시트의 게걸무 주문을 DeliveryList 양식 행으로 out_ws에 이어붙인다.

    start_index: 이미 들어있는 행 수(번호 컬럼 이어가기용).
    option_totals: 쿠팡 집계와 합칠 때 넘기면 같은 옵션끼리 합산된다.
    """
    headers = _header_map(ws)
    if option_totals is None:
        option_totals = {}

    output_row_index = start_index
    for row_number in range(2, ws.max_row + 1):
        product_text = normalize(_cell_by_header(ws, headers, row_number, "상품명"))
        option_text = normalize(_cell_by_header(ws, headers, row_number, "옵션"))
        if "게걸무" not in f"{product_text} {option_text}":
            continue

        output_row_index += 1
        order_no = _number_text(_cell_by_header(ws, headers, row_number, "주문번호"))
        qty_text = normalize(_cell_by_header(ws, headers, row_number, "수량"))
        try:
            qty_int = int(float(qty_text)) if qty_text else 1
        except (ValueError, TypeError):
            qty_int = 1

        option_name = canonical_gaegeolmu_option(product_text, option_text)
        option_display, exposed_option = _gaegeolmu_option_display(option_name)
        _collect_option_total(option_totals, option_name, order_no, qty_int)

        delivery_fee = _money_text(_cell_by_header(ws, headers, row_number, "배송비 금액"))
        delivery_fee_type = "무료" if delivery_fee in ("", "0") else "유료"
        buyer_phone = (
            normalize(_cell_by_header(ws, headers, row_number, "구매자 휴대폰"))
            or normalize(_cell_by_header(ws, headers, row_number, "구매자 전화번호"))
            or normalize(_cell_by_header(ws, headers, row_number, "수령인 휴대폰"))
        )
        recipient_phone = (
            normalize(_cell_by_header(ws, headers, row_number, "수령인 휴대폰"))
            or normalize(_cell_by_header(ws, headers, row_number, "수령인 전화번호"))
        )

        out_ws.append(
            [
                output_row_index,
                _number_text(_cell_by_header(ws, headers, row_number, "배송번호")),
                order_no,
                _normalize_carrier(_cell_by_header(ws, headers, row_number, "택배사명(발송방법)")),
                _number_text(_cell_by_header(ws, headers, row_number, "송장번호")),
                "분리배송불가",
                "",
                _format_excel_datetime(_cell_by_header(ws, headers, row_number, "발송마감일"), date_only=True),
                "",
                _format_excel_datetime(_cell_by_header(ws, headers, row_number, "주문일(결제확인전)")),
                GAEGEOLMU_PRODUCT_NAME,
                option_display,
                exposed_option,
                _number_text(_cell_by_header(ws, headers, row_number, "상품번호")),
                "",
                f"{GAEGEOLMU_PRODUCT_NAME},{option_display}",
                normalize(_cell_by_header(ws, headers, row_number, "판매자 관리코드")),
                "",
                _money_text(_cell_by_header(ws, headers, row_number, "판매금액")),
                delivery_fee_type,
                delivery_fee,
                "0",
                qty_int,
                _money_text(_cell_by_header(ws, headers, row_number, "판매단가")),
                normalize(_cell_by_header(ws, headers, row_number, "구매자명")),
                buyer_phone,
                normalize(_cell_by_header(ws, headers, row_number, "수령인명")),
                recipient_phone,
                _number_text(_cell_by_header(ws, headers, row_number, "우편번호")),
                normalize(_cell_by_header(ws, headers, row_number, "주소")),
                normalize(_cell_by_header(ws, headers, row_number, "배송시 요구사항")),
                "",
                "",
                "",
                "",
                normalize(_cell_by_header(ws, headers, row_number, "수령인 통관정보")),
                "",
                "",
                normalize(_cell_by_header(ws, headers, row_number, "판매아이디")) or "지마켓",
                "판매자 배송",
            ]
        )

    return output_row_index - start_index, option_totals


def process(
    delivery_file_bytes: bytes,
    gmarket_file_bytes: bytes | None = None,
) -> tuple[bytes, str, dict]:
    """게걸무 발주서 생성.

    delivery_file_bytes: 쿠팡 DeliveryList 또는 지마켓 신규주문(단독 업로드도 지원).
    gmarket_file_bytes: 지마켓 신규주문(선택) — 쿠팡 발주서 뒤에 이어붙여 한 장으로 합친다
        (2026-08-17 요청: 지마켓 주문도 게걸무 발주서에 합쳐서 출력).
    """
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes))
    dl_ws = dl_wb.active

    if _is_gmarket_sheet(dl_ws):
        # 첫 칸에 지마켓 파일을 올린 경우 — 종전처럼 단독 처리
        output_wb, total, options = _process_gmarket(dl_ws)
        output_ws = output_wb.active
        option_totals = {bucket["coupang_option_keyword"]: bucket for bucket in options}
        sources = ["지마켓"]
    else:
        total, options = _process_coupang_delivery(dl_ws)
        output_wb = dl_wb
        output_ws = dl_ws
        option_totals = {bucket["coupang_option_keyword"]: bucket for bucket in options}
        sources = ["쿠팡"]

    gmarket_added = 0
    if gmarket_file_bytes:
        gm_ws = load_workbook(filename=BytesIO(gmarket_file_bytes)).active
        if not _is_gmarket_sheet(gm_ws):
            raise ValueError(
                "지마켓 파일 형식이 아닙니다. 지마켓 '신규주문' 엑셀(판매아이디·주문번호·수령인명 헤더)을 올려주세요."
            )
        gmarket_added, option_totals = _append_gmarket_rows(
            gm_ws, output_ws, start_index=total, option_totals=option_totals
        )
        if gmarket_added:
            sources.append("지마켓")
        total += gmarket_added

    output = BytesIO()
    output_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_게걸무씨앗기름_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": total,
        "source": "+".join(sources),
        "product": "게걸무씨앗기름",
        "options": list(option_totals.values()),
    }
    if gmarket_file_bytes:
        stats["gmarket"] = gmarket_added
        stats["coupang"] = total - gmarket_added

    return output.read(), filename, stats
