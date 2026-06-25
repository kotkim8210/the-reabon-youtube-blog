import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def convert_quantity(option_text: str) -> str | None:
    """쿠팡 '가정용 혼합과' 옵션 → 쥬얼리팜 발주 품목명.

    2026-06부터 알뜰과 발주처가 제주다팜 → 쥬얼리팜(쥬얼리프룻)으로 전환.
    """
    text = normalize(option_text)
    if "가정용 혼합과" not in text:
        return None

    match = re.search(r"(\d+)\s*kg", text, re.IGNORECASE)
    if not match:
        return None

    weight = match.group(1)
    if weight not in ("2", "3", "5"):
        return None
    return f"가성비 혼합과 {weight}kg"


def clear_stray_header_numbers(ws) -> None:
    for col in range(14, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value in (6, "6"):
            cell.value = None


def process(delivery_file_bytes: bytes) -> tuple[bytes, str, dict] | None:
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows: list[tuple] = []
    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        converted = convert_quantity(option)
        if "성주참외" in product_name and converted:
            filtered_rows.append((row, converted))

    if not filtered_rows:
        return None

    # 쥬얼리프룻(pbfcompany) 발주서 양식으로 출력 (2026-06 발주처 전환)
    template_path = TEMPLATE_DIR / "명이나물_pbfcompany_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            "Please place the pbfcompany template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)
    now = datetime.now(KST)
    today_str = now.strftime("%Y-%m-%d")

    start_row = 2
    for row_idx in range(2, ws.max_row + 2):
        if ws.cell(row=row_idx, column=3).value is None:
            start_row = row_idx
            break

    for index, (row, product_name) in enumerate(filtered_rows):
        out_row = start_row + index

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        qty_value = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_value)
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        mapping = {
            1: today_str,           # A - 일자
            2: "아이티소프트",       # B - 거래처명
            3: name,                # C - 받는분성명
            4: phone,               # D - 받는분전화번호
            6: address,             # F - 받는분주소
            7: product_name,        # G - 품목명 (가성비 혼합과 Nkg)
            8: qty,                 # H - 수량
            9: "식품애착",           # I - 보내는분성명
            10: "010-5700-7756",    # J - 보내는분전화번호
            12: memo,               # L - 배송메시지
            13: order_no,           # M - 주문번호
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

    option_totals: dict[str, dict] = {}
    for row, product_name in filtered_rows:
        option = normalize(row[11].value) if len(row) > 11 else product_name
        qty_value = row[22].value if len(row) > 22 else 1
        order_no = normalize(row[2].value) if len(row) > 2 else ""
        try:
            qty_int = int(qty_value) if qty_value else 1
        except (TypeError, ValueError):
            qty_int = 1
        bucket = option_totals.setdefault(
            option,
            {
                "coupang_option_keyword": option,
                "vendor_option_name": product_name,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    filename = f"쥬얼리프룻_참외혼합과_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows),
        "product": "성주참외 혼합과(쥬얼리팜)",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats
