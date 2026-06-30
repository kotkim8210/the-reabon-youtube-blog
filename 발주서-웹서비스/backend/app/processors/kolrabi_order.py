import re
from datetime import datetime, timezone, timedelta
from io import BytesIO
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def convert_quantity(option_text: str) -> str | None:
    """콜라비 옵션 → '콜라비 정품 {kg}kg'.

    옵션 표기가 '3kg 1박스'·'1박스 3kg'·'로얄과 특품 3kg' 등 어떤 순서/접두여도 3·5·10kg을
    추출해 매칭한다. (process()에서 상품명에 '콜라비'가 있을 때만 호출 → 과매칭 없음)
    """
    if not option_text:
        return None
    match = re.search(r"(\d+)\s*kg", str(option_text), re.IGNORECASE)
    if match and match.group(1) in ("3", "5", "10"):
        return f"콜라비 정품 {match.group(1)}kg"
    return None


def _combined_text(product_name: object, option_text: object) -> str:
    return normalize(f"{product_name or ''} {option_text or ''}")


def is_jejudafarm_corn_order(product_name: object, option_text: object) -> bool:
    """초당옥수수(제주다팜 발주) 여부. 애플초당옥수수(쥬얼리프룻)는 제외."""
    text = re.sub(r"\s+", "", _combined_text(product_name, option_text))
    return "초당옥수수" in text and "애플" not in text


def convert_corn_option(product_name: object, option_text: object) -> str | None:
    """초당옥수수 DeliveryList → 제주다팜 발주 옵션 '초당옥수수({중품/특품}) {n}개'.

    애플초당옥수수는 제외(쥬얼리프룻). 등급(중품/특품)과 수량(개/입)이 있어야 발주 대상.
    """
    if not is_jejudafarm_corn_order(product_name, option_text):
        return None
    text = _combined_text(product_name, option_text)
    count_match = re.search(r"(\d+)\s*(?:개|입)", text)
    grade = "중품" if "중품" in text else "특품" if "특품" in text else ""
    if not count_match or not grade:
        return None
    return f"초당옥수수({grade}) {count_match.group(1)}개"


# 제주다팜 미니밤호박(보우짱 로얄과) 옵션 중량.
# 2026-06부터 3·5·10kg은 쥬얼리프룻(myeongi)이 더 싼 공급가로 발주 전환 →
# 제주다팜은 1kg만 발주(쥬얼리 시트에 1kg 규격 없음). 3/5/10kg은 myeongi_order에서 처리.
BAMHOBAK_WEIGHTS = {"1"}


def is_bamhobak_order(product_name: object, option_text: object) -> bool:
    """제주 미니밤호박(보우짱) 제주다팜 발주 여부 (1kg만; 3·5·10kg은 쥬얼리프룻)."""
    text = re.sub(r"\s+", "", _combined_text(product_name, option_text))
    return "밤호박" in text


def convert_bamhobak_option(product_name: object, option_text: object) -> str | None:
    """미니밤호박 DeliveryList → 제주다팜 발주 품목 '제주 미니밤호박 보우짱 로얄과 {n}kg'.

    DeliveryList 옵션 예: '1박스 로얄 정품 1kg' → '제주 미니밤호박 보우짱 로얄과 1kg'.
    제주다팜 발주 대상은 1kg만 (3·5·10kg은 쥬얼리프룻 발주로 이관).
    """
    if not is_bamhobak_order(product_name, option_text):
        return None
    text = _combined_text(product_name, option_text)
    match = re.search(r"(\d+)\s*kg", text, re.IGNORECASE)
    if not match or match.group(1) not in BAMHOBAK_WEIGHTS:
        return None
    return f"제주 미니밤호박 보우짱 로얄과 {match.group(1)}kg"


def clear_stray_header_numbers(ws) -> None:
    for col in range(14, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value in (6, "6"):
            cell.value = None


def process(delivery_file_bytes: bytes) -> tuple[bytes, str, dict]:
    """Process delivery list for 콜라비 orders.

    Args:
        delivery_file_bytes: Raw bytes of the DeliveryList Excel file.

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    # Load delivery list
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    # Filter rows where column K (11) contains '콜라비' and option is valid
    filtered_rows = []
    for row in dl_ws.iter_rows(min_row=2):
        k_val = normalize(row[10].value) if len(row) > 10 else ""  # Column K = index 10
        option = str(row[11].value).strip() if len(row) > 11 and row[11].value else ""
        if "콜라비" in k_val and convert_quantity(option) is not None:
            filtered_rows.append(row)

    # Load template
    template_path = TEMPLATE_DIR / "콜라비_제주다팜_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    # Keep only first sheet
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]
    clear_stray_header_numbers(ws)

    font11 = Font(size=11)

    # Find first empty row in template (look at column B for name)
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=2).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}

    for i, row in enumerate(filtered_rows):
        out_row = start_row + i

        # Column mappings (source 0-indexed):
        # AA=26 -> B(2) name
        # AB=27 -> C(3) phone
        # AD=29 -> F(6) address
        # L=11  -> H(8) product (converted)
        # W=22  -> I(9) quantity
        # AE=30 -> M(13) memo
        # AC=28 -> E(5) zipcode

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        zipcode = normalize(row[28].value) if len(row) > 28 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        product = convert_quantity(option)
        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        # Zipcode with zero-fill to 5 digits
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = zipcode.zfill(5)

        mapping = {
            2: name,       # B - name
            3: phone,      # C - phone
            5: zipcode,    # E - zipcode
            6: address,    # F - address
            8: product,    # H - product
            9: qty,        # I - quantity
            10: "식품애착",  # J - fixed
            11: "010-5700-7756",  # K - fixed
            13: memo,      # M - memo
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        if product:
            bucket = option_totals.setdefault(
                option or product,
                {
                    "coupang_option_keyword": option or product,
                    "vendor_option_name": product,
                    "quantity": 0,
                    "orders": [],
                },
            )
            bucket["quantity"] += qty_int
            if order_no:
                bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    # Save to bytes
    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"제주다팜_아이티소프트_콜라비발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows),
        "product": "콜라비",
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats


def process_corn(delivery_file_bytes: bytes) -> tuple[bytes, str, dict] | None:
    """Process DeliveryList rows for 제주다팜 초당옥수수 발주."""
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows: list[tuple] = []
    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        converted = convert_corn_option(product_name, option)
        if converted:
            filtered_rows.append((row, converted))

    if not filtered_rows:
        return None

    template_path = TEMPLATE_DIR / "콜라비_제주다팜_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]
    clear_stray_header_numbers(ws)

    font11 = Font(size=11)
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=2).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}
    for i, (row, product) in enumerate(filtered_rows):
        out_row = start_row + i

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        zipcode = normalize(row[28].value) if len(row) > 28 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = zipcode.zfill(5)

        mapping = {
            2: name,
            3: phone,
            5: zipcode,
            6: address,
            8: product,
            9: qty,
            10: "식품애착",
            11: "010-5700-7756",
            13: memo,
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            option or product,
            {
                "coupang_option_keyword": option or product,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"제주다팜_아이티소프트_초당옥수수발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows),
        "product": "초당옥수수(제주다팜)",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats


def process_bamhobak(
    delivery_file_bytes: bytes,
    toss_entries: list[dict] | None = None,
) -> tuple[bytes, str, dict] | None:
    """Process DeliveryList rows for 제주다팜 미니밤호박(보우짱 로얄과) 발주.

    toss_entries: 토스 API에서 수집한 미니밤호박 1kg 주문(선택). 각 entry는 발주
    품목명(product)이 확정돼 있어 DeliveryList 행 뒤에 그대로 추가된다.
    """
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows: list[tuple] = []
    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        converted = convert_bamhobak_option(product_name, option)
        if converted:
            filtered_rows.append((row, converted))

    if not filtered_rows and not toss_entries:
        return None

    template_path = TEMPLATE_DIR / "콜라비_제주다팜_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]
    clear_stray_header_numbers(ws)

    font11 = Font(size=11)
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=2).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}
    for i, (row, product) in enumerate(filtered_rows):
        out_row = start_row + i

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        zipcode = normalize(row[28].value) if len(row) > 28 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = zipcode.zfill(5)

        mapping = {
            2: name,
            3: phone,
            5: zipcode,
            6: address,
            8: product,
            9: qty,
            10: "식품애착",
            11: "010-5700-7756",
            13: memo,
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            option or product,
            {
                "coupang_option_keyword": option or product,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    # 토스 미니밤호박 1kg entry는 발주 품목명(product)이 확정돼 있어 DeliveryList 행 뒤에 추가.
    base_row = start_row + len(filtered_rows)
    for j, entry in enumerate(toss_entries or []):
        out_row = base_row + j
        name = normalize(entry.get("name"))
        phone = normalize(entry.get("phone"))
        zipcode = normalize(entry.get("zipcode"))
        address = normalize(entry.get("address"))
        memo = normalize(entry.get("memo"))
        product = entry.get("product") or ""
        try:
            qty_int = int(float(entry.get("qty"))) if entry.get("qty") not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        order_no = entry.get("order_id") or ""
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = zipcode.zfill(5)

        mapping = {
            2: name,
            3: phone,
            5: zipcode,
            6: address,
            8: product,
            9: str(qty_int),
            10: "식품애착",
            11: "010-5700-7756",
            13: memo,
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            entry.get("option") or product,
            {
                "coupang_option_keyword": entry.get("option") or product,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"제주다팜_아이티소프트_미니밤호박발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows) + len(toss_entries or []),
        "product": "미니밤호박(제주다팜)",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats


def process_outputs(
    delivery_file_bytes: bytes,
    toss_bamhobak_entries: list[dict] | None = None,
) -> list[tuple[bytes, str, dict]]:
    """제주다팜 발주서 목록 반환 — 콜라비 + 초당옥수수 + 미니밤호박.

    toss_bamhobak_entries: 토스 미니밤호박 1kg 주문(선택) → 미니밤호박 발주서에 합침.

    ※ 성주참외 알뜰과(쥬얼리팜)는 2026-06부터 쥬얼리팜(myeongi) 발주로 이관 —
      DeliveryList의 알뜰과는 명이나물(쥬얼리프룻) 메뉴에서 함께 출력된다.
    ※ 초당옥수수는 2026-06부터 제이비티가 아닌 제주다팜 발주로 이관(공급가 인하).
      애플초당옥수수는 계속 쥬얼리프룻(myeongi) 발주.
    """
    results: list[tuple[bytes, str, dict]] = []

    kolrabi_result = process(delivery_file_bytes)
    kolrabi_stats = kolrabi_result[2] if len(kolrabi_result) > 2 else {}
    if int((kolrabi_stats or {}).get("total") or 0) > 0:
        results.append(kolrabi_result)

    # 초당옥수수는 2026-06부터 쥬얼리프룻(myeongi) 발주로 이관 — 여기서 출력하지 않는다.

    bamhobak_result = process_bamhobak(delivery_file_bytes, toss_entries=toss_bamhobak_entries)
    if bamhobak_result and int((bamhobak_result[2] or {}).get("total") or 0) > 0:
        results.append(bamhobak_result)

    return results
