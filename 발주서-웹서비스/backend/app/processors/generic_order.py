"""Generic order processor: user-configured product→template mapping."""

import re
import logging
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from app.db import get_user_products, get_user_template_data

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return re.sub(r"\s+", " ", s)


def apply_option_template(vendor_name: str, source_option: str) -> str:
    """vendor_option_name의 placeholder를 주문 옵션 텍스트로 치환.

    - {kg}  → 옵션에서 'Nkg' 추출 (예: '5kg')
    - {n}   → 옵션에서 'N개/N입' 추출 (예: '10개')
    - {원본} → 주문 옵션 원문 그대로
    placeholder가 없으면 vendor_name을 그대로 반환(하위호환).
    예: vendor='꿀고구마 {kg}' + 주문옵션 '5kg 1박스' → '꿀고구마 5kg'
    """
    if not vendor_name or "{" not in vendor_name:
        return vendor_name
    result = vendor_name
    if "{kg}" in result:
        m = re.search(r"(\d+(?:\.\d+)?)\s*kg", source_option, re.IGNORECASE)
        result = result.replace("{kg}", f"{m.group(1)}kg" if m else "")
    if "{n}" in result:
        m = re.search(r"(\d+)\s*(?:개|입)", source_option)
        result = result.replace("{n}", f"{m.group(1)}개" if m else "")
    if "{원본}" in result:
        result = result.replace("{원본}", source_option)
    return re.sub(r"\s+", " ", result).strip()


async def process_generic_order(
    user_id: int,
    delivery_bytes: bytes,
) -> list[tuple[bytes, str, dict]]:
    """Process DeliveryList against all active user products.

    Returns list of (file_bytes, filename, stats_dict) tuples.
    """
    products = await get_user_products(user_id)
    if not products:
        return []

    # Load delivery list
    dl_wb = load_workbook(filename=BytesIO(delivery_bytes), data_only=True)
    dl_ws = dl_wb.active

    # Read all rows (skip header row 1)
    rows = list(dl_ws.iter_rows(min_row=2))

    results = []

    for product in products:
        keyword = product["coupang_product_keyword"]
        options = product.get("options", [])
        if not options:
            continue

        # Filter rows: K column (index 10) contains keyword
        matched_rows = []
        for row in rows:
            k_val = normalize(row[10].value) if len(row) > 10 else ""
            if keyword in k_val:
                matched_rows.append(row)

        if not matched_rows:
            continue

        # Determine mode: sum-to-cell vs row-append
        has_target_cells = any(o.get("template_target_cell") for o in options)

        if has_target_cells:
            result = await _process_sum_mode(product, matched_rows, user_id)
        else:
            result = await _process_row_mode(product, matched_rows, user_id)

        if result:
            results.append(result)

    return results


async def _process_sum_mode(
    product: dict,
    matched_rows: list,
    user_id: int,
) -> tuple[bytes, str, dict] | None:
    """Mode A: Sum quantities per option and write to specific cells in template."""
    options = product.get("options", [])

    # Accumulate quantities per option
    cell_totals: dict[str, int] = {}
    total_count = 0

    option_totals: dict[str, dict] = {}  # keyword → {vendor_option_name, quantity}
    for row in matched_rows:
        l_val = normalize(row[11].value) if len(row) > 11 else ""  # L column = option
        order_no = normalize(row[2].value) if len(row) > 2 else ""  # C column = order number
        qty_raw = row[22].value if len(row) > 22 else 0  # W column = quantity
        try:
            qty = int(qty_raw) if qty_raw else 1
        except (ValueError, TypeError):
            qty = 1

        for opt in options:
            opt_keyword = opt["coupang_option_keyword"]
            target_cell = opt.get("template_target_cell")
            if not target_cell:
                continue
            if opt_keyword in l_val:
                cell_totals[target_cell] = cell_totals.get(target_cell, 0) + qty
                total_count += qty
                bucket = option_totals.setdefault(opt_keyword, {
                    "coupang_option_keyword": opt_keyword,
                    "vendor_option_name": opt.get("vendor_option_name", ""),
                    "quantity": 0,
                    "orders": [],
                })
                bucket["quantity"] += qty
                if order_no:
                    bucket["orders"].append({"order_id": order_no, "quantity": qty})
                break

    if not cell_totals:
        return None

    # Load template
    template_id = product.get("template_id")
    if not template_id:
        return None

    template_data = await get_user_template_data(template_id, user_id)
    if not template_data:
        return None

    wb = load_workbook(filename=BytesIO(template_data))
    ws = wb.active

    font11 = Font(size=11)
    for cell_ref, total in cell_totals.items():
        ws[cell_ref] = total
        ws[cell_ref].font = font11

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = product["output_filename_pattern"].format(
        label=product["product_label"],
        date=now.strftime("%Y%m%d"),
    )
    stats = {
        "total": total_count,
        "product": product["product_label"],
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats


async def _process_row_mode(
    product: dict,
    matched_rows: list,
    user_id: int,
) -> tuple[bytes, str, dict] | None:
    """Mode B: Append rows to template (or new workbook) for each matched order."""
    options = product.get("options", [])

    # Build option keyword → vendor_option_name mapping
    option_map: dict[str, str] = {}
    for opt in options:
        option_map[opt["coupang_option_keyword"]] = opt["vendor_option_name"]

    # Load template or create new workbook
    template_id = product.get("template_id")
    if template_id:
        template_data = await get_user_template_data(template_id, user_id)
        if template_data:
            wb = load_workbook(filename=BytesIO(template_data))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # Find first empty row
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    font11 = Font(size=11)
    row_count = 0
    option_totals: dict[str, dict] = {}

    for row in matched_rows:
        l_val = normalize(row[11].value) if len(row) > 11 else ""
        name = normalize(row[26].value) if len(row) > 26 else ""      # AA
        phone = normalize(row[27].value) if len(row) > 27 else ""     # AB
        zipcode = normalize(row[28].value) if len(row) > 28 else ""   # AC
        address = normalize(row[29].value) if len(row) > 29 else ""   # AD
        memo = normalize(row[30].value) if len(row) > 30 else ""      # AE
        order_no = normalize(row[2].value) if len(row) > 2 else ""    # C
        qty_raw = row[22].value if len(row) > 22 else ""              # W
        qty = normalize(qty_raw)

        address = address.replace("*", "")

        # Map option + collect per-option sales
        vendor_option = l_val
        matched_kw = None
        for opt_kw, vendor_name in option_map.items():
            if opt_kw in l_val:
                vendor_option = apply_option_template(vendor_name, l_val)
                matched_kw = opt_kw
                break
        try:
            row_qty = int(qty_raw) if qty_raw else 1
        except (ValueError, TypeError):
            row_qty = 1
        if matched_kw is not None:
            bucket = option_totals.setdefault(matched_kw, {
                "coupang_option_keyword": matched_kw,
                "vendor_option_name": option_map[matched_kw],
                "quantity": 0,
                "orders": [],
            })
            bucket["quantity"] += row_qty
            if order_no:
                bucket["orders"].append({"order_id": order_no, "quantity": row_qty})

        out_row = start_row + row_count
        mapping = {
            1: name,
            2: phone,
            3: zipcode,
            4: address,
            5: vendor_option,
            6: qty,
            7: memo,
            8: order_no,
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        row_count += 1

    if row_count == 0:
        return None

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = product["output_filename_pattern"].format(
        label=product["product_label"],
        date=now.strftime("%Y%m%d"),
    )
    stats = {
        "total": row_count,
        "product": product["product_label"],
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats
