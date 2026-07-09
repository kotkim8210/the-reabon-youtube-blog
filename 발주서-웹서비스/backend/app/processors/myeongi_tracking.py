import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.processors.myeongi_order import (
    convert_option,
    is_apple_corn_order,
    is_house_watermelon_order,
    is_jewelry_bamhobak_order,
    is_jewelry_chamoe_order,
    is_jewelry_corn_order,
    is_jewelry_baekdo_order,
    is_jewelry_daegeukcheon_order,
    is_jewelry_geobando_order,
    is_jewelry_peach_order,
    is_jewelry_potato_order,
    is_mango_watermelon_order,
)
from app.processors.tracking_match import (
    name_counts,
    normalize_courier_name,
    option_key_set,
    options_match,
    requires_option_guard,
)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    """공백 완전 제거 후 문자열 반환 (매칭용)"""
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def is_jewelryfruit_tracking_target(product_name: str, option_text: str) -> bool:
    product = str(product_name or "")
    return (
        "명이나물" in product
        or is_apple_corn_order(product_name, option_text)
        or is_mango_watermelon_order(product_name, option_text)
        or is_house_watermelon_order(product_name, option_text)  # 일반 수박 6/7/8kg (2026-06 쥬얼리 전환)
        or is_jewelry_chamoe_order(product_name, option_text)    # 성주참외 로얄/중소 (2026-06 쥬얼리 전환)
        or is_jewelry_geobando_order(product_name, option_text)  # 거반도 납작복숭아 500g·1·2kg
        or is_jewelry_daegeukcheon_order(product_name, option_text)  # 대극천 복숭아 (신비복숭아 아님)
        or is_jewelry_baekdo_order(product_name, option_text)    # 백도 딱딱이복숭아 (신비복숭아 아님)
        or is_jewelry_peach_order(product_name, option_text)     # 신비복숭아 1·2kg (3·4kg은 제이비티)
        or is_jewelry_potato_order(product_name, option_text)    # 햇 홍감자 (2026 여름 쥬얼리 발주)
        or is_jewelry_bamhobak_order(product_name, option_text)  # 미니밤호박 3·5·10kg (1kg은 제주다팜)
        or is_jewelry_corn_order(product_name, option_text)      # 초당옥수수 (2026-06 제주다팜→쥬얼리 전환)
    )


def process(
    orderlist_bytes: bytes,
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """명이나물 orderlist의 운송장번호를 DeliveryList에 매핑.

    1차 매칭: 주문번호 (orderlist D열 = DeliveryList C열)
    폴백: 이름/전화/주소 매칭 (기존 tracking_input 패턴)
    """
    # Load orderlist
    ol_wb = load_workbook(filename=BytesIO(orderlist_bytes), data_only=True)
    ol_ws = ol_wb.active

    order_entries = []
    for row in ol_ws.iter_rows(min_row=2):
        order_no = normalize(row[3].value) if len(row) > 3 else ""      # D = 거래처주문번호
        name = normalize(row[10].value) if len(row) > 10 else ""        # K = 수령인
        phone = normalize(row[12].value) if len(row) > 12 else ""       # M = 수령인연락처
        address = normalize(row[14].value) if len(row) > 14 else ""     # O = 주소
        courier = normalize(row[16].value) if len(row) > 16 else ""     # Q = 택배사
        tracking = normalize(row[17].value) if len(row) > 17 else ""    # R = 운송장번호
        converted_option = convert_option(
            str(row[11].value if len(row) > 11 else ""),
            str(row[6].value if len(row) > 6 else ""),
        )
        option_keys = option_key_set(
            row[6].value if len(row) > 6 else "",
            row[8].value if len(row) > 8 else "",
            row[11].value if len(row) > 11 else "",
            converted_option,
        )

        if tracking:
            order_entries.append({
                "order_no": order_no,
                "name": name,
                "phone": phone,
                "address": address,
                "courier": courier,
                "tracking": tracking,
                "option_keys": option_keys,
            })

    # Build indexes
    # 한 주문번호에 여러 품목(예: 황혜영이 수박+참외를 한 주문번호로 주문)이면 회신에
    # 송장이 여러 건 → 주문번호별 리스트로 보관해야 각 행에 서로 다른 송장이 들어간다.
    order_by_order_no = defaultdict(list)
    order_by_name = defaultdict(list)
    for entry in order_entries:
        if entry["order_no"]:
            order_by_order_no[entry["order_no"]].append(entry)
        if entry["name"]:
            order_by_name[entry["name"]].append(entry)

    # Load DeliveryList
    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    dl_ws = dl_wb.active

    first_sheet_name = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet_name]

    used_entries = set()
    filled = 0
    skipped = 0
    has_myeongi = False
    has_apple_corn = False
    has_corn = False
    has_mango = False
    has_watermelon = False
    has_chamoe = False
    has_peach = False
    has_geobando = False
    has_daegeukcheon = False
    has_baekdo = False
    has_bamhobak = False
    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )

    for row_idx in range(2, dl_ws.max_row + 1):
        e_cell = dl_ws.cell(row=row_idx, column=5)  # E = 운송장번호

        if e_cell.value is not None and normalize(e_cell.value) != "":
            continue

        dl_order_no = normalize(dl_ws.cell(row=row_idx, column=3).value)   # C = 주문번호
        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)      # AA = 수취인이름
        dl_phone = normalize(dl_ws.cell(row=row_idx, column=28).value)     # AB = 수취인전화번호
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)   # AD = 수취인주소
        dl_product = dl_ws.cell(row=row_idx, column=11).value
        dl_option = dl_ws.cell(row=row_idx, column=12).value
        if not is_jewelryfruit_tracking_target(str(dl_product or ""), str(dl_option or "")):
            continue
        dl_option_keys = option_key_set(dl_option, convert_option(str(dl_option or ""), str(dl_product or "")))
        if not dl_option_keys:
            dl_option_keys = option_key_set(dl_product)

        if not dl_name and not dl_order_no:
            continue

        matched = None

        # 1차: 주문번호 매칭 (한 주문에 여러 품목이면 같은 주문번호가 여러 건)
        # 미사용 건 중 옵션이 맞는 것 우선, 옵션으로 못 가르면 미사용 첫 건을 쓴다.
        # → 수박+참외를 같은 주문번호로 주문해도 두 송장이 각각 두 행에 들어간다.
        if dl_order_no and dl_order_no in order_by_order_no:
            ono_avail = [c for c in order_by_order_no[dl_order_no] if id(c) not in used_entries]
            if len(ono_avail) == 1:
                matched = ono_avail[0]
            elif ono_avail:
                opt_avail = [c for c in ono_avail if options_match(c.get("option_keys"), dl_option_keys)]
                matched = opt_avail[0] if opt_avail else ono_avail[0]

        # 폴백: 이름/전화/주소 매칭
        if matched is None and dl_name:
            candidates = order_by_name.get(dl_name, [])

            if not candidates:
                MIN_PREFIX_LEN = 8
                for order_name, entries in order_by_name.items():
                    if len(order_name) < MIN_PREFIX_LEN:
                        continue
                    if dl_name.startswith(order_name) or order_name.startswith(dl_name):
                        candidates = entries
                        break

            available = [c for c in candidates if id(c) not in used_entries]

            if available:
                if requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
                    available = [
                        c for c in available
                        if options_match(c.get("option_keys"), dl_option_keys)
                    ]
                    if not available:
                        skipped += 1
                        continue

                for c in available:
                    if c["phone"] == dl_phone and c["address"] == dl_address:
                        matched = c
                        break
                if matched is None:
                    for c in available:
                        if c["phone"] == dl_phone:
                            matched = c
                            break
                if matched is None:
                    for c in available:
                        if c["address"] == dl_address:
                            matched = c
                            break
                if matched is None:
                    matched = available[0]

        if matched:
            e_cell.value = matched["tracking"]
            d_cell = dl_ws.cell(row=row_idx, column=4)
            d_cell.value = normalize_courier_name(matched.get("courier"), "롯데택배")
            used_entries.add(id(matched))
            filled += 1
            if "명이나물" in str(dl_product or ""):
                has_myeongi = True
            if is_apple_corn_order(dl_product, dl_option):
                has_apple_corn = True
            if is_jewelry_corn_order(dl_product, dl_option):
                has_corn = True
            if is_mango_watermelon_order(dl_product, dl_option):
                has_mango = True
            if is_house_watermelon_order(dl_product, dl_option):
                has_watermelon = True
            if is_jewelry_chamoe_order(dl_product, dl_option):
                has_chamoe = True
            if is_jewelry_geobando_order(dl_product, dl_option):
                has_geobando = True
            if is_jewelry_daegeukcheon_order(dl_product, dl_option):
                has_daegeukcheon = True
            if is_jewelry_baekdo_order(dl_product, dl_option):
                has_baekdo = True
            if is_jewelry_peach_order(dl_product, dl_option):
                has_peach = True
            if is_jewelry_bamhobak_order(dl_product, dl_option):
                has_bamhobak = True
        else:
            skipped += 1

    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    product_parts = []
    if has_myeongi:
        product_parts.append("명이나물")
    if has_apple_corn:
        product_parts.append("애플초당옥수수")
    if has_corn:
        product_parts.append("초당옥수수")
    if has_mango:
        product_parts.append("망고수박")
    if has_watermelon:
        product_parts.append("수박")
    if has_chamoe:
        product_parts.append("성주참외")
    if has_peach:
        product_parts.append("신비복숭아")
    if has_geobando:
        product_parts.append("거반도복숭아")
    if has_daegeukcheon:
        product_parts.append("대극천복숭아")
    if has_baekdo:
        product_parts.append("백도딱딱이복숭아")
    if has_bamhobak:
        product_parts.append("미니밤호박")
    product_label = "_".join(product_parts) if product_parts else "쥬얼리프룻"
    filename = f"DeliveryList_{product_label}_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {"filled": filled, "skipped": skipped}

    return output.read(), filename, stats
