"""당근마켓 주문 텍스트 → 쥬얼리프룻 발주서.

당근마켓은 개별 주문 '주문 상세' 화면 텍스트를 복사해 붙여넣는 방식으로 입력한다.
(엑셀/CSV가 아니라 텍스트) 여러 주문을 '주문 상세' 단위로 이어붙여도 분리 파싱한다.

쥬얼리프룻(pbfcompany) 발주 양식으로 출력. 대극천 복숭아 등 쥬얼리 취급 품목 대상.
등급 매핑: 당근 표기 대과→로얄과, 소과→소과 (쥬얼리프룻 시트 옵션명 기준).
"""

import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR

KST = timezone(timedelta(hours=9))

SENDER_NAME = "식품애착"
SENDER_PHONE = "010-5700-7756"

# 당근 등급 표기 → 쥬얼리프룻 시트 등급명
_GRADE_MAP = {
    "대과": "로얄과",
    "특대과": "로얄과",
    "로얄과": "로얄과",
    "중과": "중과",
    "중소과": "중소과",
    "소과": "소과",
    "혼합과": "혼합과",
    "랜덤과": "랜덤과",
}
_GRADES = ("특대과", "중소과", "로얄과", "혼합과", "랜덤과", "대과", "중과", "소과")
_PRODUCT_KEYWORDS = ("수박", "복숭아", "대극천")


def _clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _fmt_kg(value: object) -> str:
    text = str(value or "").strip()
    try:
        num = float(text)
    except ValueError:
        return text
    return str(int(num)) if num.is_integer() else str(num).rstrip("0").rstrip(".")


def _phone_digits(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if digits.startswith("820") and len(digits) >= 12:
        digits = digits[2:]
    elif digits.startswith("82") and len(digits) >= 11:
        digits = "0" + digits[2:]
    return digits


def _looks_like_phone(value: str) -> bool:
    digits = _phone_digits(value)
    return bool(re.fullmatch(r"01\d{8,9}", digits))


def _normalize_phone(value: str) -> str:
    digits = _phone_digits(value)
    if re.fullmatch(r"010\d{8}", digits):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if re.fullmatch(r"01\d{8,9}", digits):
        return digits
    return _clean(value)


def _clean_address(value: str) -> str:
    address = re.sub(r"^\(?\d{4,5}\)?\s*", "", _clean(value))
    return re.sub(r"\s*복사\s*$", "", address).strip()


def _value_after(lines: list[str], *labels: str) -> str:
    """라벨이 같은 줄 뒤(인라인) 또는 다음 줄에 오는 값을 찾는다."""
    for i, ln in enumerate(lines):
        for label in labels:
            if ln == label or ln.startswith(label):
                rest = ln[len(label):].strip(" :·")
                if rest:
                    return rest
                # 다음 비어있지 않은 줄
                for j in range(i + 1, min(i + 3, len(lines))):
                    if lines[j] and lines[j] not in ("복사", "채팅"):
                        return lines[j]
    return ""


def _daangn_product_label(product: str, grade: str, kg: str) -> str:
    """당근 상품/등급/kg → 쥬얼리프룻 발주 품목명."""
    compact = re.sub(r"\s+", "", product)
    if "수박" in compact or "수박" in grade:
        kg_text = _fmt_kg(kg)
        # 당근 수박 8kg 주문은 쥬얼리프룻 공급 옵션 '8~9kg 내외'로 발주한다.
        if kg_text == "8":
            return "가정용 수박 8~9kg 내외"
        return f"가정용 수박 {kg_text}kg 내외" if kg_text else "가정용 수박"
    if "대극천" in compact:
        base = "대극천"
    else:
        # '○○ 복숭아'에서 복숭아만 떼되, '신비복숭아'처럼 붙은 건 보존
        base = re.sub(r"\s*복숭아\s*$", "", product).strip() or product
    g = _GRADE_MAP.get(grade, grade)
    parts = [base]
    if g:
        parts.append(g)
    if kg:
        parts.append(f"{kg}kg")
    return " ".join(p for p in parts if p).strip()


def _fallback_name_phone_address(lines: list[str], opt_idx: int) -> tuple[str, str, str]:
    """라벨 없이 상품/이름/전화/주소만 붙인 당근 주문에서 배송 정보를 추정한다."""
    name = phone = address = ""
    phone_idx = -1
    for i, line in enumerate(lines):
        if _looks_like_phone(line):
            phone = _normalize_phone(line)
            phone_idx = i
            break

    search_start = opt_idx + 1 if opt_idx >= 0 else 0
    for i in range(search_start, len(lines)):
        line = lines[i]
        if i == phone_idx or _looks_like_phone(line):
            continue
        if re.search(r"\d{4,5}|[시도군구로길읍면동리]\b|아파트|빌라|주택|호|동", line):
            continue
        if any(keyword in line for keyword in _PRODUCT_KEYWORDS):
            continue
        name = _clean(line)
        break

    if phone_idx >= 0:
        address_lines = [
            line for line in lines[phone_idx + 1:]
            if line and not line.startswith(("배송메모", "공동현관", "비밀번호"))
        ]
    else:
        address_lines = [
            line for i, line in enumerate(lines[search_start:], start=search_start)
            if line and line != name and not any(keyword in line for keyword in _PRODUCT_KEYWORDS)
        ]
    address = _clean_address(" ".join(address_lines))
    return name, phone, address


def _parse_one(block: str) -> dict | None:
    lines = [l.strip() for l in block.splitlines() if l.strip()]
    if not lines:
        return None

    # 옵션 줄: '대과 2kg 1개' / '로얄과 2kg 1개' (등급 kg 수량)
    product = grade = kg = ""
    qty = 1
    opt_idx = -1
    for i, ln in enumerate(lines):
        m = re.match(r"(.+?)\s*(\d+(?:\.\d+)?)\s*kg\s*(\d+)\s*개", ln, re.IGNORECASE)
        if m:
            prefix = m.group(1).strip()
            grade = next((g for g in _GRADES if g in prefix), "")
            if not grade and not any(keyword in prefix for keyword in _PRODUCT_KEYWORDS):
                grade = prefix
            kg = _fmt_kg(m.group(2))
            qty = int(m.group(3))
            opt_idx = i
            break
    if opt_idx > 0:
        product = lines[opt_idx - 1]
    elif opt_idx == 0:
        prefix = re.match(r"(.+?)\s*(\d+(?:\.\d+)?)\s*kg\s*(\d+)\s*개", lines[opt_idx], re.IGNORECASE).group(1).strip()
        if any(keyword in prefix for keyword in _PRODUCT_KEYWORDS):
            product = prefix
    if not product or not kg:
        return None

    name = _value_after(lines, "받는 사람", "받는사람", "수령인")
    phone = _value_after(lines, "연락처", "전화")
    address_raw = _value_after(lines, "배송지", "주소")
    if not (name and phone and address_raw):
        fallback_name, fallback_phone, fallback_address = _fallback_name_phone_address(lines, opt_idx)
        name = name or fallback_name
        phone = phone or fallback_phone
        address_raw = address_raw or fallback_address
    phone = _normalize_phone(phone)
    # '(우편번호) 도로명주소 ... 복사' → 우편번호 괄호 제거, '복사' 꼬리 제거
    address = _clean_address(address_raw)
    order_no = _value_after(lines, "주문번호")
    memo_line = next((l for l in lines if "공동현관" in l or "비밀번호" in l or l.startswith("배송메모")), "")
    memo = memo_line.replace("배송메모", "").strip(" :")

    if not name:
        return None

    return {
        "order_id": order_no or f"daangn-{re.sub(r'[^0-9]', '', phone)[-8:]}-{kg}",
        "product": _daangn_product_label(product, grade, kg),
        "qty": str(qty),
        "name": name,
        "phone": phone,
        "address": address,
        "option": _clean(f"{grade} {kg}kg"),
        "memo": memo,
    }


def parse_daangn_orders(text: str) -> list[dict]:
    """당근 주문 텍스트(1건 또는 '주문 상세'로 이어붙인 여러 건) → entry 목록."""
    text = text or ""
    parts = re.split(r"주문\s*상세", text)
    blocks = [p for p in parts if p.strip()] or [text]
    entries: list[dict] = []
    for block in blocks:
        entry = _parse_one(block)
        if entry:
            entries.append(entry)
    return entries


def process(text: str) -> tuple[bytes, str, dict] | None:
    """당근 주문 텍스트 → 쥬얼리프룻(pbfcompany) 발주서."""
    entries = parse_daangn_orders(text)
    if not entries:
        return None

    template_path = TEMPLATE_DIR / "명이나물_pbfcompany_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)
    now = datetime.now(KST)
    today_str = now.strftime("%Y-%m-%d")

    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=3).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}
    for i, entry in enumerate(entries):
        out_row = start_row + i
        product = entry["product"]
        try:
            qty_int = int(float(entry["qty"]))
        except (TypeError, ValueError):
            qty_int = 1

        mapping = {
            1: today_str,           # A 일자
            2: "아이티소프트",       # B 거래처명
            3: entry["name"],       # C 받는분성명
            4: entry["phone"],      # D 받는분전화번호
            6: entry["address"],    # F 받는분주소
            7: product,             # G 품목명
            8: str(qty_int),        # H 수량
            9: SENDER_NAME,         # I 보내는분성명
            10: SENDER_PHONE,       # J 보내는분전화번호
            12: entry.get("memo") or "",   # L 배송메시지
            13: entry.get("order_id") or "",  # M 주문번호
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        opt_key = entry.get("option") or product
        bucket = option_totals.setdefault(opt_key, {
            "coupang_option_keyword": opt_key,
            "vendor_option_name": product,
            "quantity": 0,
            "orders": [],
        })
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    filename = f"당근_쥬얼리프룻_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(entries),
        "supplier": "쥬얼리프룻",
        "product": "당근마켓 발주",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats
