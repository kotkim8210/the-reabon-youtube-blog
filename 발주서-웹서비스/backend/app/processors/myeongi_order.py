import re
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR
from app.rules_engine import convert as rules_convert


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _combined_text(product_name: object, option_text: object) -> str:
    return normalize(f"{product_name or ''} {option_text or ''}")


def is_apple_corn_order(product_name: object, option_text: object) -> bool:
    text = _combined_text(product_name, option_text)
    return "애플초당옥수수" in text or ("애플" in text and "초당옥수수" in text)


# 망고수박 = 블랙망고수박(동일 상품). 일반수박(tomato_order)과는 반드시 구분.
# 발주 품목명은 '블랙망고수박 {kg}kg'로 통일(_blackmango_label) — 아래 테이블은 더 이상
# 출력에 쓰지 않으며, supplier_price_monitor.py(제이비티 망고수박 마진방어)의 소싱현황
# 옵션명과 병렬 참조용으로만 남겨둔다.
MANGO_WATERMELON_VENDOR_OPTIONS = {
    "2": "수박 가정용 2~3kg 내외 (2kg)",
    "3": "수박 가정용 3~4kg 내외 (3kg)",
    "5": "수박 가정용 4~5kg내외(5kg)",
    "6": "수박 가정용 5~6kg내외(6kg)",
}


def is_mango_watermelon_order(product_name: object, option_text: object) -> bool:
    # 일반 망고수박 = 블랙망고수박 (동일 상품). '망고 수박'처럼 띄어써도 잡도록 공백 무시.
    text = _combined_text(product_name, option_text).replace(" ", "")
    return "망고수박" in text


def _extract_kg(field: object) -> str:
    """한 필드(상품명 또는 옵션명)에서 'Nkg'/'N.Nkg' 무게를 추출. 없으면 ''."""
    m = re.search(r"(\d+(?:\.\d+)?)\s*kg", str(field or ""), re.IGNORECASE)
    return _fmt_kg(m.group(1)) if m else ""


def _blackmango_label(product_name: object, option_text: object) -> str:
    """블랙망고수박 → '블랙망고수박 {kg}kg'.

    블랙망고수박은 일반 망고수박('수박 가정용 N~Nkg') · 일반 수박과 반드시 구분한다.
    무게는 **상품명(K)에서 우선** 추출한다 — 옵션(L)에 '수박 가정용 3~4kg 내외 (3kg)'
    같은 엉뚱한 일반수박 SKU가 적혀 들어오는 경우가 있어 옵션의 kg은 신뢰하지 않는다.
    상품명에 무게가 없으면 옵션의 소수(2.5kg 등) 무게를, 그것도 없으면 블랙망고수박
    단일 규격인 2.5kg를 기본값으로 사용한다.
    """
    kg = _extract_kg(product_name)
    if not kg:
        # 상품명에 없을 때만 옵션을 본다. 단, '수박 가정용 N~Nkg'식 정수 표기는 무시하고
        # 명시적 소수 무게(2.5kg 등)만 인정.
        m = re.search(r"(\d+\.\d+)\s*kg", str(option_text or ""), re.IGNORECASE)
        kg = _fmt_kg(m.group(1)) if m else "2.5"
    return f"블랙망고수박 {kg}kg"


# (홍감자는 2026-07 쥬얼리프룻 품절 → 제주다팜(kolrabi_order) 발주로 이관 — 관련 함수 제거)


# 제주 미니밤호박(보우짱) — 2026-06부터 3·5·10kg은 제주다팜→쥬얼리프룻 발주 전환(공급가 인하).
# 1kg은 계속 제주다팜(kolrabi_order)에서 발주. 쥬얼리 시트 옵션: '로얄과 3/5/10kg'.
BAMHOBAK_JEWELRY_WEIGHTS = {"3", "5", "10"}


def is_jewelry_bamhobak_order(product_name: object, option_text: object) -> bool:
    """미니밤호박 쥬얼리프룻 발주 여부 (3·5·10kg만; 1kg은 제주다팜)."""
    text = _combined_text(product_name, option_text).replace(" ", "")
    if "밤호박" not in text:
        return False
    m = re.search(r"(\d+)\s*kg", text, re.IGNORECASE)
    return bool(m) and m.group(1) in BAMHOBAK_JEWELRY_WEIGHTS


def _jewelry_bamhobak_option(product_name: object, option_text: object) -> str | None:
    """미니밤호박 DeliveryList → 쥬얼리프룻 발주 품목 '미니 밤호박 보우짱 로얄과 {n}kg'.

    3·5·10kg만 대상(1kg은 제주다팜). 등급은 못난이가 명시되면 못난이, 아니면 로얄과(기본).
    """
    if not is_jewelry_bamhobak_order(product_name, option_text):
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result
    text = _combined_text(product_name, option_text)
    m = re.search(r"(\d+)\s*kg", text, re.IGNORECASE)
    kg = m.group(1) if m else ""
    grade = "못난이" if "못난이" in text else "로얄과"
    return f"미니 밤호박 보우짱 {grade} {kg}kg"


# 일반 수박 6/7/8kg — 2026-06부터 제이비티가 아닌 쥬얼리팜(쥬얼리프룻) 발주.
def is_house_watermelon_order(product_name: object, option_text: object) -> bool:
    from app.processors.tomato_order import delivery_watermelon_kg, JBT_WATERMELON_KGS
    return delivery_watermelon_kg(str(product_name or ""), str(option_text or "")) in JBT_WATERMELON_KGS


# 성주참외 전체 — 2026-06부터 제주다팜 알뜰과 발주 폐지, 전 등급 쥬얼리프룻으로 발주 통합.
# 쿠팡 옵션 → 발주 품목명:
#   가정용 로얄과 → '성주참외 로얄과 {kg}kg (R)'
#   가정용 혼합과 → '성주참외 가성비 랜덤과 {kg}kg (R)'
def is_jewelry_chamoe_order(product_name: object, option_text: object) -> bool:
    return "성주참외" in str(product_name or "")


def _chamoe_jewelry_label(product_name: object, option_text: object) -> str:
    """성주참외 → 쥬얼리프룻 발주 품목명 '성주참외 {등급} {kg}kg (R)'.

    로얄과→'로얄과', 가정용 혼합과/알뜰/랜덤→'가성비 랜덤과', 중소→'중소과'.
    """
    text = _combined_text(product_name, option_text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE)
    kg = _fmt_kg(m.group(1)) if m else ""
    if "로얄" in text:
        grade = "로얄과"
    elif "중소" in text:
        grade = "중소과"
    else:  # 혼합/알뜰/랜덤/등급미표기 → 가성비 랜덤과
        grade = "가성비 랜덤과"
    base = f"성주참외 {grade}"
    return f"{base} {kg}kg (R)" if kg else f"{base} (R)"


def _jewelry_chamoe_option(product_name: object, option_text: object) -> str | None:
    if not is_jewelry_chamoe_order(product_name, option_text):
        return None
    return _chamoe_jewelry_label(product_name, option_text)


def _house_watermelon_option(product_name: object, option_text: object) -> str | None:
    from app.processors.tomato_order import delivery_watermelon_kg, JBT_WATERMELON_KGS
    kg = delivery_watermelon_kg(str(product_name or ""), str(option_text or ""))
    if kg not in JBT_WATERMELON_KGS:
        return None
    # 테무 쥬얼리 발주와 동일한 품목명으로 통일
    return f"가정용 수박 {kg}kg 내외"


# 거반도 납작복숭아(도넛복숭아) → 쥬얼리프룻 발주. 쿠팡 500g/1kg/2kg → '거반도 {size} (과수 내외)'.
# 상품명에 '복숭아'가 들어가 신비복숭아 로직과 충돌하므로, 항상 먼저 분기하고 peach 함수에서 제외한다.
def _is_geobando(product_name: object, option_text: object) -> bool:
    t = _combined_text(product_name, option_text).replace(" ", "")
    return "거반도" in t or "납작복숭아" in t


def is_jewelry_geobando_order(product_name: object, option_text: object) -> bool:
    if not _is_geobando(product_name, option_text):
        return False
    t = _combined_text(product_name, option_text).replace(" ", "")
    return ("2kg" in t) or ("1kg" in t) or ("500g" in t)


_GEOBANDO_LABELS = {
    "2kg": "거반도 2kg (8~16과 내외)",
    "1kg": "거반도 1kg (4~10과 내외)",
    "500g": "거반도 500g (2~5과 내외)",
}


def _jewelry_geobando_option(product_name: object, option_text: object) -> str | None:
    if not is_jewelry_geobando_order(product_name, option_text):
        return None
    t = _combined_text(product_name, option_text).replace(" ", "")
    for key in ("2kg", "1kg", "500g"):  # 2kg 먼저 (1kg/500g 오매칭 방지)
        if key in t:
            return _GEOBANDO_LABELS[key]
    return None


# 노지 대극천(반납작 복숭아) → 쥬얼리프룻 발주. 쿠팡 '1박스 로얄소과 1kg' → '대극천 소과 1kg (11-16과 내외)'.
# 역시 '복숭아' 포함이라 신비복숭아보다 먼저 분기하고 peach에서 제외. (당근 텍스트 발주 daangn_order와 별개: 쿠팡 DeliveryList용)
def is_jewelry_daegeukcheon_order(product_name: object, option_text: object) -> bool:
    return "대극천" in _combined_text(product_name, option_text).replace(" ", "")


# 쿠팡 대극천 (등급, kg) → 발주 품목 과수 표기 (사용자 확정 매핑 2026-07)
_DAEGEUKCHEON_COUNTS = {
    ("로얄과", "2"): "8-20과 내외",
    ("로얄과", "1"): "4-10과 내외",
    ("소과", "2"): "21-32과 내외",
    ("소과", "1"): "11-16과 내외",
}


def _jewelry_daegeukcheon_option(product_name: object, option_text: object) -> str | None:
    if not is_jewelry_daegeukcheon_order(product_name, option_text):
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result
    text = _combined_text(product_name, option_text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE)
    kg = _fmt_kg(m.group(1)) if m else ""
    # 등급: 로얄소과/소과→소과, 로얄대과/로얄/대과→로얄과 (기본 소과)
    grade = "소과" if "소과" in text else ("로얄과" if ("로얄" in text or "대과" in text) else "소과")
    count = _DAEGEUKCHEON_COUNTS.get((grade, kg), "")
    base = f"대극천 {grade}"
    if kg and count:
        return f"{base} {kg}kg ({count})"
    return f"{base} {kg}kg" if kg else base


# 백도 딱딱이복숭아 → 쥬얼리프룻 발주. 쿠팡 '햇 백도 딱딱이복숭아 …' 옵션 '1박스 중과 1kg'
# → '백도 딱딱이 복숭아 중과 1kg (5-6과 내외)'. '복숭아' 포함이라 신비복숭아보다 먼저 분기하고
# peach에서 제외. 전 kg(1·2·4)·전 등급(중과/대과) 모두 쥬얼리(신비처럼 3·4kg 제이비티 분리 없음).
def _is_baekdo(product_name: object, option_text: object) -> bool:
    return "백도" in _combined_text(product_name, option_text).replace(" ", "")


def is_jewelry_baekdo_order(product_name: object, option_text: object) -> bool:
    return _is_baekdo(product_name, option_text)


# 쿠팡 백도 (등급, kg) → 발주 품목 과수 표기 (마진 소싱현황 기준 2026-07)
_BAEKDO_COUNTS = {
    ("중과", "1"): "5-6과 내외",
    ("중과", "2"): "11-14과 내외",
    ("중과", "4"): "20-26과 내외",
    ("대과", "1"): "3-4과 내외",
    ("대과", "2"): "6-8과 내외",
    ("대과", "4"): "12-17과 내외",
}


def _jewelry_baekdo_option(product_name: object, option_text: object) -> str | None:
    if not _is_baekdo(product_name, option_text):
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result
    text = _combined_text(product_name, option_text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE)
    kg = _fmt_kg(m.group(1)) if m else ""
    grade = "대과" if "대과" in text else "중과"  # 기본 중과
    count = _BAEKDO_COUNTS.get((grade, kg), "")
    base = "백도 딱딱이 복숭아"
    if kg and count:
        return f"{base} {grade} {kg}kg ({count})"
    return f"{base} {grade} {kg}kg" if kg else base


# '복숭아'가 들어가지만 신비복숭아가 아닌 별도 발주 품목(거반도·대극천·백도 등) — 신비 오분류 방지
def _non_shinbi_peach(product_name: object, option_text: object) -> bool:
    return (
        _is_geobando(product_name, option_text)
        or is_jewelry_daegeukcheon_order(product_name, option_text)
        or _is_baekdo(product_name, option_text)
    )


# 신비복숭아 1·2kg → 쥬얼리프룻 발주 (3·4kg은 제이비티, 800g은 제외)
def is_jewelry_peach_order(product_name: object, option_text: object) -> bool:
    from app.processors.tomato_order import peach_kg, JEWELRY_PEACH_KGS
    if _non_shinbi_peach(product_name, option_text):  # 거반도·대극천은 신비복숭아 아님
        return False
    return peach_kg(str(product_name or ""), str(option_text or "")) in JEWELRY_PEACH_KGS


def _jewelry_peach_option(product_name: object, option_text: object) -> str | None:
    if _non_shinbi_peach(product_name, option_text):  # 거반도·대극천은 신비복숭아 아님
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result
    from app.processors.tomato_order import peach_kg, JEWELRY_PEACH_KGS
    kg = peach_kg(str(product_name or ""), str(option_text or ""))
    if kg not in JEWELRY_PEACH_KGS:
        return None
    # 쿠팡 신비복숭아: (kg, 등급) → 공급사 SKU명 (예: '1kg (중소과)' → '신비 복숭아 1kg (15과 내외)')
    text = _combined_text(product_name, option_text)
    grade = "중소과" if "중소과" in text else ("대과" if "대과" in text else "")
    mapped = _PEACH_COUPANG_SKU.get((int(kg), grade))
    if mapped:
        return mapped
    return _peach_label(product_name, option_text)


def _fmt_kg(value: str) -> str:
    """'3.0' → '3', '2.5' → '2.5' (불필요한 소수점 제거)."""
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(f)) if f.is_integer() else str(f)


# 신비복숭아 등급 (긴 것 먼저 — '중소과'가 '소과'보다 먼저 매칭돼야 함)
_PEACH_GRADES = ("중소과", "중대과", "로얄과", "혼합과", "랜덤과", "대과", "소과")

# 쿠팡 신비복숭아 (kg, 등급) → 쥬얼리프룻 공급사 SKU명 (과수 내외)
_PEACH_COUPANG_SKU = {
    (1, "중소과"): "신비 복숭아 1kg (15과 내외)",
    (2, "중소과"): "신비 복숭아 2kg (30과 내외)",
    (1, "대과"): "신비 복숭아 1kg (11과 내외)",
    (2, "대과"): "신비 복숭아 2kg (22과 내외)",
}


def _peach_label(product_name: object, option_text: object) -> str:
    """신비복숭아 품목명 → '신비복숭아 {kg}kg {등급}'. (등급 없으면 생략)"""
    text = _combined_text(product_name, option_text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE)
    kg = _fmt_kg(m.group(1)) if m else ""
    grade = next((g for g in _PEACH_GRADES if g in text), "")
    parts = ["신비복숭아"]
    if kg:
        parts.append(f"{kg}kg")
    if grade:
        parts.append(grade)
    return " ".join(parts)


def _external_chamoe_product(product_name: object, option_text: object) -> str | None:
    """토스/테무 등 외부 채널의 성주참외 → 쥬얼리프룻 발주 품목명.

    쿠팡(_jewelry_chamoe_option)과 동일 표기 '성주참외 {등급} {kg}kg (R)'.
    """
    text = _combined_text(product_name, option_text)
    if "참외" not in text:
        return None
    return _chamoe_jewelry_label(product_name, option_text)


def jewelry_external_product(
    product_name: object,
    option_text: object,
    watermelon_kgs: set,
) -> str | None:
    """외부 채널(토스/테무)의 수박·성주참외 주문 → 쥬얼리팜 발주 품목명 (없으면 None).

    - 수박: watermelon_kgs 범위(토스 6/7/8, 테무 2/3/5/6/7/8)일 때만 '가정용 수박 Nkg 내외'
    - 성주참외: 로얄과/중소과/가성비 혼합과 Nkg
    - 망고수박은 제외(별도 발주)
    """
    from app.processors.tomato_order import delivery_watermelon_kg
    kg = delivery_watermelon_kg(str(product_name or ""), str(option_text or ""))
    if kg in watermelon_kgs:
        return f"가정용 수박 {kg}kg 내외"
    peach = _jewelry_peach_option(product_name, option_text)
    if peach:
        return peach
    return _external_chamoe_product(product_name, option_text)


def jewelry_passthrough_product(
    product_name: object,
    option_text: object,
    watermelon_kgs: set,
) -> str | None:
    """토스/올웨이즈 → 쥬얼리팜 발주 품목명.

    신비복숭아는 옵션 보존(소과/대과 등), 망고수박(=블랙망고수박)은 상품명(K) 기준
    '블랙망고수박 {kg}kg', 수박·성주참외는 정규화(jewelry_external_product).
    """
    text = _combined_text(product_name, option_text)

    # 거반도·대극천은 '복숭아' 포함이지만 신비복숭아 아님 → 먼저 전용 품목명으로 (토스 경로 오분류 방지)
    geobando = _jewelry_geobando_option(product_name, option_text)
    if geobando:
        return geobando

    daegeukcheon = _jewelry_daegeukcheon_option(product_name, option_text)
    if daegeukcheon:
        return daegeukcheon

    baekdo = _jewelry_baekdo_option(product_name, option_text)
    if baekdo:
        return baekdo

    if "복숭아" in text:
        # 신비복숭아: '신비복숭아 {kg}kg {등급(소과/중소과/대과...)}'
        return _peach_label(product_name, option_text)

    if "망고수박" in text.replace(" ", ""):
        # 일반 망고수박 = 블랙망고수박 (동일 상품) → 모두 블랙망고수박으로 출력.
        # 옵션(L)이 '수박 가정용 N~Nkg'로 잘못 적혀도 일반수박으로 둔갑하지 않는다.
        return _blackmango_label(product_name, option_text)


    return jewelry_external_product(product_name, option_text, watermelon_kgs)


def _mango_watermelon_option(product_name: object, option_text: object) -> str | None:
    if not is_mango_watermelon_order(product_name, option_text):
        return None
    # 일반 망고수박 = 블랙망고수박 (동일 상품) → 모두 블랙망고수박으로 출력.
    return _blackmango_label(product_name, option_text)


def _apple_corn_option(product_name: object, option_text: object) -> str | None:
    text = _combined_text(product_name, option_text)
    if not is_apple_corn_order(product_name, option_text):
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result

    count_match = re.search(r"(\d+)\s*(?:개|입)", text)
    if not count_match:
        return None
    count = count_match.group(1)
    if count not in ("5", "10", "15", "20"):
        return None
    return f"애플초당옥수수(특품) {count}개"


# 일반 초당옥수수 — 2026-06부터 제주다팜→쥬얼리프룻 발주 이관. 애플초당옥수수는 별도(_apple_corn).
def is_jewelry_corn_order(product_name: object, option_text: object) -> bool:
    text = _combined_text(product_name, option_text)
    return "초당옥수수" in text and "애플" not in text


def _jewelry_corn_option(product_name: object, option_text: object) -> str | None:
    """일반 초당옥수수 DeliveryList → 쥬얼리프룻 발주 품목 '초당옥수수({중품/특품}) {n}개'.

    등급(중품/특품)과 수량(개/입)이 있어야 발주 대상. 애플초당옥수수는 제외.
    """
    if not is_jewelry_corn_order(product_name, option_text):
        return None
    engine_result = rules_convert("jewelryfruit", product_name, option_text)
    if engine_result is not None:
        return engine_result
    text = _combined_text(product_name, option_text)
    count_match = re.search(r"(\d+)\s*(?:개|입)", text)
    grade = "중품" if "중품" in text else "특품" if "특품" in text else ""
    if not count_match or not grade:
        return None
    return f"초당옥수수({grade}) {count_match.group(1)}개"


def convert_option(option_text: str, product_name: str = "") -> str | None:
    """Convert DeliveryList option text to the vendor-facing product name."""
    apple_corn = _apple_corn_option(product_name, option_text)
    if apple_corn:
        return apple_corn

    corn = _jewelry_corn_option(product_name, option_text)
    if corn:
        return corn

    mango = _mango_watermelon_option(product_name, option_text)
    if mango:
        return mango

    house = _house_watermelon_option(product_name, option_text)
    if house:
        return house

    chamoe = _jewelry_chamoe_option(product_name, option_text)
    if chamoe:
        return chamoe

    geobando = _jewelry_geobando_option(product_name, option_text)
    if geobando:
        return geobando

    daegeukcheon = _jewelry_daegeukcheon_option(product_name, option_text)
    if daegeukcheon:
        return daegeukcheon

    baekdo = _jewelry_baekdo_option(product_name, option_text)
    if baekdo:
        return baekdo

    peach = _jewelry_peach_option(product_name, option_text)
    if peach:
        return peach

    bamhobak = _jewelry_bamhobak_option(product_name, option_text)
    if bamhobak:
        return bamhobak

    text = str(option_text).strip()
    if not text:
        # 옵션(L)이 비면 상품명(K)으로 폴백 — 품목명 공란 방지(명이나물 행 전용 분기).
        pn = str(product_name or "").strip()
        if not pn:
            return None
        wmatch = re.search(r"(\d+(?:\.\d+)?)\s*(kg|g)", pn, re.IGNORECASE)
        return f"명이나물 {wmatch.group(1)}{wmatch.group(2).lower()}" if wmatch else pn
    match = re.match(r"(\d+(?:kg|g))\s+1박스", text, re.IGNORECASE)
    if match:
        weight = match.group(1)
        return f"명이나물 {weight}"
    return text


def process(
    delivery_file_bytes: bytes,
    toss_entries: list[dict] | None = None,
) -> tuple[bytes, str, dict]:
    """쿠팡 DeliveryList + 토스 주문을 쥬얼리프룻 발주서로 출력.

    Args:
        delivery_file_bytes: 쿠팡 DeliveryList 엑셀.
        toss_entries: 토스 API에서 수집한 쥬얼리 entry 목록(선택). 각 entry는
            품목명(product)이 이미 확정돼 있어 그대로 기록된다
            (수박·성주참외·신비복숭아 등급·블랙망고수박 포함).

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows = []
    has_myeongi = False
    has_apple_corn = False
    has_corn = False
    has_mango = False
    has_watermelon = False
    has_chamoe = False
    has_peach = False
    has_geobando = False
    has_daegeukcheon = False
    has_baekdo = False
    has_bamhobak = False
    for row in dl_ws.iter_rows(min_row=2):
        k_val = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        if "명이나물" in k_val:
            filtered_rows.append(row)
            has_myeongi = True
        elif is_apple_corn_order(k_val, option):
            filtered_rows.append(row)
            has_apple_corn = True
        elif is_jewelry_corn_order(k_val, option):
            filtered_rows.append(row)
            has_corn = True
        elif is_mango_watermelon_order(k_val, option):
            # 망고수박 = 쥬얼리프룻 발주.
            filtered_rows.append(row)
            has_mango = True
        elif is_house_watermelon_order(k_val, option):
            # 일반 수박 6/7/8kg도 쥬얼리팜(쥬얼리프룻) 발주로 전환 (2026-06~)
            filtered_rows.append(row)
            has_watermelon = True
        elif is_jewelry_chamoe_order(k_val, option):
            # 성주참외(로얄/중소)도 쥬얼리팜 발주 (알뜰과는 별도 버튼)
            filtered_rows.append(row)
            has_chamoe = True
        elif is_jewelry_geobando_order(k_val, option):
            # 거반도 납작복숭아 500g·1·2kg → 쥬얼리프룻
            filtered_rows.append(row)
            has_geobando = True
        elif is_jewelry_daegeukcheon_order(k_val, option):
            # 대극천 복숭아 → 쥬얼리프룻 (신비복숭아 아님)
            filtered_rows.append(row)
            has_daegeukcheon = True
        elif is_jewelry_baekdo_order(k_val, option):
            # 백도 딱딱이복숭아 → 쥬얼리프룻 (신비복숭아 아님, 중과/대과 1·2·4kg)
            filtered_rows.append(row)
            has_baekdo = True
        elif is_jewelry_peach_order(k_val, option):
            # 신비복숭아 1·2kg → 쥬얼리프룻 (3·4kg은 제이비티)
            filtered_rows.append(row)
            has_peach = True
        elif is_jewelry_bamhobak_order(k_val, option):
            # 미니밤호박 3·5·10kg → 쥬얼리프룻 발주 (1kg은 제주다팜)
            filtered_rows.append(row)
            has_bamhobak = True

    template_path = TEMPLATE_DIR / "명이나물_pbfcompany_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)
    now = datetime.now(KST)
    today_str = now.strftime("%Y-%m-%d")

    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=3).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}

    for i, row in enumerate(filtered_rows):
        out_row = start_row + i

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        product_name = normalize(row[10].value) if len(row) > 10 else ""
        product = convert_option(option, product_name)
        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        mapping = {
            1: today_str,           # A - 일자
            2: "아이티소프트",       # B - 거래처명
            3: name,                # C - 받는분성명
            4: phone,               # D - 받는분전화번호
            # 5: "",                # E - 받는분기타연락처 (비워두기)
            6: address,             # F - 받는분주소 (우편번호 제외)
            7: product,             # G - 품목명
            8: qty,                 # H - 수량
            9: "식품애착",           # I - 보내는분성명
            10: "010-5700-7756",    # J - 보내는분전화번호
            # 11: "",              # K - 보내는분주소 (비워두기)
            12: memo,               # L - 배송메시지
            13: order_no,           # M - 주문번호
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            option or product or "명이나물",
            {
                "coupang_option_keyword": option or product or "쥬얼리프룻",
                "vendor_option_name": product or "쥬얼리프룻",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    # 토스 entry는 품목명(product)이 확정돼 있어 그대로 기록 (등급/블랙망고수박/전 kg 보존)
    base_row = start_row + len(filtered_rows)
    for j, entry in enumerate(toss_entries or []):
        out_row = base_row + j
        product = entry.get("product") or ""
        try:
            qty_int = int(float(entry.get("qty"))) if entry.get("qty") not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        order_no = entry.get("order_id") or ""
        mapping = {
            1: today_str,
            2: "아이티소프트",
            3: entry.get("name", ""),
            4: entry.get("phone", ""),
            6: entry.get("address", ""),
            7: product,
            8: str(qty_int),
            9: "식품애착",
            10: "010-5700-7756",
            12: entry.get("memo", ""),
            13: order_no,
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        if "거반도" in product:
            has_geobando = True
        elif "대극천" in product:
            has_daegeukcheon = True
        elif "백도" in product:
            has_baekdo = True
        elif "복숭아" in product:
            has_peach = True
        elif "망고수박" in product:
            has_mango = True
        elif "수박" in product:
            has_watermelon = True
        elif "과" in product:
            has_chamoe = True

        bucket = option_totals.setdefault(
            product or "쥬얼리프룻",
            {
                "coupang_option_keyword": product or "쥬얼리프룻",
                "vendor_option_name": product or "쥬얼리프룻",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    product_parts = []
    if has_myeongi:
        product_parts.append("명이나물")
    if has_apple_corn:
        product_parts.append("애플초당옥수수")
    if has_corn:
        product_parts.append("초당옥수수")
    if has_mango:
        product_parts.append("망고수박")
    if has_watermelon:
        product_parts.append("수박")
    if has_chamoe:
        product_parts.append("성주참외")
    if has_peach:
        product_parts.append("신비복숭아")
    if has_geobando:
        product_parts.append("거반도복숭아")
    if has_daegeukcheon:
        product_parts.append("대극천복숭아")
    if has_baekdo:
        product_parts.append("백도딱딱이복숭아")
    if has_bamhobak:
        product_parts.append("미니밤호박")
    product_label = "_".join(product_parts)
    product_name = "+".join(product_parts) if product_parts else "쥬얼리프룻"
    if product_label:
        filename = f"쥬얼리프룻_{product_label}_발주({now.strftime('%Y%m%d')}).xlsx"
    else:
        filename = f"쥬얼리프룻_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows) + len(toss_entries or []),
        "coupang": len(filtered_rows),
        "toss": len(toss_entries or []),
        "product": product_name,
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats


# ── 토스 + 올웨이즈 수박 → 쥬얼리팜(쥬얼리프룻) 발주서 ──────────────
# (기존 제이비티 발주에서 2026-06 발주처 전환)

def parse_alwayz_jewelry_entries(alwayz_bytes: bytes) -> list[dict]:
    """올웨이즈 주문내역에서 쥬얼리팜 발주 대상(수박 6/7/8kg + 성주참외) 주문을 추출.

    올웨이즈 컬럼: A=주문아이디, E=상품명, F=옵션, G=수량,
                  O=주소, P=우편번호, S=수령인, T=연락처
    """
    from app.processors.tomato_order import JBT_WATERMELON_KGS

    wb = load_workbook(filename=BytesIO(alwayz_bytes), data_only=True)
    ws = wb.active

    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2):
        product_name = normalize(row[4].value) if len(row) > 4 else ""
        option = normalize(row[5].value) if len(row) > 5 else ""
        name = normalize(row[18].value) if len(row) > 18 else ""
        if not name or not product_name:
            continue
        product = jewelry_passthrough_product(product_name, option, JBT_WATERMELON_KGS)
        if not product:
            continue
        entries.append({
            "order_id": normalize(row[0].value) if len(row) > 0 else "",
            "product": product,
            "qty": (normalize(row[6].value) if len(row) > 6 else "") or "1",
            "name": name,
            "phone": normalize(row[19].value) if len(row) > 19 else "",
            "address": normalize(row[14].value) if len(row) > 14 else "",
            "option": option or product,
            "memo": "",
        })
    return entries


# 하위호환 alias (기존 호출부 보호)
parse_alwayz_watermelon_entries = parse_alwayz_jewelry_entries


def _build_jewelry_order_workbook(entries: list[dict]) -> tuple[bytes, str, dict]:
    """쥬얼리팜 발주 entries(각 entry에 'product' 품목명 포함) → pbfcompany 발주서."""
    template_path = TEMPLATE_DIR / "명이나물_pbfcompany_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)
    now = datetime.now(KST)
    today_str = now.strftime("%Y-%m-%d")

    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=3).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}
    for i, entry in enumerate(entries):
        out_row = start_row + i
        product = entry["product"]
        try:
            qty_int = int(float(entry["qty"]))
        except (TypeError, ValueError):
            qty_int = 1

        mapping = {
            1: today_str,
            2: "아이티소프트",
            3: entry["name"],
            4: entry["phone"],
            6: entry["address"],
            7: product,
            8: str(qty_int),
            9: "식품애착",
            10: "010-5700-7756",
            12: entry.get("memo") or "",
            13: entry.get("order_id") or "",
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        opt_key = entry.get("option") or product
        bucket = option_totals.setdefault(opt_key, {
            "coupang_option_keyword": opt_key,
            "vendor_option_name": product,
            "quantity": 0,
            "orders": [],
        })
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    products = [e["product"] for e in entries]
    has_watermelon = any("수박" in p for p in products)
    has_baekdo = any("백도" in p for p in products)
    has_peach = any("복숭아" in p and "백도" not in p for p in products)
    has_chamoe = any("과" in p and "수박" not in p and "복숭아" not in p for p in products)
    label_parts = []
    if has_watermelon:
        label_parts.append("수박")
    if has_chamoe:
        label_parts.append("성주참외")
    if has_baekdo:
        label_parts.append("백도딱딱이복숭아")
    if has_peach:
        label_parts.append("신비복숭아")
    label = "_".join(label_parts) if label_parts else "발주"

    filename = f"쥬얼리프룻_{label}_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(entries),
        "supplier": "쥬얼리프룻",
        "product": "·".join(label_parts) if label_parts else "쥬얼리팜",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats


# 하위호환 alias
_build_jewelry_watermelon_workbook = _build_jewelry_order_workbook


async def process_jewelry_watermelon_order(
    from_date: str,
    to_date: str,
    alwayz_bytes: bytes | None = None,
    collect_toss: bool = True,
) -> list[tuple[bytes, str, dict]]:
    """토스 API + 올웨이즈 주문을 쥬얼리프룻 발주서로 출력.

    대상: 수박 6/7/8kg · 성주참외 · 신비복숭아(소과/대과 등 옵션 그대로) · (블랙)망고수박.
    반환: [(bytes, filename, stats)] (쥬얼리 발주서 1개).
    """
    from app.processors.tomato_order import collect_toss_jewelry_orders

    toss_error = ""
    toss_entries: list[dict] = []
    if collect_toss:
        try:
            toss_entries = await collect_toss_jewelry_orders(from_date, to_date)
        except Exception as exc:
            if not alwayz_bytes:
                raise
            toss_error = str(exc)

    entries: list[dict] = [{
        "order_id": t.get("order_id", ""),
        "product": t["product"],
        "qty": t.get("qty", "1"),
        "name": t.get("name", ""),
        "phone": t.get("phone", ""),
        "address": t.get("address", ""),
        "option": t.get("option") or t["product"],
        "memo": t.get("memo", ""),
    } for t in toss_entries]
    alwayz_entries = parse_alwayz_jewelry_entries(alwayz_bytes) if alwayz_bytes else []
    jewelry_entries = entries + alwayz_entries

    if not jewelry_entries:
        raise ValueError(
            f"토스/올웨이즈에서 쥬얼리프룻 발주 대상(수박·성주참외·신비복숭아·망고수박) 주문을 찾지 못했습니다. 기간: {from_date} ~ {to_date}"
        )

    output_bytes, filename, stats = _build_jewelry_order_workbook(jewelry_entries)
    stats = {
        **stats,
        "toss_orders": len(entries),
        "alwayz_orders": len(alwayz_entries),
        "period": f"{from_date} ~ {to_date}" if collect_toss else "올웨이즈 파일",
    }
    if toss_error:
        stats["toss_error"] = toss_error
    return [(output_bytes, filename, stats)]
