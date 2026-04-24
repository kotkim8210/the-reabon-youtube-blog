import re
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def convert_option(option_text: str) -> str | None:
    """Convert option like '500g 1박스' to '참두릅 500g'."""
    if not option_text:
        return None
    text = str(option_text).strip()
    match = re.match(r"(\d+(?:kg|g))\s+1박스", text, re.IGNORECASE)
    if match:
        weight = match.group(1)
        return f"참두릅 {weight}"
    return text


def process(delivery_file_bytes: bytes) -> tuple[bytes, str, dict]:
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows = []
    for row in dl_ws.iter_rows(min_row=2):
        k_val = normalize(row[10].value) if len(row) > 10 else ""
        if "참두릅" in k_val:
            filtered_rows.append(row)

    template_path = TEMPLATE_DIR / "참두릅_jaehwan0330_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)

    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=5).value is None:
            start_row = r
            break

    for i, row in enumerate(filtered_rows):
        out_row = start_row + i

        order_no = normalize(row[2].value) if len(row) > 2 else ""
        buyer = normalize(row[24].value) if len(row) > 24 else ""
        buyer_phone = normalize(row[25].value) if len(row) > 25 else ""
        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_val = row[22].value if len(row) > 22 else ""
        qty = normalize(qty_val)

        option_converted = convert_option(option)

        mapping = {
            1: order_no,            # A - 주문번호
            2: buyer,               # B - 주문자명
            3: buyer_phone,         # C - 주문자연락처
            # 4: "",                # D - 보내는사람 주소 (빈값)
            5: name,                # E - 수취인명
            6: phone,               # F - 수취인연락처
            7: address,             # G - 수취인주소
            8: product_name,        # H - 상품명
            9: option_converted,    # I - 옵션명
            10: qty,                # J - 수량
            11: memo,               # K - 배송메세지
        }

        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"참두릅_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {"total": len(filtered_rows)}

    return output.read(), filename, stats
