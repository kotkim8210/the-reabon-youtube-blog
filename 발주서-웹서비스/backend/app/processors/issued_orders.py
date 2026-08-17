"""발주 이력 기반 중복발주 방지 헬퍼.

발주서를 생성할 때 포함된 주문번호를 DB(issued_order_items)에 기록하고,
다음 생성 시 '오늘 이전 날짜'에 기록된 주문번호를 입력(DeliveryList/토스 entries)에서
미리 걸러낸다. 같은 날 재생성은 걸러지지 않는다(아침 재출력·추가발주 워크플로 보존).

키는 **주문번호 + 옵션** 복합키 'order_id|옵션(공백제거)'.
한 주문번호로 여러 옵션을 사면(쿠팡은 가능) 주문번호만으로는 구분이 안 돼,
어제 발주한 A옵션 때문에 오늘 새로 산 B옵션까지 통째로 빠졌다
(2026-08-17 김희조 청사과 소과 5kg/대과 5kg 동일 주문번호 사고).

과거에 주문번호 단독으로 기록된 이력(파이프 없는 키)은 그대로 존중해
그 주문번호 전체를 제외한다 — 키 스킴 변경으로 중복발주가 나지 않게 하는 하위호환.
"""
import re
from io import BytesIO

from openpyxl import load_workbook

KEY_SEP = "|"


def normalize_order_id(value) -> str:
    """주문번호를 문자열로 정규화. 엑셀이 숫자로 읽은 경우(float/int) 지수표기·소수점 방지."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def normalize_option(value) -> str:
    """옵션 문자열 정규화 — 공백 완전 제거.

    프로세서마다 옵션을 저장하는 형태(공백 축약/제거)가 조금씩 달라, 기록·필터 양쪽에서
    공백을 모두 지운 형태로 맞춰야 같은 옵션이 같은 키가 된다.
    """
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def make_order_key(order_id, option_text) -> str:
    """'주문번호|옵션' 복합키. 옵션이 없으면 주문번호만(과거 형태와 동일)."""
    oid = normalize_order_id(order_id)
    if not oid:
        return ""
    option = normalize_option(option_text)
    return f"{oid}{KEY_SEP}{option}" if option else oid


def _split_exclusions(exclude_keys: set[str]) -> tuple[set[str], set[str]]:
    """제외 집합을 (복합키, 주문번호 단독 레거시키)로 나눈다."""
    composite = {k for k in exclude_keys if KEY_SEP in k}
    legacy = {k for k in exclude_keys if KEY_SEP not in k}
    return composite, legacy


def filter_delivery_by_issued(
    delivery_bytes: bytes,
    exclude_order_ids: set[str],
    skipped_names: list[str] | None = None,
) -> tuple[bytes, int]:
    """DeliveryList에서 이미 발주된 주문번호(C열) 행을 제거한 bytes와 제거 건수를 반환.

    skipped_names를 넘기면 제거된 행의 수취인 이름(AA열)을 담아준다 —
    무엇이 왜 빠졌는지 사용자에게 보여 침묵 제외를 방지.
    """
    if not exclude_order_ids:
        return delivery_bytes, 0
    composite, legacy = _split_exclusions(exclude_order_ids)
    wb = load_workbook(filename=BytesIO(delivery_bytes))
    ws = wb.active
    to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        order_id = normalize_order_id(ws.cell(row=row_idx, column=3).value)  # C열 = 주문번호
        if not order_id:
            continue
        option = ws.cell(row=row_idx, column=12).value  # L열 = 옵션
        if not normalize_option(option):
            option = ws.cell(row=row_idx, column=11).value  # 옵션이 비면 K열 상품명
        # 복합키(주문번호+옵션)가 맞거나, 과거 주문번호 단독 기록이면 제외
        if make_order_key(order_id, option) in composite or order_id in legacy:
            to_delete.append(row_idx)
            if skipped_names is not None:
                name = str(ws.cell(row=row_idx, column=27).value or "").strip()  # AA열 = 수취인이름
                skipped_names.append(name or order_id)
    if not to_delete:
        return delivery_bytes, 0
    for row_idx in reversed(to_delete):
        ws.delete_rows(row_idx)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), len(to_delete)


def filter_entries_by_issued(
    entries: list[dict],
    exclude_order_ids: set[str],
    skipped_names: list[str] | None = None,
) -> tuple[list[dict], int]:
    """토스/올웨이즈/테무 entry 목록에서 이미 발주된 order_id를 제거."""
    if not exclude_order_ids or not entries:
        return entries, 0
    composite, legacy = _split_exclusions(exclude_order_ids)
    kept = []
    for entry in entries:
        order_id = normalize_order_id(entry.get("order_id"))
        option = entry.get("option") or entry.get("product")
        if order_id and (make_order_key(order_id, option) in composite or order_id in legacy):
            if skipped_names is not None:
                skipped_names.append(str(entry.get("name") or entry.get("order_id") or "").strip())
            continue
        kept.append(entry)
    return kept, len(entries) - len(kept)


def order_ids_from_stats(*stats_dicts: dict) -> list[str]:
    """프로세서 stats에서 발주서에 실제 기록된 '주문번호|옵션' 복합키를 수집.

    옵션은 bucket의 coupang_option_keyword(=DeliveryList 옵션 원문)를 쓴다 —
    다음 발주 때 DeliveryList L열과 대조하는 값이라 같은 출처여야 맞는다.
    """
    keys: set[str] = set()
    for stats in stats_dicts:
        for bucket in (stats or {}).get("options") or []:
            option = bucket.get("coupang_option_keyword") or bucket.get("vendor_option_name")
            for order in bucket.get("orders") or []:
                key = make_order_key(order.get("order_id"), option)
                if key:
                    keys.add(key)
    return sorted(keys)
