import csv
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR
from app.processors import tomato_order

JEWELRYFRUIT_WATERMELON_KGS = {2, 3, 5}
SUPPORTED_WATERMELON_KGS = JEWELRYFRUIT_WATERMELON_KGS | tomato_order.JBT_WATERMELON_KGS


@dataclass
class _VirtualCell:
    value: object = ""


@dataclass
class _VirtualWorksheet:
    rows: list[list[object]]

    @property
    def max_row(self) -> int:
        return len(self.rows)

    @property
    def max_column(self) -> int:
        return max((len(row) for row in self.rows), default=0)

    def cell(self, row: int, column: int) -> _VirtualCell:
        row_idx = row - 1
        col_idx = column - 1
        if row_idx < 0 or col_idx < 0 or row_idx >= len(self.rows):
            return _VirtualCell()
        current_row = self.rows[row_idx]
        if col_idx >= len(current_row):
            return _VirtualCell()
        return _VirtualCell(current_row[col_idx])


def _normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _header_key(value) -> str:
    return re.sub(r"\s+", "", _normalize(value)).lower()


def _int_value(value, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = _normalize(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    try:
        return int(float(match.group(0)))
    except ValueError:
        return default


def _lines(text: str) -> list[str]:
    return [_normalize(line) for line in re.split(r"[\r\n]+", text or "") if _normalize(line)]


def _value_after_label(lines: list[str], labels: tuple[str, ...]) -> str:
    normalized_labels = tuple(_normalize(label) for label in labels)
    for idx, line in enumerate(lines):
        for label in normalized_labels:
            if line == label:
                for next_line in lines[idx + 1 :]:
                    if next_line:
                        return next_line
            if line.startswith(label):
                suffix = line[len(label) :].strip(" :：")
                if suffix:
                    return suffix
    return ""


def _normalize_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if digits.startswith("82"):
        rest = digits[2:]
        # 테무는 '+82 010 8215 2937'처럼 82 뒤에 0이 이미 있는 경우가 있어 중복 0을 막는다.
        digits = rest if rest.startswith("0") else "0" + rest
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return _normalize(value)


def _extract_order_id(text: str) -> str:
    match = re.search(r"\b(PO-\d{3}-\d+)\b", text)
    if match:
        return match.group(1)
    match = re.search(r"주문\s*ID\s*[:：]?\s*([A-Z0-9-]+)", text, re.IGNORECASE)
    return match.group(1) if match else ""


def _extract_phone(text: str) -> str:
    match = re.search(r"(\+?82[\s-]*10[\s-]*\d{3,4}[\s-]*\d{4})", text)
    if match:
        return _normalize_phone(match.group(1))
    match = re.search(r"\b(01[016789][-\s]?\d{3,4}[-\s]?\d{4})\b", text)
    return _normalize_phone(match.group(1)) if match else ""


def _extract_address(lines: list[str]) -> tuple[str, str]:
    stop_keywords = (
        "상기 명시된",
        "변경 요청",
        "전화번호",
        "가상 이메일",
        "주문 내용",
        "판매자 메모",
        "매출액",
        "고객 결제",
    )
    address_parts: list[str] = []
    zipcode = ""
    capture = False

    for line in lines:
        if "배송 주소" in line:
            capture = True
            suffix = line.split("배송 주소", 1)[1].strip(" :：")
            if suffix:
                address_parts.append(suffix)
            continue
        if not capture:
            continue
        if any(keyword in line for keyword in stop_keywords) or "@" in line:
            break
        if "수박" in line or re.fullmatch(r"\d+(?:\.\d+)?\s*kg", line, re.IGNORECASE):
            break
        if re.fullmatch(r"\d{5}", line):
            zipcode = line
            continue
        if line.lower() in {"republic of korea", "south korea", "korea"}:
            continue
        address_parts.append(line)

    return zipcode, _normalize(" ".join(address_parts))


def _extract_product(lines: list[str]) -> str:
    watermelon_lines = [line for line in lines if "수박" in line]
    for line in watermelon_lines:
        if re.search(r"\d+(?:\.\d+)?\s*kg", line, re.IGNORECASE):
            return line
    return watermelon_lines[0] if watermelon_lines else ""


def _kg_from_text(text: str) -> int | None:
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*kg", text or "", re.IGNORECASE):
        try:
            kg = float(match.group(1))
        except ValueError:
            continue
        if kg.is_integer():
            kg_int = int(kg)
            if kg_int in SUPPORTED_WATERMELON_KGS:
                return kg_int
    return None


def _extract_watermelon_kg(lines: list[str], product_name: str) -> int | None:
    kg = _kg_from_text(product_name)
    if kg:
        return kg

    for idx, line in enumerate(lines):
        if line != product_name:
            continue
        # Temu often copies the option as a separate next line, e.g. product name then "6kg".
        for next_line in lines[idx + 1 : idx + 4]:
            kg = _kg_from_text(next_line)
            if kg:
                return kg

    product_idx = next((idx for idx, line in enumerate(lines) if line == product_name), -1)
    if product_idx >= 0:
        nearby = " ".join(lines[product_idx : product_idx + 6])
        kg = _kg_from_text(nearby)
        if kg:
            return kg

    return _kg_from_text("\n".join(lines))


def _extract_quantity(lines: list[str]) -> int:
    for idx, line in enumerate(lines):
        match = re.search(r"(?:수량|미발송\s*수량)\s*[:：]?\s*(\d+)", line)
        if match:
            return max(int(match.group(1)), 1)
        if line == "수량":
            for next_line in lines[idx + 1 : idx + 4]:
                match = re.search(r"\d+", next_line)
                if match:
                    return max(int(match.group(0)), 1)
    return 1


def _find_temu_header_row(ws) -> int:
    max_scan = min(ws.max_row, 20)
    for row_idx in range(1, max_scan + 1):
        headers = {_header_key(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)}
        if {"주문id", "선택사항", "발송할수량"}.issubset(headers) and "수령인이름" in headers:
            return row_idx
    raise ValueError("테무 주문 파일에서 주문 헤더 행을 찾지 못했습니다. order_export 원본 파일인지 확인해주세요.")


def _temu_header_map(ws, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _header_key(ws.cell(row=header_row, column=col).value)
        if key:
            columns.setdefault(key, col)
    return columns


def _find_col(columns: dict[str, int], *labels: str) -> int | None:
    keys = [_header_key(label) for label in labels]
    for key in keys:
        if key in columns:
            return columns[key]
    for header, col in columns.items():
        if any(key and key in header for key in keys):
            return col
    return None


def _row_text(ws, row_idx: int, columns: dict[str, int], *labels: str) -> str:
    col = _find_col(columns, *labels)
    if not col:
        return ""
    return _normalize(ws.cell(row=row_idx, column=col).value)


def _row_int(ws, row_idx: int, columns: dict[str, int], *labels: str) -> int:
    col = _find_col(columns, *labels)
    if not col:
        return 0
    return _int_value(ws.cell(row=row_idx, column=col).value)


def _temu_excel_address(ws, row_idx: int, columns: dict[str, int]) -> tuple[str, str]:
    zipcode = _row_text(ws, row_idx, columns, "배송 우편번호")
    parts = [
        _row_text(ws, row_idx, columns, "배송 주"),
        _row_text(ws, row_idx, columns, "배송 도시"),
        _row_text(ws, row_idx, columns, "지역"),
        _row_text(ws, row_idx, columns, "배송 주소 1"),
        _row_text(ws, row_idx, columns, "배송 주소 2"),
        _row_text(ws, row_idx, columns, "배송 주소 3"),
    ]
    address = _normalize(" ".join(part for part in parts if part and part.lower() not in {"republic of korea", "korea"}))
    return zipcode, address


def _parse_temu_order_excel(excel_bytes: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    ws = wb["Order report"] if "Order report" in wb.sheetnames else wb.active
    return _parse_temu_order_sheet(ws, "엑셀")


def _parse_temu_order_sheet(ws, source_label: str) -> list[dict]:
    header_row = _find_temu_header_row(ws)
    columns = _temu_header_map(ws, header_row)

    entries: list[dict] = []
    unsupported_watermelon_rows: list[str] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        order_id = _row_text(ws, row_idx, columns, "주문 ID")
        product_name = _row_text(ws, row_idx, columns, "제품 이름") or _row_text(
            ws, row_idx, columns, "고객 주문별 제품 이름"
        )
        option = _row_text(ws, row_idx, columns, "선택 사항")
        if not any([order_id, product_name, option]):
            continue

        status_text = " ".join(
            [
                _row_text(ws, row_idx, columns, "주문 상태"),
                _row_text(ws, row_idx, columns, "주문 상품 상태"),
            ]
        )
        if any(keyword in status_text for keyword in ("취소", "환불", "반품")):
            continue

        ship_qty = _row_int(ws, row_idx, columns, "발송할 수량")
        qty = ship_qty
        if qty <= 0 and "미발송" in status_text:
            qty = _row_int(ws, row_idx, columns, "구매 수량")
        if qty <= 0:
            continue

        product_text = f"{product_name} {option}"
        is_watermelon = "수박" in product_text
        is_chamoe = "참외" in product_text
        is_goguma = "고구마" in product_text
        if not (is_watermelon or is_chamoe or is_goguma):
            continue

        recipient = _row_text(ws, row_idx, columns, "수령인 이름")
        phone = _normalize_phone(_row_text(ws, row_idx, columns, "수령인 전화번호"))
        zipcode, address = _temu_excel_address(ws, row_idx, columns)
        order_product_id = _row_text(ws, row_idx, columns, "주문 상품 ID", "상품 주문 ID")

        missing = []
        if not recipient:
            missing.append("수령인 이름")
        if not phone:
            missing.append("전화번호")
        if not address:
            missing.append("배송 주소")
        if missing:
            raise ValueError(f"테무 주문 {source_label} {row_idx}행에서 {', '.join(missing)} 정보를 찾지 못했습니다.")

        if is_goguma:
            # 고구마 → 해달 발주서(한진양식). 품목명 '꿀고구마 {kg}Kg ({등급})'로 정규화.
            from app.processors.goguma_order import canonical_goguma_option

            goguma_product = canonical_goguma_option(option or product_name, product_name)
            entries.append(
                {
                    "kind": "goguma",
                    "order_id": order_id
                    or order_product_id
                    or tomato_order.stable_order_id("temu-goguma", recipient, phone, address, product_name, option, qty),
                    "order_product_id": order_product_id,
                    "name": recipient,
                    "phone": phone,
                    "zipcode": zipcode,
                    "address": address,
                    "qty": qty,
                    "product_name": product_name,
                    "product": goguma_product,
                    "memo": "테무 주문",
                }
            )
            continue

        from app.processors.myeongi_order import jewelry_external_product

        jewelry_product = jewelry_external_product(product_name, option, SUPPORTED_WATERMELON_KGS)
        if not jewelry_product:
            # 수박인데 지원 kg(2/3/5/6/7/8)이 아닌 경우만 오류로 수집
            if is_watermelon:
                unsupported_watermelon_rows.append(f"{row_idx}행 {option or product_name}")
            continue
        kg = _kg_from_text(product_text) or 0

        entries.append(
            {
                "kind": "jewelry",
                "order_id": order_id
                or order_product_id
                or tomato_order.stable_order_id("temu-jewelry", recipient, phone, address, product_name, option, qty),
                "order_product_id": order_product_id,
                "name": recipient,
                "phone": phone,
                "zipcode": zipcode,
                "address": address,
                "qty": qty,
                "product_name": product_name,
                "option": jewelry_product,
                "memo": "테무 주문",
                "kg": kg,
                "product": jewelry_product,
            }
        )

    if unsupported_watermelon_rows:
        raise ValueError(
            "테무 수박 주문 중 2kg/3kg/5kg/6kg/7kg/8kg 옵션이 아닌 행이 있습니다: "
            + ", ".join(unsupported_watermelon_rows[:5])
        )
    if not entries:
        raise ValueError("테무 주문 파일에서 발송할 수량이 있는 미발송 수박·성주참외·고구마 주문을 찾지 못했습니다.")

    return entries


def _decode_csv_text(csv_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp949", "euc-kr"):
        try:
            return csv_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("테무 주문 CSV 파일 인코딩을 읽지 못했습니다. UTF-8 또는 CP949 CSV로 저장해 다시 올려주세요.")


def _parse_temu_order_csv(csv_bytes: bytes) -> list[dict]:
    text = _decode_csv_text(csv_bytes)
    rows = [row for row in csv.reader(StringIO(text)) if any(_normalize(value) for value in row)]
    if not rows:
        raise ValueError("테무 주문 CSV 파일이 비어 있습니다.")
    return _parse_temu_order_sheet(_VirtualWorksheet(rows), "CSV")


def _parse_temu_order_text(order_text: str) -> dict:
    lines = _lines(order_text)
    joined = "\n".join(lines)

    order_id = _extract_order_id(joined)
    recipient = _value_after_label(lines, ("수령인 이름", "받는 사람", "수취인"))
    phone = _extract_phone(joined)
    zipcode, address = _extract_address(lines)
    product_name = _extract_product(lines)
    qty = _extract_quantity(lines)

    from app.processors.myeongi_order import jewelry_external_product

    is_watermelon = bool(product_name) and "수박" in product_name
    is_chamoe = bool(product_name) and "참외" in product_name
    if not (is_watermelon or is_chamoe):
        raise ValueError("현재 테무 발주 자동화는 수박·성주참외 주문을 지원합니다. 주문상세 텍스트에 상품명이 있는지 확인해주세요.")

    if is_watermelon:
        kg = _extract_watermelon_kg(lines, product_name)
    else:
        kg = 0
    option_text = " ".join(lines)
    jewelry_product = jewelry_external_product(product_name, option_text, SUPPORTED_WATERMELON_KGS)
    if not jewelry_product:
        raise ValueError("테무 수박 주문에서 2kg/3kg/5kg/6kg/7kg/8kg 옵션을 찾지 못했습니다. 상품명에 kg가 보이도록 주문상세 전체를 복사해주세요.")

    missing = []
    if not recipient:
        missing.append("수령인 이름")
    if not phone:
        missing.append("전화번호")
    if not address:
        missing.append("배송 주소")
    if missing:
        raise ValueError(f"테무 주문상세 텍스트에서 {', '.join(missing)} 정보를 찾지 못했습니다.")

    return {
        "order_id": order_id
        or tomato_order.stable_order_id("temu-jewelry", recipient, phone, address, product_name, jewelry_product, qty),
        "name": recipient,
        "phone": phone,
        "zipcode": zipcode,
        "address": address,
        "qty": qty,
        "product_name": product_name,
        "option": jewelry_product,
        "memo": "테무 주문",
        "kg": kg,
        "product": jewelry_product,
    }


def _virtual_row(entry: dict) -> list[_VirtualCell]:
    row = [_VirtualCell() for _ in range(31)]
    row[2].value = entry["order_id"]
    row[10].value = entry["product_name"]
    row[11].value = entry["option"]
    row[22].value = entry["qty"]
    row[26].value = entry["name"]
    row[27].value = entry["phone"]
    row[28].value = entry["zipcode"]
    row[29].value = entry["address"]
    row[30].value = entry["memo"]
    return row


def _jewelryfruit_product_name(kg: int) -> str:
    return f"가정용 수박 {kg}kg 내외"


def _build_jewelryfruit_workbook(entries: list[dict]) -> tuple[bytes, str, dict]:
    template_path = TEMPLATE_DIR / "명이나물_pbfcompany_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    if not entries:
        raise ValueError("쥬얼리프룻 발주로 출력할 테무 수박 주문이 없습니다.")

    wb = load_workbook(filename=str(template_path))
    first_sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames[1:]:
        del wb[name]
    ws = wb[first_sheet_name]

    font11 = Font(size=11)
    now = datetime.now(tomato_order.KST)
    excel_date = now.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)

    option_buckets: dict[str, dict] = {}
    for index, entry in enumerate(entries, start=2):
        product = entry.get("product") or _jewelryfruit_product_name(int(entry.get("kg") or 0))
        qty = int(entry["qty"])

        mapping = {
            1: excel_date,
            2: "아이티소프트",
            3: entry["name"],
            4: entry["phone"],
            6: entry["address"],
            7: product,
            8: str(qty),
            9: "식품애착",
            10: "010-5700-7756",
            12: "문 앞",
            13: entry["order_id"],
        }
        for col, value in mapping.items():
            cell = ws.cell(row=index, column=col, value=value)
            cell.font = font11
            if col == 1:
                cell.number_format = "yyyy-mm-dd"

        bucket = option_buckets.setdefault(
            entry["option"],
            {
                "coupang_option_keyword": entry["option"],
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty
        bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    products = [e.get("product") or "" for e in entries]
    has_watermelon = any("수박" in p for p in products)
    has_chamoe = any("과" in p and "수박" not in p for p in products)
    label_parts = []
    if has_watermelon:
        label_parts.append("수박")
    if has_chamoe:
        label_parts.append("성주참외")
    label = "_".join(label_parts) if label_parts else "발주"

    today = now.strftime("%Y%m%d")
    filename = f"쥬얼리프룻_{label}_발주({today}).xlsx"
    stats = {
        "total": len(entries),
        "platform": "테무",
        "supplier": "쥬얼리프룻",
        "quantity": sum(int(entry["qty"]) for entry in entries),
        "product": "·".join(label_parts) if label_parts else "쥬얼리팜",
        "options": list(option_buckets.values()),
    }
    return output.read(), filename, stats


def _build_haedal_goguma_workbook(entries: list[dict]) -> tuple[bytes, str, dict]:
    """테무 고구마 주문 → 해달 발주서(한진양식). 고구마 page의 해달 양식과 동일."""
    if not entries:
        raise ValueError("해달 발주로 출력할 테무 고구마 주문이 없습니다.")

    template_path = TEMPLATE_DIR / "해달_발주서_한진양식.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(filename=str(template_path))
    first_sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames[1:]:
        del wb[name]
    ws = wb[first_sheet_name]

    font11 = Font(size=11)
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    option_buckets: dict[str, dict] = {}
    for i, entry in enumerate(entries):
        row_idx = start_row + i
        qty = int(entry["qty"])
        product = entry["product"]
        mapping = {
            1: entry["name"],        # A 받으시는 분
            2: entry["phone"],       # B 받으시는 분 전화
            5: entry["zipcode"],     # E 받는분우편번호
            6: entry["address"],     # F 받는분총주소
            7: "식품애착",            # G 보내시는 분
            8: "010-5700-7756",      # H 보내시는 분 전화
            12: "전라남도 해남군 산이면 새상골길 ",  # L 보내는분총주소
            13: str(qty),            # M 수량
            14: product,             # N 품목명
            16: "선불",              # P 지불조건
            19: entry.get("memo", ""),  # S 메모1
        }
        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11

        bucket = option_buckets.setdefault(
            product,
            {"coupang_option_keyword": product, "vendor_option_name": product, "quantity": 0, "orders": []},
        )
        bucket["quantity"] += qty
        bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now(tomato_order.KST).strftime("%y%m%d")
    filename = f"해달 발주서 한진양식_테무_알제이시스템즈({today}).xlsx"
    stats = {
        "total": len(entries),
        "platform": "테무",
        "supplier": "해달",
        "product": "고구마",
        "quantity": sum(int(e["qty"]) for e in entries),
        "options": list(option_buckets.values()),
    }
    return output.read(), filename, stats


def _build_outputs(entries: list[dict]) -> list[tuple[bytes, str, dict]]:
    # 테무 수박·성주참외 → 쥬얼리프룻 발주서, 고구마 → 해달 발주서(한진양식).
    # 파싱 단계에서 kind('jewelry'/'goguma')와 발주 품목명을 확정해 둠.
    if not entries:
        raise ValueError("테무 발주로 출력할 수박·성주참외·고구마 주문을 찾지 못했습니다.")
    jewelry = [e for e in entries if e.get("kind") != "goguma"]
    goguma = [e for e in entries if e.get("kind") == "goguma"]
    outputs: list[tuple[bytes, str, dict]] = []
    if jewelry:
        outputs.append(_build_jewelryfruit_workbook(jewelry))
    if goguma:
        outputs.append(_build_haedal_goguma_workbook(goguma))
    if not outputs:
        raise ValueError("테무 발주로 출력할 주문을 찾지 못했습니다.")
    return outputs


def combine_output_stats(outputs: list[tuple[bytes, str, dict]]) -> dict:
    products = [str((stats or {}).get("product", "")) for _, _, stats in outputs]
    product_label = "·".join(p for p in products if p) or "수박·성주참외·고구마"
    return {
        "files": len(outputs),
        "total": sum(int((stats or {}).get("total", 0)) for _, _, stats in outputs),
        "quantity": sum(int((stats or {}).get("quantity", 0)) for _, _, stats in outputs),
        "platform": "테무",
        "product": product_label,
        "filenames": [filename for _, filename, _ in outputs],
    }


def process(order_text: str) -> tuple[bytes, str, dict]:
    outputs = _build_outputs([_parse_temu_order_text(order_text)])
    return outputs[0]


def process_excel(excel_bytes: bytes) -> list[tuple[bytes, str, dict]]:
    return _build_outputs(_parse_temu_order_excel(excel_bytes))


def process_file(file_bytes: bytes, filename: str = "") -> list[tuple[bytes, str, dict]]:
    if filename.lower().endswith(".csv"):
        return _build_outputs(_parse_temu_order_csv(file_bytes))
    try:
        return process_excel(file_bytes)
    except BadZipFile:
        return _build_outputs(_parse_temu_order_csv(file_bytes))


def extract_goguma_entries(file_bytes: bytes, filename: str = "") -> list[dict]:
    """테무 order_export에서 고구마(해달 발주) entries만 추출. 고구마 주문이 없으면 빈 리스트.

    goguma-auto(쿠팡+토스 통합 발주)에 테무 주문을 합칠 때 사용 —
    entry 필드(name/phone/zipcode/address/qty/product/memo/order_id)는 해달 양식 행 쓰기와 호환.
    """
    try:
        if filename.lower().endswith(".csv"):
            entries = _parse_temu_order_csv(file_bytes)
        else:
            try:
                entries = _parse_temu_order_excel(file_bytes)
            except BadZipFile:
                entries = _parse_temu_order_csv(file_bytes)
    except ValueError:
        return []
    return [e for e in entries if e.get("kind") == "goguma"]
