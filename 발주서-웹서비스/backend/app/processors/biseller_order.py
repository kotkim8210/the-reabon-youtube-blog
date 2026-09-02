"""비셀러 발주서 생성 (LA한입갈비) — 2026-09 신규.

쿠팡 DeliveryList에서 '한입 LA갈비' 주문을 뽑아 비셀러 발주 양식으로 채운다.
비셀러 상품명은 800g 낱개 개수를 그대로 풀어 쓴 표기다.
  쿠팡 '800g 4개' → '양념LA한입갈비 800g+800g+800g+800g (800G*4세트)'
"""

import logging
import re
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR

logger = logging.getLogger(__name__)

KST = timezone(timedelta(hours=9))

TEMPLATE_NAME = "비셀러_발주서_원본.xlsx"
# 발주서에 고정으로 들어가는 보내는 사람(주문자) 정보 — 양식 J·K열.
SENDER_NAME = "(주)아이티소프트"
SENDER_PHONE = "010-5700-7756"
PRODUCT_LABEL = "LA한입갈비(비셀러)"


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _compact(*values: object) -> str:
    return re.sub(r"\s+", "", " ".join(normalize(v) for v in values))


def is_biseller_galbi_order(product_name: object, option_text: object) -> bool:
    """비셀러 발주 대상(LA한입갈비) 여부."""
    text = _compact(product_name, option_text)
    if "갈비" not in text:
        return False
    return "한입" in text or "LA갈비" in text.upper()


def convert_galbi_option(product_name: object, option_text: object) -> str | None:
    """쿠팡 옵션 → 비셀러 상품명. 개수를 못 읽으면 None."""
    if not is_biseller_galbi_order(product_name, option_text):
        return None
    text = _compact(product_name, option_text)
    match = re.search(r"800g\s*(\d+)\s*개", text, re.IGNORECASE) or re.search(r"(\d+)\s*개", text)
    if not match:
        return None
    count = int(match.group(1))
    if count < 1:
        return None
    return f"양념LA한입갈비 {'+'.join(['800g'] * count)} (800G*{count}세트)"


def _find_total_row(ws) -> int:
    """'합계' 행 번호. 못 찾으면 마지막 행 다음."""
    for r in range(2, ws.max_row + 1):
        if normalize(ws.cell(row=r, column=7).value) == "합계":
            return r
    return ws.max_row + 1


def process(delivery_file_bytes: bytes) -> tuple[bytes, str, dict]:
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows: list[tuple] = []
    unmatched: list[str] = []
    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        converted = convert_galbi_option(product_name, option)
        if converted:
            filtered_rows.append((row, converted))
        elif is_biseller_galbi_order(product_name, option):
            recipient = normalize(row[26].value) if len(row) > 26 else ""
            unmatched.append(f"{recipient or '이름없음'}({option or product_name}) — 수량 표기를 못 읽음")

    if unmatched:
        logger.warning("비셀러 발주 미매칭 %d건: %s", len(unmatched), "; ".join(unmatched))

    template_path = TEMPLATE_DIR / TEMPLATE_NAME
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"Please place the template file in the templates directory."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    total_row = _find_total_row(ws)
    capacity = total_row - 2  # 데이터 행 수(2행 ~ 합계행 직전)
    if len(filtered_rows) > capacity:
        # 주문이 양식 칸보다 많으면 합계 행 앞에 행을 넣어 늘린다(수식은 아래에서 다시 씀).
        ws.insert_rows(total_row, len(filtered_rows) - capacity)
        total_row = _find_total_row(ws)

    font11 = Font(size=11)
    now = datetime.now(KST)
    order_date = now.strftime("%Y-%m-%d")

    option_totals: dict[str, dict] = {}
    for i, (row, product) in enumerate(filtered_rows):
        out_row = 2 + i

        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        zipcode = normalize(row[28].value) if len(row) > 28 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        order_no = normalize(row[2].value) if len(row) > 2 else ""
        qty_val = row[22].value if len(row) > 22 else ""

        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                pass

        mapping = {
            1: i + 1,        # A 순번
            2: order_date,   # B 발주일
            3: name,         # C 수취인명
            4: phone,        # D 수취인연락처
            5: zipcode,      # E 우편번호
            6: address,      # F 주소
            7: product,      # G 상품명(비셀러 상품명)
            8: qty_int,      # H 수량
            9: memo,         # I 배송메세지
            10: SENDER_NAME,   # J 주문자명
            11: SENDER_PHONE,  # K 주문자연락처
            # L 택배사 / M 송장번호는 거래처가 채운다
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            option or product,
            {
                "coupang_option_keyword": option or product,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    # 남는 빈 칸의 순번은 지우고, 합계 수식은 실제 데이터 범위로 다시 쓴다.
    for r in range(2 + len(filtered_rows), total_row):
        ws.cell(row=r, column=1, value=None)
    last_data_row = max(total_row - 1, 2)
    ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{last_data_row})")

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    filename = f"아이티소프트_비셀러발주서_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "total": len(filtered_rows),
        "product": PRODUCT_LABEL,
        "options": list(option_totals.values()),
    }
    if unmatched:
        stats["needs_check"] = unmatched

    return output.read(), filename, stats
