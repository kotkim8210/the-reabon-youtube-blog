import re
import logging
from datetime import datetime, timezone, timedelta
from io import BytesIO
from collections import defaultdict

from openpyxl import load_workbook

from app.config import TEMPLATE_DIR
from app.processors.tomato_order import (
    JBT_WATERMELON_KGS,
    JBT_WATERMELON_LABEL,
    delivery_watermelon_kg,
    is_toss_watermelon_order,
    transform_product,
)
from app.processors.tracking_match import (
    name_counts,
    normalize_courier_name,
    option_key_set,
    options_match,
    requires_option_guard,
)
from app.toss.client import toss_client


logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))
# (구)TOSS_DELIVERY_COMPANY="한진택배" 상수는 제거 — 쥬얼리·제주다팜 품목은 롯데 발송이라
# 토스 등록 기본값은 normalize_toss_courier(courier, default="롯데택배")로 지정한다.
TRACKABLE_TOSS_STATUSES = {"PAID", "PREPARING_PRODUCT", "DELIVERING"}


def normalize(value) -> str:
    """공백 완전 제거 후 문자열 반환 (이름 매칭용)"""
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def normalize_courier(name: str) -> str:
    """택배사명 정규화 — 쿠팡이 인식하는 형태로 변환.

    예: 'CJ대한통운' -> 'CJ 대한통운' (띄어쓰기 필수)
    """
    return normalize_courier_name(name)


def is_mixed_chamoe_row(product_name: str, option_text: str) -> bool:
    return "성주참외" in product_name and "가정용 혼합과" in option_text


def product_type_for(product_name: str) -> str:
    if "토마토" in product_name:
        return "토마토"
    if "참외" in product_name:
        return "참외"
    if "땅두릅" in product_name or "두릅" in product_name:
        return "땅두릅"
    if "수박" in product_name:
        return "수박"
    if "옥수수" in product_name:
        return "옥수수"
    if "복숭아" in product_name:
        return "복숭아"
    return ""


def corn_option_keys(*values: object) -> set[str]:
    text = normalize(" ".join(str(value) for value in values if value not in (None, "")))
    if "옥수수" not in text:
        return set()
    grade = "중품" if "중품" in text else "특품" if "특품" in text else ""
    counts = set(re.findall(r"(\d+)\s*(?:개|입)", text))
    if not grade or not counts:
        return set()
    return {f"corn:{grade}:{count}개" for count in counts}


def is_jbt_tracking_target(product_name: str, option_text: str) -> bool:
    if is_mixed_chamoe_row(product_name, option_text):
        return False
    product_type = product_type_for(product_name)
    if product_type in {"토마토", "땅두릅"}:
        return True
    # 신비복숭아는 3·4kg만 제이비티 (1·2kg은 쥬얼리, 800g은 제외)
    if product_type == "복숭아":
        from app.processors.tomato_order import peach_kg, JBT_PEACH_KGS
        return peach_kg(product_name, option_text) in JBT_PEACH_KGS
    # ※ 수박·성주참외는 2026-06부터 쥬얼리팜(myeongi_tracking 등) 운송장 대상 — 제이비티 제외
    if product_type == "옥수수":
        return bool(corn_option_keys(product_name, option_text))
    return False


def _toss_item_text(value) -> str:
    parts: list[str] = []

    def walk(node) -> None:
        if node is None:
            return
        if isinstance(node, dict):
            for child in node.values():
                walk(child)
            return
        if isinstance(node, (list, tuple)):
            for child in node:
                walk(child)
            return
        parts.append(str(node))

    walk(value)
    return " ".join(parts)


def _toss_order_id(item: dict) -> str:
    for key in ("orderNo", "orderId", "orderNumber", "orderSheetId", "paymentKey"):
        value = item.get(key)
        if value:
            return str(value)
    return ""


def _find_reply_col(headers: dict, predicates: list) -> int | None:
    """헤더명(정규화) 우선순위 리스트로 컬럼 번호를 찾는다. (없으면 None)"""
    for pred in predicates:
        for c in sorted(headers):
            h = headers[c]
            if h and pred(h):
                return c
    return None


def parse_reply_file(tomato_reply_bytes: bytes) -> list[dict]:
    """거래처 회신 파일 파싱 — 헤더명 기반 자동감지.

    지원 양식 (헤더 라벨로 컬럼 자동 인식):
      - 제이비티 회신: 송장번호/수령인/연락처/주소/품목명/택배사
      - 쥬얼리 orderlist: 운송장번호/수령인/수령인연락처1/주소/상품명/옵션/택배사
      - 신비복숭아 회신: 운송장번호/수령인명/수령인연락처/주소/주문상품명/주문옵션명/택배사
    헤더가 없으면 제이비티 고정 레이아웃(A=송장,B=수령인,C=연락처,D=주소,E=품목,K=택배사)으로 폴백.
    """
    tr_wb = load_workbook(filename=BytesIO(tomato_reply_bytes), data_only=True)
    tr_ws = tr_wb.active

    headers = {
        c: normalize(tr_ws.cell(row=1, column=c).value)
        for c in range(1, tr_ws.max_column + 1)
    }

    col_tracking = _find_reply_col(headers, [
        lambda h: "운송장" in h,
        lambda h: "송장" in h,
    ])
    col_name = _find_reply_col(headers, [
        lambda h: h == "수령인명",
        lambda h: h == "수령인",
        lambda h: ("수령인" in h or "수취인" in h) and "연락처" not in h and "전화" not in h,
        lambda h: h in ("받는분성명", "성함", "이름", "받는분"),
    ])
    col_phone = _find_reply_col(headers, [
        lambda h: "수령인" in h and "연락처" in h and "1" in h,
        lambda h: "수령인" in h and ("연락처" in h or "전화" in h),
        lambda h: "연락처" in h or "전화" in h,
    ])
    col_addr = _find_reply_col(headers, [lambda h: "주소" in h])
    col_courier = _find_reply_col(headers, [lambda h: "택배" in h])
    col_product = _find_reply_col(headers, [lambda h: "상품명" in h, lambda h: "품목" in h])
    col_option = _find_reply_col(headers, [lambda h: "옵션" in h])

    header_based = col_tracking is not None and col_name is not None

    def cell(r, c):
        return tr_ws.cell(row=r, column=c).value if c else None

    entries = []
    for row_idx in range(2, tr_ws.max_row + 1):
        if header_based:
            tracking = normalize(cell(row_idx, col_tracking))
            name = normalize(cell(row_idx, col_name))
            phone = normalize(cell(row_idx, col_phone))
            address = normalize(cell(row_idx, col_addr))
            product = cell(row_idx, col_product)
            option_extra = cell(row_idx, col_option)
            courier_raw = cell(row_idx, col_courier)
        else:
            # 헤더 라벨이 없는 제이비티 고정 레이아웃 폴백
            tracking = normalize(tr_ws.cell(row=row_idx, column=1).value)
            name = normalize(tr_ws.cell(row=row_idx, column=2).value)
            phone = normalize(tr_ws.cell(row=row_idx, column=3).value)
            address = normalize(tr_ws.cell(row=row_idx, column=4).value)
            product = tr_ws.cell(row=row_idx, column=5).value
            option_extra = None
            courier_raw = tr_ws.cell(row=row_idx, column=11).value
        courier = str(courier_raw).strip() if courier_raw else ""

        if not name or not tracking:
            continue

        entries.append({
            "name": name,
            "phone": phone,
            "address": address,
            "tracking": tracking,
            "courier": courier,
            "option_keys": option_key_set(product, option_extra) | corn_option_keys(product, option_extra),
        })

    return entries


async def process_toss_watermelon_tracking(tomato_reply_bytes: bytes) -> dict:
    """거래처 회신의 운송장번호를 토스 주문에 API로 자동 등록.

    대상: 수박·성주참외·신비복숭아 등 토스 주문. 회신의 수령인명으로 매칭.
    결제완료(PAID) 주문은 register_tracking이 상품준비중으로 자동 전환 후 등록한다.
    """
    reply_entries = parse_reply_file(tomato_reply_bytes)
    if not reply_entries:
        return {
            "toss_orders": 0,
            "toss_success": 0,
            "toss_fail": 0,
            "toss_skip": 0,
            "toss_error": "거래처 회신 파일에서 운송장번호를 찾을 수 없습니다.",
        }

    now = datetime.now(KST)
    from_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    to_date = now.strftime("%Y-%m-%d")

    toss_orders_raw = await toss_client.get_orders(
        start_date=from_date,
        end_date=to_date,
        status=None,
    )

    toss_orders = []
    skip_count = 0
    seen_order_product_ids = set()

    from app.processors.myeongi_order import jewelry_external_product

    for item in toss_orders_raw:
        text = _toss_item_text(item)
        # 우리 취급 토스 주문에 송장 등록 (회신 수령인명으로 매칭).
        # 쥬얼리(수박·참외·복숭아·감자) + 제주다팜(밤호박·콜라비·초당옥수수)까지 포함해야
        # 제주다팜 페이지에서도 토스 주문(밤호박 1kg 등)이 끌려와 운송장이 등록된다.
        if not any(k in text for k in ("수박", "참외", "복숭아", "감자", "밤호박", "콜라비", "옥수수")):
            continue

        option = str(item.get("optionName") or "")
        # 매칭 보조용 라벨 (복숭아 3·4kg 등 jewelry_external_product가 None이면 상품명 사용)
        label = jewelry_external_product(text, option, JBT_WATERMELON_KGS) or normalize(item.get("productName") or "")

        order_product_id = item.get("orderProductId")
        if not order_product_id or order_product_id in seen_order_product_ids:
            continue
        seen_order_product_ids.add(order_product_id)

        existing_tracking = normalize(item.get("shippingTrackingNumber") or item.get("trackingNumber") or "")
        order_status = str(item.get("orderProductStatus") or item.get("status") or item.get("orderStatus") or "").strip().upper()
        order_id = _toss_order_id(item)
        receiver_name = item.get("receiverName") or ""
        receiver_phone = item.get("receiverRealPhone") or item.get("receiverPhone") or ""
        address = item.get("address") or item.get("receiverAddress1") or ""
        detail_address = item.get("detailAddress") or item.get("receiverAddress2") or ""
        full_address = f"{address} {detail_address}".strip()

        if existing_tracking:
            skip_count += 1
            continue
        if order_status and order_status not in TRACKABLE_TOSS_STATUSES:
            skip_count += 1
            continue

        toss_orders.append({
            "order_product_id": order_product_id,
            "order_id": order_id,
            "name": normalize(receiver_name),
            "phone": normalize(receiver_phone),
            "address": normalize(full_address),
            "name_display": receiver_name,
            "option_keys": option_key_set(
                option,
                label,
                item.get("productName"),
            ),
        })

    entry_by_name = defaultdict(list)
    entry_by_phone = defaultdict(list)
    entry_by_address = defaultdict(list)
    for entry in reply_entries:
        entry_by_name[entry["name"]].append(entry)
        if entry["phone"]:
            entry_by_phone[entry["phone"]].append(entry)
        if entry["address"]:
            entry_by_address[entry["address"]].append(entry)

    used_entries = set()
    matched_pairs = []
    toss_name_counts = name_counts(order["name"] for order in toss_orders)

    for order in toss_orders:
        candidates = entry_by_name.get(order["name"], [])
        if not candidates and order["phone"]:
            candidates = entry_by_phone.get(order["phone"], [])
        if not candidates and order["address"]:
            candidates = entry_by_address.get(order["address"], [])
        available = [entry for entry in candidates if id(entry) not in used_entries]
        if not available:
            skip_count += 1
            continue

        if requires_option_guard(order["name"], toss_name_counts, len(candidates)):
            available = [
                entry for entry in available
                if options_match(entry.get("option_keys"), order.get("option_keys"))
            ]
            if not available:
                skip_count += 1
                continue

        matched = None
        for entry in available:
            if entry["phone"] == order["phone"] and entry["address"] == order["address"]:
                matched = entry
                break
        if matched is None:
            for entry in available:
                if entry["phone"] == order["phone"]:
                    matched = entry
                    break
        if matched is None:
            for entry in available:
                if entry["address"] == order["address"]:
                    matched = entry
                    break
        if matched is None:
            for entry in available:
                if options_match(entry.get("option_keys"), order.get("option_keys")):
                    matched = entry
                    break
        if matched is None:
            matched = available[0]

        used_entries.add(id(matched))
        matched_pairs.append((order, matched["tracking"], matched.get("courier") or ""))

    success_count = 0
    fail_count = 0
    courier_counts: dict[str, int] = {}
    courier_warnings: list[str] = []

    from app.processors.toss_auto import normalize_toss_courier

    for order, tracking, courier in matched_pairs:
        # 회신 택배사명을 토스 정식명으로 정규화('한진'→'한진택배' 등).
        # 이 함수의 대상(쥬얼리프룻·제주다팜 발주 품목)은 전부 롯데택배 발송이므로
        # 회신에 택배사가 비면 기본값은 롯데택배. (과거 제이비티 시절 한진 기본값 탓에
        # 수박 건이 한진으로 등록돼 배송완료가 추적 안 되던 2026-06-15 사고 재발 방지)
        delivery_company = normalize_toss_courier(courier, default="롯데택배")
        if delivery_company != "롯데택배":
            # 쥬얼리·제주다팜 품목이 롯데가 아닌 택배사로 등록되는 건 이상 신호 → 눈에 보이게
            courier_warnings.append(
                f"{order.get('name_display') or order.get('name')} → {delivery_company} {tracking} (회신 택배사 '{courier}')"
            )
        try:
            await toss_client.register_tracking(
                order_product_id=order["order_product_id"],
                delivery_company=delivery_company,
                tracking_number=tracking,
            )
            success_count += 1
            courier_counts[delivery_company] = courier_counts.get(delivery_company, 0) + 1
        except Exception as exc:
            fail_count += 1
            logger.error(
                f"토스 {JBT_WATERMELON_LABEL} 운송장 등록 실패 "
                "(orderProductId=%s, orderId=%s): %s",
                order.get("order_product_id"),
                order.get("order_id"),
                exc,
            )

    stats = {
        "toss_orders": len(toss_orders) + skip_count,
        "toss_success": success_count,
        "toss_fail": fail_count,
        "toss_skip": skip_count,
    }
    if courier_counts:
        stats["toss_couriers"] = ", ".join(f"{k} {v}건" for k, v in sorted(courier_counts.items()))
    if courier_warnings:
        stats["toss_courier_warnings"] = " / ".join(courier_warnings)
    return stats



def process(
    tomato_reply_bytes: bytes,
    delivery_bytes: bytes,
    delivery_company: str = "",
) -> tuple[bytes, str, dict]:
    """Input tracking numbers from 대저토마토 회신 파일 into DeliveryList.

    Args:
        tomato_reply_bytes: Raw bytes of the 대저토마토 회신 Excel file.
            Columns: A=송장번호, B=수령인(성함), C=수령인(연락처), D=수령인(주소), K=택배사
        delivery_bytes: Raw bytes of the DeliveryList Excel file.
        delivery_company: 택배사 이름 (미지정 시 회신 파일 K열에서 읽음)

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    tomato_entries = parse_reply_file(tomato_reply_bytes)

    # Load DeliveryList
    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    dl_ws = dl_wb.active

    # Keep only first sheet
    first_sheet_name = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet_name]

    # Build index by name
    entry_by_name = defaultdict(list)
    for entry in tomato_entries:
        entry_by_name[entry["name"]].append(entry)

    used_entries = set()
    filled = 0
    skipped = 0
    has_tomato = False
    has_chamoe = False
    has_ddureup = False
    has_watermelon = False
    has_corn = False
    has_peach = False
    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )

    for row_idx in range(2, dl_ws.max_row + 1):
        # Column E = 5 (tracking number destination)
        e_cell = dl_ws.cell(row=row_idx, column=5)

        # Only fill empty E cells
        if e_cell.value is not None and normalize(e_cell.value) != "":
            continue

        product_name = str(dl_ws.cell(row=row_idx, column=11).value or "")
        option_text = str(dl_ws.cell(row=row_idx, column=12).value or "")
        if not is_jbt_tracking_target(product_name, option_text):
            continue

        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)     # AA
        dl_phone = normalize(dl_ws.cell(row=row_idx, column=28).value)    # AB
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)  # AD
        dl_option_keys = option_key_set(
            option_text,
            transform_product(option_text, product_type_for(product_name)),
        ) | corn_option_keys(product_name, option_text)
        if not dl_option_keys:
            dl_option_keys = option_key_set(product_name)

        if not dl_name:
            continue

        candidates = entry_by_name.get(dl_name, [])
        if not candidates:
            skipped += 1
            continue

        available = [
            (i, c) for i, c in enumerate(candidates)
            if id(c) not in used_entries
        ]
        if not available:
            skipped += 1
            continue

        if requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
            available = [
                (i, c) for i, c in available
                if options_match(c.get("option_keys"), dl_option_keys)
            ]
            if not available:
                skipped += 1
                continue

        matched = None

        # Try phone + address match first
        for idx, c in available:
            if c["phone"] == dl_phone and c["address"] == dl_address:
                matched = c
                break

        # Try phone only match
        if matched is None:
            for idx, c in available:
                if c["phone"] == dl_phone:
                    matched = c
                    break

        # Try address only match
        if matched is None:
            for idx, c in available:
                if c["address"] == dl_address:
                    matched = c
                    break

        # Fall back to first available
        if matched is None:
            matched = available[0][1]

        if matched:
            e_cell.value = matched["tracking"]
            # D열(4) = 택배사: 회신 파일 K열 값을 무조건 DeliveryList D열에 덮어쓰기
            courier_val = matched.get("courier") or delivery_company
            dl_ws.cell(row=row_idx, column=4).value = normalize_courier(courier_val) if courier_val else ""
            used_entries.add(id(matched))
            filled += 1

            # K열(11) = 상품명 → 제이비티 발주 상품 판별
            if "토마토" in product_name:
                has_tomato = True
            if "참외" in product_name:
                has_chamoe = True
            if "땅두릅" in product_name:
                has_ddureup = True
            if "수박" in product_name:
                has_watermelon = True
            if "옥수수" in product_name:
                has_corn = True
            if "복숭아" in product_name:
                has_peach = True
        else:
            skipped += 1

    # Save to bytes
    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    # 파일명: 실제 처리된 상품에 따라 동적 생성
    now = datetime.now(KST)
    parts = []
    if has_tomato:
        parts.append("대저토마토")
    if has_chamoe:
        parts.append("성주참외")
    if has_ddureup:
        parts.append("남해땅두릅")
    if has_watermelon:
        parts.append("수박")
    if has_corn:
        parts.append("초당옥수수")
    if has_peach:
        parts.append("신비복숭아")
    product_label = "_".join(parts) if parts else "발주"
    filename = f"송장파일_{product_label}({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "tomato_entries": len(tomato_entries),
    }

    return output.read(), filename, stats


def _phone_digits(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def process_alwayz_tracking(
    tomato_reply_bytes: bytes,
    alwayz_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """제이비티 회신 파일의 운송장번호를 올웨이즈 주문내역 파일에 입력.

    회신: A=송장번호, B=수령인, C=연락처, D=주소, K=택배사 (parse_reply_file)
    올웨이즈: O(15)=주소, S(19)=수령인, T(20)=연락처,
             U(21)=택배사 ←입력, W(23)=운송장번호 ←입력
    매칭: 이름 → (전화번호 숫자비교 → 주소 → 첫 후보) 순.
    """
    reply_entries = parse_reply_file(tomato_reply_bytes)
    if not reply_entries:
        raise ValueError("제이비티 회신 파일에서 운송장번호를 찾을 수 없습니다.")

    entry_by_name = defaultdict(list)
    entry_by_phone = defaultdict(list)
    for entry in reply_entries:
        if entry["name"]:
            entry_by_name[entry["name"]].append(entry)
        digits = _phone_digits(entry.get("phone"))
        if digits:
            entry_by_phone[digits].append(entry)

    al_wb = load_workbook(filename=BytesIO(alwayz_bytes))
    al_ws = al_wb.active

    used: set[int] = set()
    filled = 0
    skipped = 0

    for row_idx in range(2, al_ws.max_row + 1):
        w_cell = al_ws.cell(row=row_idx, column=23)   # W = 운송장번호
        if w_cell.value is not None and normalize(w_cell.value) != "":
            continue

        al_name = normalize(al_ws.cell(row=row_idx, column=19).value)   # S
        if not al_name:
            continue
        al_phone = _phone_digits(al_ws.cell(row=row_idx, column=20).value)  # T
        al_address = normalize(al_ws.cell(row=row_idx, column=15).value)    # O

        candidates = [c for c in entry_by_name.get(al_name, []) if id(c) not in used]

        # 이름 매칭 실패 시 전화번호 fallback (CS 수취인 변경 대응)
        if not candidates and al_phone:
            candidates = [c for c in entry_by_phone.get(al_phone, []) if id(c) not in used]

        if not candidates:
            skipped += 1
            continue

        matched = None
        for c in candidates:
            if al_phone and _phone_digits(c.get("phone")) == al_phone and c.get("address") == al_address:
                matched = c
                break
        if matched is None:
            for c in candidates:
                if al_phone and _phone_digits(c.get("phone")) == al_phone:
                    matched = c
                    break
        if matched is None:
            for c in candidates:
                if al_address and c.get("address") == al_address:
                    matched = c
                    break
        if matched is None:
            matched = candidates[0]

        w_cell.value = matched["tracking"]
        courier = matched.get("courier") or ""
        al_ws.cell(row=row_idx, column=21).value = normalize_courier(courier) if courier else "롯데택배"  # U
        used.add(id(matched))
        filled += 1

    output = BytesIO()
    al_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"올웨이즈_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "reply_entries": len(reply_entries),
    }
    return output.read(), filename, stats


def process_toss_excel_tracking(
    tomato_reply_bytes: bytes,
    toss_export_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """거래처 회신의 운송장번호를 토스 '엑셀 일괄발송' 파일에 채워서 반환.

    토스 주문배송관리 다운로드 파일 구조:
      - 헤더행: A열='주문번호', 그 외 '송장번호','택배사','수령인명','수령인 연락처','배송지'
      - 헤더 다음 '수정 가능/불가' 라벨행 → 건너뜀
      - 이후 주문 데이터행
    회신(제이비티/쥬얼리 orderlist)은 parse_reply_file로 자동 감지.
    매칭: 수령인명 → 연락처(숫자) → 첫 후보. 이미 송장 있으면 스킵.
    채운 파일을 토스 셀러센터 '엑셀 일괄발송'에 그대로 업로드하면 된다.
    """
    reply_entries = parse_reply_file(tomato_reply_bytes)
    if not reply_entries:
        raise ValueError("거래처 회신 파일에서 운송장번호를 찾을 수 없습니다.")

    entry_by_name = defaultdict(list)
    entry_by_phone = defaultdict(list)
    for entry in reply_entries:
        if entry["name"]:
            entry_by_name[entry["name"]].append(entry)
        digits = _phone_digits(entry.get("phone"))
        if digits:
            entry_by_phone[digits].append(entry)

    wb = load_workbook(filename=BytesIO(toss_export_bytes))
    ws = wb.active

    # 헤더행 탐색 (A열 == '주문번호')
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        if normalize(ws.cell(row=r, column=1).value) == "주문번호":
            header_row = r
            break
    if header_row is None:
        raise ValueError(
            "토스 주문 엑셀에서 헤더(주문번호 행)를 찾지 못했습니다. "
            "토스 셀러센터 '주문배송관리'에서 받은 엑셀인지 확인해주세요."
        )

    col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = normalize(ws.cell(row=header_row, column=c).value)
        if h and h not in col:
            col[h] = c

    def _need(name: str) -> int:
        if name not in col:
            raise ValueError(f"토스 주문 엑셀에 '{name}' 열이 없습니다.")
        return col[name]

    c_tracking = _need("송장번호")
    c_name = _need("수령인명")
    c_courier = col.get("택배사")
    c_phone = col.get("수령인 연락처")

    # 데이터 시작행: 헤더 다음의 '수정 가능/불가' 라벨행은 건너뜀
    data_start = header_row + 1
    label = normalize(ws.cell(row=data_start, column=1).value).replace(" ", "")
    if label in ("수정가능", "수정불가"):
        data_start += 1

    used: set[int] = set()
    filled = 0
    skipped = 0

    for r in range(data_start, ws.max_row + 1):
        name = normalize(ws.cell(row=r, column=c_name).value)
        if not name:
            continue
        if normalize(ws.cell(row=r, column=c_tracking).value):
            skipped += 1  # 이미 송장 있음
            continue
        phone = _phone_digits(ws.cell(row=r, column=c_phone).value) if c_phone else ""

        candidates = [c for c in entry_by_name.get(name, []) if id(c) not in used]
        if not candidates and phone:
            candidates = [c for c in entry_by_phone.get(phone, []) if id(c) not in used]
        if not candidates:
            skipped += 1
            continue

        matched = None
        if phone:
            for c in candidates:
                if _phone_digits(c.get("phone")) == phone:
                    matched = c
                    break
        if matched is None:
            matched = candidates[0]

        ws.cell(row=r, column=c_tracking).value = matched["tracking"]
        if c_courier:
            courier = matched.get("courier") or ""
            ws.cell(row=r, column=c_courier).value = normalize_courier(courier) if courier else "롯데택배"
        used.add(id(matched))
        filled += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"토스_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "reply_entries": len(reply_entries),
    }
    return output.read(), filename, stats
