import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from hashlib import sha1
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color

KST = timezone(timedelta(hours=9))
JBT_SENDER_NAME = "(주)아이티소프트"
TOSS_WATERMELON_EXCLUDED_STATUS_PATTERNS = ("CANCEL", "RETURN", "REFUND", "EXCHANGE")
TOSS_SHIPPED_STATUS_PATTERNS = ("DELIVER", "SHIP")  # 배송중/배송완료 → 발주 제외
JBT_WATERMELON_KGS = {6, 7, 8}
JBT_WATERMELON_LABEL = "수박 6/7/8kg"


@dataclass
class _VirtualCell:
    value: object = ""


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def delivery_watermelon_kg(product_name: str = "", option_text: str = "") -> int | None:
    text = normalize(f"{product_name} {option_text}")
    # 망고수박은 쥬얼리프룻(myeongi_order) 발주 — 일반수박(제이비티) 로직에서 제외
    if "수박" not in text or "망고" in text:
        return None
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*kg", text, re.IGNORECASE):
        try:
            kg = float(match.group(1))
        except ValueError:
            continue
        if kg.is_integer():
            return int(kg)
    return None


# 신비복숭아 발주처 분리 (2026-06): 1·2kg → 쥬얼리프룻, 3·4kg → 제이비티, 800g 등 그 외 → 제외
JBT_PEACH_KGS = {3, 4}
JEWELRY_PEACH_KGS = {1, 2}


def peach_kg(product_name: str = "", option_text: str = "") -> float | int | None:
    """신비복숭아 주문의 무게(kg)를 반환. 복숭아 아니거나 무게 없으면 None."""
    text = normalize(f"{product_name} {option_text}")
    if "복숭아" not in text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*(kg|g)", text, re.IGNORECASE)
    if not m:
        return None
    amount = float(m.group(1))
    kg = amount / 1000 if m.group(2).lower() == "g" else amount
    return int(kg) if kg.is_integer() else kg


def is_jbt_shinbi_peach(product_name: str = "", option_text: str = "") -> bool:
    """제이비티 발주 대상 '신비복숭아 3·4kg'인지. 백도딱딱이·거반도·대극천 등
    '복숭아'가 들어간 비신비 상품(쥬얼리 발주)은 kg 무관 제외 — 백도 4kg이
    신비복숭아로 둔갑하는 오분류 방지."""
    from app.processors.myeongi_order import _non_shinbi_peach
    if _non_shinbi_peach(product_name, option_text):
        return False
    return peach_kg(product_name, option_text) in JBT_PEACH_KGS


def stable_order_id(prefix: str, *values) -> str:
    raw = "|".join(normalize(value) for value in values if normalize(value))
    if not raw:
        return ""
    return f"{prefix}:{sha1(raw.encode('utf-8')).hexdigest()[:16]}"


def toss_order_id(item: dict) -> str:
    for key in ("orderNo", "orderId", "orderNumber", "orderSheetId", "paymentKey", "orderProductId"):
        value = item.get(key)
        if value:
            return str(value)
    return ""


def _toss_item_text(value) -> str:
    parts: list[str] = []

    def walk(node) -> None:
        if node is None:
            return
        if isinstance(node, dict):
            for child in node.values():
                walk(child)
            return
        if isinstance(node, (list, tuple)):
            for child in node:
                walk(child)
            return
        parts.append(str(node))

    walk(value)
    return normalize(" ".join(parts))


def is_toss_watermelon_order(item: dict) -> bool:
    text = _toss_item_text(item)
    # 망고수박은 제이비티 일반수박 발주 대상이 아니다
    return "수박" in text and "망고" not in text


def is_toss_peach_order(item: dict) -> bool:
    return "복숭아" in _toss_item_text(item)


def _toss_peach_option(item: dict) -> str:
    """토스 신비복숭아 옵션 → 무게 표기 추출(800g/1kg 등). transform_product가 등급을 결정."""
    option = normalize(item.get("optionName") or "")
    if option:
        return option
    for source in (normalize(item.get("productName") or ""), _toss_item_text(item)):
        m = re.search(r"(\d+\.?\d*)\s*(g|kg)", source, re.IGNORECASE)
        if m:
            return f"{m.group(1)}{m.group(2)}"
    return normalize(item.get("productName") or "")


def _toss_address(item: dict) -> str:
    address = item.get("address") or item.get("receiverAddress1") or ""
    detail = item.get("detailAddress") or item.get("receiverAddress2") or ""
    return normalize(f"{address} {detail}".strip())


def _toss_zipcode(item: dict) -> str:
    zipcode = normalize(item.get("zipCode") or item.get("receiverZipCode") or "")
    if not zipcode:
        return ""
    try:
        return str(int(float(zipcode))).zfill(5)
    except (TypeError, ValueError):
        return zipcode.zfill(5)


def _toss_watermelon_option(item: dict) -> str:
    option = normalize(item.get("optionName") or "")
    if option:
        return option
    product_name = normalize(item.get("productName") or "")
    kg_match = re.search(r"(\d+\.?\d*)\s*kg", product_name, re.IGNORECASE)
    if kg_match:
        return f"{kg_match.group(1)}kg"
    kg_match = re.search(r"(\d+\.?\d*)\s*kg", _toss_item_text(item), re.IGNORECASE)
    if kg_match:
        return f"{kg_match.group(1)}kg"
    return product_name


def _virtual_toss_row(entry: dict) -> list[_VirtualCell]:
    row = [_VirtualCell() for _ in range(31)]
    row[2].value = entry.get("order_id", "")
    row[10].value = entry.get("product_name", "토스 수박")
    row[11].value = entry.get("option", "")
    row[22].value = entry.get("qty", "1")
    row[26].value = entry.get("name", "")
    row[27].value = entry.get("phone", "")
    row[28].value = entry.get("zipcode", "")
    row[29].value = entry.get("address", "")
    row[30].value = entry.get("memo", "")
    return row


def _toss_order_done(item: dict) -> bool:
    """이미 발송된(배송중/배송완료) 또는 송장이 입력된 토스 주문인지 → 발주 제외 대상."""
    status = normalize(item.get("orderProductStatus") or item.get("status") or item.get("orderStatus") or "")
    su = status.upper()
    if status and (any(p in su for p in TOSS_SHIPPED_STATUS_PATTERNS) or "배송" in status):
        return True
    return bool(normalize(item.get("shippingTrackingNumber") or item.get("trackingNumber") or ""))


async def collect_toss_peach_orders(
    from_date: str,
    to_date: str,
    kgs: set | None = None,
) -> list[dict]:
    """토스 API에서 신비복숭아 주문을 수집.

    kgs로 무게를 거른다 (제이비티=3·4kg, 쥬얼리=1·2kg). None이면 전체.
    각 entry에 'kg'(무게)를 포함한다.
    """
    from app.toss.client import toss_client

    orders = await toss_client.get_orders(
        start_date=from_date,
        end_date=to_date,
        status=None,
    )

    entries: list[dict] = []
    for item in orders:
        if not is_toss_peach_order(item):
            continue
        order_status = normalize(item.get("orderProductStatus") or item.get("status") or item.get("orderStatus") or "")
        if order_status and any(pattern in order_status.upper() for pattern in TOSS_WATERMELON_EXCLUDED_STATUS_PATTERNS):
            continue
        # 배송중/배송완료거나 송장 입력된 건은 발주서에서 제외 → 과거 주문 재발주 방지
        if _toss_order_done(item):
            continue

        option = _toss_peach_option(item)
        # 백도딱딱이·거반도·대극천 등 비신비 복숭아는 쥬얼리 수집 대상 → 제이비티 토스 수집에서 제외
        from app.processors.myeongi_order import _non_shinbi_peach
        if _non_shinbi_peach(_toss_item_text(item), option):
            continue
        kg = peach_kg(_toss_item_text(item), option)
        if kgs is not None and kg not in kgs:
            continue
        address = _toss_address(item)
        entries.append(
            {
                "name": item.get("receiverName") or "",
                "phone": item.get("receiverRealPhone") or item.get("receiverPhone") or "",
                "zipcode": _toss_zipcode(item),
                "address": address,
                "qty": str(item.get("quantity") or 1),
                "product_name": item.get("productName") or "토스 신비복숭아",
                "option": option,
                "kg": kg,
                "memo": item.get("shippingNote") or "",
                "order_id": toss_order_id(item)
                or stable_order_id(
                    "toss-peach",
                    item.get("receiverName"),
                    item.get("receiverRealPhone") or item.get("receiverPhone"),
                    address,
                    option,
                    item.get("quantity") or 1,
                ),
            }
        )

    return entries


async def collect_toss_jewelry_orders(from_date: str, to_date: str) -> list[dict]:
    """토스 API에서 쥬얼리팜 발주 대상 주문을 수집.

    대상: 수박 6/7/8kg · 성주참외 · 신비복숭아(소과/대과 등 옵션 그대로) · (블랙)망고수박.
    각 entry에 'product'(쥬얼리 발주 품목명)를 포함한다.
    """
    from app.toss.client import toss_client
    from app.processors.myeongi_order import jewelry_passthrough_product

    orders = await toss_client.get_orders(
        start_date=from_date,
        end_date=to_date,
        status=None,
    )

    entries: list[dict] = []
    for item in orders:
        text = _toss_item_text(item)
        # 수박(망고수박 포함)·성주참외·신비복숭아 토스 주문
        # (홍감자는 2026-07 쥬얼리 품절 → 제주다팜 수집 collect_toss_jejudapam_orders로 이관)
        if not any(k in text for k in ("수박", "참외", "복숭아")):
            continue
        order_status = normalize(item.get("orderProductStatus") or item.get("status") or item.get("orderStatus") or "")
        if order_status and any(pattern in order_status.upper() for pattern in TOSS_WATERMELON_EXCLUDED_STATUS_PATTERNS):
            continue
        # 배송중/배송완료거나 송장 입력된 건은 발주서에서 제외 → 과거 주문 재발주 방지
        if _toss_order_done(item):
            continue

        # 신비복숭아·망고수박은 옵션 그대로, 수박/참외는 정규화
        option = normalize(item.get("optionName") or "") or _toss_watermelon_option(item)
        # 신비복숭아 3·4kg은 제이비티 발주(토마토 메뉴)로 → 쥬얼리 수집에서 제외 (1·2kg만 쥬얼리)
        from app.processors.myeongi_order import _non_shinbi_peach
        # 신비복숭아 3·4kg만 제외(제이비티). 대극천·거반도는 크기 무관 쥬얼리 수집.
        if ("복숭아" in text and not _non_shinbi_peach(text, option)
                and peach_kg(text, option) not in JEWELRY_PEACH_KGS):
            continue
        # 백도 딱딱이복숭아 2·4kg은 제주다팜 이관 → 쥬얼리 토스 수집 제외(1kg만 쥬얼리).
        from app.processors.myeongi_order import _is_baekdo, _baekdo_kg
        if _is_baekdo(text, option) and _baekdo_kg(text, option) != "1":
            continue
        # 품목명/무게 매칭은 상품명(K)을 우선으로 본다. 옵션(L)이 비거나 잘못 적혀도
        # 상품명에서 무게·종류를 읽도록 깨끗한 productName을 넘긴다. (없을 때만 전체 text)
        product_name = normalize(item.get("productName") or "") or text
        # 방어: 토스 상품명에 품목 키워드가 누락돼도 전체 주문 텍스트(text)에 있으면 보강한다.
        # (블랙망고수박이 '수박'만 잡혀 일반수박으로 둔갑하는 실수 차단 — 무게는 그대로 추출됨)
        _compact_all = text.replace(" ", "")
        _compact_name = product_name.replace(" ", "")
        for _kw in ("블랙망고수박", "망고수박", "대극천", "거반도", "납작복숭아"):
            if _kw in _compact_all and _kw not in _compact_name:
                product_name = f"{_kw} {product_name}"
                break
        product = jewelry_passthrough_product(product_name, option, JBT_WATERMELON_KGS)
        if not product:
            continue

        address = _toss_address(item)
        entries.append(
            {
                "name": item.get("receiverName") or "",
                "phone": item.get("receiverRealPhone") or item.get("receiverPhone") or "",
                "zipcode": _toss_zipcode(item),
                "address": address,
                "qty": str(item.get("quantity") or 1),
                "product_name": item.get("productName") or "토스 주문",
                "option": option,
                "product": product,
                "memo": item.get("shippingNote") or "",
                "order_id": toss_order_id(item)
                or stable_order_id(
                    "toss-jewelry",
                    item.get("receiverName"),
                    item.get("receiverRealPhone") or item.get("receiverPhone"),
                    address,
                    option,
                    item.get("quantity") or 1,
                ),
            }
        )

    return entries


async def collect_toss_jejudapam_orders(from_date: str, to_date: str) -> dict:
    """토스 API에서 제주다팜 발주 대상(콜라비 + 미니밤호박 전 옵션 + 홍감자) 주문을 수집.

    반환: {"colrabi": [...], "bamhobak": [...], "potato": [...]} — 각 entry에
    'product'(제주다팜 발주 품목명) 포함.
    콜라비는 3/5/10kg('콜라비 정품 {kg}kg'), 미니밤호박은 전 옵션(2026-08-17 쥬얼리→제주다팜 통합),
    홍감자는 2026-07 쥬얼리 품절로 제주다팜 이관(중1→중2, 대3→특3, 대5→특5).
    배송중·송장입력 건은 자동 제외.
    """
    from app.toss.client import toss_client
    from app.processors.kolrabi_order import (
        convert_bamhobak_option,
        convert_jeju_baekdo_option,
        convert_potato_option,
        convert_quantity,
    )

    orders = await toss_client.get_orders(start_date=from_date, end_date=to_date, status=None)

    colrabi: list[dict] = []
    bamhobak: list[dict] = []
    potato: list[dict] = []
    baekdo: list[dict] = []
    for item in orders:
        text = _toss_item_text(item)
        compact = text.replace(" ", "")
        if ("콜라비" not in compact and "밤호박" not in compact
                and "홍감자" not in compact and "백도" not in compact):
            continue
        order_status = normalize(item.get("orderProductStatus") or item.get("status") or item.get("orderStatus") or "")
        if order_status and any(pattern in order_status.upper() for pattern in TOSS_WATERMELON_EXCLUDED_STATUS_PATTERNS):
            continue
        if _toss_order_done(item):
            continue

        option = normalize(item.get("optionName") or "")
        product_name = normalize(item.get("productName") or "") or text
        address = _toss_address(item)

        def _entry(product: str, tag: str) -> dict:
            return {
                "name": item.get("receiverName") or "",
                "phone": item.get("receiverRealPhone") or item.get("receiverPhone") or "",
                "zipcode": _toss_zipcode(item),
                "address": address,
                "qty": str(item.get("quantity") or 1),
                "product_name": item.get("productName") or "토스 주문",
                "option": option,
                "product": product,
                "memo": item.get("shippingNote") or "",
                "order_id": toss_order_id(item)
                or stable_order_id(
                    tag,
                    item.get("receiverName"),
                    item.get("receiverRealPhone") or item.get("receiverPhone"),
                    address,
                    option,
                    item.get("quantity") or 1,
                ),
            }

        if "콜라비" in compact:
            product = convert_quantity(f"{option} {product_name}")  # '콜라비 정품 {kg}kg' (3/5/10kg)
            if product:
                colrabi.append(_entry(product, "toss-colrabi"))
        elif "밤호박" in compact:
            product = convert_bamhobak_option(product_name, option)  # 1kg만
            if product:
                bamhobak.append(_entry(product, "toss-bamhobak"))
        elif "홍감자" in compact:
            product = convert_potato_option(product_name, option)  # 중1→중2, 대3→특3, 대5→특5
            if product:
                potato.append(_entry(product, "toss-potato"))
        elif "백도" in compact:
            product = convert_jeju_baekdo_option(product_name, option)  # 2·4kg만(1kg은 쥬얼리)
            if product:
                baekdo.append(_entry(product, "toss-baekdo"))

    return {"colrabi": colrabi, "bamhobak": bamhobak, "potato": potato, "baekdo": baekdo}


async def process_toss_watermelon_order(
    from_date: str,
    to_date: str,
    alwayz_bytes: bytes | None = None,
    collect_toss: bool = True,
) -> list[tuple[bytes, str, dict]]:
    """토스+올웨이즈 발주 — 발주처별 발주서 목록 반환.

    쥬얼리프룻(수박·성주참외·신비복숭아 1·2kg) + 제이비티(신비복숭아 3·4kg).
    myeongi_order로 위임한다.
    """
    from app.processors.myeongi_order import process_jewelry_watermelon_order

    return await process_jewelry_watermelon_order(
        from_date, to_date, alwayz_bytes=alwayz_bytes, collect_toss=collect_toss
    )


# 청사과(아오리) — 2026-08-27 제주다팜 → 제이비티 발주 이관.
# 제이비티 판매옵션(구글시트 실측): '청사과가정용A급({등급}){kg}kg({과수})'.
# 쿠팡 등급(소과/중소과/대과) → 제이비티 등급(소과/중소과/중과) 매핑:
#   제이비티엔 '대과'가 없고 가장 큰 등급이 '중과'라 대과 주문은 중과로 발주한다.
_JBT_APPLE_OPTIONS = {
    ("소과", "1"): "청사과가정용A급(소과)1kg(6-8과)",
    ("소과", "1.5"): "청사과가정용A급(소과)1.5kg(10-12과)",
    ("소과", "2"): "청사과가정용A급(소과)2kg(11-15과)",
    ("소과", "3"): "청사과가정용A급(소과)3kg(17-21과)",
    ("소과", "4"): "청사과가정용A급(소과)4kg(23-28과)",
    ("소과", "5"): "청사과가정용A급(소과)5kg(28-33과)",
    ("중소과", "1"): "청사과가정용A급(중소과)1kg(5-6과)",
    ("중소과", "1.5"): "청사과가정용A급(중소과)1.5kg(8-9과)",
    ("중소과", "2"): "청사과가정용A급(중소과)2kg(9-10과)",
    ("중소과", "3"): "청사과가정용A급(중소과)3kg(14-16과)",
    ("중소과", "4"): "청사과가정용A급(중소과)4kg(19-22과)",
    ("중소과", "5"): "청사과가정용A급(중소과)5kg(23-28과)",
    ("중과", "1.5"): "청사과가정용A급(중과)1.5kg(6-7과)",
    ("중과", "2"): "청사과가정용A급(중과)2kg(7-8과)",
    ("중과", "3"): "청사과가정용A급(중과)3kg(11-13과)",
    ("중과", "4"): "청사과가정용A급(중과)4kg(15-18과)",
    ("중과", "5"): "청사과가정용A급(중과)5kg(18-22과)",
}
# 등급 판정 순서: '중소과'가 '소과'를 포함하므로 긴 것 먼저. 쿠팡 '대과' → 제이비티 '중과'.
_JBT_APPLE_GRADES = (("중소과", "중소과"), ("소과", "소과"), ("대과", "중과"), ("중과", "중과"))


def is_jbt_apple_order(product_name: object = "", option_text: object = "") -> bool:
    """청사과(아오리) 제이비티 발주 여부. 홍로사과(가을햇사과)와는 배타."""
    text = re.sub(r"\s+", "", f"{product_name or ''} {option_text or ''}")
    if "홍로" in text or "홍사과" in text or "가을햇사과" in text:
        return False
    return "청사과" in text or "아오리" in text


def _jbt_apple_name_from_text(text: str) -> str | None:
    """등급·kg가 든 텍스트(쿠팡 옵션 등) → 제이비티 청사과 발주명."""
    compact = re.sub(r"\s+", "", text or "")
    grade = next((jbt for cou, jbt in _JBT_APPLE_GRADES if cou in compact), "")
    m = re.search(r"(\d+(?:\.\d+)?)kg", compact, re.IGNORECASE)
    if not m or not grade:
        return None
    kg = m.group(1)
    try:
        f = float(kg)
        kg = str(int(f)) if f.is_integer() else str(f)
    except ValueError:
        pass
    return _JBT_APPLE_OPTIONS.get((grade, kg))


def jbt_apple_option(product_name: object = "", option_text: object = "") -> str | None:
    """청사과 DeliveryList → 제이비티 발주명. 취급하지 않는 조합이면 None."""
    if not is_jbt_apple_order(product_name, option_text):
        return None
    return _jbt_apple_name_from_text(f"{product_name or ''} {option_text or ''}")


def transform_product(option_text: str, product_type: str = "") -> str:
    text = normalize(option_text)

    if product_type == "청사과":
        # 발주서 행 생성은 옵션(L열)만 넘어오므로 옵션에서 등급·kg를 읽어 발주명을 만든다.
        # 제이비티 청사과는 판매옵션명 자체가 발주명이라 _vip 접미사를 붙이지 않는다.
        return _jbt_apple_name_from_text(text) or text

    if product_type == "옥수수":
        grade = "중품" if "중품" in text else "특품" if "특품" in text else ""
        count_match = re.search(r"(\d+)\s*(?:개|입)", text)
        if grade and count_match:
            result = f"초당옥수수 {grade} {count_match.group(1)}개"
        else:
            result = text
        return result if result.endswith("_vip") else f"{result}_vip"

    if product_type == "수박":
        kg_match = re.search(r"(\d+\.?\d*)\s*kg", text, re.IGNORECASE)
        if kg_match:
            result = f"하우스수박(상품){kg_match.group(1)}kg이상"
        else:
            result = f"하우스수박(상품){text}".strip()
        return result if result.endswith("_vip") else f"{result}_vip"

    if product_type == "땅두릅":
        weight_match = re.search(r"(\d+\.?\d*)\s*(g|kg)", text, re.IGNORECASE)
        if weight_match:
            amount = weight_match.group(1)
            unit = weight_match.group(2).lower()
            weight_str = f"{amount}{unit}"
            if unit == "g":
                result = f"남해땅두릅(튀김용){weight_str}"
            else:
                result = f"남해땅두릅(특품){weight_str}"
        else:
            result = f"남해땅두릅{text}"
        return result if result.endswith("_vip") else f"{result}_vip"

    if product_type == "복숭아":
        # 신비복숭아: 무게로 등급 결정 (800g→혼합과, 1~4kg→중소과)
        weight_match = re.search(r"(\d+\.?\d*)\s*(g|kg)", text, re.IGNORECASE)
        if weight_match:
            amount = float(weight_match.group(1))
            unit = weight_match.group(2).lower()
            weight_kg = amount / 1000 if unit == "g" else amount
            weight_str = f"{weight_match.group(1)}{unit}"
            grade = "혼합과" if weight_kg < 1.0 else "중소과"
            result = f"신비복숭아({grade}){weight_str}"
        else:
            result = f"신비복숭아{text}"
        return result if result.endswith("_vip") else f"{result}_vip"

    kg_match = re.search(r"(\d+\.?\d*)\s*kg", text, re.IGNORECASE)
    kg_suffix = f" {kg_match.group(1)}kg" if kg_match else ""

    text = re.sub(r"^1박스\s*", "", text)

    if "가정용" in text:
        grade_match = re.search(r"(혼합과|중소과|로얄과)", text)
        grade = grade_match.group(1) if grade_match else ""
        fruit_match = re.search(r"\((\d+-\d+과)\)", text)
        fruit_suffix = f"({fruit_match.group(1)})" if fruit_match else ""
        if grade:
            kg_str = kg_suffix.strip()
            result = f"가정용참외({grade}){kg_str}{fruit_suffix}"
            return result if result.endswith("_vip") else f"{result}_vip"

    if "짭짤이" in text:
        if "로얄" in text or "S~2S" in text or "S-2S" in text:
            kg_str = kg_suffix.strip()
            result = f"대저짭짤이특품(S)과 {kg_str}"
            return result if result.endswith("_vip") else f"{result}_vip"
        return text if text.endswith("_vip") else f"{text}_vip"

    if "특품" in text:
        cleaned = re.sub(r"\d+\.?\d*\s*kg", "", text).strip()
        result = f"대저토마토{cleaned}{kg_suffix}"
        return result if result.endswith("_vip") else f"{result}_vip"

    return text if text.endswith("_vip") else f"{text}_vip"


# 제이비티 합배송 가능 품목 — 이것만 수량 N 그대로 한 행으로 발주.
# 그 외(대저토마토·신비복숭아·남해땅두릅·수박 등 합배송 불가)는 거래처 규칙상
# '수량1 고정' — 수량 N건을 N개 행(각 수량1)으로 분할해 발주한다(수량2 이상은 송장이 따로 나옴).
_JBT_COMBINABLE = ("봉지곶감", "꽃게액젓", "손질꽃게", "장어", "밤", "김치", "꿀오일", "쉐이크선식", "선식", "마")


def _jbt_combinable(product_name: object) -> bool:
    """제이비티 합배송 가능 품목이면 True(수량 유지). 그 외는 False(수량1로 분할)."""
    text = re.sub(r"\s+", "", str(product_name or ""))
    if "토마토" in text:  # '마' 오탐 방지 — 대저토마토는 합배송 불가
        return False
    return any(kw in text for kw in _JBT_COMBINABLE)


def build_main_order_workbook(
    filtered_rows: list[tuple],
    has_tomato: bool,
    has_chamoe: bool,
    has_ddureup: bool,
    has_watermelon: bool,
    has_corn: bool = False,
    has_peach: bool = False,
    has_apple: bool = False,
) -> tuple[bytes, str, dict]:
    wb = Workbook()
    ws = wb.active

    headers = [
        "송장번호",
        "수령인(성함)",
        "수령인(연락처)",
        "수령인(주소)",
        "품목명",
        "수량",
        "배송메세지",
        "보내는분 상호",
        "보내는분 연락처\n(해당업체 연락처)",
        "주문번호",
    ]

    col_widths = {
        "A": 13,
        "B": 12.375,
        "C": 15.5,
        "D": 45,
        "E": 36.75,
        "F": 13,
        "G": 26,
        "H": 15.125,
        "I": 13.375,
        "J": 16.125,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    fill_theme6 = PatternFill(patternType="solid", fgColor=Color(theme=6, tint=0.4))
    fill_theme9 = PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.4))
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="맑은 고딕", size=11, bold=True)
    data_font = Font(name="맑은 고딕", size=11)
    sales_groups: dict[str, dict[str, dict]] = {}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_idx in (1, 6):
            cell.fill = fill_theme6
        elif col_idx in (2, 3, 4, 5, 7, 8, 9):
            cell.fill = fill_theme9

    out_row = 1  # 헤더가 1행. 데이터는 2행부터(분할로 행 수가 filtered_rows보다 많을 수 있음)
    for row, product_type in filtered_rows:
        name = normalize(row[26].value) if len(row) > 26 else ""
        phone = normalize(row[27].value) if len(row) > 27 else ""
        address = normalize(row[29].value) if len(row) > 29 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        qty_value = row[22].value if len(row) > 22 else ""
        memo = normalize(row[30].value) if len(row) > 30 else ""
        order_no = normalize(row[2].value) if len(row) > 2 else ""

        address = address.replace("*", "")
        product = transform_product(option, product_type)
        try:
            qty_int = int(qty_value) if qty_value else 1
        except (TypeError, ValueError):
            qty_int = 1

        if product_type == "토마토":
            product_label = "대저토마토"
        elif product_type == "참외":
            product_label = "성주참외"
        elif product_type == "땅두릅":
            product_label = "남해땅두릅"
        elif product_type == "수박":
            product_label = "수박"
        elif product_type == "옥수수":
            product_label = "초당옥수수"
        elif product_type == "복숭아":
            product_label = "신비복숭아"
        elif product_type == "청사과":
            product_label = "청사과"
        else:
            product_label = "제이비티"

        option_bucket = sales_groups.setdefault(product_label, {}).setdefault(
            option,
            {
                "coupang_option_keyword": option,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        option_bucket["quantity"] += qty_int
        if order_no:
            option_bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

        # 제이비티 합배송 불가 품목은 수량1로 고정 → 수량 N건을 N개 행(각 수량1)으로 분할.
        # 합배송 가능 품목만 수량 N 그대로 한 행.
        row_qtys = [1] * qty_int if (qty_int >= 2 and not _jbt_combinable(product)) else [qty_int]

        for unit_qty in row_qtys:
            out_row += 1
            mapping = {
                1: "",
                2: name,
                3: phone,
                4: address,
                5: product,
                6: str(unit_qty),
                7: memo,
                8: JBT_SENDER_NAME,
                9: "010-5700-7756",
                10: order_no,
            }
            for col, value in mapping.items():
                cell = ws.cell(row=out_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border

    ws.auto_filter.ref = f"A1:K{out_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    parts = []
    if has_tomato:
        parts.append("대저토마토")
    if has_chamoe:
        parts.append("성주참외")
    if has_ddureup:
        parts.append("남해땅두릅")
    if has_watermelon:
        parts.append("수박")
    if has_corn:
        parts.append("초당옥수수")
    if has_peach:
        parts.append("신비복숭아")
    if has_apple:
        parts.append("청사과")
    product_label = "_".join(parts) if parts else "발주"
    # 신비복숭아 3·4kg(제이비티) 발주가 포함되면 파일명 맨 앞에 '제이비_' 접두
    prefix = "제이비_" if (has_peach or has_apple) else ""
    filename = f"{prefix}아이티소프트_{product_label}({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": max(out_row - 1, 0),  # 헤더 제외 실제 발주 행 수(수량1 분할 반영)
        "sales_groups": [
            {"product": label, "options": list(options.values())}
            for label, options in sales_groups.items()
        ],
    }
    return output.read(), filename, stats


def process_outputs(
    delivery_file_bytes: bytes,
    toss_watermelon_entries: list[dict] | None = None,
    toss_peach_entries: list[dict] | None = None,
) -> list[tuple[bytes, str, dict]]:
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active
    toss_watermelon_entries = toss_watermelon_entries or []
    toss_peach_entries = toss_peach_entries or []

    filtered_rows = []
    has_tomato = False
    has_chamoe = False
    has_ddureup = False
    has_watermelon = False
    has_corn = False
    has_peach = False
    has_apple = False

    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""

        if "대저토마토" in product_name:
            has_tomato = True
            filtered_rows.append((row, "토마토"))
        # ※ 성주참외(로얄/중소/알뜰)는 2026-06부터 쥬얼리팜 발주로 전환 — 제이비티 제외
        elif "땅두릅" in product_name:
            has_ddureup = True
            filtered_rows.append((row, "땅두릅"))
        # ※ 수박(6/7/8kg)은 쥬얼리팜(myeongi), 초당옥수수는 제주다팜(kolrabi_order)으로 발주 이관 — 제이비티 제외
        elif is_jbt_apple_order(product_name, option):
            # 청사과(아오리) — 2026-08-27 제주다팜 → 제이비티 이관
            if jbt_apple_option(product_name, option):
                has_apple = True
                filtered_rows.append((row, "청사과"))
        elif is_jbt_shinbi_peach(product_name, option):
            # 신비복숭아 3·4kg만 제이비티(중소과). 1·2kg은 쥬얼리프룻, 800g은 제외.
            # 백도딱딱이·거반도·대극천은 kg 무관 쥬얼리 → 여기서 걸리면 안 됨.
            has_peach = True
            filtered_rows.append((row, "복숭아"))

    for entry in toss_watermelon_entries:
        has_watermelon = True
        filtered_rows.append((_virtual_toss_row(entry), "수박"))

    for entry in toss_peach_entries:
        has_peach = True
        filtered_rows.append((_virtual_toss_row(entry), "복숭아"))

    results: list[tuple[bytes, str, dict]] = []
    if filtered_rows:
        results.append(
            build_main_order_workbook(
                filtered_rows=filtered_rows,
                has_tomato=has_tomato,
                has_chamoe=has_chamoe,
                has_ddureup=has_ddureup,
                has_watermelon=has_watermelon,
                has_corn=has_corn,
                has_peach=has_peach,
                has_apple=has_apple,
            )
        )

    return results


def process(
    delivery_file_bytes: bytes,
    toss_watermelon_entries: list[dict] | None = None,
) -> tuple[bytes, str, dict]:
    results = process_outputs(delivery_file_bytes, toss_watermelon_entries=toss_watermelon_entries)
    if not results:
        raise ValueError("제이비티 발주로 출력할 대저토마토·성주참외(중소/로얄)·남해땅두릅·수박 주문을 찾지 못했습니다.")
    return results[0]
