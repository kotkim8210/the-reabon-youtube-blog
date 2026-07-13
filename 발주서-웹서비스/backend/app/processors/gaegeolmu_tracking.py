import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.processors.tracking_match import option_key_set, options_match


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


# 운송장 줄 판정: 숫자·공백·하이픈만으로 구성 (숫자 8자리 이상은 호출부에서 확인)
_TRACKING_LINE = re.compile(r"^[\d\s\-]+$")


def parse_tracking_text(text: str) -> tuple[list[tuple[str, str]], list[str]]:
    """복붙 텍스트 → [(이름, 운송장번호)] 목록과 경고.

    형식: '이름' 줄 다음 '운송장번호' 줄 교대. 번호 숫자 사이 공백/하이픈 허용
    (예: '2552 6488 3681' → '255264883681'). '박순임 2552 6488 3681'처럼
    한 줄에 쓴 경우도 인식한다. 빈 줄·앞뒤 공백은 무시.
    """
    pairs: list[tuple[str, str]] = []
    warnings: list[str] = []
    pending_name = ""
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        digits = re.sub(r"[\s\-]", "", line)
        if _TRACKING_LINE.match(line) and len(digits) >= 8 and digits.isdigit():
            if pending_name:
                pairs.append((normalize(pending_name), digits))
                pending_name = ""
            else:
                warnings.append(f"이름 없는 운송장: {digits}")
            continue
        # 한 줄 형식: 뒤쪽의 숫자군(공백/하이픈 허용, 8자리 이상)을 운송장으로 분리
        m = re.match(r"^(.*?)[\s:]+(\d[\d\s\-]{6,}\d)$", line)
        if m:
            one_line_digits = re.sub(r"[\s\-]", "", m.group(2))
            if len(one_line_digits) >= 8 and one_line_digits.isdigit() and m.group(1).strip():
                if pending_name:
                    warnings.append(f"운송장 없는 이름: {pending_name}")
                pairs.append((normalize(m.group(1)), one_line_digits))
                pending_name = ""
                continue
        if pending_name:
            warnings.append(f"운송장 없는 이름: {pending_name}")
        pending_name = line
    if pending_name:
        warnings.append(f"운송장 없는 이름: {pending_name}")
    return pairs, warnings


def process(
    tracking_bytes: bytes | None,
    delivery_bytes: bytes,
    tracking_text: str = "",
) -> tuple[bytes, str, dict]:
    """게걸무 택배발송 파일(또는 복붙 텍스트)의 운송장번호를 DeliveryList에 매핑.

    택배발송 파일: B열=운송장번호, C열=수령인이름
    tracking_text: '이름 줄 + 운송장 줄' 교대 텍스트 (파일 대신/추가로 사용 가능)
    DeliveryList: AA열(col27)=수령인이름, AD열(col30)=수령지주소, E열(col5)=운송장번호 입력 대상

    동명이인 처리:
    - DeliveryList에 같은 이름 여러 행 + AD열 주소도 모두 동일 → 모든 행에 순차적으로 운송장 입력
    - 주소가 다르면 → 수동 처리 경고
    """
    # 같은 이름이 게걸무 파일에 여러 번 나올 수도 있음 → 옵션/주소와 함께 리스트로 저장
    src_data: dict[str, list[dict]] = defaultdict(list)
    text_warnings: list[str] = []

    if tracking_bytes:
        src_wb = load_workbook(filename=BytesIO(tracking_bytes), data_only=True)
        src_ws = src_wb.active
        for row in src_ws.iter_rows(min_row=2):
            tracking = normalize(row[1].value) if len(row) > 1 else ""  # B열
            name = normalize(row[2].value) if len(row) > 2 else ""       # C열
            if name and tracking:
                src_data[name].append(
                    {
                        "tracking": tracking,
                        "address": normalize(row[29].value) if len(row) > 29 else "",
                        "option_keys": option_key_set(row[11].value if len(row) > 11 else "")
                        or option_key_set(row[10].value if len(row) > 10 else ""),
                    }
                )

    if tracking_text and tracking_text.strip():
        pairs, text_warnings = parse_tracking_text(tracking_text)
        for name, tracking in pairs:
            # 텍스트 입력엔 주소/옵션 정보가 없음 → 이름 매칭 + 순차 입력
            src_data[name].append({"tracking": tracking, "address": "", "option_keys": set()})

    if not src_data:
        raise ValueError("게걸무 택배발송 파일을 올리거나, 이름/운송장 텍스트를 붙여넣어 주세요.")

    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    first_sheet = dl_wb.sheetnames[0]
    for s in dl_wb.sheetnames[1:]:
        del dl_wb[s]
    dl_ws = dl_wb[first_sheet]

    # DeliveryList: 이름 → [{row, 주소, 옵션키}, ...]
    dl_name_entries: dict[str, list[dict]] = defaultdict(list)
    for row_idx in range(2, dl_ws.max_row + 1):
        name = normalize(dl_ws.cell(row=row_idx, column=27).value)
        address = normalize(dl_ws.cell(row=row_idx, column=30).value)
        if name:
            dl_name_entries[name].append(
                {
                    "row": row_idx,
                    "address": address,
                    "option_keys": option_key_set(dl_ws.cell(row=row_idx, column=12).value)
                    or option_key_set(dl_ws.cell(row=row_idx, column=11).value),
                }
            )

    filled = 0
    skipped = 0
    warnings: list[str] = []

    used_rows: set[int] = set()

    for src_name, tracking_entries in src_data.items():
        if src_name not in dl_name_entries:
            warnings.append(f"미매칭: '{src_name}' - DeliveryList에 없음")
            skipped += 1
            continue

        dl_entries = dl_name_entries[src_name]
        duplicate_guard = len(dl_entries) > 1 or len(tracking_entries) > 1

        for tracking_entry in tracking_entries:
            candidates = [entry for entry in dl_entries if entry["row"] not in used_rows]
            # 텍스트 입력은 옵션 정보가 없음(option_keys 빈 set) → 옵션 필터를 건너뛰고
            # 순차 입력(원 설계 의도). 옵션 정보가 있는 파일 입력만 옵션으로 동명이인 구분.
            if duplicate_guard and tracking_entry.get("option_keys"):
                candidates = [
                    entry for entry in candidates
                    if options_match(tracking_entry.get("option_keys"), entry.get("option_keys"))
                ]
                if not candidates:
                    warnings.append(f"동명이인 옵션 불일치: '{src_name}' - 수동 처리 필요")
                    skipped += 1
                    continue

            if tracking_entry.get("address"):
                address_matches = [
                    entry for entry in candidates
                    if entry["address"] == tracking_entry["address"]
                ]
                if address_matches:
                    candidates = address_matches

            if not candidates:
                skipped += 1
                continue

            target_row = candidates[0]["row"]
            e_cell = dl_ws.cell(row=target_row, column=5)
            if e_cell.value is None or normalize(e_cell.value) == "":
                dl_ws.cell(row=target_row, column=4).value = "롯데택배"
                e_cell.value = tracking_entry["tracking"]
                used_rows.add(target_row)
                filled += 1
            else:
                skipped += 1

    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_게걸무_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats: dict = {"filled": filled, "skipped": skipped}
    all_warnings = text_warnings + warnings
    if all_warnings:
        stats["warnings"] = all_warnings

    return output.read(), filename, stats
