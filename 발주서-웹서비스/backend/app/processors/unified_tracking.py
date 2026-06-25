from __future__ import annotations

import re
from collections.abc import Callable
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.processors import (
    chamdureup_tracking,
    danharu_tracking,
    gaegeolmu_tracking,
    generic_tracking,
    myeongi_tracking,
    tomato_tracking,
    tracking_input,
)

KST = timezone(timedelta(hours=9))

Processor = Callable[[bytes, bytes], tuple[bytes, str, dict]]


def _norm(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip()).lower()


def _workbook_text(file_name: str, file_bytes: bytes, max_rows: int = 15, max_cols: int = 25) -> str:
    parts = [_norm(file_name)]
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            parts.extend(_norm(value) for value in row if value not in (None, ""))
    except Exception:
        # Detection should never block the actual processor from reporting a clearer error.
        pass
    return " ".join(part for part in parts if part)


def _processor_candidates(file_name: str, file_bytes: bytes) -> list[tuple[str, Processor]]:
    text = _workbook_text(file_name, file_bytes)

    if "게걸무" in text:
        return [("게걸무씨앗기름", gaegeolmu_tracking.process)]

    # ※ 수박·성주참외 회신은 2026-06부터 쥬얼리(orderlist) 양식 → fallback 체인에서 myeongi_tracking이 처리
    if any(token in text for token in ("대저토마토", "남해땅두릅", "옥수수", "아이티소프트")) and "회신" in text:
        return [("대저토마토/남해땅두릅/초당옥수수", tomato_tracking.process)]

    if "애플초당옥수수" in text or ("애플" in text and "초당옥수수" in text):
        return [("애플초당옥수수", myeongi_tracking.process)]

    if "명이나물" in text or "대명이" in text:
        return [("명이나물/애플초당옥수수", myeongi_tracking.process)]

    if "참두릅" in text:
        return [("참두릅", chamdureup_tracking.process)]

    if "콜라비" in text or "알뜰참외" in text or "제주다팜" in text:
        return [("콜라비+성주참외 알뜰과", tracking_input.process)]

    if "단하루" in text or "단하" in text:
        return [("단하루", danharu_tracking.process)]

    # Unknown files are tried from the most structure-specific processors to the
    # most generic one. A result is accepted only when it actually fills rows.
    return [
        ("대저토마토/성주참외/남해땅두릅/수박/초당옥수수", tomato_tracking.process),
        ("명이나물/애플초당옥수수", myeongi_tracking.process),
        ("참두릅", chamdureup_tracking.process),
        ("콜라비+성주참외 알뜰과", tracking_input.process),
        ("단하루", danharu_tracking.process),
        ("게걸무씨앗기름", gaegeolmu_tracking.process),
        ("맞춤 범용 송장", generic_tracking.process),
    ]


def _filled_count(stats: dict | None) -> int:
    if not stats:
        return 0
    value = stats.get("filled", stats.get("success", 0))
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _skipped_count(stats: dict | None) -> int:
    if not stats:
        return 0
    value = stats.get("skipped", stats.get("skip", 0))
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _cell_key(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _protect_unmatched_couriers(before_bytes: bytes, after_bytes: bytes) -> tuple[bytes, int]:
    """Keep courier cells unchanged unless the same row's tracking changed too.

    여러 회신 파일을 한 DeliveryList에 순차 적용할 때, 개별 프로세서가
    D열(택배사)을 건드리더라도 E열(운송장번호)이 실제로 입력/변경된 행만
    택배사 변경을 인정한다. 매칭되지 않은 행의 기존 택배사는 그대로 둔다.
    """
    before_wb = load_workbook(filename=BytesIO(before_bytes), data_only=False)
    after_wb = load_workbook(filename=BytesIO(after_bytes), data_only=False)
    before_ws = before_wb[before_wb.sheetnames[0]]
    after_ws = after_wb[after_wb.sheetnames[0]]

    protected = 0
    max_row = max(before_ws.max_row, after_ws.max_row)
    for row_idx in range(2, max_row + 1):
        before_tracking = _cell_key(before_ws.cell(row=row_idx, column=5).value)
        after_tracking = _cell_key(after_ws.cell(row=row_idx, column=5).value)
        before_courier = before_ws.cell(row=row_idx, column=4).value
        after_courier = after_ws.cell(row=row_idx, column=4).value

        if _cell_key(before_courier) == _cell_key(after_courier):
            continue
        if before_tracking != after_tracking:
            continue

        after_ws.cell(row=row_idx, column=4).value = before_courier
        protected += 1

    if protected <= 0:
        return after_bytes, 0

    output = BytesIO()
    after_wb.save(output)
    output.seek(0)
    return output.read(), protected


def process(
    reply_files: list[tuple[str, bytes]],
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    current_delivery = delivery_bytes
    total_filled = 0
    total_skipped = 0
    total_protected_couriers = 0
    file_results: list[dict] = []

    for file_name, file_bytes in reply_files:
        attempts: list[str] = []
        matched = False

        for label, processor in _processor_candidates(file_name, file_bytes):
            try:
                output_bytes, _filename, stats = processor(file_bytes, current_delivery)
            except Exception as exc:
                attempts.append(f"{label}: {exc}")
                continue

            filled = _filled_count(stats)
            skipped = _skipped_count(stats)
            if filled <= 0:
                attempts.append(f"{label}: 입력 0건")
                continue

            output_bytes, protected_couriers = _protect_unmatched_couriers(
                current_delivery,
                output_bytes,
            )
            current_delivery = output_bytes
            total_filled += filled
            total_skipped += skipped
            total_protected_couriers += protected_couriers
            matched = True
            file_results.append(
                {
                    "filename": file_name,
                    "processor": label,
                    "filled": filled,
                    "skipped": skipped,
                    "protected_couriers": protected_couriers,
                }
            )
            break

        if not matched:
            total_skipped += 1
            file_results.append(
                {
                    "filename": file_name,
                    "processor": "미판별",
                    "filled": 0,
                    "skipped": 1,
                    "message": " / ".join(attempts[-3:]) if attempts else "처리 가능한 송장 파일이 아닙니다.",
                }
            )

    now = datetime.now(KST)
    filename = f"DeliveryList_통합송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "filled": total_filled,
        "skipped": total_skipped,
        "reply_files": len(reply_files),
        "processed_files": sum(1 for row in file_results if row["filled"] > 0),
        "protected_couriers": total_protected_couriers,
        "file_results": file_results,
    }
    return current_delivery, filename, stats
