import re
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR
from app.processors.tracking_match import (
    name_counts,
    normalize_courier_name,
    option_key_set,
    options_match,
    requires_option_guard,
)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    """공백 완전 제거 후 문자열 반환 (이름 매칭용)"""
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def normalize_courier(value) -> str:
    """Orderlist 택배사명을 쿠팡 DeliveryList가 인식하는 표기로 맞춘다."""
    return normalize_courier_name(value)


def _phone_digits(value: object) -> str:
    """전화번호를 숫자만으로 정규화 — '010-1234-5678'과 '01012345678'을 동일 취급."""
    return re.sub(r"\D", "", str(value or ""))


def _source_option_keys(row) -> set[str]:
    # 거래처 회신 파일마다 옵션 열이 조금씩 달라서, 이름/전화/주소/송장 열을 제외한
    # 자주 쓰는 상품/옵션 후보 열만 모아 동명이인 보호용으로 비교한다.
    values = []
    for index in (4, 5, 6, 7, 8, 11, 13, 15):
        if len(row) > index:
            values.append(row[index].value)
    return option_key_set(*values) | _semantic_option_keys(*values)


def _semantic_option_keys(*values: object) -> set[str]:
    text = normalize(" ".join(str(value) for value in values if value not in (None, "")))
    if not text:
        return set()

    weights = set(re.findall(r"(\d+)kg", text, flags=re.IGNORECASE))
    keys: set[str] = set()
    if "콜라비" in text:
        keys.update(f"kolrabi:{weight}kg" for weight in weights)
    if "참외" in text and any(token in text for token in ("알뜰", "혼합", "랜덤")):
        keys.update(f"chamoe-altteul:{weight}kg" for weight in weights)
    # 초당옥수수 동명이인 구분: orderlist는 상품명에, DeliveryList는 옵션에 등급/수량이 있어
    # 통째 문자열 비교로는 안 맞는다. 등급(특품/중품/상품/정품)+수량(N개) 의미키로 양쪽을 맞춘다.
    if "옥수수" in text:
        grade = next((g for g in ("특품", "중품", "상품", "정품") if g in text), "")
        counts = set(re.findall(r"(\d+)\s*개", text))
        keys.update(f"corn:{grade}:{count}개" for count in counts)
    if "밤호박" in text:
        keys.update(f"bamhobak:{weight}kg" for weight in weights)
    if "홍감자" in text:
        # 홍감자는 발주 매핑(중1→중2, 대3→특3, 대5→특5)으로 orderlist(발주명)와
        # DeliveryList(쿠팡 원본 옵션)의 등급·kg가 달라진다 → 양쪽 모두 발주명으로
        # 정규화한 키를 쓰면 중량 업 케이스까지 정확히 묶인다.
        from app.processors.kolrabi_order import convert_potato_option
        converted = convert_potato_option(text, "")
        if converted:
            keys.add("potato:" + converted.replace(" ", ""))
    return keys


def _delivery_option_keys(product_name: object, option_name: object) -> set[str]:
    return option_key_set(product_name, option_name) | _semantic_option_keys(product_name, option_name)


def process(
    orderlist_bytes: bytes,
    delivery_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """Input tracking numbers from orderlist into DeliveryList.

    Args:
        orderlist_bytes: Raw bytes of the order list Excel file.
            Columns: K=name, M=phone, O=address, Q=courier, R=tracking
        delivery_bytes: Raw bytes of the DeliveryList Excel file.
            Columns: AA=name, AB=phone, AD=address, D=courier, E=tracking (to fill)

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    # Load orderlist
    ol_wb = load_workbook(filename=BytesIO(orderlist_bytes), data_only=True)
    ol_ws = ol_wb.active

    # Build list of order entries: (name, phone, address, tracking, option_keys)
    order_entries = []
    for row in ol_ws.iter_rows(min_row=2):
        name = normalize(row[10].value) if len(row) > 10 else ""     # K = index 10
        phone = normalize(row[12].value) if len(row) > 12 else ""    # M = index 12
        address = normalize(row[14].value) if len(row) > 14 else ""  # O = index 14
        courier = normalize_courier(row[16].value) if len(row) > 16 else "" # Q = index 16
        tracking = normalize(row[17].value) if len(row) > 17 else "" # R = index 17

        if name and tracking:
            order_entries.append({
                "name": name,
                "phone": phone,
                "address": address,
                "courier": courier,
                "tracking": tracking,
                "option_keys": _source_option_keys(row),
            })

    # Load DeliveryList
    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    dl_ws = dl_wb.active

    # Keep only first sheet
    first_sheet_name = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet_name]

    filled = 0
    skipped = 0
    skip_details: list[str] = []  # 미입력 건의 이름·사유 — 조용한 skip 금지

    # Build index of order entries by name for faster lookup
    from collections import defaultdict
    order_by_name = defaultdict(list)
    for entry in order_entries:
        order_by_name[entry["name"]].append(entry)

    # Track which order entries have been used
    used_entries = set()
    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )

    for row_idx in range(2, dl_ws.max_row + 1):
        # Column E = 5 (tracking number destination)
        e_cell = dl_ws.cell(row=row_idx, column=5)

        # Only fill empty E cells
        if e_cell.value is not None and normalize(e_cell.value) != "":
            continue

        # Get delivery list data
        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)     # AA
        dl_phone = normalize(dl_ws.cell(row=row_idx, column=28).value)    # AB
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)  # AD
        dl_option_keys = _delivery_option_keys(
            dl_ws.cell(row=row_idx, column=11).value,  # K 상품명
            dl_ws.cell(row=row_idx, column=12).value,  # L 옵션명
        )

        if not dl_name:
            continue

        candidates = order_by_name.get(dl_name, [])

        # 폴백: 이름 잘림(쿠팡 열폭 제한 등) 대응 - 접두어 매칭
        if not candidates:
            MIN_PREFIX_LEN = 8
            for order_name, entries in order_by_name.items():
                if len(order_name) < MIN_PREFIX_LEN:
                    continue
                if dl_name.startswith(order_name) or order_name.startswith(dl_name):
                    candidates = entries
                    break

        if not candidates:
            skipped += 1
            skip_details.append(f"{dl_name}(회신에 없음)")
            continue

        # Filter out already used entries
        available = [
            (i, c) for i, c in enumerate(candidates)
            if id(c) not in used_entries
        ]
        if not available:
            skipped += 1
            skip_details.append(f"{dl_name}(회신 운송장 수 부족)")
            continue

        matched = None

        # 1) 전화번호(숫자만 비교) 일치가 가장 강한 신호 — 옵션 가드보다 먼저.
        #    같은 사람이 같은 상품을 여러 건 주문한 경우(전화 동일) 미사용 건을
        #    순차 배정해 N건 전부 채워진다. 회신에 옵션 정보가 없어도 매칭됨.
        #    (2026-07-14 제주다팜: 동일인 다건이 옵션 가드에 걸려 전부 skip되던 사고 수정)
        dl_phone_d = _phone_digits(dl_phone)
        if dl_phone_d:
            phone_hits = [(i, c) for i, c in available if _phone_digits(c.get("phone")) == dl_phone_d]
            if phone_hits:
                matched = next(
                    (c for _, c in phone_hits if c["address"] == dl_address), None
                ) or next(
                    (c for _, c in phone_hits if options_match(c.get("option_keys"), dl_option_keys)), None
                ) or phone_hits[0][1]

        # 2) 전화로 못 가른 경우에만 동명이인 옵션 가드 적용 (기존 보호 유지)
        if matched is None:
            guarded = available
            if requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
                guarded = [
                    (i, c) for i, c in available
                    if options_match(c.get("option_keys"), dl_option_keys)
                ]
                if not guarded:
                    # 옵션으로도 못 가름 → 주소 일치만 허용, 아니면 오배정 위험이라 미입력+표시
                    addr_hits = [c for _, c in available if dl_address and c["address"] == dl_address]
                    if addr_hits:
                        matched = addr_hits[0]
                    else:
                        skipped += 1
                        skip_details.append(f"{dl_name}(동명이인 구분 불가 — 수동 확인)")
                        continue
            if matched is None:
                matched = next(
                    (c for _, c in guarded if c["address"] == dl_address), None
                ) or (guarded[0][1] if guarded else None)

        if matched:
            e_cell.value = matched["tracking"]
            courier = normalize_courier(matched.get("courier"))
            if courier:
                dl_ws.cell(row=row_idx, column=4).value = courier
            used_entries.add(id(matched))
            filled += 1
        else:
            skipped += 1
            skip_details.append(f"{dl_name}(매칭 실패)")

    # Save to bytes
    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_제주다팜_운송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {"filled": filled, "skipped": skipped}
    if skip_details:
        stats["skipped_names"] = ", ".join(skip_details)

    return output.read(), filename, stats
