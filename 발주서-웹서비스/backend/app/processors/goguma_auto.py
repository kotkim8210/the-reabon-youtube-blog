"""Auto-collect 고구마 orders from Coupang API + Toss API and generate 해달 발주서."""

import re
import logging
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR
from app.coupang.client import coupang_goguma_client
from app.processors import goguma_order

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

GOGUMA_KEYWORDS = ["고구마", "꿀고구마"]


def normalize(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def transform_option(option_text: str) -> str:
    """Transform option text for 고구마 orders.

    Remove '1박스 황금 ' prefix if present.
    """
    text = normalize(option_text)
    text = re.sub(r"^1박스\s*황금\s*", "", text)
    return text


def is_goguma_order(item: dict) -> bool:
    """Check if an order item is a 고구마 product."""
    seller_name = (item.get("sellerProductName") or "").lower()
    vendor_name = (item.get("vendorItemName") or "").lower()
    for kw in GOGUMA_KEYWORDS:
        if kw in seller_name or kw in vendor_name:
            return True
    return False


async def collect_orders(from_date: str, to_date: str) -> list[dict]:
    """Collect 고구마 orders from Coupang API for the given date range.

    Args:
        from_date: Start date in YYYY-MM-DD format.
        to_date: End date in YYYY-MM-DD format.

    Returns:
        List of order dicts with keys: name, phone, zipcode, address,
        quantity, product, memo, order_id.
    """
    def _dedupe_key(shipment_box_id, vendor_item_id, order_id, product, name, phone):
        return (
            str(shipment_box_id or ""),
            str(vendor_item_id or ""),
            str(order_id or ""),
            normalize(product),
            normalize(name),
            normalize(phone),
        )

    all_orders_by_key = {}

    for order_status in ("INSTRUCT", "ACCEPT"):
        next_token = None

        while True:
            result = await coupang_goguma_client.get_order_sheets(
                created_at_from=from_date,
                created_at_to=to_date,
                order_status=order_status,
                max_per_page=50,
                next_token=next_token,
            )

            if not result:
                break

            shipments = result.get("data", [])
            if not shipments:
                break

            for shipment in shipments:
                receiver = shipment.get("receiver") or {}
                shipment_box_id = shipment.get("shipmentBoxId")
                order_id = str(shipment.get("orderId", ""))
                memo = shipment.get("parcelPrintMessage") or ""

                for item in shipment.get("orderItems", []):
                    if not is_goguma_order(item):
                        continue

                    option_text = (
                        item.get("sellerProductItemName")
                        or item.get("vendorItemName")
                        or ""
                    )
                    addr1 = receiver.get("addr1") or ""
                    addr2 = receiver.get("addr2") or ""
                    address = f"{addr1} {addr2}".strip()
                    order = {
                        "name": receiver.get("name") or "",
                        "phone": (
                            receiver.get("safeNumber")
                            or receiver.get("receiverPhoneNumber1")
                            or ""
                        ),
                        "zipcode": receiver.get("postCode") or "",
                        "address": address,
                        "quantity": item.get("shippingCount") or 1,
                        "product": transform_option(option_text),
                        "memo": memo,
                        "order_id": order_id,
                    }

                    dedupe_key = _dedupe_key(
                        shipment_box_id,
                        item.get("vendorItemId"),
                        order_id,
                        order["product"],
                        order["name"],
                        order["phone"],
                    )
                    all_orders_by_key.setdefault(dedupe_key, order)

            next_token = result.get("nextToken")
            if not next_token:
                break

    all_orders = list(all_orders_by_key.values())
    logger.info(
        "고구마 주문 수집 완료: %s건 (%s ~ %s, statuses=INSTRUCT+ACCEPT)",
        len(all_orders),
        from_date,
        to_date,
    )
    return all_orders


async def process_from_api(
    from_date: str,
    to_date: str,
    include_toss: bool = False,
) -> tuple[bytes, str, dict]:
    """Collect 고구마 orders from Coupang API and generate 해달 발주서.

    Args:
        from_date: Start date (YYYY-MM-DD).
        to_date: End date (YYYY-MM-DD).

    Returns:
        Tuple of (output_bytes, filename, stats_dict).
    """
    orders = await collect_orders(from_date, to_date)

    toss_entries = []
    if include_toss:
        try:
            toss_entries = await goguma_order.collect_toss_orders(from_date, to_date)
        except Exception as e:
            logger.warning(f"토스 주문 수집 실패 (계속 진행): {e}")

    if not orders and not toss_entries:
        raise ValueError("해당 기간에 고구마 주문이 없습니다.")

    # Load template
    template_path = TEMPLATE_DIR / "해달_발주서_한진양식.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"템플릿 파일을 찾을 수 없습니다: {template_path}. "
            f"templates 디렉토리에 해달_발주서_한진양식.xlsx 파일을 넣어주세요."
        )

    tmpl_wb = load_workbook(filename=str(template_path))

    # Keep only first sheet
    first_sheet = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet]

    font11 = Font(size=11)

    # Find first empty row in template
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    option_totals: dict[str, dict] = {}

    # Write 쿠팡 orders
    row_idx = start_row
    for order in orders:
        zipcode = order["zipcode"]
        if zipcode:
            try:
                zipcode = str(int(float(zipcode))).zfill(5)
            except (ValueError, TypeError):
                zipcode = str(zipcode).zfill(5)

        mapping = {
            1: order["name"],                          # 수령인
            2: order["phone"],                         # 연락처
            5: zipcode,                                # 우편번호
            6: order["address"],                       # 주소
            7: "식품애착",                               # 보내는분
            8: "010-5700-7756",                        # 보내는분 연락처
            12: "전라남도 해남군 산이면 새상골길 ",         # 출발지
            13: str(order["quantity"]),                 # 수량
            14: order["product"],                      # 품목명
            16: "선불",                                 # 결제구분
            19: order["memo"],                         # 배송메세지
        }

        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11
        row_idx += 1

        try:
            qty_int = int(float(order["quantity"])) if order["quantity"] not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        bucket = option_totals.setdefault(
            order["product"] or "고구마",
            {
                "coupang_option_keyword": order["product"] or "고구마",
                "vendor_option_name": order["product"] or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order.get("order_id"):
            bucket["orders"].append({"order_id": order["order_id"], "quantity": qty_int})

    # Write 토스 orders
    for entry in toss_entries:
        mapping = {
            1: entry["name"],
            2: entry["phone"],
            5: entry["zipcode"],
            6: entry["address"],
            7: "식품애착",
            8: "010-5700-7756",
            12: "전라남도 해남군 산이면 새상골길 ",
            13: entry["qty"],
            14: entry["product"],
            16: "선불",
            19: entry["memo"],
        }

        for col, value in mapping.items():
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = font11
        row_idx += 1

        try:
            qty_int = int(float(entry["qty"])) if entry["qty"] not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1
        bucket = option_totals.setdefault(
            entry["product"] or "고구마",
            {
                "coupang_option_keyword": entry["product"] or "고구마",
                "vendor_option_name": entry["product"] or "고구마",
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if entry.get("order_id"):
            bucket["orders"].append({"order_id": entry["order_id"], "quantity": qty_int})

    # Save
    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    total = len(orders) + len(toss_entries)
    filename = f"해달 발주서 한진양식_알제이시스템즈({now.strftime('%y%m%d')}).xlsx"
    stats = {
        "total": total,
        "coupang": len(orders),
        "toss": len(toss_entries),
        "period": f"{from_date} ~ {to_date}",
        "product": "고구마",
        "options": list(option_totals.values()),
    }

    return output.read(), filename, stats
