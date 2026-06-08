import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.processors.tracking_match import (
    name_counts,
    normalize_courier_name,
    option_key_set,
    options_match,
    requires_option_guard,
)

KST = timezone(timedelta(hours=9))


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip()).lower()


def valid_tracking(value: object) -> str:
    text = normalize(value)
    digits = re.sub(r"\D", "", text)
    if re.fullmatch(r"01[016789]\d{7,8}", digits):
        return ""
    if 8 <= len(digits) <= 20:
        return digits
    return ""


def contains_any(value: object, needles: tuple[str, ...]) -> bool:
    text = normalize_text(value)
    return any(needle in text for needle in needles)


def detect_columns(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        "name": ("수취인", "수령인", "받는분", "받는사람", "고객명", "성명", "성함", "이름", "name"),
        "phone": ("연락처", "전화", "휴대폰", "핸드폰", "mobile", "phone"),
        "address": ("주소", "배송지", "address"),
        "tracking": ("송장", "운송장", "출고번호", "tracking", "invoice"),
        "courier": ("택배", "택배사", "배송사", "courier", "carrier"),
        "option": ("옵션", "상품명", "품목", "제품명", "상품", "option", "product"),
    }
    best_row = 1
    best: dict[str, int] = {}
    best_score = 0
    max_row = min(ws.max_row, 15)
    max_col = min(ws.max_column, 40)
    for row_idx in range(1, max_row + 1):
        found: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            for key, needles in aliases.items():
                if key not in found and contains_any(value, needles):
                    found[key] = col_idx
        score = len(found) + (2 if "tracking" in found else 0) + (1 if "name" in found else 0)
        if score > best_score:
            best_row = row_idx
            best = found
            best_score = score

    if "name" not in best:
        best["name"] = 1
    if "phone" not in best:
        best["phone"] = 2
    if "address" not in best:
        best["address"] = 4
    if "option" not in best:
        best["option"] = 5
    return best_row, best


def row_tracking(ws, row_idx: int, tracking_col: int | None, skip_cols: set[int] | None = None) -> str:
    skip_cols = skip_cols or set()
    if tracking_col:
        tracking = valid_tracking(ws.cell(row=row_idx, column=tracking_col).value)
        if tracking:
            return tracking
    max_col = min(ws.max_column, 50)
    scan_order = list(range(9, max_col + 1)) + list(range(1, min(8, max_col) + 1))
    for col_idx in scan_order:
        if col_idx in skip_cols:
            continue
        tracking = valid_tracking(ws.cell(row=row_idx, column=col_idx).value)
        if tracking:
            return tracking
    return ""


def source_option_keys(ws, row_idx: int, option_col: int | None) -> set[str]:
    values = []
    if option_col:
        values.append(ws.cell(row=row_idx, column=option_col).value)
    for col_idx in range(1, min(ws.max_column, 20) + 1):
        values.append(ws.cell(row=row_idx, column=col_idx).value)
    return option_key_set(*values)


def delivery_option_keys(ws, row_idx: int) -> set[str]:
    return option_key_set(
        ws.cell(row=row_idx, column=11).value,
        ws.cell(row=row_idx, column=12).value,
    )


def process(reply_bytes: bytes, delivery_bytes: bytes) -> tuple[bytes, str, dict]:
    """Best-effort tracking input for user-specific reply workbooks.

    The parser first tries to detect header names such as 송장번호, 수취인, 연락처,
    주소, 택배사. If headers are missing, it falls back to the generic row-mode
    order output shape: A=이름, B=전화, D=주소, E=옵션, and scans each row for a
    likely tracking number.
    """
    reply_wb = load_workbook(filename=BytesIO(reply_bytes), data_only=True)
    reply_ws = reply_wb.active
    header_row, cols = detect_columns(reply_ws)

    entries = []
    for row_idx in range(header_row + 1, reply_ws.max_row + 1):
        skip_cols = {
            cols.get("name", 0),
            cols.get("phone", 0),
            cols.get("address", 0),
            cols.get("option", 0),
            cols.get("courier", 0),
        }
        tracking = row_tracking(reply_ws, row_idx, cols.get("tracking"), skip_cols)
        if not tracking:
            continue
        name = normalize(reply_ws.cell(row=row_idx, column=cols["name"]).value)
        if not name:
            continue
        phone = normalize(reply_ws.cell(row=row_idx, column=cols.get("phone", 2)).value)
        address = normalize(reply_ws.cell(row=row_idx, column=cols.get("address", 4)).value)
        courier = ""
        if cols.get("courier"):
            courier = normalize_courier_name(reply_ws.cell(row=row_idx, column=cols["courier"]).value)
        entries.append(
            {
                "name": name,
                "phone": phone,
                "address": address,
                "courier": courier,
                "tracking": tracking,
                "option_keys": source_option_keys(reply_ws, row_idx, cols.get("option")),
            }
        )

    if not entries:
        raise ValueError("회신 파일에서 송장번호가 입력된 행을 찾지 못했습니다.")

    dl_wb = load_workbook(filename=BytesIO(delivery_bytes))
    first_sheet = dl_wb.sheetnames[0]
    for sheet_name in dl_wb.sheetnames[1:]:
        del dl_wb[sheet_name]
    dl_ws = dl_wb[first_sheet]

    by_name = defaultdict(list)
    for entry in entries:
        by_name[entry["name"]].append(entry)

    delivery_name_counts = name_counts(
        normalize(dl_ws.cell(row=row_idx, column=27).value)
        for row_idx in range(2, dl_ws.max_row + 1)
    )
    used_entries: set[int] = set()
    filled = 0
    skipped = 0

    for row_idx in range(2, dl_ws.max_row + 1):
        tracking_cell = dl_ws.cell(row=row_idx, column=5)
        if normalize(tracking_cell.value):
            continue

        dl_name = normalize(dl_ws.cell(row=row_idx, column=27).value)
        dl_phone = normalize(dl_ws.cell(row=row_idx, column=28).value)
        dl_address = normalize(dl_ws.cell(row=row_idx, column=30).value)
        dl_option_keys = delivery_option_keys(dl_ws, row_idx)

        if not dl_name:
            continue

        candidates = by_name.get(dl_name, [])
        if not candidates:
            for source_name, rows in by_name.items():
                if len(source_name) >= 3 and (dl_name.startswith(source_name) or source_name.startswith(dl_name)):
                    candidates = rows
                    break

        available = [entry for entry in candidates if id(entry) not in used_entries]
        if not available:
            skipped += 1
            continue

        matched = None
        for entry in available:
            if entry["phone"] and entry["phone"] == dl_phone and entry["address"] and entry["address"] == dl_address:
                matched = entry
                break
        if matched is None:
            for entry in available:
                if entry["phone"] and entry["phone"] == dl_phone:
                    matched = entry
                    break
        if matched is None:
            for entry in available:
                if entry["address"] and entry["address"] == dl_address:
                    matched = entry
                    break

        if matched is None and requires_option_guard(dl_name, delivery_name_counts, len(candidates)):
            option_matches = [entry for entry in available if options_match(entry.get("option_keys"), dl_option_keys)]
            if option_matches:
                matched = option_matches[0]
            else:
                skipped += 1
                continue

        if matched is None:
            matched = available[0]

        tracking_cell.value = matched["tracking"]
        if matched.get("courier"):
            dl_ws.cell(row=row_idx, column=4).value = matched["courier"]
        used_entries.add(id(matched))
        filled += 1

    output = BytesIO()
    dl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"DeliveryList_맞춤송장입력완료_{now.strftime('%Y%m%d')}.xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "reply_rows": len(entries),
        "processor": "맞춤 범용 송장",
    }
    return output.read(), filename, stats
