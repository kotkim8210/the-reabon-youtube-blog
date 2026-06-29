"""데일리푸드 홍매실(황매실) 발주서 생성기.

데일리푸드(dailyfood.adminplus.co.kr)는 2026-06 신규 공급처. 홍매실 발주를 담당.
DeliveryList에서 매실(홍매실/황매실) 주문을 추출해 데일리푸드 양식으로 출력한다.

데일리푸드 양식 열(1-indexed):
  A 보내는분주소 / B 보내는분성명 / C 보내는분전화번호 / D 받는분주소 /
  E 받는분성명 / F 받는분전화번호 / G 받는분기타연락처 / H 박스수량 /
  I 품목명 / J 배송메세지1 / K 주문번호
"""

import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.config import TEMPLATE_DIR

KST = timezone(timedelta(hours=9))

SENDER_NAME = "식품애착"
SENDER_PHONE = "010-5700-7756"
# 보내는분주소(A)는 비워둔다 — 데일리푸드가 식품애착 발송지를 알고 있어 명이나물(쥬얼리)
# 발주서와 동일하게 공란 처리(사용자 지정).


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _fmt_kg(value: str) -> str:
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(f)) if f.is_integer() else str(f)


def is_dailyfood_order(product_name: object, option_text: object) -> bool:
    """데일리푸드 홍매실(황매실) 발주 대상 여부."""
    text = re.sub(r"\s+", "", f"{product_name or ''} {option_text or ''}")
    return "매실" in text


def convert_dailyfood_option(product_name: object, option_text: object) -> str | None:
    """매실 DeliveryList → 데일리푸드 발주 품목명 '홍매실 ({등급}) {kg}kg'.

    데일리푸드 시트 등급: '특~중/혼합과' 와 '왕특' 두 종류(각 2/3/5/10kg).
    - 옵션/상품명에 '혼합/중/알뜰/랜덤' 표기 → '특~중/혼합과'
    - 그 외(대/특대/왕특/특 등) → '왕특'  (쿠팡 '1박스 대' → 왕특, 사용자 매핑)
    kg는 상품명/옵션의 'Nkg'·'N키로' 등에서 추출.
    """
    if not is_dailyfood_order(product_name, option_text):
        return None
    text = f"{product_name or ''} {option_text or ''}"
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:kg|킬로|키로|키)", text, re.IGNORECASE)
    kg = _fmt_kg(m.group(1)) if m else ""
    compact = re.sub(r"\s+", "", text)
    if any(g in compact for g in ("혼합", "중과", "알뜰", "랜덤")):
        grade = "특~중/혼합과"
    else:
        grade = "왕특"
    return f"홍매실 ({grade}) {kg}kg" if kg else f"홍매실 ({grade})"


def process(delivery_file_bytes: bytes) -> tuple[bytes, str, dict] | None:
    """DeliveryList → 데일리푸드 홍매실 발주서."""
    dl_wb = load_workbook(filename=BytesIO(delivery_file_bytes), data_only=True)
    dl_ws = dl_wb.active

    filtered_rows: list[tuple] = []
    for row in dl_ws.iter_rows(min_row=2):
        product_name = normalize(row[10].value) if len(row) > 10 else ""
        option = normalize(row[11].value) if len(row) > 11 else ""
        converted = convert_dailyfood_option(product_name, option)
        if converted:
            filtered_rows.append((row, converted))

    if not filtered_rows:
        return None

    template_path = TEMPLATE_DIR / "데일리푸드_홍매실_원본.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path}. "
            f"데일리푸드 양식 템플릿을 templates 폴더에 넣어주세요."
        )

    tmpl_wb = load_workbook(filename=str(template_path))
    first_sheet_name = tmpl_wb.sheetnames[0]
    for name in tmpl_wb.sheetnames[1:]:
        del tmpl_wb[name]
    ws = tmpl_wb[first_sheet_name]

    font11 = Font(size=11)
    start_row = 2  # 1행 = 헤더

    option_totals: dict[str, dict] = {}
    for i, (row, product) in enumerate(filtered_rows):
        out_row = start_row + i

        name = normalize(row[26].value) if len(row) > 26 else ""      # AA 수취인이름
        phone = normalize(row[27].value) if len(row) > 27 else ""     # AB 수취인전화번호
        address = normalize(row[29].value) if len(row) > 29 else ""   # AD 수취인 주소
        memo = normalize(row[30].value) if len(row) > 30 else ""      # AE 배송메세지
        order_no = normalize(row[2].value) if len(row) > 2 else ""    # C 주문번호
        option = normalize(row[11].value) if len(row) > 11 else ""    # L 등록옵션명
        qty_val = row[22].value if len(row) > 22 else ""              # W 구매수(수량)
        try:
            qty_int = int(float(qty_val)) if qty_val not in (None, "") else 1
        except (ValueError, TypeError):
            qty_int = 1

        mapping = {
            1: "",            # A 보내는분주소 (비움)
            2: SENDER_NAME,   # B 보내는분성명
            3: SENDER_PHONE,  # C 보내는분전화번호
            4: address,       # D 받는분주소
            5: name,          # E 받는분성명
            6: phone,         # F 받는분전화번호
            7: "",            # G 받는분기타연락처 (비움)
            8: qty_int,       # H 박스수량
            9: product,       # I 품목명
            10: memo,         # J 배송메세지1
            11: order_no,     # K 주문번호
        }
        for col, value in mapping.items():
            cell = ws.cell(row=out_row, column=col, value=value)
            cell.font = font11

        bucket = option_totals.setdefault(
            option or product,
            {
                "coupang_option_keyword": option or product,
                "vendor_option_name": product,
                "quantity": 0,
                "orders": [],
            },
        )
        bucket["quantity"] += qty_int
        if order_no:
            bucket["orders"].append({"order_id": order_no, "quantity": qty_int})

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"데일리푸드_홍매실_발주({now.strftime('%Y%m%d')}).xlsx"
    stats = {
        "total": len(filtered_rows),
        "product": "홍매실(데일리푸드)",
        "options": list(option_totals.values()),
    }
    return output.read(), filename, stats
