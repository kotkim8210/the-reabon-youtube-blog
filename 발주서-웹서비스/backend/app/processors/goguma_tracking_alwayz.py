"""올웨이즈 주문 파일에 해달 발주서 운송장번호를 자동 입력.

Flow:
1. 해달 발주서에서 (이름, 전화, 주소, 운송장번호) 추출
2. 올웨이즈 주문 파일에서 빈 W열(운송장번호) 행 탐색
3. 이름/전화/주소 매칭으로 운송장번호를 W열에 입력, U열에 실제 택배사 입력
   (회신파일의 택배사를 읽어 올웨이즈 표기로 정규화. 없으면 기본 한진택배)
"""

import re
from datetime import datetime, timezone, timedelta
from io import BytesIO
from collections import defaultdict

from openpyxl import load_workbook

from app.processors.goguma_order import transform_alwayz_option, transform_option
from app.processors.haedal_tracking_parser import (
    detect_haedal_columns,
    find_courier_in_row,
    find_tracking_in_row,
)
from app.processors.tracking_match import (
    name_counts,
    option_key_set,
    options_match,
    requires_option_guard,
)


KST = timezone(timedelta(hours=9))


def normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r'\s+', '', str(value).strip())


def _alwayz_courier(value: object, default: str = "한진택배") -> str:
    """회신파일 택배사 → 올웨이즈가 인식하는 택배사명. 값 없으면 기본 한진택배."""
    text = str(value).strip() if value is not None else ""
    if not text:
        return default
    c = re.sub(r"\s+", "", text).lower()
    if "대한통운" in c or "cj" in c:
        return "CJ대한통운"
    if "한진" in c:
        return "한진택배"
    if "롯데" in c:
        return "롯데택배"
    if "우체국" in c or "epost" in c:
        return "우체국택배"
    if "로젠" in c or "logen" in c:
        return "로젠택배"
    if "현대" in c:
        return "현대택배"
    return text


def process(
    haedal_bytes: bytes,
    alwayz_bytes: bytes,
) -> tuple[bytes, str, dict]:
    """해달 발주서 운송장번호를 올웨이즈 주문 파일 W열에 입력.

    Args:
        haedal_bytes: 해달 발주서 Excel (A=이름, B=전화, F=주소, P~V=운송장번호)
        alwayz_bytes: 올웨이즈 주문내역 Excel (S=수령인, T=연락처, O=주소, W=운송장번호, U=택배사)

    Returns:
        (output_bytes, filename, stats)
    """
    # 해달 발주서 파싱
    hd_wb = load_workbook(filename=BytesIO(haedal_bytes), data_only=True)
    hd_ws = hd_wb.active
    cols = detect_haedal_columns(hd_ws)

    haedal_entries = []
    for row_idx in range(cols.start_row, hd_ws.max_row + 1):
        name = normalize(hd_ws.cell(row=row_idx, column=cols.name).value)
        phone = normalize(hd_ws.cell(row=row_idx, column=cols.phone).value)
        address = normalize(hd_ws.cell(row=row_idx, column=cols.address).value)
        product = hd_ws.cell(row=row_idx, column=cols.product).value

        if not name:
            continue

        tracking = find_tracking_in_row(hd_ws, row_idx, cols.tracking)
        if tracking:
            skip_cols = tuple(
                c for c in (cols.name, cols.phone, cols.address, cols.product, cols.tracking) if c
            )
            courier = find_courier_in_row(hd_ws, row_idx, cols.courier, skip_cols)
            haedal_entries.append({
                "name": name,
                "phone": phone,
                "address": address,
                "tracking": tracking,
                "courier": courier,
                "option_keys": option_key_set(product, transform_option(str(product or ""))),
            })

    # 올웨이즈 파일 로드
    al_wb = load_workbook(filename=BytesIO(alwayz_bytes))
    al_ws = al_wb.active

    entry_by_name = defaultdict(list)
    for entry in haedal_entries:
        entry_by_name[entry["name"]].append(entry)

    used_entries = set()
    filled = 0
    skipped = 0
    alwayz_name_counts = name_counts(
        normalize(al_ws.cell(row=row_idx, column=19).value)
        for row_idx in range(2, al_ws.max_row + 1)
    )

    for row_idx in range(2, al_ws.max_row + 1):
        # W열(23) = 운송장번호 대상
        w_cell = al_ws.cell(row=row_idx, column=23)

        if w_cell.value is not None and normalize(w_cell.value) != "":
            continue

        al_name = normalize(al_ws.cell(row=row_idx, column=19).value)     # S = 수령인
        al_phone = normalize(al_ws.cell(row=row_idx, column=20).value)    # T = 수령인 연락처
        al_address = normalize(al_ws.cell(row=row_idx, column=15).value)  # O = 주소
        al_option_raw = al_ws.cell(row=row_idx, column=6).value           # F = 옵션
        al_option_keys = option_key_set(al_option_raw, transform_alwayz_option(str(al_option_raw or "")))

        if not al_name:
            continue

        candidates = entry_by_name.get(al_name, [])
        if not candidates:
            skipped += 1
            continue

        available = [
            (i, c) for i, c in enumerate(candidates)
            if id(c) not in used_entries
        ]
        if not available:
            skipped += 1
            continue

        if requires_option_guard(al_name, alwayz_name_counts, len(candidates)):
            available = [
                (i, c) for i, c in available
                if options_match(c.get("option_keys"), al_option_keys)
            ]
            if not available:
                skipped += 1
                continue

        matched = None

        # 전화+주소 매칭
        for _, c in available:
            if c["phone"] == al_phone and c["address"] == al_address:
                matched = c
                break

        # 전화만
        if matched is None:
            for _, c in available:
                if c["phone"] == al_phone:
                    matched = c
                    break

        # 주소만
        if matched is None:
            for _, c in available:
                if c["address"] == al_address:
                    matched = c
                    break

        # 첫번째
        if matched is None:
            matched = available[0][1]

        if matched:
            w_cell.value = matched["tracking"]
            al_ws.cell(row=row_idx, column=21).value = _alwayz_courier(matched.get("courier"))  # U = 택배사(회신 실제 택배사)
            used_entries.add(id(matched))
            filled += 1
        else:
            skipped += 1

    output = BytesIO()
    al_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"일상애착_주문 내역_{now.strftime('%Y-%m-%d')}.xlsx"
    stats = {
        "filled": filled,
        "skipped": skipped,
        "haedal_entries": len(haedal_entries),
    }

    return output.read(), filename, stats
