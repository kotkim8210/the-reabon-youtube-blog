"""Dedicated Danharu order sheet generator."""

from __future__ import annotations

import re
from copy import copy
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.config import TEMPLATE_DIR

KST = timezone(timedelta(hours=9))
DEFAULT_TEMPLATE = TEMPLATE_DIR / "danharu_order_template.xlsx"
DEFAULT_SENDER = {
    "name": "단하루",
    "phone": "010-7388-2346",
    "zipcode": "17857",
    "address": "비전3로115 110-502",
}


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_supply_option(option_name: str, product_name: str) -> str:
    raw = normalize(option_name)
    if raw:
        return raw
    match = re.search(r"\(([^()]+)\)\s*$", normalize(product_name))
    if match:
        return normalize(match.group(1))
    return normalize(product_name)


def clean_product_name(product_name: str) -> str:
    name = normalize(product_name)
    match = re.match(r"^(.*?)[(（][^()（）]+[)）]\s*$", name)
    if match:
        return match.group(1).strip(" ,")
    return name


def build_platform_label(product_name: str, option_name: str) -> str:
    product_text = normalize(product_name)
    option_text = normalize(option_name)
    if option_text and option_text not in product_text:
        return f"{product_text} ({option_text})"
    return product_text or option_text


def parse_quantity(value: object) -> int:
    try:
        return max(1, int(float(str(value).strip())))
    except (TypeError, ValueError):
        return 1


def _load_template(template_bytes: bytes | None):
    if template_bytes:
        return load_workbook(filename=BytesIO(template_bytes))
    if not DEFAULT_TEMPLATE.exists():
        raise FileNotFoundError(f"Danharu template not found: {DEFAULT_TEMPLATE}")
    return load_workbook(filename=str(DEFAULT_TEMPLATE))


def _sender_defaults(ws) -> dict[str, str]:
    return {
        "name": normalize(ws["A2"].value) or DEFAULT_SENDER["name"],
        "phone": normalize(ws["B2"].value) or DEFAULT_SENDER["phone"],
        "zipcode": normalize(ws["C2"].value) or DEFAULT_SENDER["zipcode"],
        "address": normalize(ws["D2"].value) or DEFAULT_SENDER["address"],
    }


def _row_styles(ws, row_index: int, max_col: int) -> list[tuple]:
    styles: list[tuple] = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_index, column=col)
        styles.append(
            (
                copy(cell._style),
                copy(cell.number_format),
                copy(cell.alignment),
                copy(cell.font),
                copy(cell.fill),
                copy(cell.border),
                copy(cell.protection),
            )
        )
    return styles


def _apply_row_styles(ws, row_index: int, styles: list[tuple]) -> None:
    for col, style in enumerate(styles, start=1):
        cell = ws.cell(row=row_index, column=col)
        (
            cell._style,
            cell.number_format,
            cell.alignment,
            cell.font,
            cell.fill,
            cell.border,
            cell.protection,
        ) = style


def process(delivery_bytes: bytes, template_bytes: bytes | None = None) -> tuple[bytes, str, dict, list[dict]]:
    """Create a Danharu order workbook from a DeliveryList workbook."""
    delivery_wb = load_workbook(filename=BytesIO(delivery_bytes), data_only=True)
    delivery_ws = delivery_wb.active

    tmpl_wb = _load_template(template_bytes)
    ws = tmpl_wb.active
    sender = _sender_defaults(ws)
    style_row = 2 if ws.max_row >= 2 else 1
    row_styles = _row_styles(ws, style_row, 15)

    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    items: list[dict] = []
    total_quantity = 0

    for row_number, row in enumerate(delivery_ws.iter_rows(min_row=2, values_only=True), start=2):
        receiver_name = normalize(row[26]) if len(row) > 26 else ""
        if not receiver_name:
            continue

        receiver_phone = normalize(row[27]) if len(row) > 27 else ""
        receiver_zipcode = normalize(row[28]) if len(row) > 28 else ""
        receiver_address = normalize(row[29]) if len(row) > 29 else ""
        platform_product_name = normalize(row[10]) if len(row) > 10 else ""
        platform_option_name = normalize(row[11]) if len(row) > 11 else ""
        external_order_id = normalize(row[2]) if len(row) > 2 else ""
        memo = normalize(row[30]) if len(row) > 30 else ""
        quantity = parse_quantity(row[22] if len(row) > 22 else 1)
        supply_option_name = clean_supply_option(platform_option_name, platform_product_name)
        product_name = clean_product_name(platform_product_name or supply_option_name)
        platform_label = build_platform_label(platform_product_name, platform_option_name)

        items.append(
            {
                "product_name": product_name,
                "platform_option_name": platform_option_name,
                "supply_option_name": supply_option_name,
                "quantity": quantity,
                "external_order_id": external_order_id,
                "receiver_name": receiver_name,
                "receiver_phone": receiver_phone,
                "receiver_address": receiver_address,
                "memo": memo,
                "source_row": row_number,
            }
        )
        total_quantity += quantity

        target_row = len(items) + 1
        _apply_row_styles(ws, target_row, row_styles)
        ws.cell(target_row, 1, sender["name"])
        ws.cell(target_row, 2, sender["phone"])
        ws.cell(target_row, 3, sender["zipcode"])
        ws.cell(target_row, 4, sender["address"])
        ws.cell(target_row, 5, receiver_name)
        ws.cell(target_row, 6, receiver_phone)
        ws.cell(target_row, 7, receiver_phone)
        ws.cell(target_row, 8, receiver_zipcode)
        ws.cell(target_row, 9, receiver_address)
        ws.cell(target_row, 10, quantity)
        ws.cell(target_row, 11, platform_label)
        ws.cell(target_row, 12, supply_option_name)
        # Column M is reserved for tracking numbers on the returned vendor sheet.
        ws.cell(target_row, 13, "")
        ws.cell(target_row, 14, memo)
        ws.cell(target_row, 15, product_name)

    if not items:
        raise ValueError("업로드한 DeliveryList에서 처리할 주문을 찾지 못했습니다.")

    ws.auto_filter.ref = f"A1:O{len(items) + 1}"

    output = BytesIO()
    tmpl_wb.save(output)
    output.seek(0)

    now = datetime.now(KST)
    filename = f"단하루_발주서_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    stats = {
        "total_rows": len(items),
        "total_quantity": total_quantity,
        "unique_options": len({item['supply_option_name'] for item in items if item['supply_option_name']}),
    }
    return output.read(), filename, stats, items
