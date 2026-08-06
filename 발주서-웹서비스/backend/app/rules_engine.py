"""상품 매칭 규칙 엔진 (Phase 1-A).

발주처별 상품 매칭(키워드·제외어·등급/kg 추출·치환맵·출력 템플릿)을
DB(product_rules)에서 읽어 해석한다. 목표: 상품 추가/이관을 코드 배포가 아닌
규칙 데이터 수정으로 처리(운영 비용 절감 + 셀프서비스 상용화의 전제).

프로세서(동기 함수)에서 쓸 수 있도록 convert()는 인메모리 캐시 기반 동기 호출이고,
캐시는 앱 시작 시(lifespan)와 규칙 CRUD 후 refresh_rules()로 갱신한다.
캐시가 로드되지 않았거나 매칭 규칙이 없으면 None을 반환해 호출부가
기존 하드코딩 로직으로 폴백하게 한다(운영 안전 최우선).
"""

import logging
import re
import threading
from datetime import datetime, timedelta, timezone

logger = logging.getLogger(__name__)
KST = timezone(timedelta(hours=9))

_lock = threading.Lock()
_rules_by_supplier: dict[str, list[dict]] = {}
_suppliers: dict[str, dict] = {}
_loaded = False
_refreshed_at: str | None = None

_KG_RE = re.compile(r"(\d+(?:\.\d+)?)\s*kg", re.IGNORECASE)
_COUNT_RE = re.compile(r"(\d+)\s*(?:개입|개|입)")


class _SafeDict(dict):
    """output_template에 없는 자리표시자는 빈 문자열로 대체."""

    def __missing__(self, key):  # noqa: D105
        return ""


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _compact(value: object) -> str:
    return re.sub(r"\s+", "", _normalize(value))


def _fmt_kg(raw: str) -> str:
    """'3.0'→'3', '2.5'→'2.5'."""
    try:
        f = float(raw)
    except (TypeError, ValueError):
        return str(raw)
    return str(int(f)) if f.is_integer() else str(f)


def _extract_grade(rule: dict, option_text: str, combined: str) -> str:
    grades = rule.get("grades") or []
    for source in (option_text, combined):
        for g in grades:
            if g and g in source:
                return g
    return ""


def _extract_kg(option_text: str, combined: str) -> str:
    for source in (option_text, combined):
        m = _KG_RE.search(source)
        if m:
            return _fmt_kg(m.group(1))
    return ""


def _extract_count(option_text: str, combined: str) -> str:
    """수량 단위 상품(옥수수 'N개/입') 지원. 없으면 빈 문자열."""
    for source in (option_text, combined):
        m = _COUNT_RE.search(source)
        if m:
            return m.group(1)
    return ""


def resolve(supplier_key: str, product_name: object, option_text: object) -> dict | None:
    """매칭 규칙 해석. 반환: {rule_id, label, output, grade, kg} 또는 None(미매칭).

    캐시 미로드 시에도 None — 호출부는 하드코딩 폴백을 유지한다.
    """
    if not _loaded:
        return None
    rules = _rules_by_supplier.get(supplier_key) or []
    if not rules:
        return None

    name_n = _normalize(product_name)
    option_n = _normalize(option_text)
    combined = f"{name_n} {option_n}".strip()
    compact = _compact(combined)

    for rule in rules:
        keywords = rule.get("name_keywords") or []
        if not any(k and _compact(k) in compact for k in keywords):
            continue
        excludes = rule.get("exclude_keywords") or []
        if any(x and _compact(x) in compact for x in excludes):
            continue

        template = rule.get("output_template") or ""
        grade = _extract_grade(rule, option_n, combined)
        kg = _extract_kg(option_n, combined)
        count = _extract_count(option_n, combined)
        if rule.get("require_grade") and not grade:
            continue
        if rule.get("require_kg") and not kg:
            continue
        if "{count}" in template and not count:
            # 수량 단위 템플릿인데 'N개/입'을 못 찾음 → 이 규칙은 미매칭
            continue
        kg_allow = [str(k) for k in (rule.get("kg_allow") or [])]
        if kg_allow and kg not in kg_allow:
            continue

        pair_map = rule.get("pair_map") or {}
        mapped = pair_map.get(f"{grade}|{kg}")
        if mapped and "|" in mapped:
            grade, kg = mapped.split("|", 1)

        extra = (rule.get("extra_map") or {}).get(f"{grade}|{kg}", "")
        output = template.format_map(
            _SafeDict(grade=grade, kg=kg, count=count, extra=extra)
        )
        output = re.sub(r"\s{2,}", " ", output).strip()
        if not output:
            continue
        return {
            "rule_id": rule.get("id"),
            "label": rule.get("label"),
            "output": output,
            "grade": grade,
            "kg": kg,
            "count": count,
        }
    return None


def convert(supplier_key: str, product_name: object, option_text: object) -> str | None:
    """발주 품목명 변환. 미매칭/캐시 미로드면 None(호출부 하드코딩 폴백)."""
    result = resolve(supplier_key, product_name, option_text)
    return result["output"] if result else None

def resolve_any(product_name: object, option_text: object) -> dict | None:
    """전체 발주처를 대상으로 첫 매칭 규칙 해석 (시뮬레이터용).

    발주처는 key 정렬 순으로 평가(결정적). 반환에 supplier_key/supplier_name 포함.
    """
    if not _loaded:
        return None
    for supplier_key in sorted(_rules_by_supplier):
        result = resolve(supplier_key, product_name, option_text)
        if result:
            supplier = _suppliers.get(supplier_key) or {}
            return {
                **result,
                "supplier_key": supplier_key,
                "supplier_name": supplier.get("name") or supplier_key,
            }
    return None


def simulate_delivery(delivery_bytes: bytes, max_rows: int = 500) -> dict:
    """DeliveryList 전 행에 규칙을 적용한 미리보기 (파일 미저장, 발주서 미생성).

    '이 규칙이면 이렇게 발주됩니다'를 보여주는 시뮬레이터의 코어.
    반환: {total, matched, unmatched, truncated, rows[], unmatched_options[]}
    rows: [{row_no, product_name, option, matched, supplier_name, label, output}]
    """
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(delivery_bytes), data_only=True, read_only=True)
    ws = wb.active

    rows: list[dict] = []
    matched = 0
    unmatched = 0
    unmatched_options: dict[str, int] = {}
    total = 0
    for row_no, row in enumerate(ws.iter_rows(min_row=2), start=2):
        product_name = _normalize(row[10].value) if len(row) > 10 else ""  # K열
        option = _normalize(row[11].value) if len(row) > 11 else ""        # L열
        if not product_name and not option:
            continue
        total += 1
        result = resolve_any(product_name, option)
        if result:
            matched += 1
        else:
            unmatched += 1
            key = f"{product_name} | {option}".strip(" |")
            unmatched_options[key] = unmatched_options.get(key, 0) + 1
        if len(rows) < max_rows:
            rows.append({
                "row_no": row_no,
                "product_name": product_name,
                "option": option,
                "matched": result is not None,
                "supplier_name": (result or {}).get("supplier_name") or "",
                "label": (result or {}).get("label") or "",
                "output": (result or {}).get("output") or "",
            })
    wb.close()
    return {
        "total": total,
        "matched": matched,
        "unmatched": unmatched,
        "truncated": total > len(rows),
        "rows": rows,
        "unmatched_options": [
            {"text": k, "count": v}
            for k, v in sorted(unmatched_options.items(), key=lambda kv: -kv[1])[:50]
        ],
    }


# ── 파일로 규칙 초안 자동 생성 (Phase 2-B: 셀프서비스 온보딩) ─────────────
# DeliveryList의 상품명(K)·옵션(L) 패턴에서 규칙 초안을 뽑아 고객이 확인·저장한다.
# 완전 무인 자동은 위험(오매핑=오배송)하므로 '초안 제시 → 사람 확인'만 담당.

_INFER_GRADE_VOCAB = [
    "로얄대과", "로얄중과", "로얄소과", "로얄과",
    "특대과", "중대과", "중소과", "왕대과",
    "특상", "중상", "상특", "특대", "왕특",
    "대과", "중과", "소과", "특품", "중품", "상품",
    "왕", "특", "대", "중", "소", "상", "로얄", "못난이", "가정용", "실속", "정품",
]
_INFER_GRADES_SORTED = sorted(set(_INFER_GRADE_VOCAB), key=len, reverse=True)
_INFER_SINGLE_GRADES = {g for g in _INFER_GRADE_VOCAB if len(g) == 1}

_INFER_NAME_NOISE = {
    "고당도", "햇", "국내산", "국산", "산지직송", "산지", "직송", "프리미엄",
    "당일수확", "당일발송", "무료배송", "선물용", "선물세트", "세척", "손질",
    "냉장", "냉동", "유기농", "친환경", "무농약", "특가", "정품", "실속", "가정용",
    "새벽배송", "당일", "명품", "프레시", "주문폭주", "인기",
}

_INFER_TOKEN_RE = re.compile(r"[,/·|]+|\s+")


def _infer_grades_in(option: str) -> list[str]:
    """옵션 텍스트에서 등급 토큰을 뽑는다. 단일글자 등급은 독립 토큰일 때만 인정."""
    found: list[str] = []
    for tok in _INFER_TOKEN_RE.split(option):
        tok = tok.strip()
        if not tok:
            continue
        matched = next((g for g in _INFER_GRADES_SORTED if len(g) >= 2 and g in tok), None)
        if matched:
            found.append(matched)
        elif tok in _INFER_SINGLE_GRADES:
            found.append(tok)
    return found


def _infer_keyword(product: str) -> str:
    """상품명에서 매칭 키워드 후보를 고른다.

    비노이즈 토큰 중 **가장 긴 것**(동률이면 앞쪽). 앞쪽 토큰을 그냥 쓰면
    '국내산 김천 피자두…'→'김천', '제주 하우스 애플망고…'→'제주'처럼 지역명이
    잡혀 발주명까지 틀어진다(2026-08-06 실측). 긴 토큰이 상품 고유명일 확률이
    높고, 더 구체적이라 다른 상품에 잘못 매칭될 위험도 작다.
    """
    tokens = [t for t in re.split(r"\s+", product) if t]
    candidates = [
        t for t in tokens
        if t not in _INFER_NAME_NOISE and not _KG_RE.search(t) and not t.isdigit()
    ]
    if not candidates:
        return tokens[0] if tokens else product
    best = max(candidates, key=lambda t: (len(t), -candidates.index(t)))
    return best


def _distinct(seq) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in seq:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def infer_rules_from_delivery(delivery_bytes: bytes, supplier_key: str, max_products: int = 100) -> dict:
    """DeliveryList → 규칙 초안 목록. 반환: {drafts, covered, product_count}.

    drafts: 기존 규칙에 안 걸리는 신규 상품의 규칙 초안(ProductRuleInput 호환 + UI meta).
    covered: 이미 규칙이 있는 상품(참고용).
    """
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(delivery_bytes), data_only=True, read_only=True)
    ws = wb.active

    groups: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2):
        product = _normalize(row[10].value) if len(row) > 10 else ""   # K열
        option = _normalize(row[11].value) if len(row) > 11 else ""    # L열
        if not product and not option:
            continue
        g = groups.setdefault(product, {"count": 0, "options": []})
        g["count"] += 1
        g["options"].append(option)
    wb.close()

    drafts: list[dict] = []
    covered: list[dict] = []
    for product, info in sorted(groups.items(), key=lambda kv: -kv[1]["count"]):
        options = info["options"]
        n = len(options)
        sample_opt = next((o for o in options if o), "")
        existing = resolve_any(product, sample_opt) if _loaded else None

        kg_have = grade_have = count_have = 0
        kg_values: set[str] = set()
        grade_values: set[str] = set()
        for opt in options:
            kgs = _KG_RE.findall(opt)
            grs = _infer_grades_in(opt)
            if kgs:
                kg_have += 1
                kg_values.update(_fmt_kg(k) for k in kgs)
            if grs:
                grade_have += 1
                grade_values.update(grs)
            if _COUNT_RE.search(opt):
                count_have += 1

        keyword = _infer_keyword(product)
        use_grade = grade_have == n and bool(grade_values)
        use_kg = kg_have == n and bool(kg_values)
        use_count = (not use_kg) and count_have == n and count_have > 0

        parts = [keyword]
        if use_grade:
            parts.append("{grade}")
        if use_kg:
            parts.append("{kg}kg")
        if use_count:
            parts.append("{count}개")
        template = " ".join(p for p in parts if p).strip() or keyword

        warnings: list[str] = []
        if 0 < grade_have < n:
            warnings.append("일부 옵션만 등급 있음 — 확인 필요")
        if 0 < kg_have < n:
            warnings.append("일부 옵션만 kg 있음 — 확인 필요")
        if not use_grade and not use_kg and not use_count:
            warnings.append("등급·kg·수량 패턴 미검출 — 출력 템플릿 직접 확인")

        def _kg_sort(v: str) -> float:
            try:
                return float(v)
            except ValueError:
                return 0.0

        draft = {
            "supplier_key": supplier_key,
            "label": (product or keyword)[:100],
            "priority": 100,
            "name_keywords": [keyword] if keyword else [],
            "exclude_keywords": [],
            "grades": sorted(grade_values) if use_grade else [],
            "kg_allow": sorted(kg_values, key=_kg_sort) if use_kg else [],
            "pair_map": {},
            "extra_map": {},
            "output_template": template,
            "require_grade": bool(use_grade),
            "require_kg": bool(use_kg),
            "active": True,
            "notes": f"파일 자동초안 · 주문 {info['count']}건",
            # ── UI 표시용 메타 (create 엔드포인트는 무시) ──
            "order_count": info["count"],
            "sample_options": _distinct(options)[:4],
            "warnings": warnings,
            "already_matched": existing is not None,
            "existing_output": (existing or {}).get("output") or "",
        }
        (covered if existing is not None else drafts).append(draft)
        if len(drafts) + len(covered) >= max_products:
            break

    return {"drafts": drafts, "covered": covered, "product_count": len(groups)}


async def refresh_rules() -> dict:
    """DB에서 활성 규칙을 다시 읽어 캐시 교체. 반환: 상태 요약."""
    global _rules_by_supplier, _suppliers, _loaded, _refreshed_at
    from app import db as database

    suppliers = await database.list_rule_suppliers(include_inactive=False)
    rules = await database.list_product_rules(active_only=True)

    by_supplier: dict[str, list[dict]] = {}
    for rule in rules:
        by_supplier.setdefault(rule["supplier_key"], []).append(rule)
    for bucket in by_supplier.values():
        bucket.sort(key=lambda r: (int(r.get("priority") or 100), int(r.get("id") or 0)))

    with _lock:
        _suppliers = {s["key"]: s for s in suppliers}
        _rules_by_supplier = by_supplier
        _loaded = True
        _refreshed_at = datetime.now(KST).isoformat(timespec="seconds")
    logger.info(
        "규칙 엔진 캐시 갱신: 발주처 %d곳, 규칙 %d건", len(suppliers), len(rules)
    )
    return get_status()


def get_status() -> dict:
    return {
        "loaded": _loaded,
        "suppliers": len(_suppliers),
        "rules": sum(len(v) for v in _rules_by_supplier.values()),
        "refreshed_at": _refreshed_at,
    }


def get_supplier(supplier_key: str) -> dict | None:
    return _suppliers.get(supplier_key)


def _set_cache_for_test(rules_by_supplier: dict[str, list[dict]], suppliers: dict[str, dict] | None = None) -> None:
    """테스트 전용: DB 없이 캐시 주입."""
    global _rules_by_supplier, _suppliers, _loaded, _refreshed_at
    with _lock:
        _rules_by_supplier = {
            k: sorted(v, key=lambda r: (int(r.get("priority") or 100), int(r.get("id") or 0)))
            for k, v in rules_by_supplier.items()
        }
        _suppliers = suppliers or {}
        _loaded = True
        _refreshed_at = "test"
