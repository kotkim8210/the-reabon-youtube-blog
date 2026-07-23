"""백도 딱딱이복숭아 2·4kg 쥬얼리→제주다팜 이관 검증 (1kg은 쥬얼리 잔류)."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import kolrabi_order, myeongi_order
from app.processors.tracking_input import _semantic_option_keys


# ── DeliveryList 라우팅: 쥬얼리(myeongi)는 1kg만, 제주다팜(kolrabi)은 2·4kg만 ──
def test_myeongi_baekdo_only_1kg():
    assert myeongi_order.is_jewelry_baekdo_order("햇 백도 딱딱이복숭아", "1박스 중과 1kg") is True
    assert myeongi_order.is_jewelry_baekdo_order("햇 백도 딱딱이복숭아", "1박스 대과 1kg") is True
    assert myeongi_order.is_jewelry_baekdo_order("햇 백도 딱딱이복숭아", "1박스 중과 2kg") is False
    assert myeongi_order.is_jewelry_baekdo_order("햇 백도 딱딱이복숭아", "1박스 대과 4kg") is False


def test_myeongi_baekdo_option_1kg_only():
    # 1kg은 과수표기 포함 발주명
    assert myeongi_order._jewelry_baekdo_option("햇 백도 딱딱이복숭아", "1박스 중과 1kg") == "백도 딱딱이 복숭아 중과 1kg (5-6과 내외)"
    # 2·4kg은 제주다팜 이관 → None
    assert myeongi_order._jewelry_baekdo_option("햇 백도 딱딱이복숭아", "1박스 중과 2kg") is None
    assert myeongi_order._jewelry_baekdo_option("햇 백도 딱딱이복숭아", "1박스 대과 4kg") is None


def test_kolrabi_baekdo_only_2_4kg():
    assert kolrabi_order.is_jeju_baekdo_order("햇 백도 딱딱이복숭아", "1박스 중과 2kg") is True
    assert kolrabi_order.is_jeju_baekdo_order("햇 백도 딱딱이복숭아", "1박스 대과 4kg") is True
    assert kolrabi_order.is_jeju_baekdo_order("햇 백도 딱딱이복숭아", "1박스 중과 1kg") is False
    assert kolrabi_order.is_jeju_baekdo_order("콜라비 정품", "3kg") is False


def test_kolrabi_baekdo_convert_matches_screenshot():
    # 제주다팜 판매옵션명 그대로: '딱딱이 복숭아 {등급} {kg}kg'
    assert kolrabi_order.convert_jeju_baekdo_option("백도 딱딱이복숭아", "1박스 대과 4kg") == "딱딱이 복숭아 대과 4kg"
    assert kolrabi_order.convert_jeju_baekdo_option("백도 딱딱이복숭아", "1박스 대과 2kg") == "딱딱이 복숭아 대과 2kg"
    assert kolrabi_order.convert_jeju_baekdo_option("백도 딱딱이복숭아", "1박스 중과 4kg") == "딱딱이 복숭아 중과 4kg"
    assert kolrabi_order.convert_jeju_baekdo_option("백도 딱딱이복숭아", "1박스 중과 2kg") == "딱딱이 복숭아 중과 2kg"
    assert kolrabi_order.convert_jeju_baekdo_option("백도 딱딱이복숭아", "1박스 중과 1kg") is None


# ── 송장 의미키: DeliveryList(쿠팡 옵션)과 orderlist(제주다팜 발주명)이 등급+kg로 묶임 ──
def test_tracking_semantic_key_matches_both_sides():
    dl = _semantic_option_keys("햇 백도 딱딱이복숭아", "1박스 중과 2kg")
    ol = _semantic_option_keys("딱딱이 복숭아 중과 2kg", "")
    assert "baekdo:중과:2kg" in dl
    assert "baekdo:중과:2kg" in ol
    # 1kg(쥬얼리)과는 안 섞임
    assert "baekdo:중과:1kg" not in dl


def _delivery(rows: list[dict]) -> bytes:
    """(order_no, product, option, qty, name, phone, zipcode, address) DeliveryList 생성."""
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))     # C
        ws.cell(row, 11, r["product"])                     # K
        ws.cell(row, 12, r["option"])                      # L
        ws.cell(row, 23, r.get("qty", 1))                  # W
        ws.cell(row, 27, r.get("name", f"수취인{i}"))       # AA
        ws.cell(row, 28, r.get("phone", "010-0000-0000"))  # AB
        ws.cell(row, 29, "12345")                          # AC zipcode
        ws.cell(row, 30, r.get("address", "서울시 어딘가")) # AD
        ws.cell(row, 31, "")                               # AE memo
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_process_baekdo_end_to_end():
    data = _delivery([
        {"product": "햇 백도 딱딱이복숭아", "option": "1박스 대과 4kg"},
        {"product": "햇 백도 딱딱이복숭아", "option": "1박스 중과 2kg"},
        {"product": "햇 백도 딱딱이복숭아", "option": "1박스 중과 1kg"},  # 쥬얼리 잔류 → 제외
        {"product": "콜라비 정품", "option": "3kg"},                     # 백도 아님 → 제외
    ])
    result = kolrabi_order.process_baekdo(data)
    assert result is not None
    _bytes, filename, stats = result
    assert "백도딱딱이복숭아" in filename
    assert stats["total"] == 2  # 2·4kg만, 1kg·콜라비 제외
    names = {o["vendor_option_name"] for o in stats["options"]}
    assert names == {"딱딱이 복숭아 대과 4kg", "딱딱이 복숭아 중과 2kg"}


def test_process_potato_parity_unchanged():
    """리팩터(_build_jejudapam_order) 후에도 홍감자 발주는 그대로."""
    data = _delivery([{"product": "햇 홍감자", "option": "1박스 중 1kg"}])
    result = kolrabi_order.process_potato(data)
    assert result is not None
    _bytes, filename, stats = result
    assert "홍감자" in filename
    assert stats["total"] == 1
    # 중 1kg → 중 2kg (중량 업 매칭 유지)
    assert stats["options"][0]["vendor_option_name"] == "홍감자 중 2kg"
