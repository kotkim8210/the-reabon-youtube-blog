"""게걸무 운송장 입력: 파일 슬롯 오배치 방어 (2026-08-17 사고).

DeliveryList 칸에 택배발송 파일을 올리면 종전엔 그 파일을 그대로 되돌려줘
쿠팡 '엑셀 대량배송'이 인식하지 못했다. 이제 명확한 오류로 막는다.
"""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.processors import gaegeolmu_tracking as GT


def _tracking_file(rows: list[tuple[str, str]]) -> bytes:
    """게걸무 택배발송 파일 (A=주문자명, B=운송장번호, C=수하인명)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "택배발송"
    ws.append(["주문자명", "운송장번호", "수하인명"])
    for name, tracking in rows:
        ws.append([name, tracking, name])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


DELIVERY_HEADERS = [
    "번호", "묶음배송번호", "주문번호", "택배사", "운송장번호", "분리배송 Y/N", "분리배송 출고예정일",
    "주문시 출고예정일", "출고일(발송일)", "주문일", "등록상품명", "등록옵션명", "노출상품명(옵션명)",
    "노출상품ID", "옵션ID", "최초등록등록상품명/옵션명", "업체상품코드", "바코드", "결제액",
    "배송비구분", "배송비", "도서산간 추가배송비", "구매수(수량)", "옵션판매가(판매단가)", "구매자",
    "구매자전화번호", "수취인이름", "수취인전화번호", "우편번호", "수취인 주소", "배송메세지",
    "상품별 추가메시지", "주문자 추가메시지", "배송완료일", "구매확정일자", "개인통관번호(PCCC)",
    "통관용수취인전화번호", "기타", "결제위치", "배송유형",
]


def _delivery_file(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery"
    ws.append(DELIVERY_HEADERS)
    for r in rows:
        row = [""] * len(DELIVERY_HEADERS)
        row[2] = r.get("order_no", "OID")
        row[10] = "식품애착 게걸무씨앗기름 폐 기침 기관지"
        row[11] = r.get("option", "2개 180ml")
        row[26] = r["name"]
        row[29] = r.get("address", "서울시 강남구 1")
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_normal_case_fills_delivery_list():
    tracking = _tracking_file([("김현규", "257765144125"), ("이한철", "257765144136")])
    delivery = _delivery_file([{"name": "김현규"}, {"name": "이한철"}])

    output, filename, stats = GT.process(tracking, delivery)

    assert stats["filled"] == 2
    assert "운송장입력완료" in filename
    ws = load_workbook(BytesIO(output)).active
    # 결과는 DeliveryList 양식 그대로 (E열=운송장)
    assert ws.cell(1, 3).value == "주문번호" and ws.cell(1, 5).value == "운송장번호"
    assert [ws.cell(r, 5).value for r in (2, 3)] == ["257765144125", "257765144136"]


def test_delivery_slot_with_tracking_file_is_rejected():
    """칸을 바꿔 올리면(둘 다 택배발송 파일) 명확한 오류 — 쿠팡 미인식 파일 방지."""
    tracking = _tracking_file([("김현규", "257765144125")])
    with pytest.raises(ValueError) as exc:
        GT.process(tracking, tracking)
    message = str(exc.value)
    assert "DeliveryList" in message
    assert "택배발송" in message


def test_error_message_lists_detected_columns():
    tracking = _tracking_file([("김현규", "257765144125")])
    with pytest.raises(ValueError) as exc:
        GT.process(tracking, tracking)
    # 무엇이 올라왔는지 알려줘야 사용자가 바로 안다
    assert "주문자명" in str(exc.value) or "수하인명" in str(exc.value)
