"""청사과(아오리사과) 제주다팜 발주·운송장 (2026-08 신규) + 백도 제주다팜 시즌아웃."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import kolrabi_order
from app.processors.tracking_input import _semantic_option_keys


def test_convert_apple_option_maps_grade_kg_to_full_jeju_name():
    f = kolrabi_order.convert_apple_option
    assert f("청사과 아오리사과", "1박스 소과 3kg") == "청사과 소과(가정용) 포장재포함 3kg(15-22과내외)"
    assert f("아오리 청사과", "1박스 중소과 5kg") == "청사과 중소과(가정용) 포장재포함 5kg(20-25과내외)"
    assert f("청사과", "1박스 대과 2kg") == "청사과 대과(가정용) 포장재포함 2kg(6과내)"
    # 취급 안 하는 kg(1kg)·비청사과·등급 없음 → None
    assert f("청사과", "1박스 소과 1kg") is None
    assert f("햇 홍감자", "1박스 중 1kg") is None
    assert f("청사과", "1박스 3kg") is None


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))     # C 주문번호
        ws.cell(row, 11, r["product"])                    # K 상품명
        ws.cell(row, 12, r["option"])                     # L 옵션
        ws.cell(row, 23, r.get("qty", 1))                 # W 수량
        ws.cell(row, 27, r.get("name", f"수취인{i}"))      # AA 수취인
        ws.cell(row, 28, r.get("phone", "010-0000-0000")) # AB 전화
        ws.cell(row, 29, "12345")                         # AC 우편번호
        ws.cell(row, 30, r.get("address", "서울시 어딘가")) # AD 주소
        ws.cell(row, 31, "")                              # AE 메모
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_process_apple_outputs_jeju_order():
    data = _delivery([
        {"product": "청사과(아오리사과) 산지직송", "option": "1박스 소과 3kg", "name": "김사과"},
        {"product": "청사과(아오리사과) 산지직송", "option": "1박스 대과 5kg", "name": "이청사"},
        {"product": "콜라비 정품", "option": "3kg"},  # 청사과 아님 → 제외
    ])
    result = kolrabi_order.process_apple(data)
    assert result is not None
    _bytes, filename, stats = result
    assert "청사과" in filename
    assert stats["total"] == 2
    names = {o["vendor_option_name"] for o in stats["options"]}
    assert names == {
        "청사과 소과(가정용) 포장재포함 3kg(15-22과내외)",
        "청사과 대과(가정용) 포장재포함 5kg(12-16과내)",
    }
    # 발주서 H열(품목명) 확인
    ws = load_workbook(BytesIO(_bytes)).active
    products = [str(ws.cell(r, 8).value or "") for r in range(2, 4)]
    assert products[0].startswith("청사과 소과")
    assert products[1].startswith("청사과 대과")


def test_apple_appears_in_process_outputs():
    data = _delivery([{"product": "청사과 아오리", "option": "1박스 중소과 4kg", "name": "박사과"}])
    results = kolrabi_order.process_outputs(data)
    labels = {st.get("product") for _, _, st in results}
    assert "청사과(제주다팜)" in labels


def test_apple_tracking_semantic_key_matches_both_sides():
    # orderlist(발주명 = 제주다팜 판매옵션 전체) ↔ DeliveryList(쿠팡 옵션)
    ol = _semantic_option_keys("청사과 소과(가정용) 포장재포함 3kg(15-22과내외)", "")
    dl = _semantic_option_keys("청사과 아오리사과 산지직송", "1박스 소과 3kg")
    assert "apple:소과:3kg" in ol
    assert "apple:소과:3kg" in dl
    # 등급 다르면 안 섞임
    assert "apple:대과:3kg" not in ol
    # 중소과가 소과로 오분류되지 않음
    js = _semantic_option_keys("청사과 중소과(가정용) 포장재포함 5kg(20-25과내외)", "")
    assert "apple:중소과:5kg" in js
    assert "apple:소과:5kg" not in js


# ── 백도 제주다팜 시즌아웃 (2026-08-10 발주 마지막) ──
def test_baekdo_jeju_order_output_gated_by_season(monkeypatch):
    data = _delivery([
        {"product": "햇 백도 딱딱이복숭아", "option": "1박스 중과 2kg", "name": "백도씨"},
    ])
    # 시즌 중이면 백도 발주서 출력
    monkeypatch.setattr(kolrabi_order, "_BAEKDO_JEJU_LAST_ORDER_DATE", "2099-12-31")
    labels_open = {st.get("product") for _, _, st in kolrabi_order.process_outputs(data)}
    assert "백도딱딱이복숭아(제주다팜)" in labels_open

    # 시즌 종료(과거 컷오프)면 미출력 — 매칭 함수는 순수 유지(송장 계속 동작)
    monkeypatch.setattr(kolrabi_order, "_BAEKDO_JEJU_LAST_ORDER_DATE", "2000-01-01")
    labels_closed = {st.get("product") for _, _, st in kolrabi_order.process_outputs(data)}
    assert "백도딱딱이복숭아(제주다팜)" not in labels_closed
    # is_jeju_baekdo_order는 시즌과 무관(송장 의미키가 의존)
    assert kolrabi_order.is_jeju_baekdo_order("햇 백도 딱딱이복숭아", "1박스 중과 2kg") is True


# ── 마진방어(apple-jeju) 연동 ──
def test_apple_margin_monitor_matches_order_mapping():
    """마진방어 옵션명 = 발주명(제주다팜 판매옵션)이어야 공급가가 붙는다."""
    from app import supplier_price_monitor as spm

    config = spm.MONITOR_CONFIGS["apple-jeju"]
    assert config.template_path.exists()
    assert config.product_code == "10001098"
    # 8종(소과·대과 2·3·4·5kg), 행 8~15
    assert [o.row for o in config.options] == [8, 9, 10, 11, 12, 13, 14, 15]
    for o in config.options:
        assert spm.option_supplier_name(o, config) == "제주다팜"
        # 발주 변환 결과와 문자열이 정확히 같아야 함(공급가 매칭 키)
        converted = kolrabi_order.convert_apple_option(o.coupang_product, o.coupang_option)
        assert converted == o.supplier_option_name, (o.label, converted, o.supplier_option_name)


def test_apple_template_rows_align_with_options():
    """템플릿 E열(옵션)·C열(발주처)이 설정 행과 맞는지 — 행 밀림 방지."""
    from app import supplier_price_monitor as spm

    config = spm.MONITOR_CONFIGS["apple-jeju"]
    ws = load_workbook(config.template_path)["쥬얼리프룻"]
    expected = ["소과 2kg", "소과 3kg", "소과 4kg", "소과 5kg", "대과 2kg", "대과 3kg", "대과 4kg", "대과 5kg"]
    assert [str(ws.cell(r, 5).value or "").strip() for r in range(8, 16)] == expected
    assert {str(ws.cell(r, 3).value or "") for r in range(8, 16)} == {"제주다팜"}
    # 공급가(I열)·쿠폰가(H열)가 채워져 있어야 마진 계산 가능
    assert all(isinstance(ws.cell(r, 8).value, (int, float)) for r in range(8, 16))
    assert all(isinstance(ws.cell(r, 9).value, (int, float)) for r in range(8, 16))
