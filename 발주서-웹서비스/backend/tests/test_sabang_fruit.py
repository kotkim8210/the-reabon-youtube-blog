"""사방넷 연동 테스트: XML 빌드/파싱 + DeliveryList 합성 → 기존 파이프라인 통과."""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.sabang import client as sabang_client
from app.processors import sabang_fruit
from app.processors.myeongi_order import process as myeongi_process


SAMPLE_ORDER_RESPONSE = (
    '<?xml version="1.0" encoding="utf-8"?>\n'
    "<SABANG_ORDER_LIST>\n"
    "<DATA>"
    "<IDX>20260718-0000001</IDX>"
    "<ORDER_ID>ITSOFT123</ORDER_ID>"
    "<MALL_ID>coupang</MALL_ID>"
    "<RECEIVE_NAME>김테스트</RECEIVE_NAME>"
    "<RECEIVE_CEL>010-1111-2222</RECEIVE_CEL>"
    "<RECEIVE_ADDR>서울시 강남구 테스트로 1</RECEIVE_ADDR>"
    "<DELV_MSG>문앞에 놔주세요</DELV_MSG>"
    "<PRODUCT_NAME>고당도 햇 백도 딱딱이복숭아 국내산 산지직송</PRODUCT_NAME>"
    "<SKU_VALUE>2kg, 1박스, 중과</SKU_VALUE>"
    "<SALE_CNT>2</SALE_CNT>"
    "</DATA>\n"
    "<DATA>"
    "<IDX>20260718-0000002</IDX>"
    "<RECEIVE_NAME>이수취</RECEIVE_NAME>"
    "<RECEIVE_TEL>02-333-4444</RECEIVE_TEL>"
    "<RECEIVE_ADDR>부산시 해운대구 바다로 2</RECEIVE_ADDR>"
    "<PRODUCT_NAME>산지직송 명이나물 장아찌</PRODUCT_NAME>"
    "<SKU_VALUE>500g 1박스</SKU_VALUE>"
    "<SALE_CNT>1</SALE_CNT>"
    "</DATA>\n"
    "</SABANG_ORDER_LIST>"
)


def test_build_order_list_xml_euckr(monkeypatch):
    monkeypatch.setattr(sabang_client.config, "SABANG_COMPANY_ID", "itsoft")
    monkeypatch.setattr(sabang_client.config, "SABANG_AUTH_KEY", "test-key")
    xml = sabang_client.build_order_list_xml("20260701", "20260718", "001")
    text = xml.decode("euc-kr")
    assert "<SEND_COMPAYNY_ID>itsoft</SEND_COMPAYNY_ID>" in text
    assert "<ORD_ST_DATE>20260701</ORD_ST_DATE>" in text
    assert "<ORDER_STATUS>001</ORDER_STATUS>" in text
    assert "IDX|ORDER_ID" in text


def test_build_invoice_xml():
    xml = sabang_client.build_invoice_xml(
        [{"idx": "20260718-0000001", "tak_code": "0019", "invoice": "255278881234"}]
    )
    text = xml.decode("euc-kr")
    assert "<SABANG_INV_REGI>" in text
    assert "<SABANGNET_IDX><![CDATA[20260718-0000001]]></SABANGNET_IDX>" in text
    assert "<TAK_INVOICE><![CDATA[255278881234]]></TAK_INVOICE>" in text
    assert "<SEND_INV_EDIT_YN>N</SEND_INV_EDIT_YN>" in text


def test_parse_order_response():
    orders = sabang_client.parse_order_response(SAMPLE_ORDER_RESPONSE)
    assert len(orders) == 2
    assert orders[0]["idx"] == "20260718-0000001"
    assert orders[0]["receive_name"] == "김테스트"
    assert orders[0]["sale_cnt"] == "2"


def test_parse_order_response_non_xml_raises():
    with pytest.raises(sabang_client.SabangApiError):
        sabang_client.parse_order_response("인증키가 올바르지 않습니다")


def test_xml_store_roundtrip():
    token = sabang_client.put_request_xml(b"<xml/>")
    assert sabang_client.get_request_xml(token) == b"<xml/>"
    assert sabang_client.get_request_xml("no-such-token") is None


def test_orders_to_delivery_xlsx_columns():
    orders = sabang_client.parse_order_response(SAMPLE_ORDER_RESPONSE)
    xlsx = sabang_fruit.orders_to_delivery_xlsx(orders)
    ws = load_workbook(filename=BytesIO(xlsx)).active
    rows = list(ws.iter_rows(min_row=2))
    assert len(rows) == 2
    row = rows[0]
    assert row[2].value == "20260718-0000001"          # C 주문번호 = 사방넷 IDX
    assert "백도 딱딱이복숭아" in row[10].value          # K 상품명
    assert row[11].value == "2kg, 1박스, 중과"          # L 옵션
    assert str(row[22].value) == "2"                    # W 수량
    assert row[26].value == "김테스트"                   # AA 수취인
    assert row[27].value == "010-1111-2222"             # AB 전화
    assert row[29].value == "서울시 강남구 테스트로 1"    # AD 주소
    assert row[30].value == "문앞에 놔주세요"            # AE 메모


def test_sabang_orders_flow_into_myeongi_process():
    """사방넷 주문 → 합성 DeliveryList → 기존 쥬얼리 발주서 파이프라인 e2e."""
    orders = sabang_client.parse_order_response(SAMPLE_ORDER_RESPONSE)
    xlsx = sabang_fruit.orders_to_delivery_xlsx(orders)
    output_bytes, filename, stats = myeongi_process(xlsx)
    ws = load_workbook(filename=BytesIO(output_bytes)).active

    found_baekdo = found_myeongi = False
    idx_in_order_no = False
    for r in range(2, ws.max_row + 1):
        product = str(ws.cell(row=r, column=7).value or "")
        order_no = str(ws.cell(row=r, column=13).value or "")
        if "백도" in product:
            found_baekdo = True
            if order_no == "20260718-0000001":
                idx_in_order_no = True
        if "명이나물" in product:
            found_myeongi = True
    assert found_baekdo, "백도 주문이 발주서에 없음"
    assert found_myeongi, "명이나물 주문이 발주서에 없음"
    assert idx_in_order_no, "사방넷 IDX가 발주서 M열 주문번호로 전달되지 않음"


def _make_orderlist(rows: list[tuple[str, str, str]]) -> bytes:
    """(수취인 C, 주문번호 D, 운송장 R) 형태의 회신 orderlist 생성."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, 20):
        ws.cell(row=1, column=col, value=f"h{col}")
    for i, (name, order_no, tracking) in enumerate(rows):
        r = i + 2
        for col in range(1, 20):
            ws.cell(row=r, column=col, value="")
        ws.cell(row=r, column=3, value=name)       # C
        ws.cell(row=r, column=4, value=order_no)   # D
        ws.cell(row=r, column=18, value=tracking)  # R
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_summarize_courier_codes_guesses_lotte():
    # 255278883171: 앞11자리(25527888317) % 7 == 1 == 마지막 자리 → 롯데 체크섬 통과
    lotte_valid = "255278883171"
    assert int(lotte_valid[:11]) % 7 == int(lotte_valid[11])
    orders = [
        {"delivery_id": "0019", "invoice_no": lotte_valid},
        {"delivery_id": "0019", "invoice_no": lotte_valid},
        {"delivery_id": "0019", "invoice_no": lotte_valid},
        {"delivery_id": "0004", "invoice_no": "1234567890"},  # 10자리 — 롯데 아님
        {"delivery_id": "0004", "invoice_no": "9876543210"},
        {"delivery_id": "", "invoice_no": "111"},  # 코드 없음 → 제외
    ]
    codes = sabang_fruit.summarize_courier_codes(orders)
    assert [c["code"] for c in codes] == ["0019", "0004"]  # 건수 많은 순
    assert codes[0]["guess"] == "롯데택배(추정)"
    assert codes[0]["count"] == 3
    assert codes[1]["guess"] == ""


def test_summarize_courier_codes_single_sample_no_guess():
    # 표본 1건이면 추정하지 않음
    codes = sabang_fruit.summarize_courier_codes(
        [{"delivery_id": "0019", "invoice_no": "255278883171"}]
    )
    assert codes[0]["guess"] == ""


def test_parse_orderlist_for_sabang():
    ol = _make_orderlist([
        ("김테스트", "20260718-0000001", "2552 7888 1234"),
        ("이수취", "20260718-0000002", "255278885678"),
        ("송장없음", "20260718-0000003", ""),
    ])
    entries = sabang_fruit.parse_orderlist_for_sabang(ol)
    assert len(entries) == 2
    assert entries[0] == {"idx": "20260718-0000001", "tracking": "255278881234", "name": "김테스트"}
    assert entries[1]["tracking"] == "255278885678"
