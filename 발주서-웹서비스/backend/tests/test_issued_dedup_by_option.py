"""중복발주 방지 키: 주문번호 단독 → '주문번호|옵션' 복합키 (2026-08-17 김희조 사고).

한 주문번호로 여러 옵션을 사면(쿠팡 가능) 주문번호만으로는 구분이 안 돼,
어제 발주한 옵션 때문에 오늘 새로 산 다른 옵션까지 통째로 빠졌다.
과거에 주문번호 단독으로 기록된 이력은 그대로 존중해야 중복발주가 안 난다.
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import issued_orders


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r["order_no"])                      # C 주문번호
        ws.cell(row, 11, r.get("product", "청사과"))          # K 상품명
        ws.cell(row, 12, r.get("option", ""))                # L 옵션
        ws.cell(row, 27, r.get("name", f"수취인{i}"))         # AA 수취인
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _names_and_options(data: bytes) -> list[tuple]:
    ws = load_workbook(BytesIO(data)).active
    return [
        (ws.cell(r, 27).value, ws.cell(r, 12).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 3).value
    ]


def test_same_order_number_different_option_is_kept():
    """김희조 사고 재현: 같은 주문번호 2건(소과 5kg/대과 5kg) 중 발주된 옵션만 제외."""
    data = _delivery([
        {"order_no": "11102305997486", "option": "1박스 가정용 소과 5kg", "name": "김희조"},
        {"order_no": "11102305997486", "option": "1박스 가정용 대과 5kg", "name": "김희조"},
        {"order_no": "3102305763339", "option": "1박스 가정용 소과 2kg", "name": "공혜정"},
    ])
    # 어제 소과 5kg만 발주됨 (복합키로 기록)
    issued = {issued_orders.make_order_key("11102305997486", "1박스 가정용 소과 5kg")}

    names = []
    filtered, removed = issued_orders.filter_delivery_by_issued(data, issued, names)

    assert removed == 1
    assert names == ["김희조"]
    assert _names_and_options(filtered) == [
        ("김희조", "1박스 가정용 대과 5kg"),   # 새로 산 옵션은 살아남는다
        ("공혜정", "1박스 가정용 소과 2kg"),
    ]


def test_legacy_order_id_only_record_still_excludes_whole_order():
    """과거 기록(주문번호 단독)은 종전대로 그 주문번호 전체 제외 — 중복발주 방지."""
    data = _delivery([
        {"order_no": "11102305997486", "option": "1박스 가정용 소과 5kg", "name": "김희조"},
        {"order_no": "11102305997486", "option": "1박스 가정용 대과 5kg", "name": "김희조"},
    ])
    filtered, removed = issued_orders.filter_delivery_by_issued(data, {"11102305997486"}, None)
    assert removed == 2
    assert _names_and_options(filtered) == []


def test_option_whitespace_differences_still_match():
    """기록·필터의 공백 표기가 달라도 같은 옵션으로 인식(프로세서별 정규화 차이 흡수)."""
    data = _delivery([{"order_no": "A1", "option": "1박스  가정용 소과 5kg", "name": "홍길동"}])
    issued = {issued_orders.make_order_key("A1", "1박스 가정용 소과 5kg")}
    _filtered, removed = issued_orders.filter_delivery_by_issued(data, issued, None)
    assert removed == 1


def test_order_keys_from_stats_include_option():
    stats = {
        "options": [
            {
                "coupang_option_keyword": "1박스 가정용 소과 5kg",
                "vendor_option_name": "청사과 소과(가정용) 포장재포함 5kg(25-40과내외)",
                "orders": [{"order_id": "11102305997486", "quantity": 1}],
            },
            {
                "coupang_option_keyword": "1박스 가정용 대과 5kg",
                "vendor_option_name": "청사과 대과(가정용) 포장재포함 5kg(12-16과내)",
                "orders": [{"order_id": "11102305997486", "quantity": 1}],
            },
        ]
    }
    keys = issued_orders.order_ids_from_stats(stats)
    assert keys == [
        "11102305997486|1박스가정용대과5kg",
        "11102305997486|1박스가정용소과5kg",
    ]


def test_toss_entries_filtered_by_option_too():
    entries = [
        {"order_id": "T1", "option": "1박스 중과 2kg", "name": "백도씨", "product": "딱딱이 복숭아 중과 2kg"},
        {"order_id": "T1", "option": "1박스 대과 4kg", "name": "백도씨", "product": "딱딱이 복숭아 대과 4kg"},
    ]
    issued = {issued_orders.make_order_key("T1", "1박스 중과 2kg")}
    kept, removed = issued_orders.filter_entries_by_issued(entries, issued, None)
    assert removed == 1
    assert [e["option"] for e in kept] == ["1박스 대과 4kg"]

    # 과거 형태(주문번호 단독)면 둘 다 제외
    kept2, removed2 = issued_orders.filter_entries_by_issued(entries, {"T1"}, None)
    assert removed2 == 2 and kept2 == []
