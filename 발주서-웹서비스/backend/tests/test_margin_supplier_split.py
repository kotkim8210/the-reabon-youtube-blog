"""마진방어: 발주처가 옵션별로 갈리는 상품을 한 파일로 뽑고, 발주 라우팅과 발주처가 일치하는지 검증.

배경(2026-07-27): 미니밤호박은 1kg(제주다팜)·3~10kg(쥬얼리)이 따로 다운로드돼 한 장에서 비교가
안 됐고, 백도 딱딱이복숭아는 발주만 제주다팜으로 이관되고 마진방어는 쥬얼리 공급가 그대로였다.
"""

import asyncio
from io import BytesIO

from openpyxl import load_workbook

from app import supplier_price_monitor as spm
from app.order_routing import JEJUDAPAM, JEWELRYFRUIT, resolve_order_supplier


# ── 발주 라우팅 판정 (발주 프로세서 재사용) ──
def test_resolve_order_supplier_split_products():
    # 미니밤호박: 1kg 제주다팜 / 3·5·10kg 쥬얼리
    assert resolve_order_supplier("제주 미니밤호박 보우짱", "1박스 로얄 정품 1kg") == JEJUDAPAM
    assert resolve_order_supplier("제주 미니밤호박 보우짱", "1박스 로얄 정품 3kg") == JEWELRYFRUIT
    assert resolve_order_supplier("제주 미니밤호박 보우짱", "1박스 로얄 정품 10kg") == JEWELRYFRUIT
    # 백도: 1kg 쥬얼리 / 2·4kg 제주다팜
    assert resolve_order_supplier("햇 백도 딱딱이복숭아", "1박스 중과 1kg") == JEWELRYFRUIT
    assert resolve_order_supplier("햇 백도 딱딱이복숭아", "1박스 중과 2kg") == JEJUDAPAM
    assert resolve_order_supplier("햇 백도 딱딱이복숭아", "1박스 대과 4kg") == JEJUDAPAM
    # 홍감자는 제주다팜 이관분
    assert resolve_order_supplier("햇 홍감자", "1박스 중 1kg") == JEJUDAPAM
    # 대상 아님
    assert resolve_order_supplier("콜라비 정품", "3kg") is None


# ── 연동 가드: 마진방어 옵션의 발주처 == 발주 라우팅 결과 ──
def test_margin_options_match_order_routing():
    """발주처가 바뀌면(발주 코드 수정) 이 테스트가 깨져 마진방어 설정도 같이 고치게 된다."""
    checked = 0
    for config in spm.MONITOR_CONFIGS.values():
        for option in config.options:
            if not option.coupang_product or not option.coupang_option:
                continue
            routed = resolve_order_supplier(option.coupang_product, option.coupang_option)
            assert routed, f"{config.key}/{option.label}: 발주 라우팅이 판정하지 못함"
            assert routed == spm.option_supplier_name(option, config), (
                f"{config.key}/{option.label}: 발주는 {routed}인데 마진방어는 "
                f"{spm.option_supplier_name(option, config)}"
            )
            checked += 1
    assert checked >= 10  # 밤호박 4 + 백도 6


# ── 통합 파일 구성 ──
def test_bamhobak_monitor_is_single_merged_file():
    config = spm.MONITOR_CONFIGS["bamhobak-jewelry"]
    assert "bamhobak-jeju" not in spm.MONITOR_CONFIGS  # 제주다팜 단독 모니터 제거(통합)
    assert config.template_path.exists()
    labels = [(o.row, spm.option_supplier_name(o, config), o.supplier_option_name) for o in config.options]
    assert labels == [
        (8, "제주다팜", "로얄과 1kg"),
        (9, "쥬얼리프룻", "로얄과 3kg"),
        (10, "쥬얼리프룻", "로얄과 5kg"),
        (11, "쥬얼리프룻", "로얄과 10kg"),
    ]
    # 템플릿 C열(상품명 자리)에 발주처가 찍혀 있어야 한 장에서 구분된다
    ws = load_workbook(config.template_path)["쥬얼리프룻"]
    assert [ws.cell(r, 3).value for r in range(8, 12)] == ["제주다팜", "쥬얼리프룻", "쥬얼리프룻", "쥬얼리프룻"]


def test_baekdo_monitor_splits_supplier_by_kg():
    config = spm.MONITOR_CONFIGS["baekdo-jewelry"]
    by_row = {o.row: spm.option_supplier_name(o, config) for o in config.options}
    assert by_row == {8: "쥬얼리프룻", 9: "제주다팜", 10: "제주다팜", 11: "쥬얼리프룻", 12: "제주다팜", 13: "제주다팜"}
    # 제주다팜 옵션명은 제주다팜 popup 표기('딱딱이 복숭아 {등급} {kg}kg')여야 매칭된다
    jeju = [o.supplier_option_name for o in config.options if o.source]
    assert jeju == [
        "딱딱이 복숭아 중과 2kg",
        "딱딱이 복숭아 중과 4kg",
        "딱딱이 복숭아 대과 2kg",
        "딱딱이 복숭아 대과 4kg",
    ]
    ws = load_workbook(config.template_path)["쥬얼리프룻"]
    assert [ws.cell(r, 3).value for r in range(8, 14)] == [
        "쥬얼리프룻", "제주다팜", "제주다팜", "쥬얼리프룻", "제주다팜", "제주다팜",
    ]


def test_supplier_source_as_config_overrides_only_given_fields():
    base = spm.MONITOR_CONFIGS["baekdo-jewelry"]
    derived = spm.JEJU_BAEKDO_SOURCE.as_config(base)
    assert derived.supplier_name == "제주다팜"
    assert derived.base_url == "https://kkangta55.adminplus.co.kr"
    assert derived.product_code == "10001059"
    assert derived.template_path == base.template_path  # 출력 양식은 그대로
    assert derived.key == base.key


# ── 실제 실행: 두 발주처 공급가가 한 워크북에 들어간다 ──
def test_run_monitor_writes_both_suppliers_into_one_workbook(monkeypatch):
    jeju_prices = {"로얄과 1kg": 5200}
    jewelry_prices = {"로얄과 3kg": 8100, "로얄과 5kg": 11700, "로얄과 10kg": 19300}

    async def fake_collect(config):
        return (jeju_prices if config.supplier_name == "제주다팜" else jewelry_prices), {}

    monkeypatch.setattr(spm, "_collect_supplier_prices", fake_collect)

    from app import db as database

    async def no_snapshots(*_args, **_kwargs):
        return []

    async def noop(*_args, **_kwargs):
        return None

    monkeypatch.setattr(database, "latest_supplier_price_snapshots_before_run_date", no_snapshots)
    monkeypatch.setattr(database, "save_supplier_price_snapshot", noop)
    monkeypatch.setattr(database, "save_supplier_price_monitor_run", noop)

    summary, output_bytes, filename = asyncio.run(spm.run_supplier_monitor("bamhobak-jewelry"))

    assert summary["total_items"] == 4
    assert [row["supplier_name"] for row in summary["rows"]] == [
        "제주다팜", "쥬얼리프룻", "쥬얼리프룻", "쥬얼리프룻",
    ]
    assert "미니밤호박" in filename
    ws = load_workbook(BytesIO(output_bytes))["쥬얼리프룻"]
    assert [ws.cell(r, 9).value for r in range(8, 12)] == [5200, 8100, 11700, 19300]
