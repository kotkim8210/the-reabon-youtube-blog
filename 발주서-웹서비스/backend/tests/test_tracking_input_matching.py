"""tracking_input 매칭 — 전화 최우선·동일인 다건·동명이인·침묵 skip 금지 (2026-07-14 사고 회귀 방지)."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors.tracking_input import process


def _orderlist(rows):
    """rows: [(name, phone, address, courier, tracking)] → orderlist bytes.
    열: K=이름, M=전화, O=주소, Q=택배사, R=운송장 (옵션 열은 비움 — 제주다팜 회신에 옵션 없던 케이스)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["헤더"] * 18)
    for name, phone, addr, courier, tracking in rows:
        row = [""] * 18
        row[10], row[12], row[14], row[16], row[17] = name, phone, addr, courier, tracking
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _delivery(rows):
    """rows: [(name, phone, address, product, option)] → DeliveryList bytes.
    열: AA=이름, AB=전화, AD=주소, K=상품명, L=옵션, E=운송장(비움)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["헤더"] * 31)
    for name, phone, addr, product, option in rows:
        row = [""] * 31
        row[26], row[27], row[29], row[10], row[11] = name, phone, addr, product, option
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _e_column(output_bytes):
    ws = load_workbook(BytesIO(output_bytes)).active
    return [ws.cell(r, 5).value for r in range(2, ws.max_row + 1)]


def test_same_person_multiple_orders_all_filled():
    """동일인 N건(전화 동일, 회신에 옵션정보 없음) → N건 전부 순차 입력.
    옵션 가드가 전화 매칭보다 먼저 걸러 전부 skip되던 2026-07-14 사고의 회귀 테스트."""
    ol = _orderlist([
        ("이현경", "010-1111-2222", "서울 A", "롯데", "255200000001"),
        ("이현경", "010-1111-2222", "서울 A", "롯데", "255200000002"),
        ("이현경", "010-1111-2222", "서울 A", "롯데", "255200000003"),
    ])
    dl = _delivery([
        ("이현경", "01011112222", "서울A", "콜라비", "3kg 1박스"),   # 하이픈 유무 차이
        ("이현경", "01011112222", "서울A", "콜라비", "3kg 1박스"),
        ("이현경", "01011112222", "서울A", "콜라비", "3kg 1박스"),
    ])
    out, _, stats = process(ol, dl)
    assert stats["filled"] == 3 and stats["skipped"] == 0, stats
    assert set(_e_column(out)) == {"255200000001", "255200000002", "255200000003"}


def test_namesake_split_by_phone():
    """동명이인은 전화번호로 정확 분리 — 운송장 뒤바뀜 금지."""
    ol = _orderlist([
        ("김철수", "010-1111-0001", "서울", "롯데", "255200000011"),
        ("김철수", "010-2222-0002", "부산", "롯데", "255200000022"),
    ])
    dl = _delivery([
        ("김철수", "01022220002", "부산", "콜라비", "5kg"),
        ("김철수", "01011110001", "서울", "콜라비", "3kg"),
    ])
    out, _, stats = process(ol, dl)
    assert stats["filled"] == 2, stats
    assert _e_column(out) == ["255200000022", "255200000011"]


def test_namesake_unresolvable_skips_with_names():
    """전화·주소·옵션 어느 것으로도 못 가르면 오배정 대신 미입력 + 이름 표시(침묵 금지)."""
    ol = _orderlist([
        ("박영희", "010-3333-0001", "대구", "롯데", "255200000031"),
        ("박영희", "010-3333-0002", "광주", "롯데", "255200000032"),
    ])
    dl = _delivery([
        ("박영희", "01099999999", "인천", "콜라비", "3kg"),
        ("박영희", "01088888888", "울산", "콜라비", "3kg"),
    ])
    _, _, stats = process(ol, dl)
    assert stats["skipped"] == 2, stats
    assert "박영희" in stats.get("skipped_names", ""), stats
