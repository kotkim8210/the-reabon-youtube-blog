"""어드민 규칙만으로 신규 상품이 쥬얼리 발주서·송장에 반영되는지 (코드 배포 불필요).

배경(2026-08-06): 사용자가 새 상품(피자두 등)이 생길 때마다 코드 수정을 기다리지 않고
어드민에서 파일 업로드(초안 자동생성) 또는 수동 등록으로 처리할 수 있어야 한다는 요구.
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app import rules_engine
from app.processors import myeongi_order
from app.processors.myeongi_tracking import is_jewelryfruit_tracking_target

NEW_PRODUCT = "제주 하우스 애플망고 고당도 산지직송"

_RULE = {
    "id": 1,
    "supplier_key": "jewelryfruit",
    "label": "애플망고",
    "priority": 100,
    "name_keywords": ["애플망고"],
    "exclude_keywords": [],
    "grades": ["특", "대"],
    "kg_allow": ["1", "2"],
    "pair_map": {},
    "extra_map": {},
    "output_template": "애플망고 {grade} {kg}kg",
    "require_grade": True,
    "require_kg": True,
    "active": True,
}


def setup_function(_fn):
    rules_engine._set_cache_for_test(
        {"jewelryfruit": [_RULE]},
        {"jewelryfruit": {"key": "jewelryfruit", "name": "쥬얼리프룻"}},
    )


def teardown_function(_fn):
    rules_engine._set_cache_for_test({}, {})


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))
        ws.cell(row, 11, r["product"])
        ws.cell(row, 12, r["option"])
        ws.cell(row, 23, r.get("qty", 1))
        ws.cell(row, 27, r.get("name", f"수취인{i}"))
        ws.cell(row, 28, "010-0000-0000")
        ws.cell(row, 29, "12345")
        ws.cell(row, 30, "서울시 어딘가")
        ws.cell(row, 31, "")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_rule_only_product_appears_in_jewelry_order():
    data = _delivery([
        {"product": NEW_PRODUCT, "option": "1박스 특 2kg", "name": "김규칙"},
        {"product": "콜라비 정품", "option": "3kg"},  # 규칙·하드코딩 모두 미매칭 → 제외
    ])
    output, filename, stats = myeongi_order.process(data)

    assert stats["total"] == 1
    assert stats["options"][0]["vendor_option_name"] == "애플망고 특 2kg"
    assert "애플망고" in filename  # 규칙 라벨이 파일명에 반영
    ws = load_workbook(BytesIO(output)).active
    assert (ws.cell(2, 3).value, ws.cell(2, 7).value) == ("김규칙", "애플망고 특 2kg")


def test_rule_product_is_tracking_target_and_converts():
    assert is_jewelryfruit_tracking_target(NEW_PRODUCT, "1박스 특 2kg") is True
    assert myeongi_order.convert_option("1박스 대 1kg", NEW_PRODUCT) == "애플망고 대 1kg"
    # 규칙 kg_allow(1·2kg) 밖은 미매칭 → 발주 대상 아님
    assert is_jewelryfruit_tracking_target(NEW_PRODUCT, "1박스 특 5kg") is False


def test_hardcoded_rules_win_over_admin_rule():
    """하드코딩 분기가 있는 상품은 규칙이 있어도 기존 동작 유지(안전)."""
    rules_engine._set_cache_for_test(
        {
            "jewelryfruit": [
                {**_RULE, "id": 2, "label": "피자두덮어쓰기", "name_keywords": ["피자두"],
                 "grades": [], "kg_allow": [], "require_grade": False, "require_kg": False,
                 "output_template": "잘못된 출력"},
            ]
        },
        {"jewelryfruit": {"key": "jewelryfruit", "name": "쥬얼리프룻"}},
    )
    assert myeongi_order.convert_option("1박스 3kg", "국내산 김천 피자두 특품") == "피자두 3kg"


def test_engine_not_loaded_is_safe():
    """캐시 미로드(엔진 꺼짐) 상태에서도 기존 발주는 그대로 동작."""
    rules_engine._rules_by_supplier = {}
    rules_engine._loaded = False
    try:
        assert myeongi_order.jewelry_rule_match(NEW_PRODUCT, "1박스 특 2kg") is None
        assert myeongi_order.convert_option("1박스 3kg", "국내산 김천 피자두 특품") == "피자두 3kg"
    finally:
        rules_engine._set_cache_for_test({}, {})
