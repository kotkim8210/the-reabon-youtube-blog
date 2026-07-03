# -*- coding: utf-8 -*-
"""회귀 테스트: 당근마켓 짧은 주문 텍스트 파싱.

PYTHONPATH=backend python tests/test_daangn_order.py
"""
from io import BytesIO

from openpyxl import load_workbook

from app.processors.daangn_order import parse_daangn_orders, process


SHORT_WATERMELON_TEXT = """수박 8kg 1개
길현석
010 5330 7059
경기도 시흥시 대은로62 대우3차 아파트 309동1403호"""


def test_short_watermelon_text_maps_to_jewelry_8_9kg():
    entries = parse_daangn_orders(SHORT_WATERMELON_TEXT)
    assert len(entries) == 1
    entry = entries[0]
    assert entry["product"] == "가정용 수박 8~9kg 내외"
    assert entry["qty"] == "1"
    assert entry["name"] == "길현석"
    assert entry["phone"] == "010-5330-7059"
    assert entry["address"] == "경기도 시흥시 대은로62 대우3차 아파트 309동1403호"


def test_short_watermelon_text_generates_valid_workbook():
    output = process(SHORT_WATERMELON_TEXT)
    assert output is not None
    output_bytes, filename, stats = output
    assert filename.startswith("당근_쥬얼리프룻_발주(")
    assert stats["options"][0]["vendor_option_name"] == "가정용 수박 8~9kg 내외"

    workbook = load_workbook(BytesIO(output_bytes), data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    assert worksheet.cell(2, 3).value == "길현석"
    assert worksheet.cell(2, 4).value == "010-5330-7059"
    assert worksheet.cell(2, 6).value == "경기도 시흥시 대은로62 대우3차 아파트 309동1403호"
    assert worksheet.cell(2, 7).value == "가정용 수박 8~9kg 내외"
    assert worksheet.cell(2, 8).value == "1"


if __name__ == "__main__":
    test_short_watermelon_text_maps_to_jewelry_8_9kg()
    test_short_watermelon_text_generates_valid_workbook()
    print("ALL TESTS PASSED")
