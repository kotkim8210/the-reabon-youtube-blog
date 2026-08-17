"""게걸무: 쿠팡 DeliveryList + 지마켓 신규주문 합치기 (2026-08-17 요청).

지마켓 파일을 따로 올리면 쿠팡 발주서 뒤에 이어붙여 한 장으로 나온다.
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import gaegeolmu_order as G

GMARKET_HEADERS = [
    "판매아이디", "주문일자(결제확인전)", "주문번호", "구매자명", "구매자아이디", "상품번호",
    "상품명", "수령인명", "발송마감일", "발송정책", "선물주문여부", "선물주문상태", "선물수락일시",
    "선물수락기한", "설치주문여부", "설치예정일", "첫구매여부", "수량", "옵션", "추가구성", "사은품",
    "사은품 관리코드", "덤", "덤 관리코드", "판매단가", "판매금액", "판매자관리코드",
    "판매자상세관리코드", "구매자 휴대폰", "구매자 전화번호", "수령인 휴대폰", "수령인 전화번호",
    "수령인 통관정보", "배송지변경 여부", "우편번호", "주소", "배송시 요구사항", "배송비 결제방법",
    "배송비 금액", "배송번호", "SKU번호 및 수량", "배송점포", "판매방식", "주문종류",
    "장바구니번호(결제번호)",
]


def _gmarket(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(GMARKET_HEADERS)
    for r in rows:
        row = [""] * len(GMARKET_HEADERS)
        row[0] = "지마켓(rjs006)"
        row[2] = r["order_no"]
        row[3] = r.get("buyer", r["name"])
        row[6] = r.get("product", "예천 게걸무씨앗기름 나복자유 오일 100")
        row[7] = r["name"]
        row[17] = r.get("qty", 1)
        row[18] = r.get("option", "")
        row[28] = r.get("phone", "010-0000-0000")
        row[30] = r.get("phone", "010-0000-0000")
        row[34] = r.get("zip", "31948")
        row[35] = r.get("address", "충청남도 서산시 어딘가")
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coupang(rows: list[dict]) -> bytes:
    """쿠팡 DeliveryList (C=주문번호, K=상품명, L=옵션, W=수량, AA=수취인)."""
    wb = Workbook()
    ws = wb.active
    ws.append(G.DELIVERY_HEADERS)
    for i, r in enumerate(rows):
        row = [""] * len(G.DELIVERY_HEADERS)
        row[2] = r["order_no"]
        row[10] = r.get("product", "식품애착 게걸무씨앗기름 폐 기침 기관지")
        row[11] = r.get("option", "1병")
        row[22] = r.get("qty", 1)
        row[26] = r["name"]
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_coupang_only_unchanged():
    data = _coupang([{"order_no": "C1", "name": "쿠팡손님"}])
    _out, filename, stats = G.process(data)
    assert stats["total"] == 1
    assert stats["source"] == "쿠팡"
    assert "게걸무씨앗기름" in filename


def test_gmarket_only_still_works_in_first_slot():
    """지마켓 파일만 있을 때 첫 칸에 올려도 종전처럼 처리된다."""
    data = _gmarket([{"order_no": "4479627418", "name": "송옥순"}])
    _out, _filename, stats = G.process(data)
    assert stats["total"] == 1
    assert stats["source"] == "지마켓"


def test_merge_coupang_and_gmarket_into_one_sheet():
    coupang = _coupang([
        {"order_no": "C1", "name": "쿠팡손님1"},
        {"order_no": "C2", "name": "쿠팡손님2", "option": "2병"},
    ])
    gmarket = _gmarket([
        {"order_no": "4479627418", "name": "송옥순"},
        {"order_no": "4479627419", "name": "지마켓손님", "option": "1+1 2병"},
    ])

    output, _filename, stats = G.process(coupang, gmarket)

    assert stats["total"] == 4
    assert stats["coupang"] == 2 and stats["gmarket"] == 2
    assert stats["source"] == "쿠팡+지마켓"

    ws = load_workbook(BytesIO(output)).active
    names = [ws.cell(r, 27).value for r in range(2, ws.max_row + 1)]
    assert names == ["쿠팡손님1", "쿠팡손님2", "송옥순", "지마켓손님"]
    # 지마켓 행도 쿠팡 양식 컬럼(주문번호 C, 수취인 AA)에 들어간다
    assert ws.cell(4, 3).value == "4479627418"
    # 번호(A열)가 이어진다
    assert ws.cell(4, 1).value == 3 and ws.cell(5, 1).value == 4

    # 옵션 집계는 1병/2병으로 합산 (쿠팡 1병+지마켓 1병, 쿠팡 2병+지마켓 2병)
    totals = {o["coupang_option_keyword"]: o["quantity"] for o in stats["options"]}
    assert totals == {"게걸무씨앗기름 1병": 2, "게걸무씨앗기름 2병": 2}


def test_non_gmarket_file_in_gmarket_slot_is_rejected():
    coupang = _coupang([{"order_no": "C1", "name": "쿠팡손님"}])
    try:
        G.process(coupang, coupang)  # 지마켓 칸에 쿠팡 파일
    except ValueError as exc:
        assert "지마켓" in str(exc)
    else:
        raise AssertionError("지마켓 형식이 아닌 파일을 걸러야 한다")


def test_gmarket_rows_skip_non_gaegeolmu():
    coupang = _coupang([{"order_no": "C1", "name": "쿠팡손님"}])
    gmarket = _gmarket([
        {"order_no": "G1", "name": "게걸무손님"},
        {"order_no": "G2", "name": "무씨앗손님", "product": "식품애착 무씨앗기름"},  # 다른 상품
    ])
    _out, _fn, stats = G.process(coupang, gmarket)
    assert stats["gmarket"] == 1
    assert stats["total"] == 2
