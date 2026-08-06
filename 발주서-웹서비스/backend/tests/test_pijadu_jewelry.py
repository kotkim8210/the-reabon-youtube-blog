"""피자두(쥬얼리프룻) 발주·운송장 (2026-08-06 신규).

쥬얼리 adminplus 실측 옵션: '피자두 1kg/2kg/3kg'. '노지 자두'는 별도 상품이라 제외.
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import myeongi_order
from app.processors.myeongi_tracking import is_jewelryfruit_tracking_target

COUPANG_PRODUCT = "국내산 김천 피자두 특품 프리미엄 제철 자두"


def test_pijadu_routing_only_jewelry_weights():
    assert myeongi_order.is_jewelry_pijadu_order(COUPANG_PRODUCT, "1박스 1kg") is True
    assert myeongi_order.is_jewelry_pijadu_order(COUPANG_PRODUCT, "1박스 2kg") is True
    assert myeongi_order.is_jewelry_pijadu_order(COUPANG_PRODUCT, "1박스 3kg") is True
    # 취급 안 하는 중량
    assert myeongi_order.is_jewelry_pijadu_order(COUPANG_PRODUCT, "1박스 5kg") is False
    # 노지자두(다른 상품)·타 품목은 제외
    assert myeongi_order.is_jewelry_pijadu_order("노지 자두 대과", "1박스 3kg") is False
    assert myeongi_order.is_jewelry_pijadu_order("햇 백도 딱딱이복숭아", "1박스 1kg") is False


def test_pijadu_vendor_option_name():
    assert myeongi_order.convert_option("1박스 3kg", COUPANG_PRODUCT) == "피자두 3kg"
    assert myeongi_order.convert_option("1박스 2kg", COUPANG_PRODUCT) == "피자두 2kg"
    assert myeongi_order.convert_option("1박스 1kg", COUPANG_PRODUCT) == "피자두 1kg"


def test_pijadu_is_tracking_target():
    assert is_jewelryfruit_tracking_target(COUPANG_PRODUCT, "1박스 3kg") is True
    assert is_jewelryfruit_tracking_target("노지 자두 대과", "1박스 3kg") is False


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))       # C 주문번호
        ws.cell(row, 11, r["product"])                       # K 상품명
        ws.cell(row, 12, r["option"])                        # L 옵션
        ws.cell(row, 23, r.get("qty", 1))                    # W 수량
        ws.cell(row, 27, r.get("name", f"수취인{i}"))         # AA 수취인
        ws.cell(row, 28, r.get("phone", "0502-5289-0408"))   # AB 전화
        ws.cell(row, 29, "12345")                            # AC 우편번호
        ws.cell(row, 30, r.get("address", "경상남도 창원시"))  # AD 주소
        ws.cell(row, 31, r.get("memo", "문 앞"))              # AE 배송메모
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_process_outputs_pijadu_rows_and_filename():
    data = _delivery([
        {"product": COUPANG_PRODUCT, "option": "1박스 3kg", "name": "김은주"},
        {"product": COUPANG_PRODUCT, "option": "1박스 1kg", "name": "박테스트"},
        {"product": "노지 자두 대과", "option": "1박스 3kg"},   # 발주 대상 아님
    ])
    output, filename, stats = myeongi_order.process(data)

    assert "피자두" in filename
    assert stats["total"] == 2
    names = {o["vendor_option_name"] for o in stats["options"]}
    assert names == {"피자두 3kg", "피자두 1kg"}

    ws = load_workbook(BytesIO(output)).active
    rows = [
        (ws.cell(r, 3).value, ws.cell(r, 7).value, ws.cell(r, 8).value)
        for r in range(2, 4)
    ]
    assert rows == [("김은주", "피자두 3kg", "1"), ("박테스트", "피자두 1kg", "1")]
    # 거래처명·보내는분은 쥬얼리 양식 고정값
    assert ws.cell(2, 2).value == "아이티소프트"
    assert ws.cell(2, 9).value == "식품애착"
