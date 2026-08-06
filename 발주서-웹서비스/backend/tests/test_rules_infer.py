"""규칙 초안 자동생성(infer_rules_from_delivery) 테스트."""

from io import BytesIO

from openpyxl import Workbook

from app import rules_engine


def _delivery(rows: list[tuple[str, str]]) -> bytes:
    """(상품명 K열, 옵션 L열) 목록으로 DeliveryList 엑셀 생성."""
    wb = Workbook()
    ws = wb.active
    for col in range(1, 13):
        ws.cell(row=1, column=col, value=f"h{col}")
    for i, (product, option) in enumerate(rows):
        r = i + 2
        ws.cell(row=r, column=11, value=product)  # K
        ws.cell(row=r, column=12, value=option)   # L
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def setup_function(_fn):
    # 빈 규칙 캐시(로드됨) — 아무것도 매칭 안 되게
    rules_engine._set_cache_for_test({}, {})


def test_infer_grade_and_kg_template():
    data = _delivery([
        ("고당도 햇 백도 딱딱이복숭아 국내산", "2kg, 1박스, 중과"),
        ("고당도 햇 백도 딱딱이복숭아 국내산", "4kg, 1박스, 대과"),
        ("고당도 햇 백도 딱딱이복숭아 국내산", "1kg, 1박스, 중과"),
    ])
    res = rules_engine.infer_rules_from_delivery(data, "jewelryfruit")
    assert res["product_count"] == 1
    assert len(res["drafts"]) == 1
    d = res["drafts"][0]
    assert d["supplier_key"] == "jewelryfruit"
    assert d["name_keywords"] == ["딱딱이복숭아"]     # 비노이즈 토큰 중 가장 긴 것(지역명 회피)
    assert d["output_template"] == "딱딱이복숭아 {grade} {kg}kg"
    assert d["require_grade"] is True
    assert d["require_kg"] is True
    assert set(d["grades"]) == {"중과", "대과"}
    assert d["kg_allow"] == ["1", "2", "4"]         # kg 오름차순
    assert d["order_count"] == 3


def test_infer_count_unit_product():
    data = _delivery([
        ("초당옥수수 실속", "특품 10개입"),
        ("초당옥수수 실속", "중품 20개입"),
    ])
    res = rules_engine.infer_rules_from_delivery(data, "jewelryfruit")
    d = res["drafts"][0]
    assert d["name_keywords"] == ["초당옥수수"]
    assert "{count}개" in d["output_template"]
    assert "{kg}" not in d["output_template"]        # kg 없으면 수량 단위
    assert d["require_kg"] is False


def test_infer_kg_only_no_grade():
    data = _delivery([
        ("산지직송 명이나물", "500g 1박스"),  # g는 kg 아님 → kg 미검출
        ("산지직송 명이나물", "1kg 1박스"),
    ])
    res = rules_engine.infer_rules_from_delivery(data, "jewelryfruit")
    d = res["drafts"][0]
    assert d["name_keywords"] == ["명이나물"]
    # kg가 일부 행에만 있음 → 경고 + require_kg False
    assert d["require_kg"] is False
    assert any("kg" in w for w in d["warnings"])


def test_infer_marks_already_covered():
    # 기존 규칙: 콜라비
    rules_engine._set_cache_for_test(
        {
            "jejudapam": [{
                "id": 1, "supplier_key": "jejudapam", "label": "콜라비", "priority": 100,
                "name_keywords": ["콜라비"], "exclude_keywords": [],
                "grades": [], "kg_allow": ["3", "5"], "pair_map": {}, "extra_map": {},
                "output_template": "콜라비 정품 {kg}kg", "require_grade": 0, "require_kg": 1,
            }],
        },
        {"jejudapam": {"key": "jejudapam", "name": "제주다팜"}},
    )
    data = _delivery([
        ("콜라비 정품", "3kg 1박스"),          # 이미 규칙 있음
        ("방울토마토 산지직송", "2kg 1박스"),   # 신규
    ])
    res = rules_engine.infer_rules_from_delivery(data, "jejudapam")
    covered_labels = [c["label"] for c in res["covered"]]
    draft_keywords = [d["name_keywords"][0] for d in res["drafts"]]
    assert any("콜라비" in c for c in covered_labels)
    assert "방울토마토" in draft_keywords
    assert res["covered"][0]["already_matched"] is True


def test_infer_empty_rows_ignored():
    data = _delivery([("", ""), ("수박", "5kg 1통")])
    res = rules_engine.infer_rules_from_delivery(data, "jewelryfruit")
    assert res["product_count"] == 1
    assert res["drafts"][0]["name_keywords"] == ["수박"]


def test_infer_keyword_skips_region_and_marketing_tokens():
    """'국내산 김천 피자두…'→'김천', '제주 하우스 애플망고…'→'제주'로 잡히던 문제(2026-08-06)."""
    data = _delivery([
        ("국내산 김천 피자두 특품 프리미엄 제철 자두", "1박스 3kg"),
        ("제주 하우스 애플망고 고당도 산지직송", "1박스 특 2kg"),
    ])
    res = rules_engine.infer_rules_from_delivery(data, "jewelryfruit")
    by_kw = {d["name_keywords"][0]: d for d in res["drafts"]}
    assert "피자두" in by_kw and "애플망고" in by_kw
    assert by_kw["피자두"]["output_template"] == "피자두 {kg}kg"
    assert by_kw["애플망고"]["output_template"] == "애플망고 {grade} {kg}kg"
