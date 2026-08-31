"""라이브 이벤트 당첨자 발주 — 2026-08-17부터 경품 무관 전량 제주다팜 1파일.

- 제주다팜 취급 품목(백도·미니밤호박·청사과·홍감자)은 거래처 판매옵션명으로 변환
- 그 밖의 경품은 경품명 그대로 (종전엔 성주참외로 둔갑했다)
- 이름·주소가 빈 당첨자는 발주에서 빼되 needs_check로 알린다
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import event_order


def _csv(rows: list[list[str]]) -> bytes:
    header = "경품명,별명,개인 식별 정보 확인,이름,연락처,주소,주문 아이디,구매 금액,환불/취소 금액,환불/날짜/시간"
    lines = [header] + [",".join(r) for r in rows]
    return "\n".join(lines).encode("utf-8-sig")


def _row(prize: str, name: str, *, address: str = "서울시 강남구 테스트로 1", order_id: str = "OID1") -> list[str]:
    return [prize, "닉", "Y", name, "010-1111-2222", address, order_id, "10000", "", ""]


def test_prizes_follow_actual_supplier():
    """경품의 실제 발주처를 따른다 — 청사과(아오리)만 제이비티, 나머지는 제주다팜."""
    for prize in ("백도 딱딱이복숭아 대과 2kg", "미니밤호박1kg", "성주참외 3kg"):
        assert event_order.event_supplier(prize) == "jejudapam", prize
    # 청사과는 2026-08-27 제이비티 이관 → 당첨자 발주도 제이비티
    assert event_order.event_supplier("아오리사과  소과 (2kg)") == "jbt"


def test_product_name_uses_jejudapam_option_names():
    # 미니밤호박·청사과는 제주다팜 판매옵션명으로 변환
    assert event_order.event_product_name("미니밤호박1kg") == "제주 미니밤호박 보우짱 로얄과 1kg"
    # 청사과는 제이비티 판매옵션명으로 변환된다(발주처 이관 반영)
    assert (
        event_order.event_product_name("아오리사과  소과 (2kg)")
        == "청사과가정용A급(소과)2kg(11-15과)"
    )
    # 백도는 1kg 경품도 최소 규격 2kg로 승급
    assert event_order.event_product_name(" 백도복숭아  중과 (1kg)") == "딱딱이 복숭아 중과 2kg"
    assert event_order.event_product_name("백도 딱딱이복숭아 대과 4kg") == "딱딱이 복숭아 대과 4kg"


def test_unknown_prize_keeps_original_option():
    """미지원 경품은 경품명 그대로 — 성주참외로 둔갑하면 안 된다(2026-08-17 사고)."""
    assert event_order.event_product_name("한라봉 3kg 특품") == "한라봉 3kg 특품"
    assert "참외" not in event_order.event_product_name("미니밤호박1kg")


def test_process_outputs_split_by_supplier():
    """제주다팜 경품 + 제이비티(청사과) 경품이 섞이면 발주서 2개."""
    results = event_order.process(_csv([
        _row("미니밤호박1kg", "순유선", order_id="12102324980320"),
        _row("아오리사과  소과 (2kg)", "심규정", order_id="5102305763420"),
    ]))
    assert len(results) == 2
    by_supplier = {st["supplier"]: (b, f, st) for b, f, st in results}
    assert set(by_supplier) == {"제주다팜", "제이비티"}

    jeju_bytes = by_supplier["제주다팜"][0]
    ws = load_workbook(BytesIO(jeju_bytes)).active
    assert (ws.cell(2, 2).value, ws.cell(2, 8).value) == ("순유선", "제주 미니밤호박 보우짱 로얄과 1kg")

    jbt_bytes = by_supplier["제이비티"][0]
    ws2 = load_workbook(BytesIO(jbt_bytes)).active
    assert (ws2.cell(2, 2).value, ws2.cell(2, 5).value) == ("심규정", "청사과가정용A급(소과)2kg(11-15과)")


def test_missing_address_is_reported_not_silently_dropped():
    """주소·이름이 빈 당첨자(취소 의심)는 발주에서 빼되 화면에 표시한다."""
    results = event_order.process(_csv([
        _row("아오리사과  소과 (2kg)", "심규정", order_id="5102305763420"),
        ["아오리사과  소과 (2kg)", "이*미", "", "", "", "", "24102306274114", "89990", "0", ""],
    ]))
    _output, _filename, stats = results[0]
    assert stats["winners"] == 1
    assert "needs_check" in stats
    assert "24102306274114" in stats["needs_check"]
    assert "이름·주소 없음" in stats["needs_check"]


def test_all_rows_incomplete_raises_with_detail():
    try:
        event_order.process(_csv([
            ["미니밤호박1kg", "이*미", "", "", "", "", "24102306274114", "89990", "0", ""],
        ]))
    except ValueError as exc:
        assert "개인정보 미기재" in str(exc) and "24102306274114" in str(exc)
    else:
        raise AssertionError("발주 가능한 당첨자가 없으면 사유를 알려야 한다")


def test_delivery_slot_rejects_winners_csv_with_guidance():
    """발주 칸에 당첨자 CSV를 올렸을 때 openpyxl 원본오류 대신 안내 (2026-07-27 사고)."""
    import pytest
    from fastapi import HTTPException

    from app.main import _require_xlsx

    csv_bytes = _csv([_row("백도 딱딱이복숭아 대과 2kg", "당첨자")])
    with pytest.raises(HTTPException) as exc:
        _require_xlsx(csv_bytes)
    assert exc.value.status_code == 400
    assert "이벤트 당첨자" in exc.value.detail
    assert "zip" not in exc.value.detail

    buf = BytesIO()
    Workbook().save(buf)
    assert _require_xlsx(buf.getvalue()) is None
