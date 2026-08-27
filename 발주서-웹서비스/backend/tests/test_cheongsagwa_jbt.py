"""청사과(아오리) 제주다팜 → 제이비티 발주·운송장 이관 (2026-08-27)."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import kolrabi_order, tomato_order
from app.processors.tomato_tracking import is_jbt_tracking_target, product_type_for

COUPANG = "[핫딜] 새콤아삭 고당도선별 아오리사과 여름 청사과 풋사과"


def test_grade_mapping_to_jbt_options():
    f = tomato_order.jbt_apple_option
    assert f(COUPANG, "1박스 가정용 소과 2kg") == "청사과가정용A급(소과)2kg(11-15과)"
    assert f(COUPANG, "1박스 가정용 중소과 3kg") == "청사과가정용A급(중소과)3kg(14-16과)"
    # 제이비티엔 '대과'가 없어 쿠팡 대과 → 중과로 발주
    assert f(COUPANG, "1박스 가정용 대과 2kg") == "청사과가정용A급(중과)2kg(7-8과)"
    assert f(COUPANG, "1박스 가정용 대과 5kg") == "청사과가정용A급(중과)5kg(18-22과)"
    # 중소과가 소과로 잘못 잡히지 않는다
    assert "(중소과)" in f(COUPANG, "1박스 가정용 중소과 2kg")


def test_hongro_stays_with_jejudapam():
    """홍로사과(가을햇사과)는 제주다팜 유지 — 청사과 로직에 걸리면 안 된다."""
    hongro = "2026 햇 안동 고당도 홍로사과 아삭한 꿀사과 산지직송 제철 햇사과"
    assert tomato_order.is_jbt_apple_order(hongro, "1박스 소과 1.5kg(7-10과내)") is False
    assert tomato_order.jbt_apple_option(hongro, "1박스 소과 1.5kg(7-10과내)") is None
    assert kolrabi_order.convert_hongro_option(hongro, "1박스 소과 1.5kg(7-10과내)") is not None


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))
        ws.cell(row, 11, r["product"])
        ws.cell(row, 12, r["option"])
        ws.cell(row, 23, r.get("qty", 1))
        ws.cell(row, 27, r.get("name", f"수취인{i}"))
        ws.cell(row, 28, "010-8244-8709")
        ws.cell(row, 29, "12345")
        ws.cell(row, 30, "부산 사상구 사상로 342번길 21")
        ws.cell(row, 31, "문 앞")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_apple_goes_to_jbt_order_sheet():
    data = _delivery([
        {"product": COUPANG, "option": "1박스 가정용 소과 2kg", "name": "홍지영"},
        {"product": "콜라비 정품", "option": "3kg", "name": "제외대상"},
    ])
    results = tomato_order.process_outputs(data)
    assert results, "제이비티 발주서가 나와야 한다"
    _bytes, filename, stats = results[0]
    assert "청사과" in filename
    ws = load_workbook(BytesIO(_bytes)).active
    assert ws.cell(2, 2).value == "홍지영"
    assert ws.cell(2, 5).value == "청사과가정용A급(소과)2kg(11-15과)"
    assert ws.cell(2, 8).value == tomato_order.JBT_SENDER_NAME


def test_apple_removed_from_jejudapam_order():
    data = _delivery([{"product": COUPANG, "option": "1박스 가정용 소과 2kg", "name": "홍지영"}])
    labels = {st.get("product") for _b, _f, st in kolrabi_order.process_outputs(data)}
    assert "청사과(제주다팜)" not in labels


def test_quantity_split_follows_jbt_rule():
    """청사과는 합배송 불가 — 수량 2면 각 수량1인 2행으로 나온다."""
    data = _delivery([{"product": COUPANG, "option": "1박스 가정용 대과 2kg", "name": "김철수", "qty": 2}])
    _bytes, _filename, _stats = tomato_order.process_outputs(data)[0]
    ws = load_workbook(BytesIO(_bytes)).active
    rows = [(ws.cell(r, 2).value, ws.cell(r, 6).value) for r in range(2, ws.max_row + 1)]
    assert rows == [("김철수", "1"), ("김철수", "1")]


def test_tracking_target_moved_to_jbt():
    assert product_type_for(COUPANG) == "청사과"
    assert is_jbt_tracking_target(COUPANG, "1박스 가정용 소과 2kg") is True
    # 취급하지 않는 조합은 대상 아님
    assert is_jbt_tracking_target(COUPANG, "1박스 가정용 소과 99kg") is False


def test_order_routing_no_longer_jejudapam():
    from app.order_routing import resolve_order_supplier
    assert resolve_order_supplier(COUPANG, "1박스 가정용 소과 2kg") is None
