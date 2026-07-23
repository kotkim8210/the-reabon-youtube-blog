"""마진계산기(소싱현황 정식 양식) 테스트."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import margin_calc


def test_compute_margin_matches_official_formula():
    # 1kg 예시(마진방어 5%): H=10800, I=5900
    c = margin_calc.compute_margin(10800, 5900, option="1kg정품")
    assert c["G"] == 11300           # H+500
    assert round(c["N"]) == 540      # H*5%
    # P = G*0.1166 + H*0.0363
    assert round(c["P"], 2) == round(11300 * 0.1166 + 10800 * 0.0363, 2)
    assert round(c["S"], 2) == round(c["R"] * 0.16, 2)  # 소득세 16%
    assert round(c["T"], 2) == 113.0                    # G*1%
    # U = H-I-L-M-N-P-Q-S-T (G 미포함)
    expected_u = 10800 - 5900 - c["N"] - c["P"] - c["Q"] - c["S"] - c["T"]
    assert round(c["U"], 2) == round(expected_u, 2)


def test_margin_defense_rate_by_kg():
    assert margin_calc._margin_defense_rate("1kg정품") == 0.05
    assert margin_calc._margin_defense_rate("2kg") == 0.05
    assert margin_calc._margin_defense_rate("3kg정품") == 0.03
    assert margin_calc._margin_defense_rate("5kg") == 0.03
    assert margin_calc._margin_defense_rate("옵션없음") == 0.05  # kg 불명 → 보수적 5%


def test_high_cost_ratio_goes_negative():
    # 원가율 높은 품목은 마이너스 마진으로 잡혀야 정상(양식 버그 아님)
    c = margin_calc.compute_margin(10000, 9000, option="3kg")
    assert c["U"] < 0


def test_build_sourcing_sheet_has_formulas_and_stats():
    rows = [
        {"product": "제주다팜", "option": "1kg정품", "qty": 1, "sale_price": 10800, "supply_price": 5900},
        {"product": "제주다팜", "option": "3kg정품", "qty": 1, "sale_price": 10000, "supply_price": 9500},
    ]
    data, filename, stats = margin_calc.build_sourcing_sheet(rows, title="테스트")
    assert filename.endswith(".xlsx")
    assert stats["총_상품수"] == 2
    assert stats["마진_마이너스"] >= 1

    ws = load_workbook(filename=BytesIO(data)).active
    # 정식 양식 수식 확인 (축소양식 금지)
    assert ws["G8"].value == "=H8+500"
    assert ws["N8"].value == "=H8*5%"
    assert ws["N9"].value == "=H9*3%"          # 3kg → 3%
    assert ws["P8"].value == "=(G8*K8)+(H8*3.63%)"
    assert ws["S8"].value == "=R8*16%"
    assert ws["U8"].value == "=H8-I8-L8-M8-N8-P8-Q8-S8-T8"
    assert ws["H8"].value == 10800
    assert ws["I8"].value == 5900


def _customer_file(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    for i, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(i, c, v)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_input_file_detects_columns():
    data = _customer_file(
        ["상품명", "옵션", "판매가", "공급가"],
        [["감귤", "3kg", 19900, 12000], ["감귤", "5kg", 26900, 18000]],
    )
    rows = margin_calc.parse_input_file(data)
    assert len(rows) == 2
    assert rows[0]["sale_price"] == 19900
    assert rows[0]["supply_price"] == 12000


def test_parse_input_file_alias_and_missing():
    # 별칭(쿠폰가/원가) 인식
    ok = _customer_file(["품명", "규격", "쿠폰가", "원가"], [["사과", "2kg", 15000, 8000]])
    assert margin_calc.parse_input_file(ok)[0]["sale_price"] == 15000

    # 판매가/공급가 없으면 에러
    bad = _customer_file(["이름", "메모"], [["x", "y"]])
    try:
        margin_calc.parse_input_file(bad)
        assert False, "should raise"
    except ValueError:
        pass


def test_blank_template():
    data = margin_calc.blank_template_bytes()
    ws = load_workbook(filename=BytesIO(data)).active
    assert ws["B2"].value == "소싱현황 관리"
    assert ws["U7"].value == "마진"
