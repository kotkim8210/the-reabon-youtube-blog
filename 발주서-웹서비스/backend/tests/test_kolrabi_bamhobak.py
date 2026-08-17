# -*- coding: utf-8 -*-
"""회귀 테스트: 미니밤호박(보우짱) 발주처.

2026-06: 1kg=제주다팜, 3·5·10kg=쥬얼리프룻으로 분리했다가
2026-08-17: **전 옵션 제주다팜(kolrabi) 발주로 통합** — 쥬얼리는 발주·송장 대상 아님.

PYTHONPATH=backend python tests/test_kolrabi_bamhobak.py
"""
from app.processors import kolrabi_order as K
from app.processors import myeongi_order as M
from app.processors import tracking_input as TI


def test_jeju_bamhobak_all_weights():
    # 제주다팜: 전 옵션 발주 대상 (adminplus 실측 1·2·3·4·5·8·10kg)
    assert (
        K.convert_bamhobak_option("[산지직송] 제주 미니 밤호박 보우짱 단호박", "1박스 로얄 정품 1kg")
        == "제주 미니밤호박 보우짱 로얄과 1kg"
    )
    for kg in ("2", "3", "4", "5", "8", "10"):
        assert (
            K.convert_bamhobak_option("제주 미니 밤호박 보우짱", f"1박스 로얄 정품 {kg}kg")
            == f"제주 미니밤호박 보우짱 로얄과 {kg}kg"
        ), kg
    # 취급하지 않는 중량은 제외
    assert K.convert_bamhobak_option("제주 미니 밤호박 보우짱", "1박스 로얄 정품 7kg") is None


def test_jewelry_bamhobak_no_longer_ordered():
    # 쥬얼리프룻: 2026-08-17 이관으로 전 옵션 발주 대상 아님
    for kg in ("1", "3", "5", "10"):
        assert M._jewelry_bamhobak_option("제주 미니 밤호박 보우짱", f"1박스 로얄 정품 {kg}kg") is None, kg
        assert M.is_jewelry_bamhobak_order("제주 미니 밤호박 보우짱", f"1박스 로얄 정품 {kg}kg") is False, kg


def test_jewelry_convert_option_skips_bamhobak():
    # convert_option(쥬얼리 발주·운송장 공용)에서도 밤호박은 쥬얼리 품목으로 변환되지 않는다
    assert M.convert_option("1박스 로얄 정품 3kg", "제주 미니 밤호박 보우짱") != "미니 밤호박 보우짱 로얄과 3kg"


def test_jewelry_tracking_target_excludes_bamhobak():
    from app.processors.myeongi_tracking import is_jewelryfruit_tracking_target
    assert is_jewelryfruit_tracking_target("제주 미니 밤호박 보우짱", "1박스 로얄 정품 5kg") is False


def test_jeju_potato_large_ships_as_special():
    # 홍감자는 2026-07-13 쥬얼리 품절→제주다팜 이관: 대→특, 중1kg→중2kg
    assert K.convert_potato_option("햇 홍감자", "3kg(대)") == "홍감자 특 3kg"
    assert K.convert_potato_option("햇 홍감자", "5kg(대)") == "홍감자 특 5kg"
    assert K.convert_potato_option("햇 홍감자", "1kg(중)") == "홍감자 중 2kg"


def test_bamhobak_excludes_other_products():
    assert M.is_jewelry_bamhobak_order("콜라비", "3kg 1박스") is False
    assert K.convert_bamhobak_option("콜라비", "3kg 1박스") is None
    assert K.convert_bamhobak_option("초당옥수수", "중품 10개") is None


def test_bamhobak_tracking_semantic_key():
    keys = TI._semantic_option_keys("제주 미니 밤호박 보우짱 단호박", "1박스 로얄 정품 3kg")
    assert "bamhobak:3kg" in keys, keys


def test_admin_rule_cannot_reintroduce_bamhobak_to_jewelry():
    """규칙 엔진에 옛 쥬얼리 밤호박 규칙이 남아 있어도 쥬얼리 발주서에 나오면 안 된다.

    (2026-08-17: 하드코딩을 껐더니 DB 규칙 'jewelryfruit/미니밤호박 3·5·10kg'이
     대신 잡아 쥬얼리 발주서가 27건 나온 실사고)
    """
    from io import BytesIO
    from openpyxl import Workbook
    from app import rules_engine
    from app.processors.myeongi_tracking import is_jewelryfruit_tracking_target

    rules_engine._set_cache_for_test(
        {"jewelryfruit": [{
            "id": 4, "supplier_key": "jewelryfruit", "label": "미니밤호박 3·5·10kg", "priority": 100,
            "name_keywords": ["밤호박"], "exclude_keywords": [], "grades": [],
            "kg_allow": ["3", "5", "10"], "pair_map": {}, "extra_map": {},
            "output_template": "미니 밤호박 보우짱 로얄과 {kg}kg",
            "require_grade": False, "require_kg": True, "active": True,
        }]},
        {"jewelryfruit": {"key": "jewelryfruit", "name": "쥬얼리프룻"}},
    )
    try:
        P, O = "제주 미니 밤호박 보우짱", "1박스 로얄 정품 5kg"
        assert M.jewelry_rule_match(P, O) is None or M.convert_option(O, P) != "미니 밤호박 보우짱 로얄과 5kg"
        assert is_jewelryfruit_tracking_target(P, O) is False

        wb = Workbook(); ws = wb.active
        ws.cell(2, 3, "OID1"); ws.cell(2, 11, P); ws.cell(2, 12, O); ws.cell(2, 23, 1)
        ws.cell(2, 27, "밤호박씨"); ws.cell(2, 28, "010-0000-0000"); ws.cell(2, 29, "12345")
        ws.cell(2, 30, "서울시"); ws.cell(2, 31, "")
        buf = BytesIO(); wb.save(buf)
        _out, filename, stats = M.process(buf.getvalue())
        assert stats["total"] == 0, (filename, stats)
        assert "밤호박" not in filename
    finally:
        rules_engine._set_cache_for_test({}, {})
