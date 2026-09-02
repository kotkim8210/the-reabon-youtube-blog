"""홍로사과(가을햇사과) 제주다팜 발주·운송장·마진방어 (2026-08-27 신규)."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import kolrabi_order
from app.processors.tracking_input import _semantic_option_keys

COUPANG_PRODUCT = "2026 햇 안동 고당도 홍로사과 아삭한 꿀사과 산지직송 제철 햇사과"


def test_all_20_options_map_to_jeju_names():
    for grade in ("소과", "중소과", "중대과", "대과"):
        for kg in ("1.5", "2", "3", "4", "5"):
            got = kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, f"1박스 {grade} {kg}kg")
            assert got, (grade, kg)
            assert grade in got and f"{kg}kg" in got


def test_real_option_text_from_coupang():
    assert (
        kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 소과 1.5kg(7-10과내)")
        == "가을햇사과(홍사과) 가정용 소과 포장재포함 1.5kg(7-10과내)"
    )
    # 중소과/중대과가 소과/대과로 잘못 잡히지 않는다
    assert "중소과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중소과 2kg")
    assert "중대과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중대과 3kg")


def test_hongro_and_cheongsagwa_are_exclusive():
    """청사과(아오리)와 홍로가 서로 섞이면 안 된다."""
    assert kolrabi_order.convert_apple_option(COUPANG_PRODUCT, "1박스 소과 2kg") is None
    green = "[핫딜] 새콤아삭 고당도선별 아오리사과 여름 청사과 풋사과"
    assert kolrabi_order.convert_hongro_option(green, "1박스 가정용 대과 4kg") is None
    assert kolrabi_order.convert_apple_option(green, "1박스 가정용 대과 4kg") is not None
    # 취급 안 하는 중량은 제외
    assert kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 소과 10kg") is None


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))
        ws.cell(row, 11, r["product"])
        ws.cell(row, 12, r["option"])
        ws.cell(row, 23, r.get("qty", 1))
        ws.cell(row, 27, r.get("name", f"수취인{i}"))
        ws.cell(row, 28, "010-0000-0000")
        ws.cell(row, 29, "12345")
        ws.cell(row, 30, "서울시 어딘가")
        ws.cell(row, 31, "")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_process_hongro_and_process_outputs():
    data = _delivery([
        {"product": COUPANG_PRODUCT, "option": "1박스 소과 1.5kg(7-10과내)", "name": "Kevin"},
        {"product": COUPANG_PRODUCT, "option": "1박스 대과 5kg", "name": "김광은"},
        {"product": "콜라비 정품", "option": "3kg"},
    ])
    result = kolrabi_order.process_hongro(data)
    assert result is not None
    _bytes, filename, stats = result
    assert "홍로사과" in filename
    assert stats["total"] == 2
    ws = load_workbook(BytesIO(_bytes)).active
    assert str(ws.cell(2, 8).value).startswith("가을햇사과(홍사과)")

    labels = {st.get("product") for _b, _f, st in kolrabi_order.process_outputs(data)}
    assert "홍로사과(제주다팜)" in labels


def test_tracking_semantic_keys_match_both_sides():
    ol = _semantic_option_keys("가을햇사과(홍사과) 가정용 소과 포장재포함 1.5kg(7-10과내)", "")
    dl = _semantic_option_keys(COUPANG_PRODUCT, "1박스 소과 1.5kg(7-10과내)")
    assert "hongro:소과:1.5kg" in ol
    assert "hongro:소과:1.5kg" in dl
    # 등급 혼동 방지
    js = _semantic_option_keys("가을햇사과(홍사과) 가정용 중소과 포장재포함 2kg(7-8과내외)", "")
    assert "hongro:중소과:2kg" in js and "hongro:소과:2kg" not in js


def test_margin_monitor_matches_order_mapping():
    from app import supplier_price_monitor as spm

    config = spm.MONITOR_CONFIGS["hongro-jeju"]
    assert config.product_code == "10001216"
    assert config.template_path.exists()
    assert [o.row for o in config.options] == list(range(8, 28))
    for o in config.options:
        assert spm.option_supplier_name(o, config) == "제주다팜"
        converted = kolrabi_order.convert_hongro_option(o.coupang_product, o.coupang_option)
        assert converted == o.supplier_option_name, (o.label, converted, o.supplier_option_name)


def test_margin_template_rows_align():
    from app import supplier_price_monitor as spm

    config = spm.MONITOR_CONFIGS["hongro-jeju"]
    ws = load_workbook(config.template_path)["쥬얼리프룻"]
    assert [str(ws.cell(r, 3).value) for r in range(8, 28)] == ["제주다팜"] * 20
    # 공급가(I열)는 실측값이 채워져 있어야 첫 실행부터 비교가 된다
    assert all(isinstance(ws.cell(r, 9).value, (int, float)) for r in range(8, 28))
    assert ws.cell(8, 9).value == 6300 and ws.cell(27, 9).value == 17500


def test_unmatched_hongro_option_is_reported():
    """거래처 판매옵션에 없는 등급·kg 주문이 조용히 사라지면 안 된다.

    2026-09 라이브 대비: 거래처에 없는 중량·등급 옵션을 쿠팡에 열면 발주명 변환이
    실패하는데, 종전엔 그 행이 아무 표시 없이 발주서에서 빠졌다.
    """
    payload = _delivery([
        {"product": COUPANG_PRODUCT, "option": "1박스 소과 3kg(17-20과내외)", "name": "정상건"},
        {"product": COUPANG_PRODUCT, "option": "1박스 소과 10kg", "name": "미등록옵션"},
    ])
    _out, _fn, stats = kolrabi_order.process_hongro(payload)
    assert stats["total"] == 1
    needs = stats.get("needs_check") or []
    assert len(needs) == 1 and "미등록옵션" in needs[0] and "10kg" in needs[0]


def test_unmatched_only_still_surfaces_in_outputs():
    """미매칭만 있고 발주 건이 0이어도 결과에 실려 사용자에게 보여야 한다."""
    payload = _delivery([
        {"product": COUPANG_PRODUCT, "option": "1박스 소과 10kg", "name": "미등록옵션"},
    ])
    results = kolrabi_order.process_outputs(payload)
    hongro = [r for r in results if "홍로" in r[1]]
    assert len(hongro) == 1
    assert hongro[0][2]["total"] == 0
    assert hongro[0][2]["needs_check"]


def test_matched_orders_have_no_needs_check():
    payload = _delivery([
        {"product": COUPANG_PRODUCT, "option": "1박스 대과 3kg(10과내외)", "name": "정상건"},
    ])
    _out, _fn, stats = kolrabi_order.process_hongro(payload)
    assert stats["total"] == 1
    assert "needs_check" not in stats


def test_coupang_junggwa_maps_to_jeju_jungdaegwa():
    """쿠팡 '중과' = 제주다팜 '중대과'(2026-09-02 사용자 확인). 과수 표기도 11-12과내외로 동일."""
    assert (
        kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "중과 3kg(11-12과내외) 1박스")
        == "가을햇사과(홍사과) 가정용 중대과 포장재포함 3kg(11-12과내외)"
    )
    # 다른 중량도 같은 규칙
    assert "중대과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중과 2kg")
    assert "중대과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중과 5kg")
    # 중소과·중대과 표기는 그대로 유지(중과 별칭에 잡아먹히면 안 됨)
    assert "중소과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중소과 3kg")
    assert "중대과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 중대과 3kg")
    # 소과·대과가 중과 별칭에 오염되지 않는다
    assert "가정용 소과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 소과 3kg")
    assert "가정용 대과" in kolrabi_order.convert_hongro_option(COUPANG_PRODUCT, "1박스 대과 3kg")


def test_junggwa_tracking_keys_match_between_coupang_and_jeju():
    coupang = _semantic_option_keys(COUPANG_PRODUCT, "중과 3kg(11-12과내외) 1박스")
    jeju = _semantic_option_keys("가을햇사과(홍사과) 가정용 중대과 포장재포함 3kg(11-12과내외)")
    assert "hongro:중대과:3kg" in coupang
    assert coupang & jeju


def test_junggwa_no_longer_reported_as_unmatched():
    payload = _delivery([
        {"product": COUPANG_PRODUCT, "option": "중과 3kg(11-12과내외) 1박스", "name": "중과주문"},
    ])
    _out, _fn, stats = kolrabi_order.process_hongro(payload)
    assert stats["total"] == 1
    assert "needs_check" not in stats
