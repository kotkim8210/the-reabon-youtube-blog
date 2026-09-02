"""비셀러 LA한입갈비 발주 (2026-09 신규)."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.processors import biseller_order

COUPANG_PRODUCT = "한입 LA갈비 양념갈비 구매 1회 특급쉐프소스 양념소갈비"


def _delivery(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = i + 2
        ws.cell(row, 3, r.get("order_no", f"OID{i}"))
        ws.cell(row, 11, r.get("product", COUPANG_PRODUCT))
        ws.cell(row, 12, r["option"])
        ws.cell(row, 23, r.get("qty", 1))
        ws.cell(row, 27, r.get("name", f"수취인{i}"))
        ws.cell(row, 28, "010-1234-5678")
        ws.cell(row, 29, r.get("zipcode", "12345"))
        ws.cell(row, 30, "서울시 어딘가 1-2")
        ws.cell(row, 31, r.get("memo", "문 앞"))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_option_to_biseller_product_name():
    """사용자 샘플 발주서의 실측 표기와 정확히 일치해야 한다."""
    assert (
        biseller_order.convert_galbi_option(COUPANG_PRODUCT, "800g 4개")
        == "양념LA한입갈비 800g+800g+800g+800g (800G*4세트)"
    )
    assert (
        biseller_order.convert_galbi_option(COUPANG_PRODUCT, "800g 2개")
        == "양념LA한입갈비 800g+800g (800G*2세트)"
    )


def test_non_galbi_orders_are_ignored():
    assert biseller_order.convert_galbi_option("콜라비(정품) 3kg", "1박스 3kg") is None
    assert biseller_order.convert_galbi_option("게걸무씨앗기름", "2개 180ml") is None
    assert not biseller_order.is_biseller_galbi_order("2026 햇 안동 홍로사과", "1박스 소과 3kg")


def test_process_fills_template():
    payload = _delivery([
        {"option": "800g 4개", "name": "김철수", "qty": 1, "zipcode": "06236"},
        {"option": "800g 2개", "name": "이영희", "qty": 2, "memo": "부재시 경비실"},
    ])
    out, filename, stats = biseller_order.process(payload)
    assert stats["total"] == 2
    assert filename.startswith("아이티소프트_비셀러발주서_")
    ws = load_workbook(BytesIO(out)).active

    assert ws.cell(1, 7).value == "상품명 (비셀러 상품명)"
    assert ws.cell(2, 1).value == 1 and ws.cell(3, 1).value == 2
    assert ws.cell(2, 3).value == "김철수"
    assert ws.cell(2, 5).value == "06236"
    assert ws.cell(2, 7).value == "양념LA한입갈비 800g+800g+800g+800g (800G*4세트)"
    assert ws.cell(2, 8).value == 1
    assert ws.cell(2, 10).value == "(주)아이티소프트"
    assert ws.cell(2, 11).value == "010-5700-7756"
    assert ws.cell(3, 7).value == "양념LA한입갈비 800g+800g (800G*2세트)"
    assert ws.cell(3, 8).value == 2
    assert ws.cell(3, 9).value == "부재시 경비실"
    # 택배사·송장번호는 거래처가 채우는 칸 → 비어 있어야 한다
    assert ws.cell(2, 12).value in (None, "")
    assert ws.cell(2, 13).value in (None, "")


def test_sum_formula_follows_row_count():
    payload = _delivery([{"option": "800g 4개"} for _ in range(40)])
    out, _fn, stats = biseller_order.process(payload)
    assert stats["total"] == 40
    ws = load_workbook(BytesIO(out)).active
    total_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, 7).value == "합계")
    assert total_row == 42  # 2행부터 40건 + 합계
    assert ws.cell(total_row, 8).value == "=SUM(H2:H41)"
    assert ws.cell(41, 7).value == "양념LA한입갈비 800g+800g+800g+800g (800G*4세트)"


def test_unreadable_quantity_is_reported():
    """개수 표기를 못 읽는 갈비 주문이 조용히 사라지면 안 된다."""
    payload = _delivery([
        {"option": "800g 4개", "name": "정상건"},
        {"option": "특대사이즈", "name": "표기이상"},
    ])
    _out, _fn, stats = biseller_order.process(payload)
    assert stats["total"] == 1
    assert len(stats["needs_check"]) == 1 and "표기이상" in stats["needs_check"][0]
