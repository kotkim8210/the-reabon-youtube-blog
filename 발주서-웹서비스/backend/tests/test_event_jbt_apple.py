"""이벤트 당첨자: 청사과(아오리)는 제이비티 양식으로 발주 (2026-08-31).

청사과가 제이비티로 이관됐으므로 당첨자 발주도 제이비티 양식이어야 한다.
다른 경품은 종전대로 제주다팜, 섞이면 파일 2개.
"""

from io import BytesIO

from openpyxl import load_workbook

from app.processors import event_order


def _csv(rows: list[list[str]]) -> bytes:
    header = "경품명,별명,개인 식별 정보 확인,이름,연락처,주소,주문 아이디,구매 금액,환불/취소 금액,환불/날짜/시간"
    return "\n".join([header] + [",".join(r) for r in rows]).encode("utf-8-sig")


def _row(prize: str, name: str, order_id: str = "3102630001234") -> list[str]:
    return [prize, "닉", "Y", name, "010-2221-9527", "서울특별시 송파구 문정동 136", order_id, "18800", "0", ""]


def test_apple_winner_routed_to_jbt():
    prize = "아오리사과  소과 (2kg)"
    assert event_order.event_supplier(prize) == "jbt"
    assert event_order.event_product_name(prize) == "청사과가정용A급(소과)2kg(11-15과)"
    # 쿠팡 대과 → 제이비티 중과 매핑도 이벤트 경로에서 동일
    assert event_order.event_product_name("아오리사과 대과 (2kg)") == "청사과가정용A급(중과)2kg(7-8과)"


def test_other_prizes_stay_jejudapam():
    assert event_order.event_supplier("미니밤호박1kg") == "jejudapam"
    assert event_order.event_supplier(" 백도복숭아  중과 (1kg)") == "jejudapam"
    # 홍로사과는 제주다팜 유지
    assert event_order.event_supplier("홍로사과 소과 1.5kg") == "jejudapam"


def test_process_outputs_jbt_workbook():
    results = event_order.process(_csv([_row("아오리사과  소과 (2kg)", "오윤영")]))
    assert len(results) == 1
    output, filename, stats = results[0]
    assert stats["supplier"] == "제이비티"
    assert "제이비티" in filename and "청사과" in filename
    ws = load_workbook(BytesIO(output)).active
    # 제이비티 발주서 양식(10열)
    assert ws.cell(1, 1).value == "송장번호" and ws.cell(1, 5).value == "품목명"
    assert ws.cell(2, 2).value == "오윤영"
    assert ws.cell(2, 5).value == "청사과가정용A급(소과)2kg(11-15과)"
    assert ws.cell(2, 8).value == "(주)아이티소프트"


def test_mixed_prizes_produce_two_files():
    results = event_order.process(_csv([
        _row("아오리사과  소과 (2kg)", "오윤영"),
        _row("미니밤호박1kg", "순유선", order_id="12102324980320"),
    ]))
    by_supplier = {st["supplier"]: (b, f, st) for b, f, st in results}
    assert set(by_supplier) == {"제이비티", "제주다팜"}
    assert by_supplier["제이비티"][2]["winners"] == 1
    assert by_supplier["제주다팜"][2]["winners"] == 1


def test_exponent_order_id_is_blanked():
    """엑셀이 망가뜨린 주문번호('3.10263E+12')는 복원 불가 → 비워서 오발주 방지."""
    results = event_order.process(_csv([_row("아오리사과  소과 (2kg)", "오윤영", order_id="3.10263E+12")]))
    output, _filename, _stats = results[0]
    ws = load_workbook(BytesIO(output)).active
    assert ws.cell(2, 10).value in (None, "")
    # 정상 주문번호는 그대로 유지
    ok = event_order.process(_csv([_row("아오리사과  소과 (2kg)", "오윤영", order_id="3102630001234")]))
    ws2 = load_workbook(BytesIO(ok[0][0])).active
    assert ws2.cell(2, 10).value == "3102630001234"
